# https://www.codeproject.com/Articles/1227268/Accessing-Financial-Reports-in-the-EDGAR-Database
# https://www.codeproject.com/Articles/1227765/Parsing-XBRL-with-Python

# Import the main functionality from the SimFin Python API.
import simfin as sf
from simfin.names import *
#

from django.apps import apps

from bs4 import BeautifulSoup
import re
import os
import shutil
from django.conf import settings
import pandas as pd
import numpy as np
from pandas_datareader import data

import string
import datetime
import time
from datetime import timedelta
from pandas.tseries.offsets import BDay
from dateutil.relativedelta import relativedelta
from django.db.models import Q

from six.moves import urllib
import xlrd
from openpyxl import Workbook, load_workbook
import statistics
from django.utils.translation import get_language
import math
from django.utils.dateparse import parse_date
import json
import asyncio
import requests
import aiohttp
import random
from tda import auth, client, orders
from tda.orders.common import Duration, Session

from channels.db import database_sync_to_async
from channels.layers import get_channel_layer

import yfinance as yf
from yahoofinancials import YahooFinancials
from ..core.utils import log_debug, clear_log_debug
from ..core.sql import SQL
from ..core.models import Debug
from ..core.OptionsAmeriTrade import BaseTDAmeriTrade

from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import SGDClassifier, LogisticRegression
from sklearn.model_selection import cross_val_score, cross_val_predict

from .models import (XBRLMainIndustryInfo, XBRLIndustryInfo, XBRLCompanyInfoInProcess,
                     XBRLCompanyInfo, XBRLValuationStatementsAccounts, XBRLValuationAccounts,
                     XBRLValuationAccountsMatch,
                     XBRLCountry, XBRLCountryYearData,
                     XBRLHistoricalReturnsSP, XBRLSPMoodys, Project,
                     XBRLRegion, XBRLRegionYearData, XBRLSPEarningForecast, XBRLSPStatistics,
                     XBRLDimTime, XBRLDimCompany, XBRLDimAccount, XBRLFactCompany,
                     XBRLRealEquityPrices, XBRLRealEquityPricesArchive,
                     ETFS, ETFWatchLists, StockPricesMinutes, StockPricesDays)

from ..acapps.ml.basic_ml_objects import BaseDataProcessing
from django.db.models import Count
from scipy.stats import norm, binom

# cik = '0000051143'
# type = '10-K'
# dateb = '20160101'

# https://github.com/alexgolec/tda-api


class GetAllUrlsProcessed(object):
    def __init__(self):
        self.name = 'GetAllUrlsProcessed'
        self.session = None
        self.agent = 'amos@drbaranes.com'

    async def get_page(self, url):
        headers = {'User-Agent': self.agent}
        async with self.session.get(url, headers=headers, timeout=30) as r:
            return await r.text()

    async def get_all_pages(self, urls, func, dic):
        tasks = []
        for url in urls:
            txt = await self.get_page(url)
            task = asyncio.create_task(func(self, txt, url, dic))
            tasks.append(task)
        results = await asyncio.gather(*tasks)
        return results

    async def process_all_pages(self, urls, func, dic):
        async with aiohttp.ClientSession() as session:
            self.session = session
            return await self.get_all_pages(urls, func, dic)


class TDAmeriTrade(BaseTDAmeriTrade):
    def __init__(self):
        super().__init__()
        self.client = None
        self.get_client()
        self.dic_share_prices = {}

    def refresh_token(self, dic):
        # Questions
        # city: Haifa
        # mother mid name: Atya
        #
        result = {"status":"ko"}
        from selenium import webdriver
        from webdriver_manager.chrome import ChromeDriverManager
        with webdriver.Chrome(ChromeDriverManager().install()) as driver:
            try:
                self.client = auth.client_from_login_flow(driver, self.api_key, self.callback_url,
                                                          self.token_path + "/token")
                result["status"] = "ok"
            except Exception as ex:
                print(ex)

        return {'data': result}

    def get_client(self):
        try:
            # print("path=", self.token_path + "/token", self.api_key)
            self.client = auth.client_from_token_file(self.token_path + "/token", self.api_key)
        except Exception as fex:
            print("fex", fex)
            try:
                from selenium import webdriver
                from webdriver_manager.chrome import ChromeDriverManager
                with webdriver.Chrome(ChromeDriverManager().install()) as driver:
                    try:
                        self.client = auth.client_from_login_flow(driver, self.api_key, self.callback_url,
                                                                  self.token_path + "/token")
                    except Exception as ex:
                        print(ex)
            except Exception as eex:
                print(eex)
        return self.client

    # ------ Queue example Should be removed --
    def run_stream_options_data(self, dic):
        asyncio.run(self.stream_options_data())

    async def worker(self, name, queue):
        l = []
        while True:
            # Get a "work item" out of the queue.
            k_ = await queue.get()
            sleep_for = k_["sleep_for"]
            n = str(k_["n"])
            l.append(n)
            # Sleep for the "sleep_for" seconds.
            await asyncio.sleep(sleep_for)

            # Notify the queue that the "work item" has been processed.
            queue.task_done()
            print(f'{n}-{name} has slept for {sleep_for:.2f} seconds')
        return l

    async def stream_options_data(self):
        # Create a queue that we will use to store our "workload".
        queue = asyncio.Queue()

        # Generate random timings and put them into the queue.
        total_sleep_time = 0
        for _ in range(20):
            sleep_for = random.uniform(0.05, 1.0)
            total_sleep_time += sleep_for
            queue.put_nowait({"n": _, "sleep_for": sleep_for})

        # Create three worker tasks to process the queue concurrently.
        tasks = []
        for i in range(3):
            task = asyncio.create_task(self.worker(f'worker-{i}', queue))
            tasks.append(task)

        # Wait until the queue is fully processed.
        started_at = time.monotonic()
        await queue.join()
        total_slept_for = time.monotonic() - started_at

        # Cancel our worker tasks.
        for task in tasks:
            task.cancel()
        # Wait until all worker tasks are cancelled.
        ll = await asyncio.gather(*tasks, return_exceptions=True)
        print(ll)
        print('====')
        print(f'3 workers slept in parallel for {total_slept_for:.2f} seconds')
        print(f'total expected sleep time: {total_sleep_time:.2f} seconds')
    # ----------------------------------------

    def set_sp500_dic(self):
        sp500_url = 'https://en.wikipedia.org/wiki/List_of_S%26P_500_companies'
        self.sp_tickers = list(pd.read_html(sp500_url)[0]['Symbol'].values)
        self.sp_tickers_list = [x.split('.')[0] for x in self.sp_tickers]
        self.sp_tickers_str = ""
        n_ = 0
        for ticker in self.sp_tickers:
            if n_ == 0:
                self.sp_tickers_str = ticker
                n_ += 1
            else:
                self.sp_tickers_str += ","+ticker
        # print(self.sp_tickers_str)
        return self.sp_tickers_str

    def update_options_statistics(self):
        log_debug("Start update_options_statistics: " + datetime.datetime.now().strftime("%H:%M:%S"))
        # print("Start update_options_statistics: " + datetime.datetime.now().strftime("%H:%M:%S"))
        start_date_ = datetime.datetime.now().date()
        end_date_ = (datetime.datetime.now() + datetime.timedelta(days=6)).date()
        # print(s.company.ticker, start_date_, end_date_)
        n = 0
        for s in XBRLSPStatistics.objects.all():
            try:
                n += 1
                if n % 100 == 0:
                    time.sleep(1)
                # print("-1"*50)
                # print(s.company.ticker)
                dic = self.get_option_chain(ticker=s.company.ticker, start_date=start_date_, end_date=end_date_)
                if dic['status'] == "ok":
                    try:
                        s.straddle_price = dic['straddle_price']
                        s.butterfly_price = dic['llc'][1]
                        s.save()
                        # log_debug("options DATA saved for: " + s.company.ticker)
                    except Exception as ex:
                        # log_debug("Error saving options for: " + s.company.ticker)
                        pass
                else:
                    # log_debug("Error getting options data for: " + s.company.ticker)
                    pass
            except Exception as ex:
                # print("error 202 save update_options_statistics: " + str(ex))
                # log_debug("Error 202 getting data 202 td: : " + str(ex) + " " + s.company.ticker)
                pass
        # print('end options')
        log_debug("End update_options_statistics: " + datetime.datetime.now().strftime("%H:%M:%S"))
        return {'status': 'ok'}

    def get_option_chain_new(self, ticker=None, start_date=None, end_date=None):
        # print("-1" * 50)
        # print("start td.get_option_chain for " + ticker)
        dic = {'status': 'ko'}
        try:
            options_ = self.client.get_option_chain(ticker, contract_type=self.client.Options.ContractType.ALL,
                                                    strike_range=self.client.Options.StrikeRange.IN_THE_MONEY,
                                                    from_date=start_date, to_date=end_date,
                                                    strategy=self.client.Options.Strategy.STRADDLE)
        except Exception as ex:
            print("Error in get_option_chain api options pull for : " + ticker)
            log_debug("Error in get_option_chain api options pull for : " + ticker)
            return dic

        # print('------options_---- ' + ticker)
        if len(options_['"putExpDateMap']) > 0 or len(options_['"callExpDateMap']) > 0:
            print(json.dumps(options_.json(), indent=4))
        # print('-----')

        return

        llp = []
        try:
            for d in options_.json()['putExpDateMap']:
                for p in options_.json()['putExpDateMap'][d]:
                    p_ = options_.json()['putExpDateMap'][d][p][0]
                    p_p = (p_['bid'] + p_['ask']) / 2
                    # print(d, p, p_p, p_['bid'], p_['ask'])
                    llp.append(p_p)
        except Exception as ex:
            return dic

        llc = []
        try:
            for d in options_.json()['callExpDateMap']:
                for c in options_.json()['callExpDateMap'][d]:
                    c_ = options_.json()['callExpDateMap'][d][c][0]
                    c_p = (c_['bid'] + c_['ask']) / 2
                    # print(d, c, c_p, c_['bid'], c_['ask'])
                    llc.append(c_p)
        except Exception as ex:
            return dic

        # print('td.get_option_chain 123')
        # print('llp')
        # print('---')
        # print(llp)
        # print('llc')
        # print('---')
        # print(llc)
        # print('---')
        # print('td.get_option_chain 124')
        # print('---')
        straddle_price = round(100 * (llp[2] + llc[2])) / 100
        butterfly_c = round(100 * (-2 * llc[2] + llc[1] + llc[3])) / 100
        butterfly_p = round(100 * (-2 * llp[2] + llp[1] + llp[3])) / 100
        dic = {'status': 'ok', 'straddle_price': straddle_price, "butterfly_c": butterfly_c, "butterfly_p": butterfly_p,
               'llc': llc, 'llp': llp}
        # log_debug("End get_option_chain: " + ticker)
        return dic

    def get_butterfly_price(self, strike_num, strikes, llc, llp):
        p0 = llc['calls']['date']['strikes'][strikes[strike_num]]["price"]
        pl = llc['calls']['date']['strikes'][strikes[strike_num - 1]]["price"]
        pr = llc['calls']['date']['strikes'][strikes[strike_num + 1]]["price"]
        bfc = round(100 * (-2 * p0 + pl + pr)) / 100
        p0 = llp['puts']['date']['strikes'][strikes[strike_num]]["price"]
        pl = llp['puts']['date']['strikes'][strikes[strike_num - 1]]["price"]
        pr = llp['puts']['date']['strikes'][strikes[strike_num + 1]]["price"]
        bfp = round(100 * (-2 * p0 + pl + pr)) / 100
        return bfc, bfp

    def get_complete_option_chain(self, ticker=None, start_date=None, end_date=None):
        print("start td.get_option_chain for 111 " + ticker)
        if not start_date:
            start_date = datetime.datetime.now().date()
            end_date = (datetime.datetime.now() + datetime.timedelta(days=360)).date()
        dic = {'status': 'ko'}
        try:
            options_ = self.client.get_option_chain(ticker, contract_type=self.client.Options.ContractType.ALL,
                                                    strike_count=5, from_date=start_date, to_date=end_date)
        except Exception as ex:
            print("Error 3456 in get_option_chain api options pull for : " + ticker + str(ex))
            return dic
        # print('------options_----')
        # print(json.dumps(options_.json(), indent=4))
        # print(123)
        try:
            llp = {'puts': {"dates": {}}}
            for d_ in options_.json()['putExpDateMap']:
                d = d_.split(":")[0]
                llp['puts']['dates'][d] = {}
                for p in options_.json()['putExpDateMap'][d_]:
                    p_ = options_.json()['putExpDateMap'][d_][p][0]
                    p_p = round(100*(p_['bid'] + p_['ask']) / 2)/100
                    llp['puts']['dates'][d][p] = {"price":p_p, "symbol":p_['symbol']}
        except Exception as ex:
            print("puts: " + str(ex))
            return dic
        # print(llp)
        # print(1234)
        try:
            llc = {'calls': {"dates": {}}}
            for d_ in options_.json()['callExpDateMap']:
                d = d_.split(":")[0]
                llc['calls']['dates'][d] = {}
                for c in options_.json()['callExpDateMap'][d_]:
                    c_ = options_.json()['callExpDateMap'][d_][c][0]
                    c_p = round(100 * (c_['bid'] + c_['ask']) / 2) / 100
                    llc['calls']['dates'][d][c] = {"price": c_p, "symbol": c_['symbol']}

        except Exception as ex:
            print("calls: " + str(ex))
            return dic
        # print(llc)
        # print(12345)
        pstrikes = []
        for d in llp['puts']['dates']:
            for k in llp['puts']['dates'][d]:
                if k not in pstrikes:
                    pstrikes.append(k)
        cstrikes = []
        for d in llc['calls']['dates']:
            for k in llc['calls']['dates'][d]:
                if k not in cstrikes:
                    cstrikes.append(k)
        # print(123456)
        try:
            #
            dic = '{"ticker": "' + ticker + '"}'
            price_data = self.get_quote(dic)['data'][ticker]
            # print(1234567)
            share_price = (price_data['bidPrice'] + price_data['askPrice']) / 2
            # print(share_price)
            # print(12345678)
            dic = {'status': 'ok', 'share_price': share_price, 'pstrikes': pstrikes, 'cstrikes': cstrikes,
                   'llc': llc, 'llp': llp}
            print(123456789)
            print(dic)

            print(1234567891)
        except Exception as ex:
            dic = {'status': 'ko'}
        return dic

    def get_option_chain_vertical(self, ticker=None, start_date=None, end_date=None):
        pass

    def get_option_chain(self, ticker=None, start_date=None, end_date=None):
        print("start td.get_option_chain for 111 " + ticker)
        if not start_date:
            n_ = 6
            if ticker in ["SPY", "QQQ", "IWM"]:
                n_ = 2
                nday = datetime.datetime.today().weekday()  # Monday = 0
                if nday in [0, 2, 4]:
                    n_ = 1
                elif nday in [1, 3, 6]:
                    n_ = 2
                else:
                    n_ = 3
            start_date = datetime.datetime.now().date()
            end_date = (datetime.datetime.now() + datetime.timedelta(days=n_)).date()
        dic = {'status': 'ko'}
        # print('-'*20)
        # print("in get_option_chain api options pull for : " + ticker)
        # print('-'*20)
        try:
            options_ = self.client.get_option_chain(ticker, contract_type=self.client.Options.ContractType.ALL,
                                                    strike_count=5, from_date=start_date, to_date=end_date)
        except Exception as ex:
            # print("Error 3456 in get_option_chain api options pull for : " + ticker + str(ex))
            return dic
        # print('------options_----')
        # print(json.dumps(options_.json(), indent=4))
        # print(123)
        try:
            llp = {'puts': {}}
            for d_ in options_.json()['putExpDateMap']:
                # print("="*100)
                # print(d_)
                # print("="*100)
                d = d_.split(":")[0]
                llp['puts']['date'] = {'date': d}
                llp['puts']['date']['strikes'] = {}
                for p in options_.json()['putExpDateMap'][d_]:
                    p_ = options_.json()['putExpDateMap'][d_][p][0]
                    p_p = (p_['bid'] + p_['ask']) / 2
                    symbol = p_['symbol']
                    llp['puts']['date']['strikes'][p] = {"price": round(100 * p_p) / 100, "symbol": symbol}
        except Exception as ex:
            # print("puts: " + str(ex))
            return dic
        # print(1234)
        # print(llp)
        try:
            llc = {'calls': {}}
            for d_ in options_.json()['callExpDateMap']:
                d = d_.split(":")[0]
                llc['calls']['date'] = {'date': d}
                llc['calls']['date']['strikes'] = {}
                for c in options_.json()['callExpDateMap'][d_]:
                    c_ = options_.json()['callExpDateMap'][d_][c][0]
                    c_p = (c_['bid'] + c_['ask']) / 2
                    symbol = c_['symbol']
                    llc['calls']['date']['strikes'][c] = {"price": round(100 * c_p) / 100, "symbol": symbol}
        except Exception as ex:
            # print("calls: " + str(ex))
            return dic
        strikes = []
        for k in llp['puts']['date']['strikes']:
            if k not in strikes:
                strikes.append(k)
        for k in llc['calls']['date']['strikes']:
            if k not in strikes:
                strikes.append(k)
        # print(12345)
        try:
            straddle_price = round(100 * (llp['puts']['date']['strikes'][strikes[2]]["price"] +
                                          llc['calls']['date']['strikes'][strikes[2]]["price"])) / 100
            bfcl, bfpl = self.get_butterfly_price(strike_num=1, strikes=strikes, llc=llc, llp=llp)
            bfc0, bfp0 = self.get_butterfly_price(strike_num=2, strikes=strikes, llc=llc, llp=llp)
            bfcr, bfpr = self.get_butterfly_price(strike_num=3, strikes=strikes, llc=llc, llp=llp)
            llc_ = [bfcl, bfc0, bfcr]
            llp_ = [bfpl, bfp0, bfpr]
            #
            dic = '{"ticker": "' + ticker + '"}'
            price_data = self.get_quote(dic)['data'][ticker]
            # print(price_data)
            share_price = round(100*(price_data['bidPrice'] + price_data['askPrice']) / 2)/100
            # print(price_data['bidPrice'], price_data['askPrice'], price)

            # self.get_quote(ticker="GOOG"):
            dic = {'status': 'ok', 'ticker': ticker, 'share_price': share_price, 'straddle_price': straddle_price, 'strikes': strikes,
                   "llc_": llc_, "llp_": llp_, 'llc': llc, 'llp': llp}
            # log_debug("End get_option_chain: " + ticker)
            # print('-3'*100)
            # print(dic)
            # print('-3'*100)
        except Exception as ex:
            dic = {'status': 'ko'}
        # print(dic)
        return dic

    # Not used to be deleted --
    # pull data for all strikes of contract for type=(put or call)
    def get_option_strategy_lh_(self, options_, dic, option_type, l_=0.1, h_=0.9, ticker=""):
        # print(option_type)
        # print(ticker)
        # print(dic)
        try:
            for d in options_.json()[option_type + 'ExpDateMap']:
                # print(d)
                if 'date' not in dic:
                    dic['date'] = str(d).split(":")[0]
                    dic['tickers'] = {}
                elif dic['date'] != str(d).split(":")[0]:
                    dic['ticker'] = ticker
                    dic['underlyingPrice'] = options_.json()['underlyingPrice']
                    dic['date'] = str(d).split(":")[0]
                    dic['tickers'] = {}
                for t in options_.json()[option_type + 'ExpDateMap'][d]:
                    # print(options_.json()[option_type+'ExpDateMap'][d][t][0]['delta'])
                    if options_.json()[option_type + 'ExpDateMap'][d][t][0]['delta'] != "NaN":
                        if l_ < abs(options_.json()[option_type + 'ExpDateMap'][d][t][0]['delta']) < h_:
                            if t not in dic['tickers']:
                                dic['tickers'][t] = {}
                            dic['tickers'][t][option_type] = {}
                            dic['tickers'][t][option_type]['price'] = round(100 * (
                                        options_.json()[option_type + 'ExpDateMap'][d][t][0]['bid'] +
                                        options_.json()[option_type + 'ExpDateMap'][d][t][0]['ask']) / 2) / 100
                            dic['tickers'][t][option_type]['delta'] = round(
                                100 * options_.json()[option_type + 'ExpDateMap'][d][t][0]['delta']) / 100
                            dic['tickers'][t][option_type]['theta'] = round(
                                100 * options_.json()[option_type + 'ExpDateMap'][d][t][0]['theta']) / 100
                            dic['tickers'][t][option_type]['symbol'] = options_.json()[option_type + 'ExpDateMap'][d][t][0]['symbol']
                            # print(options_.json()[option_type+'ExpDateMap'][d][t][0])
        except Exception as ex:
            log_debug("Error 201 in get_option_chain api options pull for : " + ticker + " = " + str(ex))
            # print("Error 201 in get_option_chain api options pull for : " + ticker + " = " + str(ex))
        # print("in get_option_statistics_for_ticker 211 : for options: " + option_type)
        return dic

    a = {'symbol': 'AAPL', 'status': 'SUCCESS', 'underlying': None, 'strategy': 'SINGLE', 'interval': 0.0, 'isDelayed': False, 'isIndex': False, 'interestRate': 5.625, 'underlyingPrice': 193.125, 'volatility': 29.0, 'daysToExpiration': 0.0, 'numberOfContracts': 124,
         'callExpDateMap': {'2023-12-15:4': {
             '65.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C65', 'description': 'AAPL Dec 15 2023 65 Call', 'exchangeName': 'OPR', 'bid': 128.1, 'ask': 128.25, 'last': 126.25, 'mark': 128.18, 'bidSize': 1, 'askSize': 39, 'bidAskSize': '1X39', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 130.78, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701460164769, 'quoteTimeInLong': 1702327940042, 'netChange': -4.53, 'volatility': 293.404, 'delta': 1.0, 'gamma': 0.0, 'theta': -0.023, 'vega': 0.0, 'rho': 0.008, 'openInterest': 51, 'timeValue': -1.87, 'theoreticalOptionValue': 128.198, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 65.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -3.46, 'markChange': -2.61, 'markPercentChange': -1.99, 'intrinsicValue': 128.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}],
             '70.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C70', 'description': 'AAPL Dec 15 2023 70 Call', 'exchangeName': 'OPR', 'bid': 123.1, 'ask': 123.25, 'last': 122.38, 'mark': 123.18, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 125.79, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701895891260, 'quoteTimeInLong': 1702327939719, 'netChange': -3.41, 'volatility': 274.102, 'delta': 1.0, 'gamma': 0.0, 'theta': -0.024, 'vega': 0.0, 'rho': 0.009, 'openInterest': 121, 'timeValue': -0.74, 'theoreticalOptionValue': 123.202, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 70.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -2.71, 'markChange': -2.61, 'markPercentChange': -2.08, 'intrinsicValue': 123.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}],
             '75.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C75', 'description': 'AAPL Dec 15 2023 75 Call', 'exchangeName': 'OPR', 'bid': 118.1, 'ask': 118.25, 'last': 117.41, 'mark': 118.18, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 120.79, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701895891260, 'quoteTimeInLong': 1702327934594, 'netChange': -3.38, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 248, 'timeValue': -0.71, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 75.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -2.8, 'markChange': -2.62, 'markPercentChange': -2.17, 'intrinsicValue': 118.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '80.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C80', 'description': 'AAPL Dec 15 2023 80 Call', 'exchangeName': 'OPR', 'bid': 113.05, 'ask': 113.25, 'last': 108.89, 'mark': 113.15, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 115.8, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701712107897, 'quoteTimeInLong': 1702327936800, 'netChange': -6.91, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 205, 'timeValue': -4.23, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 80.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -5.96, 'markChange': -2.65, 'markPercentChange': -2.28, 'intrinsicValue': 113.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}],
             '85.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C85', 'description': 'AAPL Dec 15 2023 85 Call', 'exchangeName': 'OPR', 'bid': 108.05, 'ask': 108.3, 'last': 107.51, 'mark': 108.18, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 107.51, 'lowPrice': 107.51, 'openPrice': 0.0, 'closePrice': 110.8, 'totalVolume': 1, 'tradeDate': None, 'tradeTimeInLong': 1702309893081, 'quoteTimeInLong': 1702327927986, 'netChange': -3.29, 'volatility': 223.576, 'delta': 1.0, 'gamma': 0.0, 'theta': -0.026, 'vega': 0.0, 'rho': 0.011, 'openInterest': 559, 'timeValue': -0.61, 'theoreticalOptionValue': 108.212, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 85.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -2.97, 'markChange': -2.63, 'markPercentChange': -2.37, 'intrinsicValue': 108.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}],
             '90.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C90', 'description': 'AAPL Dec 15 2023 90 Call', 'exchangeName': 'OPR', 'bid': 103.05, 'ask': 103.3, 'last': 102.5, 'mark': 103.18, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 102.5, 'lowPrice': 102.5, 'openPrice': 0.0, 'closePrice': 105.81, 'totalVolume': 1, 'tradeDate': None, 'tradeTimeInLong': 1702309893081, 'quoteTimeInLong': 1702327926085, 'netChange': -3.31, 'volatility': 208.695, 'delta': 1.0, 'gamma': 0.0, 'theta': -0.026, 'vega': 0.0, 'rho': 0.012, 'openInterest': 800, 'timeValue': -0.62, 'theoreticalOptionValue': 103.216, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 90.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -3.12, 'markChange': -2.63, 'markPercentChange': -2.49, 'intrinsicValue': 103.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}],
             '95.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C95', 'description': 'AAPL Dec 15 2023 95 Call', 'exchangeName': 'OPR', 'bid': 98.0, 'ask': 98.35, 'last': 97.2, 'mark': 98.18, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 97.2, 'lowPrice': 97.2, 'openPrice': 0.0, 'closePrice': 100.81, 'totalVolume': 5, 'tradeDate': None, 'tradeTimeInLong': 1702309022440, 'quoteTimeInLong': 1702327926275, 'netChange': -3.61, 'volatility': 194.608, 'delta': 1.0, 'gamma': 0.0, 'theta': -0.027, 'vega': 0.0, 'rho': 0.012, 'openInterest': 558, 'timeValue': -0.92, 'theoreticalOptionValue': 98.22, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 95.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -3.58, 'markChange': -2.64, 'markPercentChange': -2.62, 'intrinsicValue': 98.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}],
             '100.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C100', 'description': 'AAPL Dec 15 2023 100 Call', 'exchangeName': 'OPR', 'bid': 93.05, 'ask': 93.3, 'last': 92.84, 'mark': 93.18, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 92.84, 'lowPrice': 92.84, 'openPrice': 0.0, 'closePrice': 95.82, 'totalVolume': 1, 'tradeDate': None, 'tradeTimeInLong': 1702321138173, 'quoteTimeInLong': 1702327926057, 'netChange': -2.98, 'volatility': 181.234, 'delta': 1.0, 'gamma': 0.0, 'theta': -0.028, 'vega': 0.0, 'rho': 0.013, 'openInterest': 1042, 'timeValue': -0.28, 'theoreticalOptionValue': 93.223, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 100.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -3.11, 'markChange': -2.64, 'markPercentChange': -2.76, 'intrinsicValue': 93.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '105.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C105', 'description': 'AAPL Dec 15 2023 105 Call', 'exchangeName': 'OPR', 'bid': 88.1, 'ask': 88.3, 'last': 86.4, 'mark': 88.2, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 90.82, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1700579278204, 'quoteTimeInLong': 1702327940703, 'netChange': -4.42, 'volatility': 168.491, 'delta': 1.0, 'gamma': 0.0, 'theta': -0.029, 'vega': 0.0, 'rho': 0.013, 'openInterest': 221, 'timeValue': -1.72, 'theoreticalOptionValue': 88.227, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 105.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -4.87, 'markChange': -2.62, 'markPercentChange': -2.89, 'intrinsicValue': 88.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '110.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C110', 'description': 'AAPL Dec 15 2023 110 Call', 'exchangeName': 'OPR', 'bid': 83.1, 'ask': 83.35, 'last': 84.57, 'mark': 83.23, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 85.83, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701974284100, 'quoteTimeInLong': 1702327937467, 'netChange': -1.26, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 131, 'timeValue': 1.45, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 110.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -1.46, 'markChange': -2.6, 'markPercentChange': -3.03, 'intrinsicValue': 83.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '115.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C115', 'description': 'AAPL Dec 15 2023 115 Call', 'exchangeName': 'OPR', 'bid': 78.1, 'ask': 78.35, 'last': 76.33, 'mark': 78.22, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 80.83, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701273438975, 'quoteTimeInLong': 1702327930742, 'netChange': -4.5, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 1008, 'timeValue': -1.79, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 115.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -5.57, 'markChange': -2.61, 'markPercentChange': -3.23, 'intrinsicValue': 78.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '120.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C120', 'description': 'AAPL Dec 15 2023 120 Call', 'exchangeName': 'OPR', 'bid': 73.1, 'ask': 73.35, 'last': 75.35, 'mark': 73.22, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 75.84, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1702059415828, 'quoteTimeInLong': 1702327940354, 'netChange': -0.49, 'volatility': 133.481, 'delta': 1.0, 'gamma': 0.0, 'theta': -0.031, 'vega': 0.0, 'rho': 0.015, 'openInterest': 162, 'timeValue': 2.23, 'theoreticalOptionValue': 73.238, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 120.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -0.64, 'markChange': -2.61, 'markPercentChange': -3.44, 'intrinsicValue': 73.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '125.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C125', 'description': 'AAPL Dec 15 2023 125 Call', 'exchangeName': 'OPR', 'bid': 68.1, 'ask': 68.4, 'last': 69.26, 'mark': 68.25, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 70.84, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701982020357, 'quoteTimeInLong': 1702327931509, 'netChange': -1.58, 'volatility': 122.723, 'delta': 0.999, 'gamma': 0.0, 'theta': -0.031, 'vega': 0.001, 'rho': 0.016, 'openInterest': 428, 'timeValue': 1.14, 'theoreticalOptionValue': 68.241, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 125.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -2.23, 'markChange': -2.59, 'markPercentChange': -3.66, 'intrinsicValue': 68.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '130.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C130', 'description': 'AAPL Dec 15 2023 130 Call', 'exchangeName': 'OPR', 'bid': 63.1, 'ask': 63.35, 'last': 62.95, 'mark': 63.23, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 62.95, 'lowPrice': 62.6, 'openPrice': 0.0, 'closePrice': 65.85, 'totalVolume': 3, 'tradeDate': None, 'tradeTimeInLong': 1702321262136, 'quoteTimeInLong': 1702327925491, 'netChange': -2.9, 'volatility': 112.349, 'delta': 0.999, 'gamma': 0.0, 'theta': -0.032, 'vega': 0.001, 'rho': 0.017, 'openInterest': 1291, 'timeValue': -0.17, 'theoreticalOptionValue': 63.245, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 130.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -4.4, 'markChange': -2.62, 'markPercentChange': -3.98, 'intrinsicValue': 63.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '135.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C135', 'description': 'AAPL Dec 15 2023 135 Call', 'exchangeName': 'OPR', 'bid': 58.1, 'ask': 58.35, 'last': 57.5, 'mark': 58.23, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 57.5, 'lowPrice': 57.05, 'openPrice': 0.0, 'closePrice': 60.85, 'totalVolume': 73, 'tradeDate': None, 'tradeTimeInLong': 1702309750102, 'quoteTimeInLong': 1702327929968, 'netChange': -3.35, 'volatility': 102.331, 'delta': 0.999, 'gamma': 0.0, 'theta': -0.033, 'vega': 0.001, 'rho': 0.017, 'openInterest': 2277, 'timeValue': -0.62, 'theoreticalOptionValue': 58.248, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 135.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -5.51, 'markChange': -2.63, 'markPercentChange': -4.32, 'intrinsicValue': 58.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '140.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C140', 'description': 'AAPL Dec 15 2023 140 Call', 'exchangeName': 'OPR', 'bid': 53.1, 'ask': 53.35, 'last': 52.89, 'mark': 53.23, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 53.26, 'lowPrice': 51.8, 'openPrice': 0.0, 'closePrice': 55.86, 'totalVolume': 25, 'tradeDate': None, 'tradeTimeInLong': 1702321114168, 'quoteTimeInLong': 1702327922593, 'netChange': -2.97, 'volatility': 92.632, 'delta': 0.999, 'gamma': 0.0, 'theta': -0.034, 'vega': 0.001, 'rho': 0.018, 'openInterest': 5350, 'timeValue': -0.23, 'theoreticalOptionValue': 53.252, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 140.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -5.31, 'markChange': -2.63, 'markPercentChange': -4.71, 'intrinsicValue': 53.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '145.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C145', 'description': 'AAPL Dec 15 2023 145 Call', 'exchangeName': 'OPR', 'bid': 48.1, 'ask': 48.35, 'last': 47.84, 'mark': 48.23, 'bidSize': 30, 'askSize': 1, 'bidAskSize': '30X1', 'lastSize': 0, 'highPrice': 47.84, 'lowPrice': 47.02, 'openPrice': 0.0, 'closePrice': 50.86, 'totalVolume': 2, 'tradeDate': None, 'tradeTimeInLong': 1702319510972, 'quoteTimeInLong': 1702327923672, 'netChange': -3.02, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 2706, 'timeValue': -0.28, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 145.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -5.94, 'markChange': -2.64, 'markPercentChange': -5.19, 'intrinsicValue': 48.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '150.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C150', 'description': 'AAPL Dec 15 2023 150 Call', 'exchangeName': 'OPR', 'bid': 43.1, 'ask': 43.35, 'last': 43.1, 'mark': 43.23, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 43.1, 'lowPrice': 42.02, 'openPrice': 0.0, 'closePrice': 45.87, 'totalVolume': 43, 'tradeDate': None, 'tradeTimeInLong': 1702325787238, 'quoteTimeInLong': 1702327929789, 'netChange': -2.77, 'volatility': 74.085, 'delta': 0.999, 'gamma': 0.0, 'theta': -0.035, 'vega': 0.001, 'rho': 0.019, 'openInterest': 8491, 'timeValue': -0.02, 'theoreticalOptionValue': 43.259, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 150.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -6.04, 'markChange': -2.64, 'markPercentChange': -5.76, 'intrinsicValue': 43.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '152.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C152.5', 'description': 'AAPL Dec 15 2023 152.5 Call', 'exchangeName': 'OPR', 'bid': 40.6, 'ask': 40.85, 'last': 0.0, 'mark': 40.73, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 43.37, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 0, 'quoteTimeInLong': 1702327922794, 'netChange': 0.0, 'volatility': 69.606, 'delta': 0.999, 'gamma': 0.0, 'theta': -0.035, 'vega': 0.001, 'rho': 0.019, 'openInterest': 0, 'timeValue': 0.1, 'theoreticalOptionValue': 40.761, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 152.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 0.0, 'markChange': -2.65, 'markPercentChange': -6.1, 'intrinsicValue': 40.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '155.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C155', 'description': 'AAPL Dec 15 2023 155 Call', 'exchangeName': 'OPR', 'bid': 38.15, 'ask': 38.35, 'last': 38.05, 'mark': 38.25, 'bidSize': 56, 'askSize': 56, 'bidAskSize': '56X56', 'lastSize': 0, 'highPrice': 38.05, 'lowPrice': 36.79, 'openPrice': 0.0, 'closePrice': 40.88, 'totalVolume': 95, 'tradeDate': None, 'tradeTimeInLong': 1702326566770, 'quoteTimeInLong': 1702327940345, 'netChange': -2.83, 'volatility': 65.183, 'delta': 0.999, 'gamma': 0.0, 'theta': -0.036, 'vega': 0.001, 'rho': 0.02, 'openInterest': 20422, 'timeValue': -0.07, 'theoreticalOptionValue': 38.263, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 155.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -6.91, 'markChange': -2.63, 'markPercentChange': -6.42, 'intrinsicValue': 38.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '157.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C157.5', 'description': 'AAPL Dec 15 2023 157.5 Call', 'exchangeName': 'OPR', 'bid': 35.65, 'ask': 35.9, 'last': 35.45, 'mark': 35.78, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 35.6, 'lowPrice': 34.95, 'openPrice': 0.0, 'closePrice': 38.38, 'totalVolume': 15, 'tradeDate': None, 'tradeTimeInLong': 1702323373664, 'quoteTimeInLong': 1702327933103, 'netChange': -2.93, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.001, 'rho': 'NaN', 'openInterest': 6, 'timeValue': -0.17, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 157.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -7.64, 'markChange': -2.61, 'markPercentChange': -6.79, 'intrinsicValue': 35.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '160.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C160', 'description': 'AAPL Dec 15 2023 160 Call', 'exchangeName': 'OPR', 'bid': 33.15, 'ask': 33.35, 'last': 33.26, 'mark': 33.25, 'bidSize': 12, 'askSize': 30, 'bidAskSize': '12X30', 'lastSize': 0, 'highPrice': 33.26, 'lowPrice': 32.13, 'openPrice': 0.0, 'closePrice': 35.89, 'totalVolume': 128, 'tradeDate': None, 'tradeTimeInLong': 1702327938140, 'quoteTimeInLong': 1702327940891, 'netChange': -2.63, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 28101, 'timeValue': 0.14, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 160.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -7.33, 'markChange': -2.64, 'markPercentChange': -7.36, 'intrinsicValue': 33.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '162.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C162.5', 'description': 'AAPL Dec 15 2023 162.5 Call', 'exchangeName': 'OPR', 'bid': 30.6, 'ask': 30.9, 'last': 30.75, 'mark': 30.75, 'bidSize': 36, 'askSize': 36, 'bidAskSize': '36X36', 'lastSize': 0, 'highPrice': 30.75, 'lowPrice': 30.69, 'openPrice': 0.0, 'closePrice': 33.4, 'totalVolume': 2, 'tradeDate': None, 'tradeTimeInLong': 1702327884403, 'quoteTimeInLong': 1702327928871, 'netChange': -2.65, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 30, 'timeValue': 0.13, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 162.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -7.93, 'markChange': -2.65, 'markPercentChange': -7.93, 'intrinsicValue': 30.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '165.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C165', 'description': 'AAPL Dec 15 2023 165 Call', 'exchangeName': 'OPR', 'bid': 28.15, 'ask': 28.4, 'last': 28.23, 'mark': 28.28, 'bidSize': 32, 'askSize': 30, 'bidAskSize': '32X30', 'lastSize': 0, 'highPrice': 28.3, 'lowPrice': 27.06, 'openPrice': 0.0, 'closePrice': 30.91, 'totalVolume': 102, 'tradeDate': None, 'tradeTimeInLong': 1702327894857, 'quoteTimeInLong': 1702327925116, 'netChange': -2.68, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 12869, 'timeValue': 0.11, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 165.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -8.66, 'markChange': -2.63, 'markPercentChange': -8.51, 'intrinsicValue': 28.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '167.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C167.5', 'description': 'AAPL Dec 15 2023 167.5 Call', 'exchangeName': 'OPR', 'bid': 25.65, 'ask': 25.9, 'last': 25.26, 'mark': 25.78, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 25.75, 'lowPrice': 25.26, 'openPrice': 0.0, 'closePrice': 28.41, 'totalVolume': 9, 'tradeDate': None, 'tradeTimeInLong': 1702307076887, 'quoteTimeInLong': 1702327929390, 'netChange': -3.15, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.005, 'rho': 'NaN', 'openInterest': 37, 'timeValue': -0.36, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 167.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -11.1, 'markChange': -2.64, 'markPercentChange': -9.29, 'intrinsicValue': 25.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '170.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C170', 'description': 'AAPL Dec 15 2023 170 Call', 'exchangeName': 'OPR', 'bid': 23.2, 'ask': 23.35, 'last': 23.2, 'mark': 23.28, 'bidSize': 136, 'askSize': 132, 'bidAskSize': '136X132', 'lastSize': 0, 'highPrice': 23.49, 'lowPrice': 21.65, 'openPrice': 0.0, 'closePrice': 25.92, 'totalVolume': 1096, 'tradeDate': None, 'tradeTimeInLong': 1702327752725, 'quoteTimeInLong': 1702327940071, 'netChange': -2.72, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.008, 'rho': 'NaN', 'openInterest': 30222, 'timeValue': 0.08, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 170.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -10.5, 'markChange': -2.65, 'markPercentChange': -10.22, 'intrinsicValue': 23.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '172.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C172.5', 'description': 'AAPL Dec 15 2023 172.5 Call', 'exchangeName': 'OPR', 'bid': 20.65, 'ask': 20.85, 'last': 20.55, 'mark': 20.75, 'bidSize': 319, 'askSize': 105, 'bidAskSize': '319X105', 'lastSize': 0, 'highPrice': 20.95, 'lowPrice': 20.5, 'openPrice': 0.0, 'closePrice': 23.43, 'totalVolume': 14, 'tradeDate': None, 'tradeTimeInLong': 1702326148196, 'quoteTimeInLong': 1702327940475, 'netChange': -2.88, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.009, 'rho': 'NaN', 'openInterest': 105, 'timeValue': -0.07, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 172.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -12.3, 'markChange': -2.68, 'markPercentChange': -11.44, 'intrinsicValue': 20.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '175.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C175', 'description': 'AAPL Dec 15 2023 175 Call', 'exchangeName': 'OPR', 'bid': 18.2, 'ask': 18.35, 'last': 18.27, 'mark': 18.27, 'bidSize': 297, 'askSize': 207, 'bidAskSize': '297X207', 'lastSize': 0, 'highPrice': 18.52, 'lowPrice': 16.86, 'openPrice': 0.0, 'closePrice': 20.94, 'totalVolume': 895, 'tradeDate': None, 'tradeTimeInLong': 1702327911960, 'quoteTimeInLong': 1702327940574, 'netChange': -2.67, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.009, 'rho': 'NaN', 'openInterest': 26751, 'timeValue': 0.15, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 175.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -12.75, 'markChange': -2.66, 'markPercentChange': -12.72, 'intrinsicValue': 18.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '177.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C177.5', 'description': 'AAPL Dec 15 2023 177.5 Call', 'exchangeName': 'OPR', 'bid': 15.7, 'ask': 15.95, 'last': 15.3, 'mark': 15.83, 'bidSize': 219, 'askSize': 221, 'bidAskSize': '219X221', 'lastSize': 0, 'highPrice': 15.76, 'lowPrice': 14.5, 'openPrice': 0.0, 'closePrice': 18.45, 'totalVolume': 509, 'tradeDate': None, 'tradeTimeInLong': 1702318655886, 'quoteTimeInLong': 1702327940879, 'netChange': -3.15, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.01, 'rho': 'NaN', 'openInterest': 433, 'timeValue': -0.32, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 177.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -17.08, 'markChange': -2.63, 'markPercentChange': -14.23, 'intrinsicValue': 15.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '180.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C180', 'description': 'AAPL Dec 15 2023 180 Call', 'exchangeName': 'OPR', 'bid': 13.25, 'ask': 13.4, 'last': 13.3, 'mark': 13.33, 'bidSize': 241, 'askSize': 160, 'bidAskSize': '241X160', 'lastSize': 0, 'highPrice': 13.71, 'lowPrice': 11.83, 'openPrice': 0.0, 'closePrice': 15.96, 'totalVolume': 1574, 'tradeDate': None, 'tradeTimeInLong': 1702327891245, 'quoteTimeInLong': 1702327940417, 'netChange': -2.66, 'volatility': 31.346, 'delta': 0.979, 'gamma': 0.007, 'theta': -0.07, 'vega': 0.011, 'rho': 0.022, 'openInterest': 40277, 'timeValue': 0.18, 'theoreticalOptionValue': 13.331, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 180.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -16.69, 'markChange': -2.64, 'markPercentChange': -16.53, 'intrinsicValue': 13.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '182.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C182.5', 'description': 'AAPL Dec 15 2023 182.5 Call', 'exchangeName': 'OPR', 'bid': 10.75, 'ask': 10.9, 'last': 10.68, 'mark': 10.83, 'bidSize': 158, 'askSize': 133, 'bidAskSize': '158X133', 'lastSize': 0, 'highPrice': 11.1, 'lowPrice': 9.3, 'openPrice': 0.0, 'closePrice': 13.48, 'totalVolume': 2703, 'tradeDate': None, 'tradeTimeInLong': 1702326881202, 'quoteTimeInLong': 1702327940639, 'netChange': -2.8, 'volatility': 26.869, 'delta': 0.972, 'gamma': 0.011, 'theta': -0.074, 'vega': 0.014, 'rho': 0.023, 'openInterest': 1402, 'timeValue': 0.06, 'theoreticalOptionValue': 10.843, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 182.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -20.75, 'markChange': -2.65, 'markPercentChange': -19.68, 'intrinsicValue': 10.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '185.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C185', 'description': 'AAPL Dec 15 2023 185 Call', 'exchangeName': 'OPR', 'bid': 8.3, 'ask': 8.45, 'last': 8.33, 'mark': 8.38, 'bidSize': 449, 'askSize': 152, 'bidAskSize': '449X152', 'lastSize': 0, 'highPrice': 9.09, 'lowPrice': 6.8, 'openPrice': 0.0, 'closePrice': 11.01, 'totalVolume': 11847, 'tradeDate': None, 'tradeTimeInLong': 1702327831833, 'quoteTimeInLong': 1702327940663, 'netChange': -2.68, 'volatility': 23.974, 'delta': 0.948, 'gamma': 0.02, 'theta': -0.092, 'vega': 0.023, 'rho': 0.022, 'openInterest': 33994, 'timeValue': 0.21, 'theoreticalOptionValue': 8.394, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 185.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -24.33, 'markChange': -2.63, 'markPercentChange': -23.93, 'intrinsicValue': 8.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '187.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C187.5', 'description': 'AAPL Dec 15 2023 187.5 Call', 'exchangeName': 'OPR', 'bid': 5.95, 'ask': 6.05, 'last': 5.95, 'mark': 6.0, 'bidSize': 381, 'askSize': 96, 'bidAskSize': '381X96', 'lastSize': 0, 'highPrice': 6.33, 'lowPrice': 4.5, 'openPrice': 0.0, 'closePrice': 8.57, 'totalVolume': 1795, 'tradeDate': None, 'tradeTimeInLong': 1702327847679, 'quoteTimeInLong': 1702327940842, 'netChange': -2.62, 'volatility': 21.51, 'delta': 0.896, 'gamma': 0.038, 'theta': -0.121, 'vega': 0.039, 'rho': 0.021, 'openInterest': 14134, 'timeValue': 0.33, 'theoreticalOptionValue': 6.016, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 187.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -30.54, 'markChange': -2.57, 'markPercentChange': -29.96, 'intrinsicValue': 5.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '190.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C190', 'description': 'AAPL Dec 15 2023 190 Call', 'exchangeName': 'OPR', 'bid': 3.8, 'ask': 3.85, 'last': 3.78, 'mark': 3.83, 'bidSize': 246, 'askSize': 24, 'bidAskSize': '246X24', 'lastSize': 0, 'highPrice': 4.2, 'lowPrice': 2.63, 'openPrice': 0.0, 'closePrice': 6.2, 'totalVolume': 16216, 'tradeDate': None, 'tradeTimeInLong': 1702327893185, 'quoteTimeInLong': 1702327940937, 'netChange': -2.42, 'volatility': 19.746, 'delta': 0.782, 'gamma': 0.068, 'theta': -0.163, 'vega': 0.064, 'rho': 0.019, 'openInterest': 74495, 'timeValue': 0.66, 'theoreticalOptionValue': 3.828, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 190.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -39.06, 'markChange': -2.38, 'markPercentChange': -38.34, 'intrinsicValue': 3.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '192.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C192.5', 'description': 'AAPL Dec 15 2023 192.5 Call', 'exchangeName': 'OPR', 'bid': 2.04, 'ask': 2.05, 'last': 2.04, 'mark': 2.05, 'bidSize': 25, 'askSize': 11, 'bidAskSize': '25X11', 'lastSize': 0, 'highPrice': 2.41, 'lowPrice': 1.29, 'openPrice': 0.0, 'closePrice': 4.04, 'totalVolume': 52422, 'tradeDate': None, 'tradeTimeInLong': 1702327936459, 'quoteTimeInLong': 1702327940676, 'netChange': -2.0, 'volatility': 18.62, 'delta': 0.581, 'gamma': 0.096, 'theta': -0.19, 'vega': 0.085, 'rho': 0.014, 'openInterest': 16619, 'timeValue': 1.42, 'theoreticalOptionValue': 2.043, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 192.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -49.54, 'markChange': -2.0, 'markPercentChange': -49.42, 'intrinsicValue': 0.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '195.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C195', 'description': 'AAPL Dec 15 2023 195 Call', 'exchangeName': 'OPR', 'bid': 0.86, 'ask': 0.87, 'last': 0.87, 'mark': 0.87, 'bidSize': 460, 'askSize': 72, 'bidAskSize': '460X72', 'lastSize': 0, 'highPrice': 1.25, 'lowPrice': 0.52, 'openPrice': 0.0, 'closePrice': 2.26, 'totalVolume': 97357, 'tradeDate': None, 'tradeTimeInLong': 1702327940791, 'quoteTimeInLong': 1702327940908, 'netChange': -1.39, 'volatility': 17.992, 'delta': 0.336, 'gamma': 0.093, 'theta': -0.165, 'vega': 0.08, 'rho': 0.008, 'openInterest': 40325, 'timeValue': 0.87, 'theoreticalOptionValue': 0.865, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 195.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -61.57, 'markChange': -1.4, 'markPercentChange': -61.79, 'intrinsicValue': -1.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '197.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C197.5', 'description': 'AAPL Dec 15 2023 197.5 Call', 'exchangeName': 'OPR', 'bid': 0.3, 'ask': 0.31, 'last': 0.3, 'mark': 0.31, 'bidSize': 807, 'askSize': 305, 'bidAskSize': '807X305', 'lastSize': 0, 'highPrice': 0.73, 'lowPrice': 0.2, 'openPrice': 0.0, 'closePrice': 1.03, 'totalVolume': 25531, 'tradeDate': None, 'tradeTimeInLong': 1702327928863, 'quoteTimeInLong': 1702327940542, 'netChange': -0.74, 'volatility': 18.183, 'delta': 0.15, 'gamma': 0.059, 'theta': -0.104, 'vega': 0.051, 'rho': 0.004, 'openInterest': 18803, 'timeValue': 0.3, 'theoreticalOptionValue': 0.305, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 197.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -71.01, 'markChange': -0.73, 'markPercentChange': -70.53, 'intrinsicValue': -4.38, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '200.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C200', 'description': 'AAPL Dec 15 2023 200 Call', 'exchangeName': 'OPR', 'bid': 0.1, 'ask': 0.11, 'last': 0.11, 'mark': 0.11, 'bidSize': 892, 'askSize': 334, 'bidAskSize': '892X334', 'lastSize': 0, 'highPrice': 0.19, 'lowPrice': 0.05, 'openPrice': 0.0, 'closePrice': 0.41, 'totalVolume': 28747, 'tradeDate': None, 'tradeTimeInLong': 1702327938227, 'quoteTimeInLong': 1702327940921, 'netChange': -0.29, 'volatility': 19.161, 'delta': 0.059, 'gamma': 0.028, 'theta': -0.055, 'vega': 0.026, 'rho': 0.001, 'openInterest': 68695, 'timeValue': 0.11, 'theoreticalOptionValue': 0.105, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 200.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -72.84, 'markChange': -0.3, 'markPercentChange': -74.07, 'intrinsicValue': -6.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '202.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C202.5', 'description': 'AAPL Dec 15 2023 202.5 Call', 'exchangeName': 'OPR', 'bid': 0.04, 'ask': 0.05, 'last': 0.05, 'mark': 0.05, 'bidSize': 557, 'askSize': 1243, 'bidAskSize': '557X1243', 'lastSize': 0, 'highPrice': 0.08, 'lowPrice': 0.03, 'openPrice': 0.0, 'closePrice': 0.15, 'totalVolume': 13011, 'tradeDate': None, 'tradeTimeInLong': 1702327902840, 'quoteTimeInLong': 1702327940130, 'netChange': -0.1, 'volatility': 21.046, 'delta': 0.026, 'gamma': 0.013, 'theta': -0.031, 'vega': 0.013, 'rho': 0.001, 'openInterest': 6926, 'timeValue': 0.05, 'theoreticalOptionValue': 0.045, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 202.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -66.67, 'markChange': -0.11, 'markPercentChange': -70.0, 'intrinsicValue': -9.38, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '205.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C205', 'description': 'AAPL Dec 15 2023 205 Call', 'exchangeName': 'OPR', 'bid': 0.02, 'ask': 0.03, 'last': 0.03, 'mark': 0.03, 'bidSize': 910, 'askSize': 1119, 'bidAskSize': '910X1119', 'lastSize': 0, 'highPrice': 0.05, 'lowPrice': 0.02, 'openPrice': 0.0, 'closePrice': 0.06, 'totalVolume': 3575, 'tradeDate': None, 'tradeTimeInLong': 1702327938748, 'quoteTimeInLong': 1702327940487, 'netChange': -0.04, 'volatility': 23.509, 'delta': 0.014, 'gamma': 0.007, 'theta': -0.02, 'vega': 0.008, 'rho': 0.0, 'openInterest': 56396, 'timeValue': 0.03, 'theoreticalOptionValue': 0.025, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 205.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -53.85, 'markChange': -0.04, 'markPercentChange': -61.54, 'intrinsicValue': -11.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '207.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C207.5', 'description': 'AAPL Dec 15 2023 207.5 Call', 'exchangeName': 'OPR', 'bid': 0.01, 'ask': 0.02, 'last': 0.01, 'mark': 0.02, 'bidSize': 690, 'askSize': 797, 'bidAskSize': '690X797', 'lastSize': 0, 'highPrice': 0.03, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.04, 'totalVolume': 870, 'tradeDate': None, 'tradeTimeInLong': 1702327876566, 'quoteTimeInLong': 1702327940159, 'netChange': -0.03, 'volatility': 25.908, 'delta': 0.008, 'gamma': 0.004, 'theta': -0.014, 'vega': 0.005, 'rho': 0.0, 'openInterest': 1708, 'timeValue': 0.01, 'theoreticalOptionValue': 0.015, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 207.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -71.43, 'markChange': -0.02, 'markPercentChange': -57.14, 'intrinsicValue': -14.38, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '210.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C210', 'description': 'AAPL Dec 15 2023 210 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 113, 'bidAskSize': '0X113', 'lastSize': 0, 'highPrice': 0.02, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.03, 'totalVolume': 2240, 'tradeDate': None, 'tradeTimeInLong': 1702327319744, 'quoteTimeInLong': 1702327935269, 'netChange': -0.01, 'volatility': 26.424, 'delta': 0.003, 'gamma': 0.002, 'theta': -0.006, 'vega': 0.002, 'rho': 0.0, 'openInterest': 24471, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 210.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -60.0, 'markChange': -0.02, 'markPercentChange': -80.0, 'intrinsicValue': -16.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '212.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C212.5', 'description': 'AAPL Dec 15 2023 212.5 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 1988, 'bidAskSize': '0X1988', 'lastSize': 0, 'highPrice': 0.02, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.02, 'totalVolume': 3471, 'tradeDate': None, 'tradeTimeInLong': 1702325090408, 'quoteTimeInLong': 1702327939134, 'netChange': -0.01, 'volatility': 29.79, 'delta': 0.003, 'gamma': 0.001, 'theta': -0.006, 'vega': 0.002, 'rho': 0.0, 'openInterest': 623, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 212.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -34.21, 'markChange': -0.01, 'markPercentChange': -67.11, 'intrinsicValue': -19.38, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '215.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C215', 'description': 'AAPL Dec 15 2023 215 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 4460, 'bidAskSize': '0X4460', 'lastSize': 0, 'highPrice': 0.02, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.01, 'totalVolume': 17, 'tradeDate': None, 'tradeTimeInLong': 1702320480728, 'quoteTimeInLong': 1702327939136, 'netChange': 0.0, 'volatility': 33.081, 'delta': 0.002, 'gamma': 0.001, 'theta': -0.006, 'vega': 0.002, 'rho': 0.0, 'openInterest': 15626, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 215.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 85.19, 'markChange': 0.0, 'markPercentChange': -7.41, 'intrinsicValue': -21.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '217.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C217.5', 'description': 'AAPL Dec 15 2023 217.5 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 610, 'bidAskSize': '0X610', 'lastSize': 0, 'highPrice': 0.01, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 2, 'tradeDate': None, 'tradeTimeInLong': 1702305006505, 'quoteTimeInLong': 1702327920417, 'netChange': 0.01, 'volatility': 36.304, 'delta': 0.002, 'gamma': 0.001, 'theta': -0.006, 'vega': 0.001, 'rho': 0.0, 'openInterest': 160, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 217.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 525.0, 'markChange': 0.0, 'markPercentChange': 212.5, 'intrinsicValue': -24.38, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '220.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C220', 'description': 'AAPL Dec 15 2023 220 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 3478, 'bidAskSize': '0X3478', 'lastSize': 0, 'highPrice': 0.01, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 3, 'tradeDate': None, 'tradeTimeInLong': 1702315245524, 'quoteTimeInLong': 1702327939136, 'netChange': 0.01, 'volatility': 39.463, 'delta': 0.002, 'gamma': 0.001, 'theta': -0.006, 'vega': 0.001, 'rho': 0.0, 'openInterest': 18473, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 220.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 2400.0, 'markChange': 0.0, 'markPercentChange': 1150.0, 'intrinsicValue': -26.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '222.5': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C222.5', 'description': 'AAPL Dec 15 2023 222.5 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 1035, 'bidAskSize': '0X1035', 'lastSize': 0, 'highPrice': 0.01, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 1, 'tradeDate': None, 'tradeTimeInLong': 1702305071995, 'quoteTimeInLong': 1702327920417, 'netChange': 0.01, 'volatility': 42.562, 'delta': 0.002, 'gamma': 0.001, 'theta': -0.006, 'vega': 0.001, 'rho': 0.0, 'openInterest': 0, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 222.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -29.38, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '225.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C225', 'description': 'AAPL Dec 15 2023 225 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 1236, 'bidAskSize': '0X1236', 'lastSize': 0, 'highPrice': 0.01, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 2, 'tradeDate': None, 'tradeTimeInLong': 1702305033739, 'quoteTimeInLong': 1702327920412, 'netChange': 0.01, 'volatility': 45.605, 'delta': 0.002, 'gamma': 0.001, 'theta': -0.006, 'vega': 0.001, 'rho': 0.0, 'openInterest': 18488, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 225.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -31.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '230.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C230', 'description': 'AAPL Dec 15 2023 230 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 1460, 'bidAskSize': '0X1460', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1702045817813, 'quoteTimeInLong': 1702327920407, 'netChange': 0.01, 'volatility': 51.537, 'delta': 0.002, 'gamma': 0.0, 'theta': -0.006, 'vega': 0.001, 'rho': 0.0, 'openInterest': 13426, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 230.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -36.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '235.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C235', 'description': 'AAPL Dec 15 2023 235 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 1665, 'bidAskSize': '0X1665', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701794810935, 'quoteTimeInLong': 1702327920416, 'netChange': 0.01, 'volatility': 57.276, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.006, 'vega': 0.001, 'rho': 0.0, 'openInterest': 11153, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 235.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -41.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '240.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C240', 'description': 'AAPL Dec 15 2023 240 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 1865, 'bidAskSize': '0X1865', 'lastSize': 0, 'highPrice': 0.01, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 5, 'tradeDate': None, 'tradeTimeInLong': 1702305001184, 'quoteTimeInLong': 1702327920412, 'netChange': 0.01, 'volatility': 62.844, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.006, 'vega': 0.001, 'rho': 0.0, 'openInterest': 4209, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 240.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -46.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '245.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C245', 'description': 'AAPL Dec 15 2023 245 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 2065, 'bidAskSize': '0X2065', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1700146861173, 'quoteTimeInLong': 1702327920753, 'netChange': 0.01, 'volatility': 68.249, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.006, 'vega': 0.001, 'rho': 0.0, 'openInterest': 1979, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 245.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -51.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '250.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C250', 'description': 'AAPL Dec 15 2023 250 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 2465, 'bidAskSize': '0X2465', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701447441246, 'quoteTimeInLong': 1702327920412, 'netChange': 0.01, 'volatility': 73.505, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 4191, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 250.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -56.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '255.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C255', 'description': 'AAPL Dec 15 2023 255 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 387, 'bidAskSize': '0X387', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1702046214313, 'quoteTimeInLong': 1702327939246, 'netChange': 0.01, 'volatility': 78.622, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 107, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 255.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -61.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '260.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C260', 'description': 'AAPL Dec 15 2023 260 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 384, 'bidAskSize': '0X384', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1698864323110, 'quoteTimeInLong': 1702327939246, 'netChange': 0.01, 'volatility': 83.608, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 1787, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 260.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -66.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '265.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C265', 'description': 'AAPL Dec 15 2023 265 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 387, 'bidAskSize': '0X387', 'lastSize': 0, 'highPrice': 0.01, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 1, 'tradeDate': None, 'tradeTimeInLong': 1702305006507, 'quoteTimeInLong': 1702327939246, 'netChange': 0.01, 'volatility': 88.47, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 1825, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 265.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -71.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '270.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C270', 'description': 'AAPL Dec 15 2023 270 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 791, 'bidAskSize': '0X791', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1698864278485, 'quoteTimeInLong': 1702327906837, 'netChange': 0.01, 'volatility': 93.216, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 1056, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 270.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -76.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '275.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C275', 'description': 'AAPL Dec 15 2023 275 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 388, 'bidAskSize': '0X388', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1697117671294, 'quoteTimeInLong': 1702327939246, 'netChange': 0.01, 'volatility': 97.85, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 1112, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 275.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -81.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '280.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C280', 'description': 'AAPL Dec 15 2023 280 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 634, 'bidAskSize': '0X634', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701804459322, 'quoteTimeInLong': 1702327939136, 'netChange': 0.01, 'volatility': 102.381, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 161, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 280.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -86.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '285.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C285', 'description': 'AAPL Dec 15 2023 285 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 622, 'bidAskSize': '0X622', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1696955239129, 'quoteTimeInLong': 1702327939136, 'netChange': 0.01, 'volatility': 106.812, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 265, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 285.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -91.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '290.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C290', 'description': 'AAPL Dec 15 2023 290 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 588, 'bidAskSize': '0X588', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701807594214, 'quoteTimeInLong': 1702327939136, 'netChange': 0.01, 'volatility': 111.147, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 936, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 290.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -96.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '295.0': [{'putCall': 'CALL', 'symbol': 'AAPL_121523C295', 'description': 'AAPL Dec 15 2023 295 Call', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 4064, 'bidAskSize': '0X4064', 'lastSize': 0, 'highPrice': 0.01, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 2, 'tradeDate': None, 'tradeTimeInLong': 1702319614886, 'quoteTimeInLong': 1702327920417, 'netChange': 0.01, 'volatility': 115.39, 'delta': 0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 4851, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 295.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -101.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}]}},
         'putExpDateMap': {'2023-12-15:4': {'65.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P65', 'description': 'AAPL Dec 15 2023 65 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 3412, 'bidAskSize': '0X3412', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701969434172, 'quoteTimeInLong': 1702327920416, 'netChange': 0.01, 'volatility': 293.404, 'delta': 0.0, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.0, 'rho': 0.0, 'openInterest': 2893, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 65.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -128.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '70.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P70', 'description': 'AAPL Dec 15 2023 70 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 3212, 'bidAskSize': '0X3212', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1699041011092, 'quoteTimeInLong': 1702327920409, 'netChange': 0.01, 'volatility': 274.102, 'delta': 0.0, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.0, 'rho': 0.0, 'openInterest': 3238, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 70.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -123.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '75.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P75', 'description': 'AAPL Dec 15 2023 75 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 3012, 'bidAskSize': '0X3012', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701374704309, 'quoteTimeInLong': 1702327920416, 'netChange': 0.01, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 1202, 'timeValue': 0.01, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 75.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -118.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '80.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P80', 'description': 'AAPL Dec 15 2023 80 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 2812, 'bidAskSize': '0X2812', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701700240171, 'quoteTimeInLong': 1702327920417, 'netChange': 0.01, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 6740, 'timeValue': 0.01, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 80.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -113.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '85.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P85', 'description': 'AAPL Dec 15 2023 85 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 2612, 'bidAskSize': '0X2612', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701700240259, 'quoteTimeInLong': 1702327920412, 'netChange': 0.01, 'volatility': 223.576, 'delta': 0.0, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.0, 'rho': 0.0, 'openInterest': 1121, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 85.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -108.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '90.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P90', 'description': 'AAPL Dec 15 2023 90 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 2412, 'bidAskSize': '0X2412', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701450242984, 'quoteTimeInLong': 1702327920409, 'netChange': 0.01, 'volatility': 208.695, 'delta': 0.0, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.0, 'rho': 0.0, 'openInterest': 4154, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 90.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -103.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '95.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P95', 'description': 'AAPL Dec 15 2023 95 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 2212, 'bidAskSize': '0X2212', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701700206403, 'quoteTimeInLong': 1702327920417, 'netChange': 0.01, 'volatility': 194.608, 'delta': 0.0, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.0, 'rho': 0.0, 'openInterest': 4773, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 95.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -98.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '100.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P100', 'description': 'AAPL Dec 15 2023 100 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 5534, 'bidAskSize': '0X5534', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701887667977, 'quoteTimeInLong': 1702327939136, 'netChange': 0.01, 'volatility': 181.234, 'delta': 0.0, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.0, 'rho': 0.0, 'openInterest': 8374, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 100.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -93.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '105.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P105', 'description': 'AAPL Dec 15 2023 105 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 5679, 'bidAskSize': '0X5679', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701969824749, 'quoteTimeInLong': 1702327939136, 'netChange': 0.01, 'volatility': 168.491, 'delta': 0.0, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.0, 'rho': 0.0, 'openInterest': 5860, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 105.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -88.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '110.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P110', 'description': 'AAPL Dec 15 2023 110 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 3544, 'bidAskSize': '0X3544', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1702060305558, 'quoteTimeInLong': 1702327939136, 'netChange': 0.01, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 10755, 'timeValue': 0.01, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 110.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -83.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '115.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P115', 'description': 'AAPL Dec 15 2023 115 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 1412, 'bidAskSize': '0X1412', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701806671085, 'quoteTimeInLong': 1702327920416, 'netChange': 0.01, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 3602, 'timeValue': 0.01, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 115.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -78.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '120.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P120', 'description': 'AAPL Dec 15 2023 120 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 1212, 'bidAskSize': '0X1212', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701969844160, 'quoteTimeInLong': 1702327920412, 'netChange': 0.01, 'volatility': 133.481, 'delta': -0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.0, 'rho': 0.0, 'openInterest': 8412, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 120.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -73.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '125.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P125', 'description': 'AAPL Dec 15 2023 125 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 4541, 'bidAskSize': '0X4541', 'lastSize': 0, 'highPrice': 0.01, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 6, 'tradeDate': None, 'tradeTimeInLong': 1702320704833, 'quoteTimeInLong': 1702327939136, 'netChange': 0.01, 'volatility': 122.723, 'delta': -0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 24459, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 125.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -68.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '130.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P130', 'description': 'AAPL Dec 15 2023 130 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 3544, 'bidAskSize': '0X3544', 'lastSize': 0, 'highPrice': 0.01, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 2, 'tradeDate': None, 'tradeTimeInLong': 1702323070075, 'quoteTimeInLong': 1702327939136, 'netChange': 0.01, 'volatility': 112.349, 'delta': -0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 4632, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 130.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -63.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '135.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P135', 'description': 'AAPL Dec 15 2023 135 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 4541, 'bidAskSize': '0X4541', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1702066980128, 'quoteTimeInLong': 1702327939136, 'netChange': 0.01, 'volatility': 102.331, 'delta': -0.001, 'gamma': 0.0, 'theta': -0.007, 'vega': 0.001, 'rho': 0.0, 'openInterest': 132851, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 135.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -58.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '140.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P140', 'description': 'AAPL Dec 15 2023 140 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 4465, 'bidAskSize': '0X4465', 'lastSize': 0, 'highPrice': 0.01, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 65, 'tradeDate': None, 'tradeTimeInLong': 1702322737605, 'quoteTimeInLong': 1702327939136, 'netChange': 0.01, 'volatility': 92.632, 'delta': -0.001, 'gamma': 0.0, 'theta': -0.006, 'vega': 0.001, 'rho': 0.0, 'openInterest': 15990, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 140.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -53.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '145.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P145', 'description': 'AAPL Dec 15 2023 145 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 3473, 'bidAskSize': '0X3473', 'lastSize': 0, 'highPrice': 0.02, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 173, 'tradeDate': None, 'tradeTimeInLong': 1702327304711, 'quoteTimeInLong': 1702327939136, 'netChange': 0.01, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.0, 'rho': 'NaN', 'openInterest': 18949, 'timeValue': 0.01, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 145.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 9900.0, 'markChange': 0.0, 'markPercentChange': 4900.0, 'intrinsicValue': -48.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '150.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P150', 'description': 'AAPL Dec 15 2023 150 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 1653, 'bidAskSize': '0X1653', 'lastSize': 0, 'highPrice': 0.01, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 162, 'tradeDate': None, 'tradeTimeInLong': 1702327763025, 'quoteTimeInLong': 1702327906832, 'netChange': 0.01, 'volatility': 74.085, 'delta': -0.001, 'gamma': 0.0, 'theta': -0.006, 'vega': 0.001, 'rho': 0.0, 'openInterest': 36682, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 150.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 1900.0, 'markChange': 0.0, 'markPercentChange': 900.0, 'intrinsicValue': -43.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '152.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P152.5', 'description': 'AAPL Dec 15 2023 152.5 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.01, 'mark': 0.01, 'bidSize': 0, 'askSize': 113, 'bidAskSize': '0X113', 'lastSize': 0, 'highPrice': 0.01, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 1809, 'tradeDate': None, 'tradeTimeInLong': 1702327914794, 'quoteTimeInLong': 1702327935272, 'netChange': 0.01, 'volatility': 69.592, 'delta': -0.001, 'gamma': 0.0, 'theta': -0.006, 'vega': 0.001, 'rho': 0.0, 'openInterest': 1477, 'timeValue': 0.01, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 152.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 733.33, 'markChange': 0.0, 'markPercentChange': 316.67, 'intrinsicValue': -40.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '155.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P155', 'description': 'AAPL Dec 15 2023 155 Put', 'exchangeName': 'OPR', 'bid': 0.0, 'ask': 0.01, 'last': 0.02, 'mark': 0.01, 'bidSize': 0, 'askSize': 77, 'bidAskSize': '0X77', 'lastSize': 0, 'highPrice': 0.02, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.0, 'totalVolume': 2152, 'tradeDate': None, 'tradeTimeInLong': 1702327304711, 'quoteTimeInLong': 1702327935269, 'netChange': 0.02, 'volatility': 65.183, 'delta': -0.001, 'gamma': 0.0, 'theta': -0.006, 'vega': 0.001, 'rho': 0.0, 'openInterest': 40156, 'timeValue': 0.02, 'theoreticalOptionValue': 0.005, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 155.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 640.74, 'markChange': 0.0, 'markPercentChange': 85.19, 'intrinsicValue': -38.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '157.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P157.5', 'description': 'AAPL Dec 15 2023 157.5 Put', 'exchangeName': 'OPR', 'bid': 0.01, 'ask': 0.02, 'last': 0.02, 'mark': 0.02, 'bidSize': 86, 'askSize': 834, 'bidAskSize': '86X834', 'lastSize': 0, 'highPrice': 0.02, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.01, 'totalVolume': 113, 'tradeDate': None, 'tradeTimeInLong': 1702326656815, 'quoteTimeInLong': 1702327936405, 'netChange': 0.01, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.001, 'rho': 'NaN', 'openInterest': 776, 'timeValue': 0.02, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 157.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 233.33, 'markChange': 0.01, 'markPercentChange': 150.0, 'intrinsicValue': -35.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '160.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P160', 'description': 'AAPL Dec 15 2023 160 Put', 'exchangeName': 'OPR', 'bid': 0.01, 'ask': 0.02, 'last': 0.01, 'mark': 0.02, 'bidSize': 126, 'askSize': 699, 'bidAskSize': '126X699', 'lastSize': 0, 'highPrice': 0.02, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.01, 'totalVolume': 82, 'tradeDate': None, 'tradeTimeInLong': 1702326556998, 'quoteTimeInLong': 1702327938072, 'netChange': 0.0, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.004, 'rho': 'NaN', 'openInterest': 67784, 'timeValue': 0.01, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 160.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -13.79, 'markChange': 0.0, 'markPercentChange': 29.31, 'intrinsicValue': -33.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '162.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P162.5', 'description': 'AAPL Dec 15 2023 162.5 Put', 'exchangeName': 'OPR', 'bid': 0.01, 'ask': 0.02, 'last': 0.01, 'mark': 0.02, 'bidSize': 1038, 'askSize': 721, 'bidAskSize': '1038X721', 'lastSize': 0, 'highPrice': 0.02, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.02, 'totalVolume': 142, 'tradeDate': None, 'tradeTimeInLong': 1702309093203, 'quoteTimeInLong': 1702327940131, 'netChange': -0.01, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.004, 'rho': 'NaN', 'openInterest': 949, 'timeValue': 0.01, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 162.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -42.2, 'markChange': 0.0, 'markPercentChange': -13.29, 'intrinsicValue': -30.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '165.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P165', 'description': 'AAPL Dec 15 2023 165 Put', 'exchangeName': 'OPR', 'bid': 0.01, 'ask': 0.02, 'last': 0.02, 'mark': 0.02, 'bidSize': 544, 'askSize': 505, 'bidAskSize': '544X505', 'lastSize': 0, 'highPrice': 0.03, 'lowPrice': 0.01, 'openPrice': 0.0, 'closePrice': 0.02, 'totalVolume': 318, 'tradeDate': None, 'tradeTimeInLong': 1702326619258, 'quoteTimeInLong': 1702327940159, 'netChange': 0.0, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.005, 'rho': 'NaN', 'openInterest': 40464, 'timeValue': 0.02, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 165.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -12.66, 'markChange': -0.01, 'markPercentChange': -34.5, 'intrinsicValue': -28.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '167.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P167.5', 'description': 'AAPL Dec 15 2023 167.5 Put', 'exchangeName': 'OPR', 'bid': 0.02, 'ask': 0.03, 'last': 0.02, 'mark': 0.03, 'bidSize': 10, 'askSize': 1658, 'bidAskSize': '10X1658', 'lastSize': 0, 'highPrice': 0.03, 'lowPrice': 0.02, 'openPrice': 0.0, 'closePrice': 0.03, 'totalVolume': 120, 'tradeDate': None, 'tradeTimeInLong': 1702327255565, 'quoteTimeInLong': 1702327940132, 'netChange': -0.01, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.005, 'rho': 'NaN', 'openInterest': 1290, 'timeValue': 0.02, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 167.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -30.07, 'markChange': 0.0, 'markPercentChange': -12.59, 'intrinsicValue': -25.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '170.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P170', 'description': 'AAPL Dec 15 2023 170 Put', 'exchangeName': 'OPR', 'bid': 0.02, 'ask': 0.03, 'last': 0.02, 'mark': 0.03, 'bidSize': 678, 'askSize': 458, 'bidAskSize': '678X458', 'lastSize': 0, 'highPrice': 0.05, 'lowPrice': 0.02, 'openPrice': 0.0, 'closePrice': 0.03, 'totalVolume': 1889, 'tradeDate': None, 'tradeTimeInLong': 1702327397541, 'quoteTimeInLong': 1702327940488, 'netChange': -0.01, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.008, 'rho': 'NaN', 'openInterest': 60785, 'timeValue': 0.02, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 170.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -41.52, 'markChange': -0.01, 'markPercentChange': -26.9, 'intrinsicValue': -23.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '172.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P172.5', 'description': 'AAPL Dec 15 2023 172.5 Put', 'exchangeName': 'OPR', 'bid': 0.03, 'ask': 0.04, 'last': 0.04, 'mark': 0.04, 'bidSize': 2, 'askSize': 1772, 'bidAskSize': '2X1772', 'lastSize': 0, 'highPrice': 0.04, 'lowPrice': 0.02, 'openPrice': 0.0, 'closePrice': 0.04, 'totalVolume': 344, 'tradeDate': None, 'tradeTimeInLong': 1702327728796, 'quoteTimeInLong': 1702327940487, 'netChange': 0.0, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.009, 'rho': 'NaN', 'openInterest': 6406, 'timeValue': 0.04, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 172.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 0.25, 'markChange': 0.0, 'markPercentChange': -12.28, 'intrinsicValue': -20.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '175.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P175', 'description': 'AAPL Dec 15 2023 175 Put', 'exchangeName': 'OPR', 'bid': 0.03, 'ask': 0.04, 'last': 0.03, 'mark': 0.04, 'bidSize': 1151, 'askSize': 620, 'bidAskSize': '1151X620', 'lastSize': 0, 'highPrice': 0.06, 'lowPrice': 0.03, 'openPrice': 0.0, 'closePrice': 0.05, 'totalVolume': 754, 'tradeDate': None, 'tradeTimeInLong': 1702327923363, 'quoteTimeInLong': 1702327940796, 'netChange': -0.02, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.009, 'rho': 'NaN', 'openInterest': 59339, 'timeValue': 0.03, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 175.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -34.07, 'markChange': -0.01, 'markPercentChange': -23.08, 'intrinsicValue': -18.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '177.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P177.5', 'description': 'AAPL Dec 15 2023 177.5 Put', 'exchangeName': 'OPR', 'bid': 0.03, 'ask': 0.04, 'last': 0.05, 'mark': 0.04, 'bidSize': 1309, 'askSize': 141, 'bidAskSize': '1309X141', 'lastSize': 0, 'highPrice': 0.06, 'lowPrice': 0.04, 'openPrice': 0.0, 'closePrice': 0.05, 'totalVolume': 725, 'tradeDate': None, 'tradeTimeInLong': 1702326224421, 'quoteTimeInLong': 1702327940487, 'netChange': 0.0, 'volatility': 'NaN', 'delta': 'NaN', 'gamma': 'NaN', 'theta': 'NaN', 'vega': 0.01, 'rho': 'NaN', 'openInterest': 6578, 'timeValue': 0.05, 'theoreticalOptionValue': 'NaN', 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 177.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -9.09, 'markChange': -0.02, 'markPercentChange': -36.36, 'intrinsicValue': -15.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '180.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P180', 'description': 'AAPL Dec 15 2023 180 Put', 'exchangeName': 'OPR', 'bid': 0.05, 'ask': 0.06, 'last': 0.06, 'mark': 0.06, 'bidSize': 227, 'askSize': 820, 'bidAskSize': '227X820', 'lastSize': 0, 'highPrice': 0.14, 'lowPrice': 0.05, 'openPrice': 0.0, 'closePrice': 0.06, 'totalVolume': 3174, 'tradeDate': None, 'tradeTimeInLong': 1702327923363, 'quoteTimeInLong': 1702327940131, 'netChange': 0.0, 'volatility': 31.346, 'delta': -0.021, 'gamma': 0.007, 'theta': -0.037, 'vega': 0.011, 'rho': -0.001, 'openInterest': 81388, 'timeValue': 0.06, 'theoreticalOptionValue': 0.055, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 180.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -7.69, 'markChange': -0.01, 'markPercentChange': -15.38, 'intrinsicValue': -13.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '182.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P182.5', 'description': 'AAPL Dec 15 2023 182.5 Put', 'exchangeName': 'OPR', 'bid': 0.06, 'ask': 0.07, 'last': 0.07, 'mark': 0.07, 'bidSize': 1184, 'askSize': 494, 'bidAskSize': '1184X494', 'lastSize': 0, 'highPrice': 0.12, 'lowPrice': 0.07, 'openPrice': 0.0, 'closePrice': 0.08, 'totalVolume': 3657, 'tradeDate': None, 'tradeTimeInLong': 1702327914794, 'quoteTimeInLong': 1702327940796, 'netChange': 0.0, 'volatility': 26.869, 'delta': -0.028, 'gamma': 0.011, 'theta': -0.04, 'vega': 0.014, 'rho': -0.001, 'openInterest': 10210, 'timeValue': 0.07, 'theoreticalOptionValue': 0.065, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 182.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': -6.67, 'markChange': -0.01, 'markPercentChange': -13.33, 'intrinsicValue': -10.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '185.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P185', 'description': 'AAPL Dec 15 2023 185 Put', 'exchangeName': 'OPR', 'bid': 0.11, 'ask': 0.12, 'last': 0.11, 'mark': 0.12, 'bidSize': 772, 'askSize': 1144, 'bidAskSize': '772X1144', 'lastSize': 0, 'highPrice': 0.19, 'lowPrice': 0.11, 'openPrice': 0.0, 'closePrice': 0.1, 'totalVolume': 9642, 'tradeDate': None, 'tradeTimeInLong': 1702327893185, 'quoteTimeInLong': 1702327940269, 'netChange': 0.0, 'volatility': 23.974, 'delta': -0.052, 'gamma': 0.02, 'theta': -0.057, 'vega': 0.023, 'rho': -0.001, 'openInterest': 38185, 'timeValue': 0.11, 'theoreticalOptionValue': 0.115, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 185.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 4.76, 'markChange': 0.01, 'markPercentChange': 9.52, 'intrinsicValue': -8.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '187.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P187.5', 'description': 'AAPL Dec 15 2023 187.5 Put', 'exchangeName': 'OPR', 'bid': 0.23, 'ask': 0.24, 'last': 0.23, 'mark': 0.24, 'bidSize': 30, 'askSize': 1010, 'bidAskSize': '30X1010', 'lastSize': 0, 'highPrice': 0.42, 'lowPrice': 0.23, 'openPrice': 0.0, 'closePrice': 0.16, 'totalVolume': 12441, 'tradeDate': None, 'tradeTimeInLong': 1702327931225, 'quoteTimeInLong': 1702327940292, 'netChange': 0.07, 'volatility': 21.51, 'delta': -0.104, 'gamma': 0.038, 'theta': -0.087, 'vega': 0.039, 'rho': -0.003, 'openInterest': 21950, 'timeValue': 0.23, 'theoreticalOptionValue': 0.235, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 187.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 43.75, 'markChange': 0.08, 'markPercentChange': 46.88, 'intrinsicValue': -5.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '190.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P190', 'description': 'AAPL Dec 15 2023 190 Put', 'exchangeName': 'OPR', 'bid': 0.54, 'ask': 0.55, 'last': 0.54, 'mark': 0.55, 'bidSize': 678, 'askSize': 324, 'bidAskSize': '678X324', 'lastSize': 0, 'highPrice': 1.02, 'lowPrice': 0.5, 'openPrice': 0.0, 'closePrice': 0.29, 'totalVolume': 31958, 'tradeDate': None, 'tradeTimeInLong': 1702327935685, 'quoteTimeInLong': 1702327940640, 'netChange': 0.25, 'volatility': 19.746, 'delta': -0.218, 'gamma': 0.068, 'theta': -0.128, 'vega': 0.064, 'rho': -0.005, 'openInterest': 40451, 'timeValue': 0.54, 'theoreticalOptionValue': 0.545, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 190.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 83.05, 'markChange': 0.25, 'markPercentChange': 84.75, 'intrinsicValue': -3.12, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '192.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P192.5', 'description': 'AAPL Dec 15 2023 192.5 Put', 'exchangeName': 'OPR', 'bid': 1.25, 'ask': 1.27, 'last': 1.27, 'mark': 1.26, 'bidSize': 137, 'askSize': 44, 'bidAskSize': '137X44', 'lastSize': 0, 'highPrice': 2.17, 'lowPrice': 0.8, 'openPrice': 0.0, 'closePrice': 0.63, 'totalVolume': 27522, 'tradeDate': None, 'tradeTimeInLong': 1702327936974, 'quoteTimeInLong': 1702327940374, 'netChange': 0.63, 'volatility': 18.62, 'delta': -0.421, 'gamma': 0.097, 'theta': -0.156, 'vega': 0.085, 'rho': -0.01, 'openInterest': 19298, 'timeValue': 1.27, 'theoreticalOptionValue': 1.26, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 192.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 100.0, 'markChange': 0.63, 'markPercentChange': 98.43, 'intrinsicValue': -0.62, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': False, 'mini': False}], '195.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P195', 'description': 'AAPL Dec 15 2023 195 Put', 'exchangeName': 'OPR', 'bid': 2.58, 'ask': 2.61, 'last': 2.61, 'mark': 2.6, 'bidSize': 43, 'askSize': 277, 'bidAskSize': '43X277', 'lastSize': 0, 'highPrice': 3.95, 'lowPrice': 2.35, 'openPrice': 0.0, 'closePrice': 1.36, 'totalVolume': 9674, 'tradeDate': None, 'tradeTimeInLong': 1702327938587, 'quoteTimeInLong': 1702327940805, 'netChange': 1.25, 'volatility': 17.992, 'delta': -0.671, 'gamma': 0.096, 'theta': -0.131, 'vega': 0.079, 'rho': -0.015, 'openInterest': 27657, 'timeValue': 0.73, 'theoreticalOptionValue': 2.589, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 195.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 91.91, 'markChange': 1.23, 'markPercentChange': 90.81, 'intrinsicValue': 1.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '197.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P197.5', 'description': 'AAPL Dec 15 2023 197.5 Put', 'exchangeName': 'OPR', 'bid': 4.5, 'ask': 4.65, 'last': 4.5, 'mark': 4.58, 'bidSize': 338, 'askSize': 564, 'bidAskSize': '338X564', 'lastSize': 0, 'highPrice': 6.15, 'lowPrice': 4.3, 'openPrice': 0.0, 'closePrice': 2.64, 'totalVolume': 1032, 'tradeDate': None, 'tradeTimeInLong': 1702327921276, 'quoteTimeInLong': 1702327940763, 'netChange': 1.86, 'volatility': 18.183, 'delta': -0.868, 'gamma': 0.064, 'theta': -0.07, 'vega': 0.046, 'rho': -0.014, 'openInterest': 6293, 'timeValue': 0.12, 'theoreticalOptionValue': 4.555, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 197.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 70.19, 'markChange': 1.93, 'markPercentChange': 73.03, 'intrinsicValue': 4.38, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '200.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P200', 'description': 'AAPL Dec 15 2023 200 Put', 'exchangeName': 'OPR', 'bid': 6.85, 'ask': 7.0, 'last': 6.97, 'mark': 6.93, 'bidSize': 179, 'askSize': 130, 'bidAskSize': '179X130', 'lastSize': 0, 'highPrice': 8.25, 'lowPrice': 6.2, 'openPrice': 0.0, 'closePrice': 4.54, 'totalVolume': 1065, 'tradeDate': None, 'tradeTimeInLong': 1702327683472, 'quoteTimeInLong': 1702327940639, 'netChange': 2.43, 'volatility': 19.161, 'delta': -0.965, 'gamma': 0.029, 'theta': -0.023, 'vega': 0.015, 'rho': -0.006, 'openInterest': 2596, 'timeValue': 0.09, 'theoreticalOptionValue': 6.898, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 200.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 53.45, 'markChange': 2.38, 'markPercentChange': 52.46, 'intrinsicValue': 6.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '202.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P202.5', 'description': 'AAPL Dec 15 2023 202.5 Put', 'exchangeName': 'OPR', 'bid': 9.35, 'ask': 9.45, 'last': 10.63, 'mark': 9.4, 'bidSize': 24, 'askSize': 83, 'bidAskSize': '24X83', 'lastSize': 0, 'highPrice': 11.0, 'lowPrice': 9.0, 'openPrice': 0.0, 'closePrice': 6.83, 'totalVolume': 78, 'tradeDate': None, 'tradeTimeInLong': 1702312810152, 'quoteTimeInLong': 1702327940379, 'netChange': 3.8, 'volatility': 21.046, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 1913, 'timeValue': 1.25, 'theoreticalOptionValue': 9.38, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 202.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 55.64, 'markChange': 2.57, 'markPercentChange': 37.63, 'intrinsicValue': 9.38, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '205.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P205', 'description': 'AAPL Dec 15 2023 205 Put', 'exchangeName': 'OPR', 'bid': 11.8, 'ask': 11.95, 'last': 12.0, 'mark': 11.88, 'bidSize': 126, 'askSize': 110, 'bidAskSize': '126X110', 'lastSize': 0, 'highPrice': 13.25, 'lowPrice': 11.7, 'openPrice': 0.0, 'closePrice': 9.29, 'totalVolume': 3239, 'tradeDate': None, 'tradeTimeInLong': 1702323830887, 'quoteTimeInLong': 1702327940745, 'netChange': 2.71, 'volatility': 23.509, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 87, 'timeValue': 0.12, 'theoreticalOptionValue': 11.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 205.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 29.17, 'markChange': 2.59, 'markPercentChange': 27.83, 'intrinsicValue': 11.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '207.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P207.5', 'description': 'AAPL Dec 15 2023 207.5 Put', 'exchangeName': 'OPR', 'bid': 14.35, 'ask': 14.5, 'last': 14.65, 'mark': 14.43, 'bidSize': 22, 'askSize': 109, 'bidAskSize': '22X109', 'lastSize': 0, 'highPrice': 14.65, 'lowPrice': 14.6, 'openPrice': 0.0, 'closePrice': 11.79, 'totalVolume': 101, 'tradeDate': None, 'tradeTimeInLong': 1702323242011, 'quoteTimeInLong': 1702327940528, 'netChange': 2.86, 'volatility': 25.908, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 14, 'timeValue': 0.27, 'theoreticalOptionValue': 14.38, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 207.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 24.26, 'markChange': 2.64, 'markPercentChange': 22.35, 'intrinsicValue': 14.38, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '210.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P210', 'description': 'AAPL Dec 15 2023 210 Put', 'exchangeName': 'OPR', 'bid': 16.85, 'ask': 17.0, 'last': 16.98, 'mark': 16.93, 'bidSize': 98, 'askSize': 322, 'bidAskSize': '98X322', 'lastSize': 0, 'highPrice': 18.35, 'lowPrice': 16.98, 'openPrice': 0.0, 'closePrice': 14.29, 'totalVolume': 619, 'tradeDate': None, 'tradeTimeInLong': 1702327314962, 'quoteTimeInLong': 1702327940745, 'netChange': 2.69, 'volatility': 26.424, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 65, 'timeValue': 0.1, 'theoreticalOptionValue': 16.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 210.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 18.82, 'markChange': 2.64, 'markPercentChange': 18.44, 'intrinsicValue': 16.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '212.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P212.5', 'description': 'AAPL Dec 15 2023 212.5 Put', 'exchangeName': 'OPR', 'bid': 19.3, 'ask': 19.5, 'last': 19.4, 'mark': 19.4, 'bidSize': 120, 'askSize': 88, 'bidAskSize': '120X88', 'lastSize': 0, 'highPrice': 19.4, 'lowPrice': 19.4, 'openPrice': 0.0, 'closePrice': 16.79, 'totalVolume': 1, 'tradeDate': None, 'tradeTimeInLong': 1702305504253, 'quoteTimeInLong': 1702327940468, 'netChange': 2.61, 'volatility': 29.79, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 3, 'timeValue': 0.02, 'theoreticalOptionValue': 19.38, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 212.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 15.54, 'markChange': 2.61, 'markPercentChange': 15.54, 'intrinsicValue': 19.38, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '215.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P215', 'description': 'AAPL Dec 15 2023 215 Put', 'exchangeName': 'OPR', 'bid': 21.85, 'ask': 21.95, 'last': 22.5, 'mark': 21.9, 'bidSize': 25, 'askSize': 130, 'bidAskSize': '25X130', 'lastSize': 0, 'highPrice': 22.5, 'lowPrice': 22.5, 'openPrice': 0.0, 'closePrice': 19.29, 'totalVolume': 2, 'tradeDate': None, 'tradeTimeInLong': 1702305315807, 'quoteTimeInLong': 1702327940933, 'netChange': 3.21, 'volatility': 33.081, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 0.62, 'theoreticalOptionValue': 21.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 215.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 16.64, 'markChange': 2.61, 'markPercentChange': 13.53, 'intrinsicValue': 21.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '217.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P217.5', 'description': 'AAPL Dec 15 2023 217.5 Put', 'exchangeName': 'OPR', 'bid': 24.3, 'ask': 24.5, 'last': 28.11, 'mark': 24.4, 'bidSize': 108, 'askSize': 79, 'bidAskSize': '108X79', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 21.79, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701723178197, 'quoteTimeInLong': 1702327940171, 'netChange': 6.32, 'volatility': 36.304, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 3.73, 'theoreticalOptionValue': 24.38, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 217.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 29.0, 'markChange': 2.61, 'markPercentChange': 11.98, 'intrinsicValue': 24.38, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '220.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P220', 'description': 'AAPL Dec 15 2023 220 Put', 'exchangeName': 'OPR', 'bid': 26.8, 'ask': 26.95, 'last': 24.42, 'mark': 26.88, 'bidSize': 30, 'askSize': 60, 'bidAskSize': '30X60', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 24.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1702061924555, 'quoteTimeInLong': 1702327940485, 'netChange': 0.13, 'volatility': 39.463, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 1, 'timeValue': -2.46, 'theoreticalOptionValue': 26.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 220.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 0.54, 'markChange': 2.58, 'markPercentChange': 10.64, 'intrinsicValue': 26.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '222.5': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P222.5', 'description': 'AAPL Dec 15 2023 222.5 Put', 'exchangeName': 'OPR', 'bid': 29.25, 'ask': 29.5, 'last': 0.0, 'mark': 29.38, 'bidSize': 60, 'askSize': 30, 'bidAskSize': '60X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 26.79, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 0, 'quoteTimeInLong': 1702327939444, 'netChange': 0.0, 'volatility': 42.562, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 0.0, 'theoreticalOptionValue': 29.38, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 222.5, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 0.0, 'markChange': 2.58, 'markPercentChange': 9.65, 'intrinsicValue': 29.38, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '225.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P225', 'description': 'AAPL Dec 15 2023 225 Put', 'exchangeName': 'OPR', 'bid': 31.8, 'ask': 32.0, 'last': 51.15, 'mark': 31.9, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 29.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1696447545529, 'quoteTimeInLong': 1702327933701, 'netChange': 21.86, 'volatility': 45.605, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 19.27, 'theoreticalOptionValue': 31.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 225.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 74.63, 'markChange': 2.61, 'markPercentChange': 8.91, 'intrinsicValue': 31.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '230.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P230', 'description': 'AAPL Dec 15 2023 230 Put', 'exchangeName': 'OPR', 'bid': 36.8, 'ask': 37.0, 'last': 54.25, 'mark': 36.9, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 34.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1699036258216, 'quoteTimeInLong': 1702327928698, 'netChange': 19.96, 'volatility': 51.537, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 17.37, 'theoreticalOptionValue': 36.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 230.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 58.21, 'markChange': 2.61, 'markPercentChange': 7.61, 'intrinsicValue': 36.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '235.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P235', 'description': 'AAPL Dec 15 2023 235 Put', 'exchangeName': 'OPR', 'bid': 41.75, 'ask': 42.0, 'last': 63.75, 'mark': 41.88, 'bidSize': 60, 'askSize': 30, 'bidAskSize': '60X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 39.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1698847200225, 'quoteTimeInLong': 1702327938194, 'netChange': 24.46, 'volatility': 57.276, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 21.87, 'theoreticalOptionValue': 41.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 235.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 62.26, 'markChange': 2.58, 'markPercentChange': 6.58, 'intrinsicValue': 41.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '240.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P240', 'description': 'AAPL Dec 15 2023 240 Put', 'exchangeName': 'OPR', 'bid': 46.75, 'ask': 47.0, 'last': 61.1, 'mark': 46.88, 'bidSize': 60, 'askSize': 30, 'bidAskSize': '60X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 44.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1694527140870, 'quoteTimeInLong': 1702327940603, 'netChange': 16.81, 'volatility': 62.844, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 14.22, 'theoreticalOptionValue': 46.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 240.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 37.95, 'markChange': 2.58, 'markPercentChange': 5.84, 'intrinsicValue': 46.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '245.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P245', 'description': 'AAPL Dec 15 2023 245 Put', 'exchangeName': 'OPR', 'bid': 51.8, 'ask': 52.0, 'last': 55.55, 'mark': 51.9, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 49.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701377400242, 'quoteTimeInLong': 1702327928794, 'netChange': 6.26, 'volatility': 68.249, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 3.67, 'theoreticalOptionValue': 51.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 245.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 12.7, 'markChange': 2.61, 'markPercentChange': 5.3, 'intrinsicValue': 51.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '250.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P250', 'description': 'AAPL Dec 15 2023 250 Put', 'exchangeName': 'OPR', 'bid': 56.75, 'ask': 57.0, 'last': 55.83, 'mark': 56.88, 'bidSize': 54, 'askSize': 30, 'bidAskSize': '54X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 54.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701979262526, 'quoteTimeInLong': 1702327934196, 'netChange': 1.54, 'volatility': 73.505, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': -1.05, 'theoreticalOptionValue': 56.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 250.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 2.84, 'markChange': 2.58, 'markPercentChange': 4.76, 'intrinsicValue': 56.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '255.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P255', 'description': 'AAPL Dec 15 2023 255 Put', 'exchangeName': 'OPR', 'bid': 61.75, 'ask': 62.0, 'last': 65.22, 'mark': 61.88, 'bidSize': 30, 'askSize': 60, 'bidAskSize': '30X60', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 59.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1701118097867, 'quoteTimeInLong': 1702327930287, 'netChange': 5.93, 'volatility': 78.622, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 3.34, 'theoreticalOptionValue': 61.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 255.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 10.0, 'markChange': 2.58, 'markPercentChange': 4.36, 'intrinsicValue': 61.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '260.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P260', 'description': 'AAPL Dec 15 2023 260 Put', 'exchangeName': 'OPR', 'bid': 66.75, 'ask': 67.0, 'last': 84.27, 'mark': 66.88, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 64.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1697744319442, 'quoteTimeInLong': 1702327926690, 'netChange': 19.98, 'volatility': 83.608, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 17.39, 'theoreticalOptionValue': 66.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 260.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 31.08, 'markChange': 2.58, 'markPercentChange': 4.02, 'intrinsicValue': 66.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '265.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P265', 'description': 'AAPL Dec 15 2023 265 Put', 'exchangeName': 'OPR', 'bid': 71.75, 'ask': 72.0, 'last': 0.0, 'mark': 71.88, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 69.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 0, 'quoteTimeInLong': 1702327927608, 'netChange': 0.0, 'volatility': 88.47, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 0.0, 'theoreticalOptionValue': 71.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 265.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 0.0, 'markChange': 2.58, 'markPercentChange': 3.73, 'intrinsicValue': 71.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '270.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P270', 'description': 'AAPL Dec 15 2023 270 Put', 'exchangeName': 'OPR', 'bid': 76.75, 'ask': 77.0, 'last': 96.55, 'mark': 76.88, 'bidSize': 81, 'askSize': 30, 'bidAskSize': '81X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 74.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1683573772394, 'quoteTimeInLong': 1702327939590, 'netChange': 22.26, 'volatility': 93.216, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 19.67, 'theoreticalOptionValue': 76.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 270.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 29.96, 'markChange': 2.58, 'markPercentChange': 3.48, 'intrinsicValue': 76.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '275.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P275', 'description': 'AAPL Dec 15 2023 275 Put', 'exchangeName': 'OPR', 'bid': 81.75, 'ask': 82.0, 'last': 101.5, 'mark': 81.88, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 79.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1696258111292, 'quoteTimeInLong': 1702327925204, 'netChange': 22.21, 'volatility': 97.85, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 19.62, 'theoreticalOptionValue': 81.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 275.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 28.01, 'markChange': 2.58, 'markPercentChange': 3.26, 'intrinsicValue': 81.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '280.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P280', 'description': 'AAPL Dec 15 2023 280 Put', 'exchangeName': 'OPR', 'bid': 86.75, 'ask': 87.0, 'last': 104.45, 'mark': 86.88, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 84.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1699020000090, 'quoteTimeInLong': 1702327926540, 'netChange': 20.16, 'volatility': 102.381, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 17.57, 'theoreticalOptionValue': 86.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 280.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 23.92, 'markChange': 2.58, 'markPercentChange': 3.07, 'intrinsicValue': 86.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '285.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P285', 'description': 'AAPL Dec 15 2023 285 Put', 'exchangeName': 'OPR', 'bid': 91.8, 'ask': 91.95, 'last': 93.6, 'mark': 91.88, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 89.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1700683744535, 'quoteTimeInLong': 1702327925714, 'netChange': 4.31, 'volatility': 106.812, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 1.72, 'theoreticalOptionValue': 91.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 285.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 4.83, 'markChange': 2.58, 'markPercentChange': 2.9, 'intrinsicValue': 91.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '290.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P290', 'description': 'AAPL Dec 15 2023 290 Put', 'exchangeName': 'OPR', 'bid': 96.8, 'ask': 96.95, 'last': 111.4, 'mark': 96.88, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 94.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1694448577647, 'quoteTimeInLong': 1702327926020, 'netChange': 17.11, 'volatility': 111.147, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 14.52, 'theoreticalOptionValue': 96.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 290.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 18.15, 'markChange': 2.58, 'markPercentChange': 2.74, 'intrinsicValue': 96.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}], '295.0': [{'putCall': 'PUT', 'symbol': 'AAPL_121523P295', 'description': 'AAPL Dec 15 2023 295 Put', 'exchangeName': 'OPR', 'bid': 101.8, 'ask': 101.95, 'last': 120.0, 'mark': 101.88, 'bidSize': 30, 'askSize': 30, 'bidAskSize': '30X30', 'lastSize': 0, 'highPrice': 0.0, 'lowPrice': 0.0, 'openPrice': 0.0, 'closePrice': 99.29, 'totalVolume': 0, 'tradeDate': None, 'tradeTimeInLong': 1694630942539, 'quoteTimeInLong': 1702327926020, 'netChange': 20.71, 'volatility': 115.39, 'delta': -1.0, 'gamma': 0.0, 'theta': 0.0, 'vega': 0.0, 'rho': 0.0, 'openInterest': 0, 'timeValue': 18.12, 'theoreticalOptionValue': 101.88, 'theoreticalVolatility': 29.0, 'optionDeliverablesList': None, 'strikePrice': 295.0, 'expirationDate': 1702674000000, 'daysToExpiration': 4, 'expirationType': 'R', 'lastTradingDay': 1702688400000, 'multiplier': 100.0, 'settlementType': ' ', 'deliverableNote': '', 'isIndexOption': None, 'percentChange': 20.86, 'markChange': 2.58, 'markPercentChange': 2.6, 'intrinsicValue': 101.88, 'pennyPilot': True, 'nonStandard': False, 'inTheMoney': True, 'mini': False}]}}}

    def get_option_statistics_for_ticker(self, ticker):
        dic = {'status': 'ko'}
        # print("-1"*50)
        # print("atm: 90551: "+ticker)
        # print("-2"*50)
        # ticker = "^SPX" ticker = "^GSPC" print("="*100)
        n_ = 6
        if ticker in ["SPY", "QQQ", "IWM"]:
            n_ = 2
            nday = datetime.datetime.today().weekday()  # Monday = 0
            if nday in [0, 2, 4]:
                n_ = 1
            elif nday in [1, 3, 6]:
                n_ = 2
            else:
                n_ = 3
        # print("nday=", n_, "ticker=", ticker)
        try:
            start_date_ = datetime.datetime.now().date()
            end_date_ = (datetime.datetime.now() + datetime.timedelta(days=n_)).date()
            # print("-3"*50)
            # print(start_date_, end_date_)
            # print("-4"*50)

            options_ = self.client.get_option_chain(ticker, contract_type=self.client.Options.ContractType.ALL,
                                                    from_date=start_date_, to_date=end_date_)
            # print("-5"*50)

        except Exception as ex:
            print("Error 2345 in get_option_chain api options pull for : " + ticker + " = " + str(ex))
            log_debug("Error 2345 in get_option_chain api options pull for : " + ticker + " = " + str(ex))
            return dic
        # print("get_option_statistics_for_ticker\n", options_.json())
        ll = ["bid", "ask", "last", "mark", "bidSize", "askSize", "totalVolume", "volatility", "delta", "theta", "vega",
              "rho", "openInterest",
              "theoreticalVolatility", "strikePrice"]
        o_ = options_.json()
        for k in o_:
            if k not in ["putExpDateMap", "callExpDateMap"]:
                print("k", k, o_[k])
            else:
                pass
                # for k1 in o_[k]:
                #     print("="*30, "\nExpiration date:", k1, "\n", "="*30)
                #     for w in o_[k][k1]:
                #         if float(w) < 66.0:
                #             print("Strike:", w)
                #             print("---")
                #             # print("AAA\n", o_[k][k1][w])
                #             ll_ = []
                #             for z in o_[k][k1][w]:
                #                 print("Option:\n", "-"*30, "\n", z, "\n", "-"*30)
                #                 for v in ll:
                #                     ll_.append(z[v])
                #                 print(ll,"\n",ll_)
                #                 for x in z:
                #                     print("XXX", x, z[x])


        return {'status': 'ok', 'option_data_ticker': options_.json()}

    # new functions for Django-CMF --
    def get_option_statistics_for_ticker_cmf(self, dic):
        # print("-"*50)
        # print(dic)
        ticker = dic["ticker"]
        return self.get_option_statistics_for_ticker(ticker)


    # need to check to delete
    def get_prices_minutes(self, dic):
        # print("9010")
        # print(dic)
        # print("9011")

        try:
            dic = eval(dic)
        except Exception as ex:
            pass
            # print("er 9015: "+str(ex))
        ticker_ = dic['ticker']
        print("9012 ticker:  "+ticker_)
        try:
            r = self.client.get_price_history(ticker_,
                                              period_type=client.Client.PriceHistory.PeriodType.DAY,
                                              period=client.Client.PriceHistory.Period.TEN_DAYS,# .ONE_DAY,
                                              frequency_type=client.Client.PriceHistory.FrequencyType.MINUTE,
                                              frequency=client.Client.PriceHistory.Frequency.EVERY_MINUTE)
            assert r.status_code == 200, r.raise_for_status()
            # print(json.dumps(r.json(), indent=4))
        except Exception as ex:
            pass
            # print("9022 "+str(ex))

        dic = {'data': r.json()}
        log_debug("End get_prices.")
        return dic

    # need to check to delete
    def update_prices(self, dic):
        # print("9010")
        # print(dic)
        # print("9011")
        try:
            dic = eval(dic)
        except Exception as ex:
            pass
            # print("er 9015: "+str(ex))
        ticker_ = dic['ticker']
        # print("9012 ticker:  "+ticker_)
        try:
            r = self.client.get_price_history(ticker_,
                                              # period_type=client.Client.PriceHistory.PeriodType.MONTH,
                                              # period=client.Client.PriceHistory.Period.THREE_MONTHS,
                                              period_type=client.Client.PriceHistory.PeriodType.YEAR,
                                              period=client.Client.PriceHistory.Period.TEN_YEARS,
                                              frequency_type=client.Client.PriceHistory.FrequencyType.DAILY,  # .MINUTE,
                                              frequency=client.Client.PriceHistory.Frequency.DAILY
                                              )  #  .EVERY_MINUTE
            assert r.status_code == 200, r.raise_for_status()
            # print(json.dumps(r.json(), indent=4))
        except Exception as ex:
            pass
            # print("9022 "+str(ex))
        # StockPricesDays

        result = {"x": [], "o": [], "h": [], "l": [], "c": [], "v": []}
        dic = r.json()["candles"]
        # print(dic)
        for i in dic:
            print("-"*50)
            print(i["datetime"])
            print(datetime.datetime.fromtimestamp(i["datetime"]))
            print("-"*50)
            result["x"].append(i["datetime"])
            result["o"].append(i["open"])
            result["h"].append(i["high"])
            result["l"].append(i["low"])
            result["c"].append(i["close"])
            result["v"].append(i["volume"])
        # print(result)

        #
        # # print("9010 input dic: \n", dic, "\n"+"-"*30)
        # # print(dic)
        # # print(dic["ticker"])
        # l_f = str(dic["letter_from"])
        # l_t = str(dic["letter_to"])
        # w_ = int(dic["numer_of_weeks"])
        # d_ = int(dic["numer_of_days"])  # d_ = 7
        # # StockPricesMinutes.truncate();
        #
        # watch_list = ETFWatchLists.objects.all()
        # for w in watch_list:
        #     companies = XBRLCompanyInfo.objects.filter(etfwatchlist=w).all()
        #     for c in companies:
        #         # print("="*50)
        #         # print("="*50)
        #         # print("="*50)
        #         if l_t >= c.company_letter >= l_f:
        #             # print(w.symbol, c.ticker, c.company_letter)
        #                 # print("="*50)
        #             # print("="*50)
        #             # print("="*50)
        #             obj = yf.Ticker(c.ticker)
        #             i = w_  # 5 for 1m, 45 for 1d
        #
        #             i -= 1
        #
        #             end_ = str(date_e_.year)+"-"+str(date_e_.month)+"-"+str(date_e_.day)
        #             beg_ = str(date_b_.year)+"-"+str(date_b_.month)+"-"+str(date_b_.day)
        #             hist = obj.history(interval="1m", start=beg_, end=end_)
        #             # n_tail=10
        #             hist = hist[:-1]  # .tail(n_tail+1)
        #             # print("-1"*50)
        #             # print(hist)
        #             # print("-2"*50)
        #             for index, row in hist.iterrows():
        #                 idx = pd.Timestamp(index)
        #                 idx_ = idx.year*100000000+idx.month*1000000+idx.day*10000+idx.hour*100+idx.minute
        #                 # print("-2"*50)
        #                 # print(idx_)
        #                 if row["Volume"] > 0:
        #                     # print("-3"*50)
        #                     # print(row["Volume"])
        #                     # print("-4"*50)
        #                     open_ = round(100*row["Open"])/100
        #                     high_ = round(100*row["High"])/100
        #                     low_ = round(100*row["Low"])/100
        #                     close_ = round(100*row["Close"])/100
        #                     volume_ = round(row["Volume"])
        #                     dividends_ = round(100*row["Dividends"])/100
        #                     stock_splits_ = round(100*row["Stock Splits"])/100
        #                     try:
        #                         # print("-5"*50)
        #                         sp, is_created = StockPricesMinutes.objects.get_or_create(company=c, idx=idx_)
        #                         # print("-6"*50)
        #                         # print(is_created)
        #                         # print("-7"*50)
        #                         if is_created:
        #                             sp.open = open_
        #                             sp.high = high_
        #                             sp.low = low_
        #                             sp.close = close_
        #                             sp.volume = volume_
        #                             sp.dividends = dividends_
        #                             sp.stock_splits = stock_splits_
        #                             # print("-8"*50)
        #                             sp.save()
        #                             # print("-9"*50)
        #                     except Exception as ex:
        #                         print(ex)
        #
        # dic = {'data': {"status": "ok"}}
        #


        return result

    def get_prices(self, dic):
        # print("9010")
        # print(dic)
        # print("9011")
        try:
            dic = eval(dic)
        except Exception as ex:
            pass
            # print("er 9015: "+str(ex))
        ticker_ = dic['ticker']
        # print("9012 ticker:  "+ticker_)
        try:
            r = self.client.get_price_history(ticker_,
                                              # period_type=client.Client.PriceHistory.PeriodType.MONTH,
                                              # period=client.Client.PriceHistory.Period.THREE_MONTHS,
                                              period_type=client.Client.PriceHistory.PeriodType.YEAR,
                                              period=client.Client.PriceHistory.Period.TEN_YEARS,
                                              frequency_type=client.Client.PriceHistory.FrequencyType.DAILY,  # .MINUTE,
                                              frequency=client.Client.PriceHistory.Frequency.DAILY
                                              )  #  .EVERY_MINUTE
            assert r.status_code == 200, r.raise_for_status()
            # print(json.dumps(r.json(), indent=4))
        except Exception as ex:
            pass
            # print("9022 "+str(ex))

        result = {"x": [], "o": [], "h": [], "l": [], "c": [], "v": []}
        dic = r.json()["candles"]
        # print(dic)
        for i in dic:
            result["x"].append(i["datetime"])
            result["o"].append(i["open"])
            result["h"].append(i["high"])
            result["l"].append(i["low"])
            result["c"].append(i["close"])
            result["v"].append(i["volume"])
        # print(result)
        return result

    def get_quote(self, dic):
        # print(9024)
        try:
            dic = eval(dic)
        except Exception as ex:
            pass
        # print(dic)
        ticker_ = dic['ticker']
        url = r"https://api.tdameritrade.com/v1/marketdata/{}/quotes".format(ticker_.upper())
        # print(url)
        pay_load = {'apikey': self.customer_key + '@AMER.OAUTHAP'}
        # print(pay_load)
        try:
            r = requests.get(url=url, params=pay_load).json()
            # print(r)
        except Exception as ex:
            print(ex)
        return {'data': r}

    def get_iron_condor(self, rdic):
        ticker = rdic['ticker']
        try:
            log_debug("Start ICondor for: " + ticker)
            rdic['is_success'] = False
            start_date_ = datetime.datetime.now().date()
            end_date_ = (datetime.datetime.now() + datetime.timedelta(days=6)).date()
            options_ = self.client.get_option_chain(ticker, contract_type=self.client.Options.ContractType.ALL,
                                                    from_date=start_date_, to_date=end_date_)
            # print('options_.json()')
            # print(options_.json())
            dicd = {}
            dicd = self.get_option_strategy_lh_(options_=options_, dic=dicd, option_type='call', l_=0.1, h_=0.4)
            # print(dicd)
            delta_call_low = -1
            strike_call_low = -1
            price_call_low = -1
            delta_put_high = -1
            strike_put_high = -1
            price_put_high = -1
            d_ = -1
            sk = -1
            sk1 = -1
            try:
                for k in dicd['tickers']:
                    if sk == -1:
                        sk = float(k)
                    else:
                        sk1 = float(k)
                        d_ = abs(sk1 - sk)
                        sk = sk1
                    delta_ = dicd['tickers'][k]['call']['delta']
                    price_ = dicd['tickers'][k]['call']['price']
                    if 0.3 >= delta_ >= 0.2:
                        if delta_ > delta_call_low:
                            delta_call_low = delta_
                            strike_call_low = k
                            price_call_low = price_
            except Exception as ex:
                log_debug(str(ex)+ " 1 " + ticker)
                return rdic
            try:
                strike_call_high = str(float(strike_call_low) + d_)
                if float(strike_call_low) < rdic['p']:
                    log_debug("float(strike_call_low) < rdic['p'] "+ticker)
                    return rdic
                delta_call_high = dicd['tickers'][strike_call_high]['call']['delta']
                price_call_high = dicd['tickers'][strike_call_high]['call']['price']
                dicd = {}
                dicd = self.get_option_strategy_lh_(options_=options_, dic=dicd, option_type='put', l_=0.1, h_=0.4)
            except Exception as ex:
                log_debug(str(ex) + " 2 " + ticker)
                return rdic
            try:
                for k in dicd['tickers']:
                    delta_ = abs(dicd['tickers'][k]['put']['delta'])
                    price_ = dicd['tickers'][k]['put']['price']
                    if 0.3 >= delta_ >= 0.2:
                        if delta_ > delta_put_high:
                            delta_put_high = delta_
                            strike_put_high = k
                            price_put_high = price_
            except Exception as ex:
                log_debug(str(ex) + " 3 " + ticker)
                return rdic
            # print(delta_put_high, strike_put_high, price_put_high, d_)
            if delta_put_high == -1:
                log_debug("delta_put_high == -1")
                return rdic
            try:
                strike_put_low = str(float(strike_put_high) - d_)
                # print(strike_put_low)
                delta_put_low = abs(dicd['tickers'][strike_put_low]['put']['delta'])
                price_put_low = dicd['tickers'][strike_put_low]['put']['price']
                # print(delta_put_low, strike_put_low, price_put_low, d_)
            except Exception as ex:
                log_debug(str(ex) + " 4 " + ticker)
                return rdic
            if delta_put_low == -1:
                log_debug("delta_put_low == -1")
                return rdic
            try:
                condor_price = round(100 * (price_call_low - price_call_high + price_put_high - price_put_low)) / 100
                result = {'call': {'low': {'delta': delta_call_low, 'strike': strike_call_low, 'price': price_call_low},
                                   'high': {'delta': delta_call_high, 'strike': strike_call_high,
                                            'price': price_call_high}},
                          'put': {'low': {'delta': delta_put_low, 'strike': strike_put_low, 'price': price_put_low},
                                  'high': {'delta': delta_put_high, 'strike': strike_put_high, 'price': price_put_high}}}
                rdic['condor_price'] = condor_price
                rdic['condor'] = result
                rdic['is_success'] = True
            except Exception as ex:
                log_debug(str(ex) + " 5 " + ticker)
                return rdic
            # print(rdic)
        except Exception as ex:
            print("Error 7891 in get_option_chain api options pull for : " + str(ex) + " " + ticker)
            return rdic
        # print('End -5000  ' + ticker)
        return rdic

    def get_quotes(self, dic):
        # print(dic)
        try:
            dic = eval(dic)
        except Exception as ex:
            pass
        ticker_ = dic['ticker']
        #
        # need to pull the list of all stocks
        list1 = "SPX,RUT,IWM,QQQ,NDX,SPY,TSLAS,PYPL,AMZN,BABA,FB" # "MMM,ABT,ABBV,ABMD,ACN,ATVI,ADBE"

        if ticker_ == "ALL":
            ticker_ = list1

        # print(self.sp_tickers_str)
        log_debug("get_quotes: " + ticker_)
        url = r"https://api.tdameritrade.com/v1/marketdata/quotes"
        pay_load = {'apikey': self.customer_key + "@AMER.OAUTHAP", 'symbol': ticker_}
        # print(pay_load)
        r = requests.get(url=url, params=pay_load)
        data = r.json()
        # print(data)
        log_debug("get_quotes 100: data pulled")
        results = {}
        for k in data:
            # print(k)
            log_debug("Start ticker: "+k)
            d = datetime.datetime.fromtimestamp(data[k]['quoteTimeInLong'] / 999.999451)
            m = d.minute
            h = d.hour
            date_ = d.date()
            dic_k = {"ticker": k, "date": str(date_), "h": h, "m": m, "p": data[k]['lastPrice']}
            result = self.get_iron_condor(dic_k)
            if result['is_success']:
                # print('result')
                # print(result)
                cld = result['condor']['call']['low']['delta']
                cls = float(result['condor']['call']['low']['strike'])
                clp = result['condor']['call']['low']['price']
                chd = result['condor']['call']['high']['delta']
                chs = float(result['condor']['call']['high']['strike'])
                chp = result['condor']['call']['high']['price']

                pld = result['condor']['put']['low']['delta']
                pls = float(result['condor']['put']['low']['strike'])
                plp = result['condor']['put']['low']['price']

                phd = result['condor']['put']['high']['delta']
                phs = float(result['condor']['put']['high']['strike'])
                php = result['condor']['put']['high']['price']

                ep = round(100*((clp - chp) + (php - plp)))/100
                elc = round(100*((chs - cls) * (cld - chd) / 2 + chd * (chs - cls)))/100
                elp = round(100*((phs - pls) * (phd - pld) / 2 + pld * (phs - pls)))/100
                # print(ep, elc, elp)
                # print('-' * 100)
                result['expected_profit'] = ep - elc - elp
                results[k] = result
            else:
                continue

        dic = {'data': results}
        # print(dic)
        # print("End get_Quotes.")
        return dic

    # ------
    async def get_iron_condora(self, rdic):
        ticker = rdic['ticker']
        try:
            rdic['is_success'] = False
            start_date_ = datetime.datetime.now().date()
            end_date_ = (datetime.datetime.now() + datetime.timedelta(days=6)).date()
            options_ = self.client.get_option_chain(ticker, contract_type=self.client.Options.ContractType.ALL,
                                                    from_date=start_date_, to_date=end_date_)
            # print('options_.json()')
            # print(options_.json())
            dicd = {}
            dicd = self.get_option_strategy_lh_(options_=options_, dic=dicd, option_type='call', l_=0.1, h_=0.4)
            # print(dicd)
            delta_call_low = -1
            strike_call_low = -1
            price_call_low = -1
            delta_put_high = -1
            strike_put_high = -1
            price_put_high = -1
            d_ = -1
            sk = -1
            sk1 = -1
            try:
                for k in dicd['tickers']:
                    if sk == -1:
                        sk = float(k)
                    else:
                        sk1 = float(k)
                        d_ = abs(sk1 - sk)
                        sk = sk1
                    delta_ = dicd['tickers'][k]['call']['delta']
                    price_ = dicd['tickers'][k]['call']['price']
                    if 0.3 >= delta_ >= 0.2:
                        if delta_ > delta_call_low:
                            delta_call_low = delta_
                            strike_call_low = k
                            price_call_low = price_
            except Exception as ex:
                return rdic
            try:
                strike_call_high = str(float(strike_call_low) + d_)
                if float(strike_call_low) < rdic['p']:
                    return rdic
                delta_call_high = dicd['tickers'][strike_call_high]['call']['delta']
                price_call_high = dicd['tickers'][strike_call_high]['call']['price']
                dicd = {}
                dicd = self.get_option_strategy_lh_(options_=options_, dic=dicd, option_type='put', l_=0.1, h_=0.4)
            except Exception as ex:
                return rdic
            try:
                for k in dicd['tickers']:
                    delta_ = abs(dicd['tickers'][k]['put']['delta'])
                    price_ = dicd['tickers'][k]['put']['price']
                    if 0.3 >= delta_ >= 0.2:
                        if delta_ > delta_put_high:
                            delta_put_high = delta_
                            strike_put_high = k
                            price_put_high = price_
            except Exception as ex:
                return rdic
            # print(delta_put_high, strike_put_high, price_put_high, d_)
            if delta_put_high == -1:
                return rdic
            try:
                strike_put_low = str(float(strike_put_high) - d_)
                # print(strike_put_low)
                delta_put_low = abs(dicd['tickers'][strike_put_low]['put']['delta'])
                price_put_low = dicd['tickers'][strike_put_low]['put']['price']
                # print(delta_put_low, strike_put_low, price_put_low, d_)
            except Exception as ex:
                return rdic
            if delta_put_low == -1:
                return rdic
            try:
                condor_price = round(100 * (price_call_low - price_call_high + price_put_high - price_put_low)) / 100
                result = {'call': {'low': {'delta': delta_call_low, 'strike': strike_call_low, 'price': price_call_low},
                                   'high': {'delta': delta_call_high, 'strike': strike_call_high,
                                            'price': price_call_high}},
                          'put': {'low': {'delta': delta_put_low, 'strike': strike_put_low, 'price': price_put_low},
                                  'high': {'delta': delta_put_high, 'strike': strike_put_high, 'price': price_put_high}}}
                rdic['condor_price'] = condor_price
                rdic['condor'] = result
                rdic['is_success'] = True
            except Exception as ex:
                return rdic
        except Exception as ex:
            print("Error 7891 in get_option_chain api options pull for : " + str(ex) + " " + ticker)
            return rdic
        return rdic

    async def get_quotesa(self, i):
        # print(dic)
        # print(type(dic))
        ticker_ = self.get_tickers(i)
        #
        print(ticker_)

        url = r"https://api.tdameritrade.com/v1/marketdata/quotes"
        pay_load = {'apikey': self.customer_key + "@AMER.OAUTHAP", 'symbol': ticker_}
        url = url+"?apikey="+self.customer_key+"@AMER.OAUTHAP&symbol="+ticker_
        async with aiohttp.ClientSession as session:
            r = await aiohttp.get(url=url, params=pay_load)

        data = await r.json()
        # print(json.dumps(data))

        for k in data:
            print(k)
            d = datetime.datetime.fromtimestamp(data[k]['quoteTimeInLong'] / 999.999451)
            m = d.minute
            h = d.hour
            date_ = d.date()
            dic_k = {"ticker": k, "date": str(date_), "h": h, "m": m, "p": data[k]['lastPrice']}
            try:
                print(datetime.datetime.now().strftime("%H:%M:%S"))
                result = await self.get_iron_condora(dic_k)
                print(datetime.datetime.now().strftime("%H:%M:%S"))
                print(result['is_success'])
            except Exception as ex:
                print(str(ex))
            if result['is_success']:
                # print('result')
                # print(result)
                cld = result['condor']['call']['low']['delta']
                cls = float(result['condor']['call']['low']['strike'])
                clp = result['condor']['call']['low']['price']
                chd = result['condor']['call']['high']['delta']
                chs = float(result['condor']['call']['high']['strike'])
                chp = result['condor']['call']['high']['price']

                pld = result['condor']['put']['low']['delta']
                pls = float(result['condor']['put']['low']['strike'])
                plp = result['condor']['put']['low']['price']

                phd = result['condor']['put']['high']['delta']
                phs = float(result['condor']['put']['high']['strike'])
                php = result['condor']['put']['high']['price']

                ep = round(100*((clp - chp) + (php - plp)))/100
                elc = round(100*((chs - cls) * (cld - chd) / 2 + chd * (chs - cls)))/100
                elp = round(100*((phs - pls) * (phd - pld) / 2 + pld * (phs - pls)))/100
                # print(ep, elc, elp)
                # print('-' * 100)
                result['expected_profit'] = ep - elc - elp
                print(json.dumps(result))
            else:
                continue
        # print("End get_Quotes.")
        return dic

    def get_tickers(self, i):
        n = 1
        k_ = self.sp_tickers_list[i*n: (i+1)*n]
        k = ""
        n_ = 0
        for ticker in k_:
            if n_ == 0:
                k = ticker
                n_ += 1
            else:
                k += ","+ticker
        return k

    def run_stream_options_dataa(self):
        asyncio.run(self.stream_options_dataa())

    async def workera(self, name, queue):
        while True:
            # Get a "work item" out of the queue.
            k_ = await queue.get()
            queue.task_done()
            n = str(k_["n"])
            await self.get_quotesa(int(n))

            # Notify the queue that the "work item" has been processed.
            # print(f'{n}-{name} has slept for {sleep_for:.2f} seconds')

    async def stream_options_dataa(self):
        # Create a queue that we will use to store our "workload".
        queue = asyncio.Queue()

        # Generate random timings and put them into the queue.
        n_sets_of_tickers = 11
        total_sleep_time = 0
        for sleep_for in range(n_sets_of_tickers):
            # sleep_for = random.uniform(0.05, 1.0)
            total_sleep_time += sleep_for
            queue.put_nowait({"n": sleep_for, "sleep_for": sleep_for})

        # Create three worker tasks to process the queue concurrently.
        tasks = []
        for i in range(n_sets_of_tickers):
            task = asyncio.create_task(self.workera(f'worker-{i}', queue))
            tasks.append(task)

        # Wait until the queue is fully processed.
        started_at = time.monotonic()
        await queue.join()
        total_slept_for = time.monotonic() - started_at

        # Cancel our worker tasks.
        for task in tasks:
            task.cancel()
        # Wait until all worker tasks are cancelled.
        await asyncio.gather(*tasks, return_exceptions=True)
        await self.stream_options_dataa()

        # print('====')
        # print(f'3 workers slept in parallel for {total_slept_for:.2f} seconds')
        # print(f'total expected sleep time: {total_sleep_time:.2f} seconds')

    # ------  SP Option streaming handlers ---
    async def nasdaq_order_book_handler(self, msg):
        print("="*50)
        print("nasdaq_order_book_handler")
        print("-"*40)
        print(msg)
        print("="*50)

        for t_ in msg["content"]:
            bb = []
            aa = []
            for b in t_["BIDS"]:
                bb.append(b["BID_PRICE"])
            if len(bb) == 0:
                bb = 0
            else:
                bb = sum(bb)/len(bb)
            for a in t_["ASKS"]:
                aa.append(a["ASK_PRICE"])
            if len(aa) == 0:
                aa = 0
            else:
                aa = sum(aa)/len(aa)
            price = round(100*(bb+aa)/2)/100
            # print(price)

            if not t_["key"] in self.dic_share_prices:
                self.dic_share_prices["n__"] = 0
                self.dic_share_prices[t_["key"]] = {}

                d = datetime.datetime.fromtimestamp(t_["BOOK_TIME"] / 999.999451)
                h = d.hour
                m = d.minute
                date_ = d.date()
                time_ = str(date_) + ":" + str(h) + ":" + str(m)

                self.dic_share_prices[t_["key"]]["BOOK_TIME"] = time_
                self.dic_share_prices[t_["key"]]["LOW_PRICE"] = price
                self.dic_share_prices[t_["key"]]["HIGH_PRICE"] = price
                self.dic_share_prices[t_["key"]]["OPEN_PRICE"] = price
                self.dic_share_prices[t_["key"]]["CLOSE_PRICE"] = 0
            self.dic_share_prices[t_["key"]]["CLOSE_PRICE"] = price

            if price < self.dic_share_prices[t_["key"]]["LOW_PRICE"]:
                self.dic_share_prices[t_["key"]]["LOW_PRICE"] = price
            elif price > self.dic_share_prices[t_["key"]]["LOW_PRICE"]:
                self.dic_share_prices[t_["key"]]["HIGH_PRICE"] = price
        # print(self.dic_share_prices)
        n__ = self.dic_share_prices["n__"]
        n__ += 1
        self.dic_share_prices["n__"] = n__
        if n__ < 2:
            return
        n__ = self.dic_share_prices.pop("n__")

        print("="*30)
        print("processed nasdaq_order_book_handler")
        print("-"*20)
        print(self.dic_share_prices)
        print("="*50)

        return

        try:
            msg = json.dumps(self.dic_share_prices)
            print(json.dumps(self.dic_share_prices, indent=4))
            response_ = {
                'msg': msg,
                'type': "data_received_nasdaq_order_book"
            }
            channel_layer = get_channel_layer()
            await channel_layer.group_send(
                'option1', {
                    'type': 'chat_message',
                    'text': json.dumps(response_)
                })
        except Exception as ex:
            pass
            # print(ex)
        self.dic_share_prices = {}

    async def option_order_book_handler(self, msg):
        print("="*50)
        print("option_order_book_handler")
        print("="*50)
        print(msg)

    async def level_one_option_handler(self, msg):
        print("="*50)
        print("level_one_option_handler")
        print("="*50)
        print(msg)

    async def timesale_options(self, msg):
        print("="*50)
        print("timesale_options_subs")
        print("="*50)
        print(msg)

    # for now this is the only function I use.
    async def chart_equity_handler(self, msg):
        # print("="*100)
        # print("chart_equity_handler")
        # await self.log_debug_async("1000 - chart_equity_handler")
        # print("-"*40)
        # print(msg)
        # print("="*100)
        dic = {}
        for t_ in msg["content"]:
            # d = datetime.datetime.fromtimestamp(t_["CHART_TIME"] / 999.999451)
            # h = d.hour
            # m = d.minute
            # # date_ = d.date()
            # time_ = str(d.day) + " " + calendar.month_abbr[d.month] + " " + str(d.year)+" "
            # time_ += self.add_zero(str(h)) + ":" + self.add_zero(str(m)) + " GMT"

            time_ = t_["CHART_TIME"]

            # '06 Nov 1994 20:49 GMT'
            o = t_["OPEN_PRICE"]
            h = t_["HIGH_PRICE"]
            l = t_["LOW_PRICE"]
            c = t_["CLOSE_PRICE"]
            v = round(t_["VOLUME"])
            dic[t_["key"]] = [time_, o, h, l, c, v]
        try:
            response_ = {
                'msg': dic,
                'type': "data_received_chart_equity"
            }

            # print(response_)

            await self.log_debug_async("1100 - chart_equity_handler - sending response")

            channel_layer = get_channel_layer()
            await channel_layer.group_send(
                'option1', {
                    'type': 'chat_message',
                    'text': json.dumps(response_)
                })
            await self.record_data_received_chart_equity(dic)
        except Exception as ex:
            pass
            # print(ex)

    @database_sync_to_async
    def record_data_received_chart_equity(self, dic):
        for k in dic:
            ll = dic[k]
            try:
                r = XBRLRealEquityPrices.objects.create(ticker=k, t=ll[0], o=ll[1], h=ll[2], l=ll[3], c=ll[4], v=ll[5])
            except Exception as ex:
                pass

    def get_archived_data_from_db(self):
        msg_ = {}
        for q in XBRLRealEquityPrices.objects.all():
            try:
                if q.ticker not in msg_:
                    msg_[q.ticker] = []
                point_data = {'x': q.t, 'o': float(q.o), 'h': float(q.h), 'l': float(q.l), 'c': float(q.c)};
                msg_[q.ticker].append(point_data)
            except Exception as ex:
                log_debug("DataPointError: " + str() + str(ex))
        log_debug("get_archived_data_from_db 222")
        return msg_

    async def level_one_equity_handler(self, msg):
        print("="*50)
        print("level_one_equity_handler")
        print("-"*40)
        print(msg)
        print("="*50)

    # ---
    def activate_several_stream(self, dic):
        services = eval(dic)
        asyncio.run(self.read_several_stream(services))

    # -------------------------------------------------
    def place_order(self, dic):
        dic = json.loads(dic)
        print("="*100)
        print(dic)
        print("-"*10)
        print(dic["dic"]["q"])
        print("-"*50)

        order = {
            "orderStrategyType": "TRIGGER",
            "orderType": "LIMIT",
            "quantity": dic["dic"]["q"],
            "price": dic['limits']['sell_limit'],
            "duration": "DAY",
            "session": "NORMAL",
            "complexOrderStrategyType": "CUSTOM",
            "orderLegCollection": [],
            "childOrderStrategies": [
                {
                    "orderStrategyType": "SINGLE",
                    "orderType": "LIMIT",
                    "quantity": dic["dic"]["q"],
                    "price": dic['limits']['buy_limit'],
                    "duration": "GOOD_TILL_CANCEL",
                    "session": "NORMAL",
                    "complexOrderStrategyType": "CUSTOM",
                    "orderLegCollection": []
                }
            ]
        }
        # print(order)
        dic = dic["order"]
        for type_ in dic:
            if type_ in ["put", "call"]:
                for s_ in dic[type_]:
                    if dic[type_][s_]["q"] > 0:
                        order["orderLegCollection"].append(
                            {"instrument": {"assetType": "OPTION", "symbol": s_},
                             "instruction": "SELL_TO_OPEN", "quantity": dic[type_][s_]["q"]})
                        order["childOrderStrategies"][0]["orderLegCollection"].append(
                            {"instrument": {"assetType": "OPTION", "symbol": s_},
                             "instruction": "BUY_TO_CLOSE", "quantity": dic[type_][s_]["q"]})
                    else:
                        order["orderLegCollection"].append(
                            {"instrument": {"assetType": "OPTION", "symbol": s_},
                             "instruction": "BUY_TO_OPEN", "quantity": -dic[type_][s_]["q"]})
                        order["childOrderStrategies"][0]["orderLegCollection"].append(
                            {"instrument": {"assetType": "OPTION", "symbol": s_},
                             "instruction": "SELL_TO_CLOSE", "quantity": -dic[type_][s_]["q"]})

        print("-1"*50)
        print(order)
        print("-1"*50)
        return {'data': order}

    def place_order_test(self, dic):
        dic = json.loads(dic)
        print("="*100)
        print(dic)
        print("-"*50)

        # from tda.orders.equities import equity_buy_limit
        eq = orders.equities.equity_buy_limit(dic["ticker"], 2, 1250.0).set_duration(Duration.GOOD_TILL_CANCEL).set_session(Session.SEAMLESS).build()
        print(eq)
        print(type(eq))
        print("-"*50)
        o = orders.options.option_buy_to_open_limit("TSLA_022522C985", 1, 5).build()
        print(o)
        print("-"*50)
        ca = orders.options.bear_call_vertical_open("TSLA_022522C985", "TSLA_022522C990", 5, 1.3)
        ca.set_duration(Duration.GOOD_TILL_CANCEL)
        ca.set_quantity(10)
        print(ca.build())
        print("-"*50)
        # c = self.get_stream_client()
        # c.place_order(self.account_id, ca.build())
        return {'data': order}

    def place_order_(self, dic):
        print(dic)
        print("="*50)
        # a = orders.options.option_buy_to_open_market("TSLA_021122P840", 1)
        ca = orders.options.bear_call_vertical_open("TSLA_021122C985", "TSLA_021122C990", 100, 1.3)
        ca.set_duration(Duration.GOOD_TILL_CANCEL)
        # ac.set_quantity(10)
        print(ca.build())
        print("-"*30)
        cb = orders.options.bear_call_vertical_close("TSLA_021122C985", "TSLA_021122C990", 100, 0.8)
        cb.set_duration(Duration.GOOD_TILL_CANCEL)
        print(cb.build())

        print("="*40)
        pa = orders.options.bull_put_vertical_open("TSLA_021122P845", "TSLA_021122P850", 10, 1.1)
        pa.set_duration(Duration.GOOD_TILL_CANCEL)
        print(pa.build())
        print("-"*30)
        pb = orders.options.bull_put_vertical_close("TSLA_021122P845", "TSLA_021122P850", 10, 0.95)
        pb.set_duration(Duration.GOOD_TILL_CANCEL)
        print(pb.build())
        # self.client.place_order(self.account_id, order_spec=a.build())

        print("="*100)
        cab = orders.common.first_triggers_second(ca, cb)
        print(cab.build())
        print("="*100)
        pab = orders.common.first_triggers_second(pa, pb)
        print(pab.build())
        print("="*100)

        return {'status': 'ok'}

    # def place_order_test(self, dic):
    #     dic = json.loads(dic)
    #     print("="*100)
    #     print(dic)
    #     print("-"*50)
    #
    #     o = orders.options.option_buy_to_open_limit("TSLA_022522C985", 1, 5).build()
    #     print(o)
    #     print("-"*50)
    #     ca = orders.options.bear_call_vertical_open("TSLA_022522C985", "TSLA_022522C990", 5, 1.3)
    #     ca.set_duration(Duration.GOOD_TILL_CANCEL)
    #     ca.set_quantity(10)
    #     print(ca.build())
    #     print("-"*50)
    #
    #     # c = self.get_stream_client()
    #     # c.place_order(self.account_id, ca.build())
    #
    #     return {'data': order}

    def account_test(self, dic):
        # print(dic)
        # print(dic['app'])
        result = {}
        try:
            c = self.get_client()

            so_ = {
                "orderType": "NET_CREDIT",
                "session": "NORMAL",
                "quantity": 1,
                "price": "1.20",
                "duration": "DAY",
                "orderStrategyType": "SINGLE",
                "orderLegCollection": [
                    {
                        "instruction": "BUY_TO_OPEN",
                        "quantity": 1,
                        "instrument": {
                            "symbol": "TSLA_022522C990",
                            "assetType": "OPTION"
                        }
                    },
                    {
                        "instruction": "SELL_TO_OPEN",
                        "quantity": 1,
                        "instrument": {
                            "symbol": "TSLA_022522C985",
                            "assetType": "OPTION"
                        }
                    }
                ]
            }
            try:
                so = c.create_saved_order(self.account_id, so_)
            except Exception as ex:
                print(ex)
            try:
                co = c.get_saved_orders_by_path(self.account_id)
            except Exception as ex:
                print(ex)
        except Exception as ex:
            print(ex)

        return {'data': result}

    def create_option_symbol(self, ticker):
        symbol = ticker
        return symbol


class StockPrices(object):
    def __init__(self):
        pass

    def record_option_strategy(self, dic):
        print("9044-40 input dic: \n", dic, "\n"+"-"*30)
        app_ = dic["app"]
        ticker = dic["ticker"]
        idx = int(dic["idx"])
        strike = dic["strike"]
        stock_price = dic["stock_price"]
        strategy_price = dic["strategy_price"]
        call_strategy_price = dic["call_strategy_price"]
        put_strategy_price = dic["put_strategy_price"]

        model_c = apps.get_model(app_label=app_, model_name="XBRLCompanyInfo")
        model_s = apps.get_model(app_label=app_, model_name="TwoSpreadStrategy")
        c = model_c.objects.get(ticker=ticker)
        # print("c=", c)
        count = model_s.objects.filter(company=c, strike=strike, is_open_position=True).all().count()
        # print("count=", count)
        if count == 0:
            # print("record strike for", c, strike)
            s, is_created = model_s.objects.get_or_create(company=c, strategy_idx=idx, strike=strike)
            s.stock_price = stock_price
            s.strategy_price = strategy_price
            s.call_strategy_price = call_strategy_price
            s.put_strategy_price = put_strategy_price
            s.save()
            # print(s)

        dic = {'data': "ok"}
        # dic = {}
        # print("9099 output dic: \n", dic, "\n"+"="*30)
        return dic

    def record_option_strategy_detail(self, dic):
        print("9077-70-7 input dic: \n", dic, "\n"+"-"*100)
        app_ = dic["app"]
        id_ = int(dic["id"])
        idx = int(dic["idx"])
        seconds = int(dic["seconds"])
        stock_price = dic["stock_price"]
        strategy_price = dic["strategy_price"]
        put_strategy_price = dic["put_strategy_price"]
        call_strategy_price = dic["call_strategy_price"]

        model_t = apps.get_model(app_label=app_, model_name="TwoSpreadStrategy")
        tw = model_t.objects.get(id=id_)
        # print(tw)
        model_td = apps.get_model(app_label=app_, model_name="TwoSpreadStrategyDetails")
        c, is_created = model_td.objects.get_or_create(two_spread_strategy=tw, idx = idx, seconds=seconds)
        c.stock_price = stock_price
        c.strategy_price = strategy_price
        c.call_strategy_price = call_strategy_price
        c.put_strategy_price = put_strategy_price
        c.save()
        print("c", c)

        dic = {'data': "ok"}
        # dic = {}
        # print("9099 output dic: \n", dic, "\n"+"="*30)
        return dic

    def get_option_strategies(self, dic):
        print("9055-50 input dic: \n", dic, "\n"+"-"*30)
        app_ = dic["app"]
        ticker_ = ""
        try:
            ticker_ = dic["ticker"]
        except Exception as ex:
            pass

        s = ""
        is_all_ = "True"
        try:
            is_all_ = dic["is_all"]
            if is_all_ != "all":
                if is_all_ == "false":
                    is_all_ = "False"
                else:
                    is_all_ = "True"
                s += ".filter(is_open_position=" + is_all_ + ")"
        except Exception as ex:
            s += ".filter(is_open_position=" + is_all_ + ")"

        if ticker_ !="":
            s += ".filter(company__ticker='"+ticker_+"')"
        if s == "":
            s = ".all"
        s = "model_s.objects" + s
        model_s = apps.get_model(app_label=app_, model_name="TwoSpreadStrategy")
        # print(s)
        objs = eval(s)
        strategies = {}
        for o in objs:
            # print(o.company.ticker, o.strategy_idx, o.strike, o.id)
            if o.company.ticker not in strategies:
                strategies[o.company.ticker] = {}
            # print(o.strike)
            strategies[o.company.ticker][int(o.strike)] = {"strategy_idx": o.strategy_idx, "id": o.id}
        # print(strategies)
        dic = {'data': "ok", "strategies": strategies}
        # dic = {}
        # print("9099 output dic: \n", dic, "\n"+"="*30)
        return dic

    def get_option_strategies_detail(self, dic):
        print("9066-60 input dic: \n", dic, "\n"+"-"*30)
        app_ = dic["app"]
        ticker = dic["ticker"]
        strategy_id = dic["strategy_id"]
        model_ = apps.get_model(app_label=app_, model_name="TwoSpreadStrategyDetails")
        objs = model_.objects.filter(two_spread_strategy__company__ticker=ticker,
                                     two_spread_strategy__strategy_idx=strategy_id)
        dic = {}
        for o in objs:
            m = round(o.idx/10000)
            if m not in dic:
                dic[m] = {"x":[], "y":[]}
            dic[m]["x"].append(float(o.stock_price))
            dic[m]["y"].append(float(o.strategy_price))

        dic = {'data': "ok", "dic": dic}
        # print("9099 output dic: \n", dic, "\n"+"="*30)
        return dic

    def analyze_prices_minutes(self, dic):
        # print("9010 input dic: \n", dic, "\n"+"-"*30)
        # ^GSPC
        ticker = dic["ticker"]
        num_of_bars = 99999999
        # num_of_bars = int(dic["num_of_bars"])
        # print("ticker", ticker)

        sp = StockPricesMinutes.objects.filter(company__ticker=ticker)[:num_of_bars]
        # df = pd.DataFrame(list(sp.values()))
        df = pd.DataFrame(list(sp.values("idx", "open", "close", "high", "low", "volume")))
        # print("AAA\n", df)
        #
        df['idx'] = pd.to_datetime(df['idx'], format='%Y%m%d%H%M')
        df.set_index('idx', inplace=True)
        df.sort_index(inplace=True)
        df['A'] = df.index
        mux = pd.MultiIndex.from_arrays([df['A'].dt.date, df['A'].dt.time], names=('date', 'time'))
        df = df.set_index(mux).drop('A', 1)
        # print("AAA Data\n", "-"*50, "\n", df)

        dfg = df.groupby(level = 0)['open'].first()
        idx = dfg.index.get_level_values(0)
        # print("AAA Original data collected - idx\n", idx, "\n", len(idx))
        # print(df.index.min()[0], df.index.max()[0])

        dfi = pd.date_range(start=df.index.min()[0], end=df.index.max()[0] + timedelta(days=1), freq='T')
        dfb = pd.DataFrame({'a': range(len(dfi))}, index=dfi)
        # print("dfb11\n", dfb)
        dfb = dfb.between_time('09:31', '15:59')
        #
        dfb['A'] = dfb.index
        mux = pd.MultiIndex.from_arrays([dfb['A'].dt.date, dfb['A'].dt.time], names=('date', 'time'))
        dfb = dfb.set_index(mux).drop('A', 1)
        # print("BBB dfb33\n", dfb)
        r = pd.bdate_range(dfb.index.min()[0], dfb.index.max()[0])
        dfb = dfb[pd.DatetimeIndex(dfb.index.get_level_values(0)).isin(r).astype(int) > 0]
        # print("BBB Business DateTime\n", "-"*50, "\n", dfb)
        #
        dfm = dfb.join(df, on=['date', 'time'], how='left')
        # print("CCC Merged Data\n", "-"*50, "\n", dfm)
        # dfs = dfm.copy() # .shift(periods=1)
        dfs = dfm.shift(periods=1)
        # print("dfs\n", dfs)
        null_dfs = dfs.isnull()
        dfs_null = dfm[null_dfs['open']]
        dfs_null['a'] = 1
        # print("DDD dfs_null\n", dfs_null)
        dfs_null = pd.DataFrame(dfs_null.groupby(level = 0)['a'].sum())
        dfs_null = dfs_null[dfs_null["a"] > 20]
        # print("DDD1 dfs_null\n", dfs_null, dfs_null['a'].min(), dfs_null['a'].max())
        # print("DD2 dfs_null.index\n", dfs_null.index)
        dates_to_remove_s = dfs_null.index
        # print("DDD3 dates_to_remove_s\n", dates_to_remove_s)
        dfs_ = dfs[pd.DatetimeIndex(dfs.index.get_level_values(0)).isin(dates_to_remove_s).astype(int) == 0].drop('a', 1)
        # print("DDD3 dfs_\n", , dfs_)
        dfs_g = dfs_.groupby(level = 0)['open'].first()
        idxs_ = dfs_g.index.get_level_values(0)
        # print("EEE Business DateTime\n", idxs_, "\n", len(idxs_))
        #
        logic = {'open': 'first',
                 'high': 'max',
                 'low': 'min',
                 'close': 'last',
                 'volume': 'sum'}
        df = dfs_.groupby(level=0).agg(logic)
        # print("Agg WWW\n", df)

        df['up'] = df['high']-df['open']
        df['dn'] = df['open']-df['low']
        dr = 0
        df_ = df[((df["up"] >= dr) | (df["dn"] >= dr)) == True]
        df_['mm'] = df_[["up", "dn"]].max(axis=1)
        df_['mm_ch'] = df_.apply(lambda x: float(x["up"]) if x["up"] > x["dn"] else float(-x["dn"]), axis=1)
        c_all = int(df_["mm"].count())

        res_dic = {"date":[], "mm":[], "mm_ch":[], "pr_over":{"c":[], "p":[]}}

        for i in range(7, 50, 1):
            c = round(100*int(df_[df_["mm"] > i]["mm"].count())/c_all)
            res_dic["pr_over"]["c"].append(i)
            res_dic["pr_over"]["p"].append(c)

        n = 0
        for index, row in df_.iterrows():
            n += 1
            # i_ = int(str(index).split("-"))
            # i_ = i_[0]+i_[1]+i_[2]
            # res_dic["date"].append(i_)
            res_dic["date"].append(n)
            res_dic["mm"].append(row['mm'])
            res_dic["mm_ch"].append(row['mm_ch'])
        # print(res_dic)

        res_ = {"begin date":str(df_.index[0]),"end date":str(df_.index[-1]),
                "count":round(100*df_['mm'].count())/100,
                "min":round(100*df_['mm'].min())/100,
                "max": round(100*df_['mm'].max())/100,
                "mean":round(100*df_['mm'].mean())/100,
                "median":round(100*df_['mm'].median())/100,
                "std":round(100*df_['mm'].std())/100}

        dic = {'data': "ok", "result": res_, "res_dic":res_dic}
        # dic = {}
        # print("9099 output dic: \n", dic, "\n"+"="*30)
        return dic

    def get_prices_minutes(self, dic):
        # print("9010 input dic: \n", dic, "\n"+"-"*30)
        # ^GSPC
        ticker = dic["ticker"]
        num_of_bars = int(dic["num_of_bars"])
        print(ticker)
        result = {"idx": [], "open": [], "high": [], "low": [], "close": [], "volume": [], "dividends": [], "stock_splits": []}

        sp = StockPricesMinutes.objects.filter(company__ticker=ticker)[:num_of_bars]
        # print(sp)
        for s in sp:
            # print("-1"*50)
            # print(s)
            # print("-2"*50)
            result["idx"].append(s.idx)
            result["open"].append(float(s.open))
            result["high"].append(float(s.high))
            result["low"].append(float(s.low))
            result["close"].append(float(s.close))
            result["volume"].append(int(s.volume))
        dic = {'data': result}
        # print("9099 output dic: \n", dic, "\n"+"="*30)
        return dic

    def get_prices_days(self, dic):
        # print("9010 input dic: \n", dic, "\n"+"-"*30)
        # ^GSPC
        ticker = dic["ticker"]
        num_of_bars = int(dic["num_of_bars"])
        # print(ticker)
        result = {"idx": [], "open": [], "high": [], "low": [], "close": [], "volume": [], "dividends": [], "stock_splits": []}

        sp = StockPricesDays.objects.filter(company__ticker=ticker)[:num_of_bars]
        # print(sp)
        for s in sp:
            # print("-1"*50)
            # print(s)
            # print("-2"*50)
            result["idx"].append(s.idx)
            result["open"].append(float(s.open))
            result["high"].append(float(s.high))
            result["low"].append(float(s.low))
            result["close"].append(float(s.close))
            result["volume"].append(int(s.volume))
        dic = {'data': result}
        # print("9099 output dic: \n", dic, "\n"+"="*30)
        return dic

    def remove_prices_minutes_duplicates(self, dic):

        def remove_duplicates(stockpricesminutes, c):
            try:
                duplicates = stockpricesminutes.objects.filter(company=c).values('idx').annotate(idx_count=Count('idx')).filter(idx_count__gt=1)
                # print(duplicates)
                for item in duplicates:
                    stockpricesminutes.objects.filter(company=c, idx=item['idx']).all()[0].delete()
            except Exception as ex:
                print("Error 22-22-22-22-1", ex)

        # print("9060-60 input dic: \n", dic, "\n"+"-"*30)
        # print(dic)
        # print(dic["ticker"])
        l_f = str(dic["letter_from"])
        l_t = str(dic["letter_to"])

        # StockPricesMinutes.truncate();

        watch_list = ETFWatchLists.objects.all()
        for w in watch_list:
            companies = XBRLCompanyInfo.objects.filter(etfwatchlist=w).all()
            for c in companies:
                if l_t >= c.company_letter >= l_f or (c.ticker == "$SPX.X" and w == "HighV"):
                    ticker = c.ticker
                    if ticker == "$SPX.X":
                        ticker = "^GSPC"
                    # if ticker == "CMS":
                    print("90151 company info: ", w.symbol, ticker, c.company_letter, "\n"+"-"*30)
                    log_debug(ticker)
                    remove_duplicates(StockPricesMinutes, c)


        dic = {'data': {"status": "ok"}}
        # print("9099 output dic: \n", dic, "\n"+"="*30)
        return dic

    def update_prices_minutes(self, dic):

        def get_data(stockpricesminutes, c, idx_, n=0):
            try:
                # print("-5"*50)
                sp, is_created = stockpricesminutes.objects.get_or_create(company=c, idx=idx_)
                # print("-6"*50)
                # print(is_created)
                # print("-7"*50)
                if is_created:
                    sp.open = open_
                    sp.high = high_
                    sp.low = low_
                    sp.close = close_
                    sp.volume = volume_
                    sp.dividends = dividends_
                    sp.stock_splits = stock_splits_
                    # print("-8"*50)
                    sp.save()
                    # print("-9"*50)
            except Exception as ex:
                # print("Error 9088-77-66 idx_=", idx_, ex)
                spm = stockpricesminutes.objects.filter(company=c, idx=idx_).all()
                spm[0].delete()
                if n==0:
                    get_data(stockpricesminutes, c, idx_, n=1)
                else:
                    print("Error 22-22-22-22-1", idx_, ex)

        print("9010 input dic: \n", dic, "\n"+"-"*30)
        # print(dic)
        # print(dic["ticker"])
        l_f = str(dic["letter_from"])
        l_t = str(dic["letter_to"])
        w_ = int(dic["numer_of_weeks"])
        d_ = int(dic["numer_of_days"])  # d_ = 7
        # StockPricesMinutes.truncate();

        watch_list = ETFWatchLists.objects.all()
        for w in watch_list:
            companies = XBRLCompanyInfo.objects.filter(etfwatchlist=w).all()
            for c in companies:
                # print("="*50)
                # print("="*50)
                # print("="*50)
                if l_t >= c.company_letter >= l_f or (c.ticker == "$SPX.X" and w == "HighV"):
                        # print("="*50)
                    # print("="*50)
                    # print("="*50)
                    ticker = c.ticker
                    if ticker == "$SPX.X":
                        ticker = "^GSPC"

                    print("90151 company info: \n", w.symbol, ticker, c.company_letter, "\n"+"-"*30)
                    obj = yf.Ticker(ticker)
                    i = w_  # 5 for 1m, 45 for 1d
                    nn = 0
                    while i > 0:
                        if nn == 0:
                            date_e = (datetime.datetime.now() + datetime.timedelta(days=-((i-1)*d_)))
                            date_b = (datetime.datetime.now() + datetime.timedelta(days=-(i*d_)))
                            nn = 1
                        else:
                            date_b = (date_e + datetime.timedelta(days=1))
                            date_e = (date_b + datetime.timedelta(days=d_))
                        i -= 1
                        date_b_ = date_b.date()
                        date_e_ = date_e.date()
                        end_ = str(date_e_.year)+"-"+str(date_e_.month)+"-"+str(date_e_.day)
                        beg_ = str(date_b_.year)+"-"+str(date_b_.month)+"-"+str(date_b_.day)
                        hist = obj.history(interval="1m", start=beg_, end=end_)
                        # n_tail=10
                        hist = hist[:-1]  # .tail(n_tail+1)
                        # print("-1"*50)
                        # print(hist)
                        # print("-2"*50)
                        for index, row in hist.iterrows():
                            idx = pd.Timestamp(index)
                            idx_ = idx.year*100000000+idx.month*1000000+idx.day*10000+idx.hour*100+idx.minute
                            # print("-2"*50)
                            # print(idx_)
                            if row["Volume"] > 0:
                                # print("-3"*50)
                                # print(row["Volume"])
                                # print("-4"*50)
                                open_ = round(100*row["Open"])/100
                                high_ = round(100*row["High"])/100
                                low_ = round(100*row["Low"])/100
                                close_ = round(100*row["Close"])/100
                                volume_ = round(row["Volume"])
                                dividends_ = round(100*row["Dividends"])/100
                                stock_splits_ = round(100*row["Stock Splits"])/100
                                get_data(StockPricesMinutes, c, idx_, n=0)

        dic = {'data': {"status": "ok"}}
        # print("9099 output dic: \n", dic, "\n"+"="*30)
        return dic

    def update_prices_days(self, dic):
        print("9015 input dic: \n", dic, "\n"+"-"*30)
        l_f = str(dic["letter_from"])
        l_t = str(dic["letter_to"])
        y_ = int(dic["numer_of_years"]) # I use w for years
        d_ = int(dic["numer_of_days"])  # d_ = 7

        watch_list = ETFWatchLists.objects.all()
        for w in watch_list:
            companies = XBRLCompanyInfo.objects.filter(etfwatchlist=w).all()
            for c in companies:
                if l_t >= c.company_letter >= l_f or (c.ticker == "$SPX.X" and w == "HighV"):
                    ticker = c.ticker
                    if ticker == "$SPX.X":
                        ticker = "^GSPC"
                    print("90152 company info: \n", w.symbol, ticker, c.company_letter, "\n"+"-"*30)
                    obj = yf.Ticker(ticker)
                    date_e = datetime.datetime.now()
                    date_b = (datetime.datetime.now() + datetime.timedelta(days=-d_))
                    date_b = (date_b + relativedelta(years = -y_))
                    date_b_ = date_b.date()
                    date_e_ = date_e.date()
                    # print("9016 dates: \n", date_b_, date_e_, "\n"+"-"*30)
                    end_ = str(date_e_.year)+"-"+str(date_e_.month)+"-"+str(date_e_.day)
                    beg_ = str(date_b_.year)+"-"+str(date_b_.month)+"-"+str(date_b_.day)
                    hist = obj.history(interval="1d", start=beg_, end=end_)

                    # hist = hist[:-1]
                    # print("9016 hist: \n", hist, "\n"+"-"*30)
                    # print("-2"*50)
                    for index, row in hist.iterrows():
                        idx = pd.Timestamp(index)
                        idx_ = idx.year*100000000+idx.month*1000000+idx.day*10000+idx.hour*100+idx.minute
                        # print("9017 idx_: \n", idx_, row, "\n"+"-"*30)
                        if row["Volume"] > 0:
                            # print(row["Volume"])
                            # print("9019 dates: \n", round(row["Volume"]/100), "\n"+"-"*30)
                            open_ = round(100*row["Open"])/100
                            high_ = round(100*row["High"])/100
                            low_ = round(100*row["Low"])/100
                            close_ = round(100*row["Close"])/100
                            volume_ = round(row["Volume"]/100)
                            dividends_ = round(100*row["Dividends"])/100
                            stock_splits_ = round(100*row["Stock Splits"])/100
                            try:
                                sp, is_created = StockPricesDays.objects.get_or_create(company=c, idx=idx_)
                                if is_created:
                                    sp.open = open_
                                    sp.high = high_
                                    sp.low = low_
                                    sp.close = close_
                                    sp.volume = volume_
                                    sp.dividends = dividends_
                                    sp.stock_splits = stock_splits_
                                    sp.save()
                            except Exception as ex:
                                print(ex)

        dic = {'data': {"status": "ok"}}
        # print("9099 output dic: \n", dic, "\n"+"="*30)
        return dic

    def create_return_statistics_minutes(self, dic):
        # print("create_return_statistics_minutes\n90600")
        # print(dic)
        ticker = dic["ticker"]
        # print("90601  "+ticker)

        def get_y(x):
            y = round(x/10000)
            yd = x-y
            return y

        def get_yd(x):
            y = round(x/10000)*10000
            yd = x-y
            return yd

        sp = StockPricesMinutes.objects.filter(company__ticker=ticker).all().values('idx', 'open', 'high',
                                                                                    'low', 'close', 'volume')
        # print(sp)
        df = pd.DataFrame(list(sp))
        # print(df)
        df["y"] = df["idx"].apply(get_y)
        df["yd"] = df["idx"].apply(get_yd)
        df = df[['y', 'yd', 'open', 'high', 'low', 'close', 'volume']]
        # print('-'*10, df, '-'*10)
        try:
            table = pd.pivot_table(df, values=['open', 'high', 'low', 'close', 'volume'], index=['yd'],
                                   columns=['y'],
                                   aggfunc={'open': np.sum,
                                            'high': np.sum,
                                            'low': np.sum,
                                            'close': np.sum,
                                            'volume': np.sum})
            table.fillna(method='bfill', inplace=True)
        except Exception as ex:
            print("10001"+str(ex))
        try:
            table1 = ((table.div(table.iloc[0])-1).astype(float)).round(4)
            table1 = table1.reset_index().to_dict(orient='list')
            table11 = {}
            for k in table1:
                if k[0] not in ['index', 'yd']:
                    if k[0] not in table11:
                        table11[k[0]] = {}
                    table11[k[0]][k[1]] = table1[k]
                    table11[k[0]][k[1]] = [0 if pd.isna(x) else x for x in table11[k[0]][k[1]]]
                elif k[0] == 'yd':
                    table11[k[0]] = table1[k]
        except Exception as ex:
            print("err10003 k="+ "\n" +str(ex))
        try:
            table2 = ((1/table.div(table.iloc[-1])-1).astype(float)).round(4)
            table2 = table2.reset_index().to_dict(orient='list')
            table22 = {}
            for k in table2:
                if k[0] not in ['index', 'yd']:
                    if k[0] not in table22:
                        table22[k[0]] = {}
                    table22[k[0]][k[1]] = table2[k]
                    # It means: if nan I treat it as no change.
                    table22[k[0]][k[1]] = [0 if pd.isna(x) else x for x in table22[k[0]][k[1]]]
                elif k[0] == 'yd':
                    table22[k[0]] = table2[k]
        except Exception as ex:
            print("10005"+str(ex))
        result = {"table1": table11, "table2": table22}
        return result

    def create_return_statistics_minutes_bu(self, dic):
        ticker = dic["ticker"]
        print("90600  "+ticker)

        def get_y(x):
            y = round(x/10000)
            yd = x-y
            return y

        def get_yd(x):
            y = round(x/10000)*10000
            yd = x-y
            return yd

        sp = StockPricesMinutes.objects.filter(company__ticker=ticker).all().values('idx', 'close')
        # print(sp)
        df = pd.DataFrame(list(sp))
        ndf = df.to_numpy(dtype=float)
        print(ndf.shape)
        print(ndf)

        # print(df)
        df["y"] = df["idx"].apply(get_y)
        df["yd"] = df["idx"].apply(get_yd)
        df = df[['y', 'yd', 'close']]
        # print('-'*10)
        # print('-'*10)
        # print(df)
        # print('-'*10)
        # print('-'*10)
        try:
            table = pd.pivot_table(df, values='close', index=['yd'], columns=['y'], aggfunc=np.sum)
            table.fillna(method='bfill', inplace=True)
        except Exception as ex:
            print("10001"+str(ex))
        # print('-'*10)
        # print('-'*10)
        # print('table')
        # print(table)
        # print('-'*10)
        # print('-'*10)
        try:
            table1 = ((table.div(table.iloc[0])-1).astype(float)).round(4)
        except Exception as ex:
            print("10002"+str(ex))
        # print('table1')
        # print(table1)
        # print('-'*10)
        try:
            table1 = table1.reset_index().to_dict(orient='list')
        except Exception as ex:
            print("10003"+str(ex))
        # print('table1')
        # print(table1)
        # print('-'*10)
        try:
            table2 = ((1/table.div(table.iloc[-1])-1).astype(float)).round(4)
        except Exception as ex:
            print("10004"+str(ex))
        # print('table2')
        # print(table2)
        # print('-'*10)
        try:
            table2 = table2.reset_index().to_dict(orient='list')
        except Exception as ex:
            print("10005"+str(ex))
        # print('table2')
        # print(type(table2))
        # print(table2)
        # print("end of data")
        result = {"table1": table1, "table2": table2}
        # print(result)
        # print('table1')
        # print(table1[20221110])
        # print("9099 output dic: \n", dic, "\n"+"="*30)
        return result

    def get_prices_to_excel(self, dic):
        # print("9010 input dic: \n", dic, "\n" + "-" * 30)
        dic = dic["dic"]
        idx_from = int(dic["from_date"]+"0000")
        idx_to = int(dic["to_date"]+"0000")
        # print("9015 : \nfrom: ", idx_from, "\nto: ", idx_to, "\n"+"-"*30)

        # ^GSPC
        ticker = dic["ticker"]
        # print(ticker)
        result = {"idx": [], "open": [], "high": [], "low": [], "close": [], "volume": [], "dividends": [], "stock_splits": []}

        sp = StockPricesMinutes.objects.filter(company__ticker=ticker, idx__gte=idx_from, idx__lte=idx_to)
        # print(sp)
        for s in sp:
            # print("-1"*50)
            # print(s)
            # print("-2"*50)
            result["idx"].append(s.idx)
            result["open"].append(float(s.open))
            result["high"].append(float(s.high))
            result["low"].append(float(s.low))
            result["close"].append(float(s.close))
            result["volume"].append(int(s.volume))
        dic = {'data': result}
        print("9099 output dic: \n", dic, "\n"+"="*30)
        return dic

    # https://analyzingalpha.com/yfinance-python
    def explore_yfinance(self, dic):
        print("9010 input dic: \n", dic, "\n"+"-"*30)
        ticker = dic["ticker"]
        obj = yf.Ticker(ticker)
        date_e = datetime.datetime.now()
        date_b = datetime.datetime.now() + datetime.timedelta(days=-5)
        date_b_ = date_b.date()
        date_e_ = date_e.date()
        end_ = str(date_e_.year)+"-"+str(date_e_.month)+"-"+str(date_e_.day)
        beg_ = str(date_b_.year)+"-"+str(date_b_.month)+"-"+str(date_b_.day)
        hist = obj.history(interval="1m", start=beg_, end=end_)
        hist = hist[:-1]  # .tail(n_tail+1)
        print("-1"*50)
        print("hist")
        print(hist)
        print("-2"*50)

        print('obj.calendar')
        print(obj.calendar)
        print("-"*10)
        print('obj.news')
        print(obj.news)
        print("-"*10)
        print('obj.actions')
        print(obj.actions)
        print("-"*10)
        print('obj.info')
        print(obj.info)
        print("-"*10)
        dic = {'data': {"status": "ok"}}

    # https://towardsdatascience.com/principal-component-analysis-for-dimensionality-reduction-115a3d157bad
    # https://scikit-learn.org/stable/modules/generated/sklearn.linear_model.LogisticRegression.html
    def get_pca(self, dic):
        ticker = dic["ticker"]
        print(ticker)

        def get_y(x):
            y = round(x/10000)
            yd = x-y
            return y

        def get_yd(x):
            y = round(x/10000)*10000
            yd = x-y
            return yd

        sp = StockPricesMinutes.objects.filter(company__ticker=ticker).all().values('idx', 'close')
        df = pd.DataFrame(list(sp))
        df["y"] = df["idx"].apply(get_y)
        df["yd"] = df["idx"].apply(get_yd)
        df = df[['y', 'yd', 'close']]
        table = pd.pivot_table(df, values='close', index=['yd'], columns=['y'], aggfunc=np.sum)
        table.fillna(method='bfill', inplace=True)
        table1 = ((table.div(table.iloc[0])-1).astype(float)).round(4)
        table1 = table1.T
        test_size = .10
        test_size_ = int(table1.shape[0]*test_size)
        xtest = table1[table1.shape[0]-test_size_:]
        xtrain = table1[:table1.shape[0]-test_size_]
        # print('table1')
        # print(xtrain, xtest)
        # print('-'*10)
        table2 = ((1/table.div(table.iloc[-1])-1).astype(float)).round(4)
        table2 = table2.T
        ytest = table2[table2.shape[0]-test_size_:]
        ytrain = table2[:table2.shape[0]-test_size_]
        # print('table2')
        # print(ytrain, ytest)
        # print(table2)
        # print('-'*10)

        max_ = 0
        data_ = {}
        for j in range(10, table1.shape[1]):
            col = xtrain.columns[j]
            # print('col')
            # print(col)
            xtrain_np = table1.to_numpy()[:, 1:j]
            ytrain_np = table2.to_numpy()[:, j:j+1]
            x1 = table1.to_numpy()[0:1, 1:j].copy()
            # print('x1', '1\n', xtrain_np.shape, x1.shape)
            # print(j, xtrain_np, ytrain_np)
            rr = np.max(ytrain_np) - np.min(ytrain_np)
            ub = 0.1
            bins = np.array([rr*ub+np.min(ytrain_np),
                             rr*ub*2+np.min(ytrain_np),
                             rr*(1-2*ub)+np.min(ytrain_np),
                             rr*(1-ub)+np.min(ytrain_np)])
            ytrain_np = np.digitize(ytrain_np, bins)
            # print(j, ytrain_np)
            scaler = StandardScaler()
            scaler.fit(xtrain_np)
            X = scaler.transform(xtrain_np)
            # print('x1', '22\n', x1)
            x1 = scaler.transform(x1)
            # print('x1', '23\n', x1)
            np_ = min(xtrain.shape[0], xtrain_np.shape[1])

            pca = PCA(n_components=np_)
            pca.fit(X)
            # print('pca.explained_variance_ratio_')
            sg = pca.explained_variance_ratio_
            # print(sg[sg > 0.01])
            # print(sg[sg > 0.01].shape[0])
            np_ = sg[sg > 0.005].shape[0]
            if max_ < np_:
                max_ = np_
            pca = PCA(n_components=np_)
            pca.fit(X)
            sg = pca.explained_variance_ratio_
            # print('pca.explained_variance_ratio_', sg, np_)
            xtrain_pca = pca.transform(X)
            x1 = pca.transform(x1)
            data_[col] = [scaler, xtrain_pca, ytrain_np]   # , xtrain_np

            # logistic regression
            # print('logistic regression')
            # print(xtrain_pca)
            # print(ytrain_np.reshape(-1))
            # print(ytrain_np.reshape(1, -1))
            # print(ytrain_np.reshape(-1, 1))
            lr = LogisticRegression(multi_class='auto', solver='liblinear')
            # lr.fit(xtrain_pca, ytrain_np.reshape(-1))
            sgc = SGDClassifier(random_state=42)
            ytrain_np = ytrain_np.reshape(-1, 1)
            print(j, "\n", xtrain_pca.shape, ytrain_np.shape)
            # sgc.fit(xtrain_pca, ytrain_np)
            # print('logistic regression 1')
            # print(ytrain_np.T, "\n", sgc.predict(xtrain_pca))

            # print('sgc.decision_function(x1)')
            # print(sgc.decision_function(x1))
            # print(sgc.classes_)
            # print(np.argmax(sgc.decision_function(x1)))
            print('-'*10)

            ncv = 20
            # y_scores = cross_val_predict(sgc, xtrain_pca, ytrain_np, method="predict", cv=ncv)
            y_scores = cross_val_predict(lr, xtrain_pca, ytrain_np, method="predict", cv=ncv)
            print('y_scores--')
            print("actual:\n", ytrain_np.T, "\npredic\n", y_scores.reshape(1, -1), "\n", ytrain_np.T - y_scores.reshape(1,-1))
            print('y_scores--2')

            # test
            # xtest_np = table1.to_numpy()[:, :j]
            # ytest_np = table2.to_numpy()[j:j+1]
            #
            # xtest = scaler.transform(xtest_np)
            # xtest_pca = pca.transform(xtest)
            # ytest_np = np.digitize(ytest_np, bins)

        # print(data_)
        # print(max_)

        dic = {'data': {"status": "ok"}}

    # application auxiliary functions


class TF(object):
    def __init__(self):
        pass

    def mnist(self, dic):
        print("90555-5531-12\n", dic)
        from keras.datasets import mnist

        # load dataset
        (x_train, y_train), (x_test, y_test) = mnist.load_data()

        unique, counts = np.unique(y_train, return_counts=True)
        print("\nTrain labels: \n", dict(zip(unique, counts)))

        # a = np.arange(15).reshape(3, 5)
        # print("A\n", a, "\n")
        # a = np.arange(15).reshape(1, 3, 5)
        # print("B\n", a, "\n")
        # a = np.arange(15).reshape(1, 3, 5, 1)
        # print("C\n", a, "\n")
        # a = np.arange(30).reshape(1, 3, 5, 2)
        # print("D\n", a, "\n")
        # a = np.arange(30).reshape(2, 3, 5, 1)
        # print("E\n", a, "\n")
        # a = np.arange(60).reshape(2, 3, 5, 2)
        # print("F\n", a, "\n")

        result = {"status": "ok"}
        return result


class AcademyCityXBRL(object):
    def __init__(self):

        # Should remove it later on
        try:
            clear_log_debug()
        except Exception as ex:
            pass
        #
        self.PROJECT_ROOT_DIR = os.path.join(settings.WEB_DIR, "data", "corporatevaluation")
        os.makedirs(self.PROJECT_ROOT_DIR, exist_ok=True)

        self.TO_DATA_PATH = os.path.join(self.PROJECT_ROOT_DIR, "datasets")
        os.makedirs(self.TO_DATA_PATH, exist_ok=True)
        # print(self.TO_DATA_PATH)

        self.IMAGES_PATH = os.path.join(self.TO_DATA_PATH, "images")
        os.makedirs(self.IMAGES_PATH, exist_ok=True)
        # print(self.IMAGES_PATH)

        self.MODELS_PATH = os.path.join(self.TO_DATA_PATH, "models")
        os.makedirs(self.MODELS_PATH, exist_ok=True)
        # print(self.MODELS_PATH)

        self.EXCEL_PATH = os.path.join(self.TO_DATA_PATH, "excel")
        os.makedirs(self.EXCEL_PATH, exist_ok=True)
        # print(self.EXCEL_PATH)

        self.TEXT_PATH = os.path.join(self.TO_DATA_PATH, "text")
        os.makedirs(self.TEXT_PATH, exist_ok=True)
        # print(self.TEXT_PATH)

        self.XBRL_PATH = os.path.join(self.TO_DATA_PATH, "xbrl")
        os.makedirs(self.XBRL_PATH, exist_ok=True)
        # print(self.XBRL_PATH)

        self.Data = None
        self.sp_tickers = []
        self.xbrl_base_year = 2020
        self.xbrl_start_year = 2012
        self.today_year = datetime.datetime.now().year
        # log_debug("AcademyCityXBRL was created")`
        self.matching_accounts = None

    # Valuation Functions
    def get_risk_premium(self, year10=1928, year50=1928, cv_project_id=None, is_update='no'):
        log_debug("Start get_risk_premium.")
        base_year = 1928
        project = Project.objects.filter(translations__language_code=get_language()).get(id=int(cv_project_id))
        # print(is_update)
        if is_update != 'yes':
            if project.dic_data:
                return project.dic_data
        dic = {'Arithmetic': {}, 'Geometric': {}}
        try:
            # l_years = [base_year, year50, year10]
            l_years = range(base_year, datetime.datetime.now().year)
            for y in l_years:
                # print(y)
                # for rp in XBRLHistoricalReturnsSP.objects.filter(year__gte=y).all():
                #     try:
                #         print(rp)
                #         print(rp.return_on_sp500, rp.tb3ms_rate, rp.return_on_tbond)
                #     except Exception as ex:
                #         print(ex)
                #
                # print(2222222)

                log_debug('Start process year : ' + str(y))
                ll = [[rp.return_on_sp500, rp.tb3ms_rate, rp.return_on_tbond] for rp in
                      XBRLHistoricalReturnsSP.objects.filter(year__gte=y).all()]
                df = pd.DataFrame(ll)
                df.columns = ['M', 'TB', 'B']
                df['M_TB'] = df['M'] - df['TB']
                df['M_B'] = df['M'] - df['B']

                llb = [[rp.return_on_sp500, rp.tb3ms_rate, rp.return_on_tbond] for rp in
                       XBRLHistoricalReturnsSP.objects.filter(year__lte=y).filter(year__gte=base_year).all()]
                dfb = pd.DataFrame(llb)
                dfb.columns = ['M', 'TB', 'B']
                dfb['M_TB'] = dfb['M'] - dfb['TB']
                dfb['M_B'] = dfb['M'] - dfb['B']
                log_debug('process year : ' + str(y) + " ll done")

                # print(df)
                # print(df.mean())
                # print(df.std())
                # print(df.std()/((df.shape[0]+1)**(1/2)))
                dic['Arithmetic'][y] = {}
                dic['Geometric'][y] = {}
                dic['Arithmetic'][y]['Stocks-TBills'] = {}
                dic['Arithmetic'][y]['Stocks-TBonds'] = {}
                dic['Geometric'][y]['Stocks-TBills'] = {}
                dic['Geometric'][y]['Stocks-TBonds'] = {}

                dic['Arithmetic'][y]['Stocks-TBills']['value'] = round(10000 * df.mean()['M_TB']) / 10000
                dic['Arithmetic'][y]['Stocks-TBonds']['value'] = round(10000 * df.mean()['M_B']) / 10000

                yrp = XBRLHistoricalReturnsSP.objects.get(year=y)

                log_debug('process year : ' + str(y) + " yrp done")

                dic['Arithmetic'][y]['Stocks-TBonds']['sp500'] = yrp.return_on_sp500
                dic['Arithmetic'][y]['Stocks-TBonds']['tb3ms'] = yrp.tb3ms_rate
                dic['Arithmetic'][y]['Stocks-TBonds']['tbond'] = yrp.return_on_tbond
                dic['Arithmetic'][y]['Stocks-TBonds']['m_b_median'] = round(10000 * df.median()['M_B']) / 10000
                #
                dic['Arithmetic'][y]['Stocks-TBonds']['b_value'] = round(10000 * dfb.mean()['M_B']) / 10000
                dic['Arithmetic'][y]['Stocks-TBonds']['b_m_b_median'] = round(10000 * dfb.median()['M_B']) / 10000
                #
                log_debug('process year : ' + str(y) + " calculation 1 done")

                try:
                    dic['Arithmetic'][y]['Stocks-TBills']['std'] = round(
                        10000 * df.std()['M_TB'] / ((df.shape[0] + 1) ** (1 / 2))) / 10000
                    dic['Arithmetic'][y]['Stocks-TBonds']['std'] = round(
                        10000 * df.std()['M_B'] / ((df.shape[0] + 1) ** (1 / 2))) / 10000

                    dic['Arithmetic'][y]['Stocks-TBonds']['b_std'] = round(
                        10000 * dfb.std()['M_B'] / ((dfb.shape[0] + 1) ** (1 / 2))) / 10000
                    log_debug('process year : ' + str(y) + " std calculation 2 done")
                except Exception as exx:
                    log_debug('Error 10 : ' + " " + str(y) + " " + str(exx))
                gm = 1
                gtb = 1
                gb = 1
                for i, r in df.iterrows():
                    gm *= (1 + r['M'])
                    gtb *= (1 + r['TB'])
                    gb *= (1 + r['B'])
                gm = gm ** (1 / df.shape[0]) - 1
                gtb = gtb ** (1 / df.shape[0]) - 1
                gb = gb ** (1 / df.shape[0]) - 1

                dic['Geometric'][y]['Stocks-TBills']['value'] = round(10000 * (gm - gtb)) / 10000
                dic['Geometric'][y]['Stocks-TBonds']['value'] = round(10000 * (gm - gb)) / 10000
                dic['Geometric'][y]['Stocks-TBills']['std'] = ''
                dic['Geometric'][y]['Stocks-TBonds']['std'] = ''

                try:
                    bgm = 1
                    bgtb = 1
                    bgb = 1
                    for i, r in dfb.iterrows():
                        bgm *= (1 + r['M'])
                        # bgtb *= (1+r['TB'])
                        bgb *= (1 + r['B'])
                    bgm = bgm ** (1 / dfb.shape[0]) - 1
                    # bgtb = bgtb**(1/dfb.shape[0]) - 1
                    bgb = bgb ** (1 / dfb.shape[0]) - 1
                    dic['Geometric'][y]['Stocks-TBonds']['b_value'] = round(10000 * (bgm - bgb)) / 10000
                except Exception as ex:
                    print('ex')
                    print(ex)
                    print('ex')

                log_debug('End process year : ' + str(y))

            # Damodarad uses geometric:  dic['Geometric'][1928]['Stocks-TBonds']['value']
            project.mature_marker_risk_premium = dic['Arithmetic'][1928]['Stocks-TBonds']['value']
            project.dic_data = dic
            project.save()
            # print(dic)

        except Exception as ex:
            log_debug('Error 10 : ' + " " + str(ex))

        log_debug("End get_risk_premium.")
        return dic

    # Assisting functions
    def download_excel_file(self, url, file, ext='xlsx'):
        path = os.path.join(self.EXCEL_PATH, file + "." + ext)
        if not os.path.isfile(path):
            urllib.request.urlretrieve(url, path)
        return path

    def load_excel_data(self, file, sheet_name=None):
        excel_path = os.path.join(self.EXCEL_PATH, file + ".xlsx")
        if sheet_name:
            self.DATA = pd.read_excel(excel_path, engine='openpyxl', sheet_name=sheet_name)
        else:
            self.DATA = pd.read_excel(excel_path, engine='openpyxl')
        return self.DATA

    # I use ticker as cik
    def get_statements(self, company):
        statements = {}
        for statement in XBRLValuationStatementsAccounts.objects.all():
            statements[statement.order] = {'name': statement.statement, 'accounts': {}}
            sic_ = company.industry.sic_code
            sic__ = company.industry.main_sic.sic_code
            accounts = statement.xbrl_valuation_statements.filter(Q(sic=0) | Q(sic=sic_) | Q(sic=sic__)).all()

            for a in accounts:
                statements[statement.order]['accounts'][a.order] = [a.account, a.type, a.scale]
        # print(statements)
        return statements

    # --
    async def process_page(self, obj, text, url, dic):
        print("process_page")
        print(obj.name)
        # print(text)
        print(url)
        print(dic)
        n = np.random.randn()
        dic['url'+str(n)] = url
        asyncio.sleep(1)
        return {"k1": "amob"}

    def run_all_pages(self, dic):
        print('='*50)
        dic = eval(dic)
        print(dic)
        print('='*50)

        urls = ['https://www.sec.gov/Archives/edgar/data/320193/000119312512444068/aapl-20120929.xml',
                'https://www.sec.gov/Archives/edgar/data/320193/000119312513416534/aapl-20130928.xml']

        gaup = GetAllUrlsProcessed()
        k = asyncio.run(gaup.process_all_pages(urls, func=self.process_page, dic=dic))
        print('='*50)
        print(dic)
        print('='*50)
        print('k')
        print(k)

    # --
    def get_data_ticker(self, ticker, is_update='no', is_updateq='no'):
        try:
            company = XBRLCompanyInfo.objects.get(ticker=ticker)
        except Exception as ex:
            print("Error 666" + str(ex))
        dataq = ""
        if is_update == "no" and company.financial_data:
            dic_company_info = company.financial_data
            # print(dic_company_info)
            if is_updateq == "no" and company.financial_dataq:
                # print('dic_company_info q 2')
                dataq = company.financial_dataq
            elif is_updateq == "yes":
                dataq = self.get_dic_company_info_q_(company, dic_company_info)
                company.financial_dataq = dataq
                company.save()
        else:
            dic_company_info = self.get_dic_company_info(company)
            # print('-2'*50)
            company.financial_data = dic_company_info
            # print('-3'*50)
            company.save()
        try:
            countries_regions_ = company.get_countries_regions()
            dic_company_info["countries_regions"] = countries_regions_
        except Exception as ex:
            log_debug("get_dic_company_info ex1: " + str(ex))
        result = {'dic_company_info': dic_company_info, 'dataq': dataq}
        # print(result)
        return result

    def get_dic_company_info(self, company, type_="10-K"):
        dic_company_info = self.get_dic_company_info_(company, type_)
        # print("22222:\n", "dic_company_info:\n", dic_company_info)
        dic_data = dic_company_info['data']
        if len(dic_data) == 0:
            try:
                dic_company_info = self.get_dic_company_info_(company, "20-F")
                dic_data = dic_company_info['data']
            except Exception as ex:
                pass
            if len(dic_data) == 0:
                return dic_company_info
        if 'statements' not in dic_company_info:
            dic_company_info['statements'] = self.get_statements(company=company)
        dic_company_info = self.process_dic_company_info_(dic_company_info)
        print("33333:\n", "dic_company_info:\n", dic_company_info)
        return dic_company_info

    def get_dic_company_info_(self, company, type_="10-K"):
        clear_log_debug()
        # print("get_dic_company_info_ 1")
        base_url = "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={}&type={}&count=100"  # &dateb={}"
        url = base_url.format(company.cik, type_)
        print("11111:\n", url, "\n", '-'*10)
        dic_company_info = {"company_info": {'ticker': company.ticker,
                                             'cik': company.cik,
                                             'sic_code': company.industry.sic_code,
                                             'marginal_tax_rate': str(company.tax_rate),
                                             '10k_url': url,
                                             },
                            'data': {}}
        # print(dic_company_info)
        headers = {'User-Agent': 'amos@drbaranes.com'}
        edgar_resp = requests.get(url, headers=headers, timeout=30)
        # print("Current Time 11 =", datetime.datetime.now().strftime("%H:%M:%S"))
        edgar_str = edgar_resp.text
        # print(edgar_str)
        #
        cik0 = ''
        cik_re = re.compile(r'.*CIK=(\d{10}).*')
        results = cik_re.findall(edgar_str)
        if len(results):
            results[0] = int(re.sub('\.[0]*', '.', results[0]))
            cik0 = str(results[0])
        dic_company_info['company_info']['cik'] = cik0
        #
        sic0 = ''
        cik_re = re.compile(r'.*SIC=(\d{4}).*')
        results = cik_re.findall(edgar_str)
        if len(results):
            results[0] = int(re.sub('\.[0]*', '.', results[0]))
            sic0 = str(results[0])
        dic_company_info['company_info']['SIC'] = sic0
        # print('-3'*10)
        # print(dic_company_info)
        # print('-3'*10)
        #
        # Find the document links
        soup = BeautifulSoup(edgar_str, 'html.parser')
        table_tag = soup.find('table', class_='tableFile2')
        try:
            rows = table_tag.find_all('tr')
        except Exception as ex:
            return dic_company_info

        # Obtain HTML for document page
        dic_data = {}
        for row in rows:
            try:
                cells = row.find_all('td')
                if len(cells) > 3:
                    if cells[0].text.lower() != type_.lower():
                        continue
                    # for filing_year in range(2019, 2020):
                    for filing_year in range(self.xbrl_start_year, self.today_year + 1):
                        if str(filing_year) in cells[3].text:
                            # print(str(filing_year), cells[3].text, cells[1].a['href'], cells[0].text)
                            if filing_year not in dic_data and (cells[0].text == "10-K" or cells[0].text == "20-F"):
                                dic_data[filing_year] = {}
                                dic_data[filing_year]['href'] = 'https://www.sec.gov' + cells[1].a['href']
            except Exception as ex:
                pass

        dic_company_info['data'] = dic_data
        # print(dic_company_info)
        return dic_company_info

    def process_dic_company_info_(self, dic_company_info):
        urls = []
        max_y = 0
        for y in dic_company_info['data']:
            # print(dic_company_info['data'][y])
            urls.append(dic_company_info['data'][y]["href"])
            if max_y < y:
                max_y = y
        max_y += 1
        dic_matching_accounts = self.get_matching_accounts_(dic_company_info, max_y)
        dic_ = {"dic_company_info": dic_company_info, "dic_matching_accounts": dic_matching_accounts}
        # --
        gaup = GetAllUrlsProcessed()
        k = asyncio.run(gaup.process_all_pages(urls, func=self.get_data_for_years_async, dic=dic_))
        # --
        dic_company_info = dic_["dic_company_info"]
        return dic_company_info

    async def get_data_for_years_async(self, obj, txt, url, dic):
        dic_company_info = dic["dic_company_info"]
        dic_matching_accounts = dic["dic_matching_accounts"]
        try:
            value = None
            key = None
            for y in dic_company_info["data"]:
                if dic_company_info["data"][y]["href"] == url:
                    value = dic_company_info["data"][y]
                    key = y
            doc_str = txt
            xbrl_link = ''
            soup = BeautifulSoup(doc_str, 'html.parser')
            table_tag = soup.find('table', class_='tableFile', summary='Data Files')
        except Exception as ex:
            print("Error 22: " + str(ex))
        try:
            rows = table_tag.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) > 3:
                    if 'INS' in cells[3].text or 'XML' in cells[3].text:
                        #
                        # print(cells[3].text)
                        #
                        xbrl_link = cells[2].a['href']
            value['xbrl_link'] = 'https://www.sec.gov' + xbrl_link
            accession_number = xbrl_link.split('/')
            view_link = 'https://www.sec.gov/cgi-bin/viewer?action=view&cik='
            view_link += accession_number[4] + '&accession_number=' + accession_number[5] + '&xbrl_type=v#'

            r_link = "https://www.sec.gov/Archives/edgar/data/" + accession_number[4] + "/" + accession_number[5] + "/R"
            value['r_link'] = r_link
            value['view_link'] = view_link
        except Exception as ex:
            # print("Error 67: " + str(ex))
            return value

        xbrl_str = await obj.get_page(value['xbrl_link'])
        soup = BeautifulSoup(xbrl_str, 'lxml')
        value['dei'] = {}
        for tag in soup.find_all(re.compile("dei:")):
            name_ = tag.name.split(":")
            value['dei'][name_[1]] = tag.text

        # print(str(key) + '-11'*20)
        # print(str(key) + '-11'*20)
        # print(str(key) + '-11'*20)
        # print(str(key) + '-11'*20)
        # print(str(key) + '-11'*20)
        # print(str(key) + '-11'*20)
        # print(value)
        # print(key)
        # # print(txt)
        # print(str(key) + '-11'*20)
        # print(str(key) + '-11'*20)
        # print(str(key) + '-11'*20)
        # print(str(key) + '-11'*20)
        # print(str(key) + '-11'*20)
        # print(str(key) + '-11'*20)

        documentperiodenddate = value['dei']['documentperiodenddate']
        entitycentralindexkey = value['dei']['entitycentralindexkey']
        y_ = int(value['dei']['documentfiscalyearfocus'])
        # For ticker ANTM there is a problem for year 2014 it shows documentfiscalyearfocus=2013 when it should be 2014
        # I left that error, so I miss data for 2014
        # y_ = int(str(documentperiodenddate).split("-")[0])
        try:
            s_cik = "NA"
            flow_context_id = ""
            for tag in soup.find_all(name=re.compile('enddate'), string=documentperiodenddate):
                # print(tag.name)
                try:
                    context = tag.find_parent(re.compile('context'))
                    context_name = context.name.split(":")
                    if len(context_name) > 1:
                        identifier = context.find(context_name[0] + ':identifier')
                        segment = context.find(context_name[0] + ':segment')
                        startdate = context.find(context_name[0] + ':startdate')
                    else:
                        identifier = context.find('identifier')
                        segment = context.find('segment')
                        startdate = context.find('startdate')

                    end_date = tag.text.split('-')
                    start_date = startdate.text.split('-')
                    start_date = start_date[0] + '-' + start_date[1]

                    start_date_should = str((int(end_date[0]) - 1)) + '-' + end_date[1]
                    start_date0_should = str((int(end_date[0]) - 2)) + '-12'
                    if int(end_date[1]) > 10 or int(end_date[1]) < 3:
                        start_date1_should = end_date[0] + '-01'
                    else:
                        start_date1_should = 0
                    start_date2_should = str((int(end_date[0]) - 1)) + '-' + self.add_zero(str((int(end_date[1]) + 1)))
                    start_date3_should = str((int(end_date[0]) - 1)) + '-' + self.add_zero(str((int(end_date[1]) - 1)))

                    if (not segment) and (identifier.text == entitycentralindexkey) \
                            and (
                            start_date == start_date_should or start_date == start_date0_should or
                            start_date == start_date1_should or start_date == start_date2_should or
                            start_date == start_date3_should):
                        # if int(y_) == 2014:
                        #     print("-=3"*3)
                        #     print(end_date, context['id'], start_date_should, start_date0_should, start_date1_should,start_date2_should,start_date3_should)
                        #     print(context)
                        flow_context_id = context['id']
                    if flow_context_id == "":
                        s_cik = dic_company_info['company_info']['cik']
                        while len(s_cik) < 10:
                            s_cik = "0" + s_cik
                        if (not segment) and (identifier.text == s_cik) and (
                                start_date == start_date_should or start_date == start_date0_should or
                                start_date == start_date1_should or start_date == start_date2_should or
                                start_date == start_date3_should
                        ):
                            flow_context_id = context['id']
                    else:
                        break

                except Exception as ex:
                    # print(ex)
                    continue

            instant_context_id = ""
            for tag in soup.find_all(name=re.compile('instant'), string=documentperiodenddate):
                context = tag.find_parent(re.compile('context'))

                context_name = context.name.split(":")

                if len(context_name) > 1:
                    identifier = context.find(context_name[0] + ':identifier')
                    segment = context.find(context_name[0] + ':segment')
                else:
                    identifier = context.find('identifier')
                    segment = context.find('segment')
                if not segment and identifier.text == entitycentralindexkey:
                    # print(context)
                    instant_context_id = context['id']

                if instant_context_id == "":
                    s_cik = dic_company_info['company_info']['cik']
                    while len(s_cik) < 10:
                        s_cik = "0" + s_cik
                    if not segment and identifier.text == s_cik:
                        # print(context)
                        instant_context_id = context['id']
        except Exception as ex:
            return value
        try:
            matching_accounts = dic_matching_accounts[y_]["matching_accounts"]
            accounts_ = dic_matching_accounts[y_]["accounts_"]
            used_accounting_standards = dic_matching_accounts[y_]["used_accounting_standards"]

            value['matching_accounts'] = matching_accounts
            year_data = {}

            accounts_flow = {}
            for k in accounts_['flow']:
                accounts_flow[k[0:92]] = accounts_['flow'][k]

            accounts_instant = {}
            for k in accounts_['instant']:
                accounts_instant[k[0:92]] = accounts_['instant'][k]

            for std in used_accounting_standards:
                tag_list = soup.find_all(re.compile(std + ":"))
                for tag in tag_list:
                    name_ = tag.name.split(":")
                    try:
                        if instant_context_id != "":
                            if name_[1][0:92] in accounts_instant and tag['contextref'] == instant_context_id:
                                if accounts_instant[name_[1][0:92]][2] == std:
                                    order = accounts_instant[name_[1][0:92]][0]
                                    scale = accounts_instant[name_[1][0:92]][1]
                                    if scale == 1:
                                        year_data[order] = tag.text
                                    else:
                                        year_data[order] = int(round(float(tag.text))) / scale
                        if name_[1][0:92] in accounts_flow:
                            if tag['contextref'] == flow_context_id and accounts_flow[name_[1][0:92]][2] == std:
                                order = accounts_flow[name_[1][0:92]][0]
                                scale = accounts_flow[name_[1][0:92]][1]
                                if scale == 1:
                                    year_data[order] = tag.text
                                else:
                                    year_data[order] = int(round(float(tag.text))) / scale
                    except Exception as ex:
                        pass
                        # print("Error: " + str(ex))
            value['year_data'] = year_data
            dic_company_info['data'][key] = value
        except Exception as ex:
            return value
        return value

    def get_matching_accounts_(self, dic_company_info, max_y):
        ticker = dic_company_info['company_info']['ticker']
        sic = dic_company_info['company_info']['SIC']
        statements = dic_company_info['statements']
        dic = {}
        for year in range(2011, max_y):
            dic[year] = {}
            try:
                if year <= self.xbrl_base_year:
                    years = sorted(range(year, self.xbrl_base_year + 1), reverse=True)
                else:
                    years = range(self.xbrl_base_year, year + 1)
            except Exception as ex:
                print(ex)
            matches = XBRLValuationAccountsMatch.objects.filter((Q(year=0) & Q(company__industry__sic_code=sic)) |
                                                                (Q(year__in=years) & Q(company__ticker=ticker))).all()
            used_accounting_standards = []
            dic_matches = {}
            for m in matches:
                if m.year == 0:
                    dic_matches[m.account.order] = [m.match_account, m.accounting_standard]
                    if m.accounting_standard not in used_accounting_standards:
                        used_accounting_standards.append(m.accounting_standard)

            for y in years:
                for m in matches:
                    if m.year == y:
                        dic_matches[m.account.order] = [m.match_account, m.accounting_standard]
                        if m.accounting_standard not in used_accounting_standards:
                            used_accounting_standards.append(m.accounting_standard)

            matching_accounts = {}
            accounts_ = {'instant': {}, 'flow': {}}

            for st_ in statements:
                for a_order in statements[st_]['accounts']:
                    try:
                        if int(a_order) in dic_matches:
                            ma_, ma_std = dic_matches[int(a_order)][0].lower(), dic_matches[int(a_order)][1].lower()
                        else:
                            ma_, ma_std = "", ""
                    except Exception as ex:
                        pass

                    if ma_ != '':
                        if statements[st_]['accounts'][a_order][1] == 1:
                            accounts_['instant'][ma_] = (a_order, statements[st_]['accounts'][a_order][2], ma_std)
                        else:
                            accounts_['flow'][ma_] = (a_order, statements[st_]['accounts'][a_order][2], ma_std)
                    matching_accounts[a_order] = [ma_, ma_std]
            dic[year]["matching_accounts"] = matching_accounts
            dic[year]["accounts_"] = accounts_
            dic[year]["used_accounting_standards"] = used_accounting_standards
        return dic

    #  --
    def get_dic_company_info_q_(self, company, dic_company_info):

        headers = {'User-Agent': 'amos@drbaranes.com'}
        base_url = "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={}&type={}&count=100"  # &dateb={}"
        type_ = '10-Q'
        url_q = base_url.format(company.cik, type_)
        dic_data = {'10q_url': url_q, 'data': {}}
        try:
            edgar_resp_q = requests.get(url_q, headers=headers, timeout=30)
        except Exception as ex:
            print(str(ex))
        edgar_str_q = edgar_resp_q.text
        soup_q = BeautifulSoup(edgar_str_q, 'html.parser')
        table_tag_q = soup_q.find('table', class_='tableFile2')
        try:
            rows = table_tag_q.find_all('tr')
        except Exception as ex:
            return dic_company_info
        for row in rows:
            try:
                cells = row.find_all('td')
                if len(cells) > 3:
                    if cells[0].text.lower() != type_.lower():
                        continue
                    for filing_year in range(self.xbrl_start_year, self.today_year + 1):
                        if str(filing_year) in cells[3].text:
                            m = cells[3].text.split("-")[1]
                            try:
                                if filing_year not in dic_data['data']:
                                    dic_data['data'][filing_year] = {}
                            except Exception as exx:
                                print(str(exx))
                            dic_data['data'][filing_year][m] = {'href': 'https://www.sec.gov' + cells[1].a['href']}
            except Exception as ex:
                pass
        urls = []
        max_y = 0
        for key in dic_data['data']:
            if max_y<key:
                max_y=key
            try:
                url_ = dic_company_info['data'][str(key)]['href']
                dic_data['data'][key]['Q4'] = {'href': url_}
            except Exception as ex:
                dic_data['data'][key]['Q4'] = {'href': ""}
            for m in dic_data['data'][key]:
                if dic_data['data'][key][m]['href'] == "":
                    continue
                else:
                    urls.append(dic_data['data'][key][m]['href'])
        max_y += 1
        dic_matching_accounts = self.get_matching_accounts_(dic_company_info, max_y)
        s_cik = str(company.cik)
        while len(s_cik) < 10:
            s_cik = "0" + s_cik
        dic_ = {"dic_data": dic_data, "dic_matching_accounts": dic_matching_accounts, "cik": s_cik}
        # --
        gaup = GetAllUrlsProcessed()
        k = asyncio.run(gaup.process_all_pages(urls, func=self.get_data_for_one_year_q_async, dic=dic_))
        return dic_["dic_data"]

    async def get_data_for_one_year_q_async(self, obj, txt, url, dic):
        dic_data = dic["dic_data"]
        s_cik = dic["cik"]
        # print(s_cik)
        dic_matching_accounts = dic["dic_matching_accounts"]
        value = None
        key = None
        m = None
        for y in dic_data["data"]:
            for m_ in dic_data["data"][y]:
                try:
                    if dic_data["data"][y][m_]["href"] == url:
                        value = dic_data["data"][y][m_]
                        key = y
                        m = m_
                except Exception as ex:
                    continue
        xbrl_link = ''
        soup = BeautifulSoup(txt, 'html.parser')
        table_tag = soup.find('table', class_='tableFile', summary='Data Files')
        # print(table_tag)
        try:
            rows = table_tag.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) > 3:
                    if 'INS' in cells[3].text or 'XML' in cells[3].text:
                        # print(cells[3].text)
                        xbrl_link = cells[2].a['href']
            value['xbrl_link'] = 'https://www.sec.gov' + xbrl_link
            accession_number = xbrl_link.split('/')
            view_link = 'https://www.sec.gov/cgi-bin/viewer?action=view&cik='
            view_link += accession_number[4] + '&accession_number=' + accession_number[5] + '&xbrl_type=v#'

            r_link = "https://www.sec.gov/Archives/edgar/data/" + accession_number[4] + "/" + accession_number[5] + "/R"
            value['r_link'] = r_link
            value['view_link'] = view_link
        except Exception as ex:
            return value

        # xbrl_resp = requests.get(value['xbrl_link'], headers=headers, timeout=30)
        # xbrl_str = xbrl_resp.text
        # print(value['xbrl_link'])
        xbrl_str = await obj.get_page(value['xbrl_link'])

        soup = BeautifulSoup(xbrl_str, 'lxml')
        value['dei'] = {}
        for tag in soup.find_all(re.compile("dei:")):
            name_ = tag.name.split(":")
            value['dei'][name_[1]] = tag.text

        documentperiodenddate = value['dei']['documentperiodenddate']
        entitycentralindexkey = value['dei']['entitycentralindexkey']
        if m == "Q4":
            documentfiscalperiodfocus = "Q4"
            value['dei']['documentfiscalperiodfocus'] = "Q4"
        else:
            documentfiscalperiodfocus = value['dei']['documentfiscalperiodfocus']

        flow_context_id = ""
        for tag in soup.find_all(name=re.compile('enddate'), string=documentperiodenddate):
            try:
                context = tag.find_parent(re.compile('context'))
                # print(context)
                context_name = context.name.split(":")

                if len(context_name) > 1:
                    identifier = context.find(context_name[0] + ':identifier')
                    segment = context.find(context_name[0] + ':segment')
                    startdate = context.find(context_name[0] + ':startdate')
                else:
                    identifier = context.find('identifier')
                    segment = context.find('segment')
                    startdate = context.find('startdate')

                end_date = tag.text.split('-')
                start_date = startdate.text.split('-')
                start_date = start_date[0] + '-' + start_date[1]

                if 12 >= int(end_date[1]) > 4:
                    start_date_should = end_date[0] + '-' + self.add_zero(str(int(end_date[1]) - 3))
                    start_date0_should = end_date[0] + '-' + self.add_zero(str(int(end_date[1]) - 2))
                    start_date1_should = end_date[0] + '-' + self.add_zero(str(int(end_date[1]) - 4))
                    start_date2_should = "none"
                    # print(start_date_should)
                elif int(end_date[1]) <= 4:
                    start_date_should = end_date[0] + '-01'
                    start_date0_should = end_date[0] + '-02'
                    start_date1_should = str((int(end_date[0]) - 1)) + '-12'
                    start_date2_should = str((int(end_date[0]) - 1)) + '-11'

                if (not segment) and (identifier.text == entitycentralindexkey) \
                        and (start_date == start_date_should or start_date == start_date0_should or
                             start_date == start_date1_should or start_date == start_date2_should):
                    flow_context_id = context['id']
                if flow_context_id == "":
                    if (not segment) and (identifier.text == s_cik) \
                            and (start_date == start_date_should or start_date == start_date0_should or
                                 start_date == start_date1_should or start_date == start_date2_should):
                        flow_context_id = context['id']

            except Exception as ex:
                # print(ex)
                continue

        # print('flow_context_id')
        # print(flow_context_id)
        # print('entitycentralindexkey')
        # print(entitycentralindexkey)
        # print('s_cik')
        # print(s_cik)

        instant_context_id = ""
        for tag in soup.find_all(name=re.compile('instant'), string=documentperiodenddate):
            context = tag.find_parent(re.compile('context'))
            context_name = context.name.split(":")
            if len(context_name) > 1:
                identifier = context.find(context_name[0] + ':identifier')
                segment = context.find(context_name[0] + ':segment')
            else:
                identifier = context.find('identifier')
                segment = context.find('segment')
            if not segment and identifier.text == entitycentralindexkey:
                # print(context)
                instant_context_id = context['id']
                if instant_context_id == "":
                    if not segment and identifier.text == s_cik:
                        instant_context_id = context['id']
        # print(instant_context_id)

        matching_accounts = dic_matching_accounts[key]["matching_accounts"]
        accounts_ = dic_matching_accounts[key]["accounts_"]
        used_accounting_standards = dic_matching_accounts[key]["used_accounting_standards"]

        value['matching_accounts'] = matching_accounts
        year_data = {}
        accounts_flow = {}
        for k in accounts_['flow']:
            accounts_flow[k[0:92]] = accounts_['flow'][k]

        accounts_instant = {}
        for k in accounts_['instant']:
            accounts_instant[k[0:92]] = accounts_['instant'][k]

        for std in used_accounting_standards:
            tag_list = soup.find_all(re.compile(std + ":"))
            for tag in tag_list:
                name_ = tag.name.split(":")
                try:
                    if name_[1][0:92] in accounts_instant and tag['contextref'] == instant_context_id:
                        if accounts_instant[name_[1][0:92]][2] == std:
                            order = accounts_instant[name_[1][0:92]][0]
                            scale = accounts_instant[name_[1][0:92]][1]
                            if scale == 1:
                                year_data[order] = tag.text
                            else:
                                year_data[order] = int(round(float(tag.text))) / scale
                    if name_[1][0:92] in accounts_flow and tag['contextref'] == flow_context_id:
                        if accounts_flow[name_[1][0:92]][2] == std:
                            order = accounts_flow[name_[1][0:92]][0]
                            scale = accounts_flow[name_[1][0:92]][1]
                            if scale == 1:
                                year_data[order] = tag.text
                            else:
                                year_data[order] = int(round(float(tag.text))) / scale
                except Exception as ex:
                    pass
                    # print("Error: year=" + str(dic_data_year[0]) + "   dic=" + dic_data_year[1]['href'] + "   " + str(ex) + tag.text)

        value['year_data'] = year_data
        dic_data['data'][key][m] = value
        return value

    # --
    # to be deleted
    def get_data_for_one_year(self, key, dic_company_info):
        value = dic_company_info['data'][key]
        temp = [key, value,
                dic_company_info['company_info']['SIC'], dic_company_info['company_info']['ticker'],
                dic_company_info['statements']
                ]
        dic_company_info['data'][key] = self.get_data_for_years(dic_data_year=temp)[1]
        return dic_company_info

    # to be deleted
    def get_data_for_years(self, dic_data_year):
        # print(dic_data_year)
        log_debug("start get_data_for_years: " + dic_data_year[3] + ", " + str(dic_data_year[0]))
        # print("Current Time Start get data = " + str(dic_data_year[0]), datetime.datetime.now().strftime("%H:%M:%S"))
        # dic_data_year[3] = ticker
        # dic_data_year[0] = year
        # dic_data_year[2] = sic
        #
        # print('-30'*10)
        # print(str(dic_data_year[0]) + dic_data_year[1]['href'])
        # print('-30'*10)

        headers = {'User-Agent': 'amos@drbaranes.com'}
        # print(dic_data_year[1])

        doc_resp = requests.get(dic_data_year[1]['href'], headers=headers, timeout=30)
        doc_str = doc_resp.text

        # print("get_data_for_years (after getting file_name): "+dic_data_year[3]+", "+str(dic_data_year[0]))

        # Find the XBRL link
        xbrl_link = ''
        soup = BeautifulSoup(doc_str, 'html.parser')
        table_tag = soup.find('table', class_='tableFile', summary='Data Files')

        # print(dic_data_year[0])
        # print(dic_data_year[1]['href'])

        try:
            rows = table_tag.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) > 3:
                    if 'INS' in cells[3].text or 'XML' in cells[3].text:
                        #
                        # print(cells[3].text)
                        #
                        xbrl_link = cells[2].a['href']

            dic_data_year[1]['xbrl_link'] = 'https://www.sec.gov' + xbrl_link
            accession_number = xbrl_link.split('/')
            # print('-12-'*10)
            # print(accession_number)
            # print('-12-'*10)

            view_link = 'https://www.sec.gov/cgi-bin/viewer?action=view&cik='
            view_link += accession_number[4] + '&accession_number=' + accession_number[5] + '&xbrl_type=v#'

            r_link = "https://www.sec.gov/Archives/edgar/data/" + accession_number[4] + "/" + accession_number[5] + "/R"
            dic_data_year[1]['r_link'] = r_link
            dic_data_year[1]['view_link'] = view_link

        except Exception as ex:
            return dic_data_year

        xbrl_file_name = "xbrl_" + dic_data_year[3] + "_" + str(dic_data_year[0])
        # try:
        #     xbrl_str = request.session[xbrl_file_name]
        # except Exception as exc:
        xbrl_resp = requests.get(dic_data_year[1]['xbrl_link'], headers=headers, timeout=30)
        xbrl_str = xbrl_resp.text
        # request.session[xbrl_file_name] = xbrl_str

        # log_debug("get_data_for_years (after getting xbrl_file_name): "+dic_data_year[3]+", "+str(dic_data_year[0]))
        soup = BeautifulSoup(xbrl_str, 'lxml')
        dic_data_year[1]['dei'] = {}
        for tag in soup.find_all(re.compile("dei:")):
            name_ = tag.name.split(":")
            dic_data_year[1]['dei'][name_[1]] = tag.text

        documentperiodenddate = dic_data_year[1]['dei']['documentperiodenddate']
        entitycentralindexkey = dic_data_year[1]['dei']['entitycentralindexkey']

        # print(documentperiodenddate)
        # print(entitycentralindexkey)

        # print("Current Time 4 =", datetime.datetime.now().strftime("%H:%M:%S"))
        flow_context_id = ""
        for tag in soup.find_all(name=re.compile('enddate'), string=documentperiodenddate):
            # print(tag.name)
            try:
                context = tag.find_parent(re.compile('context'))
                # print(context)
                context_name = context.name.split(":")
                # print('=-3-'*50)
                # print(context.name + ' : ' + str(len(context_name)))
                # print('=-3-'*50)

                if len(context_name) > 1:
                    identifier = context.find(context_name[0] + ':identifier')
                    segment = context.find(context_name[0] + ':segment')
                    startdate = context.find(context_name[0] + ':startdate')
                else:
                    identifier = context.find('identifier')
                    segment = context.find('segment')
                    startdate = context.find('startdate')

                end_date = tag.text.split('-')
                start_date = startdate.text.split('-')
                start_date = start_date[0] + '-' + start_date[1]

                start_date_should = str((int(end_date[0]) - 1)) + '-' + end_date[1]
                start_date0_should = str((int(end_date[0]) - 2)) + '-12'
                if int(end_date[1]) > 10 or int(end_date[1]) < 3:
                    start_date1_should = end_date[0] + '-01'
                else:
                    start_date1_should = 0
                start_date2_should = str((int(end_date[0]) - 1)) + '-' + self.add_zero(str((int(end_date[1]) + 1)))
                start_date3_should = str((int(end_date[0]) - 1)) + '-' + self.add_zero(str((int(end_date[1]) - 1)))

                # if end_date[0] == "2016":
                #     print('-6'*10)
                #     print('end_date')
                #     print(end_date)
                #     print('end_date')
                #     print(start_date_should)
                #     print(start_date0_should)
                #     print(start_date1_should)
                #     print(start_date2_should)
                #     print(start_date3_should)
                #     print(start_date)
                #     print('=5'*10)
                #     print('=6'*10)

                if (not segment) and (identifier.text == entitycentralindexkey) \
                        and (
                        start_date == start_date_should or start_date == start_date0_should or
                        start_date == start_date1_should or start_date == start_date2_should or
                        start_date == start_date3_should):
                    flow_context_id = context['id']

            except Exception as ex:
                # print(ex)
                continue

        # if end_date[0] == "2016":
        #     print('-flow'*20)
        #     print(flow_context_id)
        #     print(documentperiodenddate)
        #     print('-flow'*20)

        # print('=bs'*50)
        # print("Current Time 5 =", datetime.datetime.now().strftime("%H:%M:%S"))
        for tag in soup.find_all(name=re.compile('instant'), string=documentperiodenddate):
            context = tag.find_parent(re.compile('context'))

            context_name = context.name.split(":")
            # print('=--'*50)
            # print(context.name + ' : ' + str(len(context_name)))
            # print('=--'*50)

            if len(context_name) > 1:
                identifier = context.find(context_name[0] + ':identifier')
                segment = context.find(context_name[0] + ':segment')
            else:
                identifier = context.find('identifier')
                segment = context.find('segment')
            if not segment and identifier.text == entitycentralindexkey:
                # print(context)
                instant_context_id = context['id']

        # print('-instant'*10)
        # print(instant_context_id)
        # print(documentperiodenddate)
        # print('-instant'*10)

        matching_accounts, accounts_, used_accounting_standards = \
            self.get_matching_accounts(dic_data_year[3], int(dic_data_year[1]['dei']['documentfiscalyearfocus']),
                                       dic_data_year[2], dic_data_year[4])

        dic_data_year[1]['matching_accounts'] = matching_accounts
        year_data = {}

        # log_debug("get_data_for_years (after getting matching_accounts): "+dic_data_year[3]+", "+str(dic_data_year[0]))

        # print("Current Time 6 =", datetime.datetime.now().strftime("%H:%M:%S"))

        accounts_flow = {}
        for k in accounts_['flow']:
            accounts_flow[k[0:92]] = accounts_['flow'][k]

        accounts_instant = {}
        for k in accounts_['instant']:
            accounts_instant[k[0:92]] = accounts_['instant'][k]

        # if str(dic_data_year[0]) == "2020":
        #     print(int(dic_data_year[1]['dei']['documentfiscalyearfocus']))
        #     print(str(dic_data_year[0]))
        #     print(accounts_instant)

        for std in used_accounting_standards:
            tag_list = soup.find_all(re.compile(std + ":"))
            for tag in tag_list:
                name_ = tag.name.split(":")
                try:
                    if name_[1][0:92] in accounts_instant and tag['contextref'] == instant_context_id:
                        if accounts_instant[name_[1][0:92]][2] == std:
                            order = accounts_instant[name_[1][0:92]][0]
                            scale = accounts_instant[name_[1][0:92]][1]
                            if scale == 1:
                                year_data[order] = tag.text
                            else:
                                year_data[order] = int(tag.text) / scale
                    if name_[1][0:92] in accounts_flow and tag['contextref'] == flow_context_id:
                        order = accounts_flow[name_[1][0:92]][0]
                        scale = accounts_flow[name_[1][0:92]][1]
                        if scale == 1:
                            year_data[order] = tag.text
                        else:
                            year_data[order] = int(tag.text) / scale

                except Exception as ex:
                    pass
                    # print("Error: year=" + str(dic_data_year[0]) + "   dic=" + dic_data_year[1]['href'] + "   " + str(ex) + tag.text)

        dic_data_year[1]['year_data'] = year_data

        log_debug("End get_data_for_years: " + dic_data_year[3] + ", " + str(dic_data_year[0]))
        # print("Current Time End get data = " + str(dic_data_year[0]), datetime.datetime.now().strftime("%H:%M:%S"))
        return dic_data_year

    def get_dic_company_info_q(self, company, dic_company_info):
        # print('get_dic_company_info_q')
        # print(dic_company_info)
        # for key in dic_company_info['data']:
        #     print(key)
        #     u = dic_company_info['data'][key]
        #     print(u)
        #     u = dic_company_info['data'][key]['href']
        #     print(u)

        # print('get_dic_company_info_q')
        headers = {'User-Agent': 'amos@drbaranes.com'}
        base_url = "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={}&type={}&count=100"  # &dateb={}"
        type_ = '10-Q'
        url_q = base_url.format(company.cik, type_)
        # print(url_q)
        dic_data = {'10q_url': url_q, 'data': {}}
        try:
            edgar_resp_q = requests.get(url_q, headers=headers, timeout=30)
        except Exception as ex:
            print(str(ex))
        edgar_str_q = edgar_resp_q.text
        soup_q = BeautifulSoup(edgar_str_q, 'html.parser')
        table_tag_q = soup_q.find('table', class_='tableFile2')
        try:
            rows = table_tag_q.find_all('tr')
        except Exception as ex:
            return dic_company_info

        # print('dic_data 1110000')
        # print(dic_data)
        # print('-'*100)

        for row in rows:
            try:
                cells = row.find_all('td')
                # print(cells[0].text.lower())
                if len(cells) > 3:
                    if cells[0].text.lower() != type_.lower():
                        continue
                    # for filing_year in range(2019, 2020):
                    for filing_year in range(self.xbrl_start_year, self.today_year + 1):
                        if str(filing_year) in cells[3].text:
                            # print('---')
                            m = cells[3].text.split("-")[1]
                            # print(cells[3].text, 'filing_year', filing_year, 'm', m)
                            try:
                                # print('='*20)
                                if filing_year not in dic_data['data']:
                                    dic_data['data'][filing_year] = {}
                            except Exception as exx:
                                print(str(exx))
                            dic_data['data'][filing_year][m] = {'href': 'https://www.sec.gov' + cells[1].a['href']}
            except Exception as ex:
                pass

        # print('dic_data111')
        # print(dic_data)
        # print('='*100)

        for key in dic_data['data']:
            # print(key)
            try:
                # print(dic_company_info['data'][str(key)])
                url_ = dic_company_info['data'][str(key)]['href']
                # print(url_)
                # print("dic_data['data'][key]")
                # print(dic_data['data'][key])
                dic_data['data'][key]['Q4'] = {'href': url_}
            except Exception as ex:
                dic_data['data'][key]['Q4'] = {'href': ""}
                # print("Error 300: " + str(ex))
                log_debug("Error 300: " + str(ex))
            for m in dic_data['data'][key]:
                # print(1111)
                # print('='*150)
                # print(key, m)
                if dic_data['data'][key][m]['href'] == "":
                    # print("empty: "+m)
                    continue
                dic_data = self.get_data_for_one_year_q(key, m, dic_data, dic_company_info)
                # print('='*150)

        # print(dic_data)
        # print('-='*50)
        return dic_data

    def get_data_for_one_year_q(self, key, m, dic_data, dic_company_info):
        # print(key, m)
        value = dic_data['data'][key][m]
        temp = [key, m, value,
                dic_company_info['company_info']['SIC'], dic_company_info['company_info']['ticker'],
                dic_company_info['statements']
                ]
        # print('temp')
        # print(temp)
        d = self.get_data_for_years_q(dic_data_year=temp)[2]
        # print('-d'*50)
        # print(d)
        # print('-d'*50)
        dic_data['data'][key][m] = d
        # print(dic_data)
        return dic_data

    def get_data_for_years_q(self, dic_data_year):
        # print("start get_data_for_years_q: "+dic_data_year[4]+", "+str(dic_data_year[0])+", "+str(dic_data_year[1]))
        # print("start get_data_for_years_q: "+dic_data_year[4]+", "+str(dic_data_year[0])+", "+str(dic_data_year[1]))
        headers = {'User-Agent': 'amos@drbaranes.com'}
        doc_resp = requests.get(dic_data_year[2]['href'], headers=headers, timeout=30)
        doc_str = doc_resp.text
        # print('-'*100)
        # print('-'*100)
        # print("get_data_for_years_q: "+str(dic_data_year[4])+", "+str(dic_data_year[0])+":"+str(dic_data_year[1]),
        #       dic_data_year[2]['href'])

        # Find the XBRL link
        xbrl_link = ''
        soup = BeautifulSoup(doc_str, 'html.parser')
        table_tag = soup.find('table', class_='tableFile', summary='Data Files')
        try:
            rows = table_tag.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) > 3:
                    if 'INS' in cells[3].text or 'XML' in cells[3].text:
                        #
                        # print(cells[3].text)
                        #
                        xbrl_link = cells[2].a['href']

            dic_data_year[2]['xbrl_link'] = 'https://www.sec.gov' + xbrl_link
            accession_number = xbrl_link.split('/')
            # print('-12-'*3, 'accession_number', accession_number)

            view_link = 'https://www.sec.gov/cgi-bin/viewer?action=view&cik='
            view_link += accession_number[4] + '&accession_number=' + accession_number[5] + '&xbrl_type=v#'

            r_link = "https://www.sec.gov/Archives/edgar/data/" + accession_number[4] + "/" + accession_number[5] + "/R"
            dic_data_year[2]['r_link'] = r_link
            dic_data_year[2]['view_link'] = view_link
            # print(dic_data_year)
        except Exception as ex:
            return dic_data_year

        xbrl_resp = requests.get(dic_data_year[2]['xbrl_link'], headers=headers, timeout=30)
        xbrl_str = xbrl_resp.text

        # print("get_data_for_years q: "+dic_data_year[4]+", "+str(dic_data_year[0])+":"+str(dic_data_year[1]))
        # print('-'*20)

        soup = BeautifulSoup(xbrl_str, 'lxml')
        dic_data_year[2]['dei'] = {}
        for tag in soup.find_all(re.compile("dei:")):
            name_ = tag.name.split(":")
            dic_data_year[2]['dei'][name_[1]] = tag.text

        documentperiodenddate = dic_data_year[2]['dei']['documentperiodenddate']
        entitycentralindexkey = dic_data_year[2]['dei']['entitycentralindexkey']
        if str(dic_data_year[1]) == "Q4":
            documentfiscalperiodfocus = "Q4"
            dic_data_year[2]['dei']['documentfiscalperiodfocus'] = "Q4"
        else:
            documentfiscalperiodfocus = dic_data_year[2]['dei']['documentfiscalperiodfocus']

        # print(dic_data_year[2]['dei'])
        # print(documentperiodenddate, entitycentralindexkey, documentfiscalperiodfocus)
        # print('-'*20)

        # print("Current Time 4 =", datetime.datetime.now().strftime("%H:%M:%S"))
        flow_context_id = ""
        for tag in soup.find_all(name=re.compile('enddate'), string=documentperiodenddate):
            # print(tag.name)
            try:
                context = tag.find_parent(re.compile('context'))
                # print(context)
                context_name = context.name.split(":")
                # print('=-3-'*50)
                # print(context.name + ' : ' + str(len(context_name)))
                # print('=-3-'*50)

                if len(context_name) > 1:
                    identifier = context.find(context_name[0] + ':identifier')
                    segment = context.find(context_name[0] + ':segment')
                    startdate = context.find(context_name[0] + ':startdate')
                else:
                    identifier = context.find('identifier')
                    segment = context.find('segment')
                    startdate = context.find('startdate')

                end_date = tag.text.split('-')
                start_date = startdate.text.split('-')
                start_date = start_date[0] + '-' + start_date[1]

                # print(end_date[1])
                if 12 >= int(end_date[1]) > 4:
                    start_date_should = end_date[0] + '-' + self.add_zero(str(int(end_date[1]) - 3))
                    start_date0_should = end_date[0] + '-' + self.add_zero(str(int(end_date[1]) - 2))
                    start_date1_should = end_date[0] + '-' + self.add_zero(str(int(end_date[1]) - 4))
                    start_date2_should = "none"
                    # print(start_date_should)
                elif int(end_date[1]) <= 4:
                    start_date_should = end_date[0] + '-01'
                    start_date0_should = end_date[0] + '-02'
                    start_date1_should = str((int(end_date[0]) - 1)) + '-12'
                    start_date2_should = str((int(end_date[0]) - 1)) + '-11'

                # if str(dic_data_year[1]) == "Q4":
                #     if end_date[0] == "2021":
                #         print('-6-'*50)
                #         print('end_date')
                #         print(end_date)
                #         print('end_date')
                #         print(start_date_should)
                #         print(start_date0_should)
                #         print(start_date1_should)
                #         print('start_date')
                #         print(start_date)
                #         print('=5'*10)
                #         print('=6'*10)

                if (not segment) and (identifier.text == entitycentralindexkey) \
                        and (start_date == start_date_should or start_date == start_date0_should or
                             start_date == start_date1_should or start_date == start_date2_should):
                    flow_context_id = context['id']

            except Exception as ex:
                # print(ex)
                continue

        # if end_date[0] == "2021" or end_date == "2020":
        #     print('-flow'*20)
        #     print(documentfiscalperiodfocus)
        #     print(end_date)
        #     print(flow_context_id, documentperiodenddate, start_date_should)
        #     print('-flow'*20)

        # print('-'*50)

        # print('=bs'*50)
        # print("Current Time 5 =", datetime.datetime.now().strftime("%H:%M:%S"))
        for tag in soup.find_all(name=re.compile('instant'), string=documentperiodenddate):
            context = tag.find_parent(re.compile('context'))

            context_name = context.name.split(":")
            # print('=--'*50)
            # print(context.name + ' : ' + str(len(context_name)))
            # print('=--'*50)

            if len(context_name) > 1:
                identifier = context.find(context_name[0] + ':identifier')
                segment = context.find(context_name[0] + ':segment')
            else:
                identifier = context.find('identifier')
                segment = context.find('segment')
            if not segment and identifier.text == entitycentralindexkey:
                # print(context)
                instant_context_id = context['id']

        # if end_date[0] == "2021":
        #     print('-instant'*10)
        #     print(instant_context_id)
        #     print(documentperiodenddate)
        #     print('-instant'*10)

        # temp = [key, m, value,
        #         dic_company_info['company_info']['SIC'], dic_company_info['company_info']['ticker'],
        #         dic_company_info['statements']
        #         ]

        # print(dic_data_year[4], int(dic_data_year[2]['dei']['documentfiscalyearfocus']), dic_data_year[3], dic_data_year[5])

        matching_accounts, accounts_, used_accounting_standards = \
            self.get_matching_accounts(dic_data_year[4], int(dic_data_year[2]['dei']['documentfiscalyearfocus']),
                                       dic_data_year[3], dic_data_year[5])

        # print(222222222222222222222)
        dic_data_year[2]['matching_accounts'] = matching_accounts
        # print(matching_accounts)
        # print(3333333333333333333)
        year_data = {}
        # log_debug("get_data_for_years (after getting matching_accounts): "+dic_data_year[3]+", "+str(dic_data_year[0]))
        # print("Current Time 6 =", datetime.datetime.now().strftime("%H:%M:%S"))
        accounts_flow = {}
        for k in accounts_['flow']:
            accounts_flow[k[0:92]] = accounts_['flow'][k]

        accounts_instant = {}
        for k in accounts_['instant']:
            accounts_instant[k[0:92]] = accounts_['instant'][k]

        # if str(dic_data_year[0]) == "2021":
        #     print(int(dic_data_year[2]['dei']['documentfiscalyearfocus']))
        #     print(str(dic_data_year[0]))
        #     print(accounts_instant)

        for std in used_accounting_standards:
            tag_list = soup.find_all(re.compile(std + ":"))
            for tag in tag_list:
                name_ = tag.name.split(":")
                try:
                    if name_[1][0:92] in accounts_instant and tag['contextref'] == instant_context_id:
                        if accounts_instant[name_[1][0:92]][2] == std:
                            order = accounts_instant[name_[1][0:92]][0]
                            scale = accounts_instant[name_[1][0:92]][1]
                            if scale == 1:
                                year_data[order] = tag.text
                            else:
                                year_data[order] = int(tag.text) / scale
                    if name_[1][0:92] in accounts_flow and tag['contextref'] == flow_context_id:
                        order = accounts_flow[name_[1][0:92]][0]
                        scale = accounts_flow[name_[1][0:92]][1]
                        if scale == 1:
                            year_data[order] = tag.text
                        else:
                            year_data[order] = int(tag.text) / scale

                except Exception as ex:
                    pass
                    # print("Error: year=" + str(dic_data_year[0]) + "   dic=" + dic_data_year[1]['href'] + "   " + str(ex) + tag.text)

        # print(year_data)
        # print(4444444444444444444)

        dic_data_year[2]['year_data'] = year_data

        log_debug("End get_data_for_years: " + dic_data_year[3] + ", " + str(dic_data_year[0]))
        # print("Current Time End get data = " + str(dic_data_year[0]), datetime.datetime.now().strftime("%H:%M:%S"))
        return dic_data_year

    # to be deleted --
    def get_matching_accounts(self, ticker, year, sic, statements):
        try:
            if year <= self.xbrl_base_year:
                years = sorted(range(year, self.xbrl_base_year + 1), reverse=True)
            else:
                years = range(self.xbrl_base_year, year + 1)
        except Exception as ex:
            print(ex)
        matches = XBRLValuationAccountsMatch.objects.filter((Q(year=0) & Q(company__industry__sic_code=sic)) |
                                                            (Q(year__in=years) & Q(company__ticker=ticker))).all()
        used_accounting_standards = []
        dic_matches = {}
        for m in matches:
            if m.year == 0:
                dic_matches[m.account.order] = [m.match_account, m.accounting_standard]
                if m.accounting_standard not in used_accounting_standards:
                    used_accounting_standards.append(m.accounting_standard)

        # print('dic_matches 0')
        # print(year)
        # print(dic_matches)
        # print('dic_matches 0')

        for y in years:
            for m in matches:
                if m.year == y:
                    dic_matches[m.account.order] = [m.match_account, m.accounting_standard]
                    if m.accounting_standard not in used_accounting_standards:
                        used_accounting_standards.append(m.accounting_standard)

        # print('dic_matches')
        # print(year)
        # print(dic_matches)
        # print('dic_matches')

        matching_accounts = {}
        accounts_ = {'instant': {}, 'flow': {}}

        for st_ in statements:
            for a_order in statements[st_]['accounts']:
                # statements[st_]['accounts'][a_order] = [a.account, a.type, a.scale]
                # dic_matches[a_order] = [m.match_account, m.accounting_standard]
                # print(a_order)
                # if int(a_order) in dic_matches:
                #     print('in')
                # else:
                #     print('not')
                # print('str')
                # print(dic_matches)
                # if str(a_order) in dic_matches:
                #     print('in')
                # else:
                #     print('not')
                # print('str')

                try:
                    if int(a_order) in dic_matches:
                        ma_, ma_std = dic_matches[int(a_order)][0].lower(), dic_matches[int(a_order)][1].lower()
                    else:
                        ma_, ma_std = "", ""
                except Exception as ex:
                    pass
                    # print('ex')
                    # print(ex)

                if ma_ != '':
                    if statements[st_]['accounts'][a_order][1] == 1:
                        accounts_['instant'][ma_] = (a_order, statements[st_]['accounts'][a_order][2], ma_std)
                    else:
                        accounts_['flow'][ma_] = (a_order, statements[st_]['accounts'][a_order][2], ma_std)
                matching_accounts[a_order] = [ma_, ma_std]

        # print('matching_accounts')
        # print(year)
        # print(matching_accounts)
        # print('matching_accounts')

        return matching_accounts, accounts_, used_accounting_standards

    def save_industry_default(self, year, ticker, sic):
        dic = {'status': 'ok'}
        if year == self.xbrl_base_year:
            # print('-1'*20)
            industry = XBRLIndustryInfo.objects.get(sic_code=sic)
            c, created = XBRLCompanyInfo.objects.get_or_create(industry=industry, company_name=sic, ticker=sic, cik=sic)
            try:
                # print('-2'*20)
                # zero_company = XBRLValuationAccountsMatch.objects.filter(Q(company__ticker=ticker) & Q(year=year)).all()
                # zero_company.update(company=c, year=0)
                for m in XBRLValuationAccountsMatch.objects.filter(Q(company__ticker=ticker) & Q(year=year)).all():
                    cs, created_s = XBRLValuationAccountsMatch.objects.get_or_create(year=0, company=c,
                                                                                     account=m.account)
                    if created_s:
                        cs.match_account = m.match_account
                        cs.accounting_standard = m.accounting_standard
                        cs.save()
                        m.delete()
            except Exception as ex:
                print(ex)
                dic = {'status': 'ko'}
        else:
            dic = {'status': 'You can update only data of year 2020.'}
        return dic

    # # #
    # Functions for Forecasted EPS
    def get_sp500(self):
        sp500_url = 'https://en.wikipedia.org/wiki/List_of_S%26P_500_companies'
        self.sp_tickers = list(pd.read_html(sp500_url)[0]['Symbol'].values)
        self.sp_tickers = [x.split('.')[0] for x in self.sp_tickers]
        for ticker in self.sp_tickers:
            try:
                XBRLCompanyInfo.objects.get(ticker=ticker)
            except Exception as ex:
                try:
                    dic = self.create_company_by_ticker(ticker=ticker)
                except Exception as exx:
                    self.sp_tickers.remove(ticker)
                    # print('exx')
                    # print(exx)

        dic = {'status': 'ok', 'sp_tickers': self.sp_tickers}
        log_debug("End load_sp_returns.")
        return dic

    def month_to_num(self, month):
        return {
            'January': '01',
            'February': '02',
            'March': '03',
            'April': '04',
            'May': '05',
            'June': '06',
            'July': '07',
            'August': '08',
            'September': '09',
            'October': '10',
            'November': '11',
            'December': '12'
        }[month]

    def add_zero(self, s):
        if len(s) == 1:
            s = "0" + s
        return s

    def upload_old_earning_forecast_sp500(self):
        # XBRLSPEarningForecast.truncate()
        # print('upload_old_earning_forecast_sp500')
        sp_tickers = self.get_sp500()['sp_tickers']
        for ticker_ in sp_tickers:
            # print(ticker_)
            try:
                company_ = XBRLCompanyInfo.objects.get(ticker=ticker_)
                s, c = XBRLSPStatistics.objects.get_or_create(company=company_)
            except Exception as eex:
                pass
                # print("error sp_tickers: 111" + ticker_ + " : " + str(eex))
        # print("End check sp500")
        for file in os.listdir(self.TEXT_PATH):
            # log_debug("Start Processing file: " + file)
            # print(file)
            log_debug("Start file: " + file +" wait for it to finish/")
            file_path = f"{self.TEXT_PATH}/{file}"
            with open(file_path, 'r') as f:
                text = f.read()
                soup = BeautifulSoup(text, 'html.parser')
                rows = soup.find_all('tr')
                for row in rows:
                    try:
                        cells = row.find_all('td')
                        if len(cells) < 3:
                            dd = cells[0].text.split(",")
                            ddd = dd[1].strip().split(" ")
                            mm = self.month_to_num(ddd[0])
                            yy = dd[2].strip()
                            dd = ddd[1]
                            date_str = yy + "-" + mm + "-" + self.add_zero(dd)
                            date_ = parse_date(date_str)
                        else:
                            # if cells[2].text != '--':

                            ticker = cells[1].find('a').text
                            # print(ticker)
                            if ticker in sp_tickers:
                                # print('-'*20)
                                # print('in sp')
                                # print(date_str)
                                # print(ticker)
                                # log_debug("Ticker: " + ticker)
                                # print('-'*5)
                                actual = cells[2].text
                                forecast = cells[3].text.split('/')[1].lstrip()
                                # print("actual: " + str(actual) + " forecast: " + str(forecast))
                                # print('-'*20)

                                company = XBRLCompanyInfo.objects.get(ticker=ticker)
                                year = date_.year
                                quarter = math.ceil(date_.month / 3)
                                ef, ct = XBRLSPEarningForecast.objects.get_or_create(company=company, year=year,
                                                                                     quarter=quarter)
                                # print(ct)
                                # print(ef)
                                # print("Got or created record")
                                # print('-'*20)
                                try:
                                    forecast_ = float(forecast)
                                    ef.forecast = forecast_
                                    ef.save()
                                    # print("forecast saved")
                                except Exception as eex:
                                    # print("error forecast 123: " + str(eex))
                                    pass

                                try:
                                    actual_ = float(actual)
                                    ef.actual = actual_
                                    ef.save()
                                    # print("Actual saved")
                                except Exception as eex:
                                    # print("error actual 234: " + str(eex))
                                    ef.actual = None
                                    # pass

                                try:
                                    ef.date = date_
                                    ef.save()
                                    # print("Date saved")
                                except Exception as eex:
                                    pass
                                    # print("error Date act for 444: " + str(eex))

                                # try:
                                #     ef.save()
                                # except Exception as eex:
                                #     pass
                                #     # print("error Save act for: " + str(eex))

                                # print('ticker1')
                                try:
                                    self.get_ticker_prices(earning_forecast=ef)
                                except Exception as eex:
                                    pass
                                    # print("error Price act for 555: " + str(eex))

                                try:
                                    next_release_date = self.get_next_relealse_date(cells[1], date_)
                                    # print(next_release_date)
                                    ef.next_release_date = next_release_date
                                    s = XBRLSPStatistics.objects.get(company__ticker=ticker)
                                    s.next_release_date = next_release_date
                                    s.save()
                                    ef.save()
                                    # print("next_release_date saved")
                                except Exception as eex:
                                    pass
                                    # print("error next_release_date: " + str(eex))

                                # print(ticker)
                                # print('ticker2')

                                try:
                                    company.company_statistic.set_company_statistics()
                                except Exception as eex:
                                    pass
                                    # print("error company.set_company_statisitcs 666: " + str(eex))
                    except Exception as ex:
                        # if ticker in sp_tickers:
                        # print('ex')
                        # print(ex)
                        # print(ticker)
                        # print(cells[2].text)
                        # print('ex')
                        pass
            print("End Processing file: " + file)
            log_debug("End Processing file: " + file)
        log_debug("End Processing all files1: ")
        log_debug("End Processing all files2: ")

    def get_next_relealse_date(self, cell, date):
        # print("get_next_relealse_date ")
        # print("get_next_relealse_date ")
        # print("get_next_relealse_date ")
        # print(cell)
        # print(cell.find('a'))
        # print(cell.find('a')['href'])
        url = cell.find('a')['href']
        url = "https://www.investing.com/" + url
        # print(url)
        id_ = "earningsHistory" + cell['_r_pid']
        # print(id_)
        headers = {'User-Agent': 'amos@drbaranes.com'}
        sp_resp = requests.get(url, headers=headers, timeout=30)
        sp_str = sp_resp.text
        soup = BeautifulSoup(sp_str, 'html.parser')
        table_tag = soup.find('table', id=id_)
        # print(table_tag)
        tbody_tag = table_tag.find('tbody')
        # print("-"*100)
        # print(date)
        # print("-"*10)
        try:
            # date_str = tbody_tag.find_all('tr')[0]['event_timestamp']
            next_date = datetime.datetime.now()
            next_date = (next_date + timedelta(days=180)).date()
            # print(next_date)
            for k in tbody_tag.find_all('tr'):
                date_str = k['event_timestamp']
                # print(date_str)
                date_ = parse_date(date_str)
                if date < date_ < next_date:
                    next_date = date_
                # print("-"*5)
        except Exception as ex:
            print("Error 105" + str(ex))
        # print(next_date)
        # print("-"*100)

        return next_date

    def get_earning_forecast_sp500(self):
        # print('get_earning_forecast_sp500')
        sp_tickers = self.get_sp500()['sp_tickers']
        for ticker_ in sp_tickers:
            try:
                # print(ticker_)
                company_ = XBRLCompanyInfo.objects.get(ticker=ticker_)
                # print(company_)
                s, c = XBRLSPStatistics.objects.get_or_create(company=company_)
                # print(c)
            except Exception as eex:
                pass
                # print("error 345 sp_tickers: " + ticker_ + " : " + str(eex))
        headers = {'User-Agent': 'amos@drbaranes.com'}
        url = "https://www.investing.com/earnings-calendar/"
        sp_resp = requests.get(url, headers=headers, timeout=30)
        # print("Current Time 11 =", datetime.datetime.now().strftime("%H:%M:%S"))
        sp_str = sp_resp.text
        soup = BeautifulSoup(sp_str, 'html.parser')
        form_tag = soup.find('form', id='earningsCalendarForm')
        # print(form_tag)
        form_cells = form_tag.find_all('input')
        # print(form_cells)
        date_str = form_cells[0]['value']
        date_ = parse_date(date_str)

        table_tag = soup.find('table', id='earningsCalendarData')
        tbody_tag = table_tag.find('tbody')
        try:
            rows = tbody_tag.find_all('tr')
            for row in rows:
                try:
                    cells = row.find_all('td')
                    # print(row)
                    if len(cells) > 3:
                        # print('row 3')
                        # if cells[2].text != '--':
                        ticker = cells[1].find('a').text
                        # print(ticker)
                        if ticker in sp_tickers:
                            # print('ticker in sp')
                            # print(ticker)
                            year = date_.year
                            quarter = math.ceil(date_.month / 3)
                            company = XBRLCompanyInfo.objects.get(ticker=ticker)
                            ef, ct = XBRLSPEarningForecast.objects.get_or_create(company=company, year=year,
                                                                                 quarter=quarter)
                            if cells[2].text != '--':
                                actual = cells[2].text
                                forecast = cells[3].text.split('/')[1].lstrip()
                                ef.forecast = forecast
                                ef.actual = actual
                                ef.date = date_
                                next_release_date = self.get_next_relealse_date(cells[1], date_)
                                ef.next_release_date = next_release_date
                                ef.save()
                                s, c = XBRLSPStatistics.objects.get_or_create(company__ticker=ticker)
                                s.next_release_date = next_release_date
                                s.save()
                                self.get_ticker_prices(ef)
                                try:
                                    company.company_statistic.set_company_statistics()
                                except Exception as eex:
                                    pass
                                    # print("error company.set_company_statisitcs: " + str(eex))
                except Exception as ex:
                    print("error 102: " + str(ex))
        except Exception as ex:
            return print("Error 101")
        dic = {'status': 'ok'}
        return dic

    def get_announcement_time_day(self, s_day):
        try:
            s_from = "2021-01-01"
            s_to = "2021-12-31"
            url = "https://finance.yahoo.com/calendar/earnings?from=" + s_from + "&to=" + s_to + "&day=" + s_day
            log_debug(url)
            headers = {'User-Agent': 'amos@drbaranes.com'}
            sp_resp = requests.get(url, headers=headers, timeout=30)
            sp_str = sp_resp.text
            # print(sp_str)
            soup = BeautifulSoup(sp_str, 'html.parser')
            id_ = "cal-res-table"
            div_tag = soup.find('div', id=id_)
            # print('--12--')
            tbody_tag = div_tag.find('tbody')
            # print('--13--')
            # print("-"*100)
            # print(tbody_tag)
            # print("-"*10)
            sp_tickers = self.get_sp500()['sp_tickers']
            rows = tbody_tag.find_all('tr')
            # print('--14--')
            for row in rows:
                try:
                    cells = row.find_all('td')
                    if len(cells) > 2:
                        sa = ""
                        # print(cells[2])
                        sa_ = cells[2].text
                        if sa_ == "Time Not Supplied":
                            sa = "N"
                        elif sa_ == "TAS":
                            sa = "T"
                        elif sa_ == "Before Market Open":
                            sa = "B"
                        elif sa_ == "After Market Close":
                            sa = "A"
                        else:
                            sa = "O"
                        ticker = cells[0].find('a').text
                        if ticker in sp_tickers:
                            # log_debug(ticker +":" + sa_ + ":" + sa)
                            s = XBRLSPStatistics.objects.get(company__ticker=ticker)
                            s.announcement_time = sa
                            s.save()
                            log_debug("Saved: " + ticker + " : " + sa)
                except Exception as ex:
                    log_debug("Error 102: " + ticker + " : " + str(ex))
        except Exception as ex:
            log_debug("Error 101: " + str(ex))

    def get_announcement_time(self, m=""):
        try:
            m = int(m)
            nd = 31
            if m in [1, 3, 5, 7, 8, 10, 12]:
                nd = 32
            elif m == 2:  # do not care about 29
                nd = 28
            sm = self.add_zero(str(m))
            for d in range(1, nd):
                sd = self.add_zero(str(d))
                s_day = "2021-" + sm + "-" + sd
                # print(s_day)
                log_debug("Start day: " + s_day)
                # print('-----------')
                try:
                    self.get_announcement_time_day(s_day=s_day)
                except Exception as ex:
                    log_debug("Error 202 time: " + str(ex))
            # print(s_day)
            # self.get_announcement_time_day(s_day="2021-07-20")
        except Exception as ex:
            log_debug("Error 201 time: " + str(ex))
        dic = {'status': 'ok'}
        return dic

    def get_ticker_prices(self, earning_forecast):
        # https://pypi.org/project/yahoofinancials/
        # https://www.analyticsvidhya.com/blog/2021/06/download-financial-dataset-using-yahoo-finance-in-python-a-complete-guide/

        # print(earning_forecast.company.ticker)
        yahoo_financials = YahooFinancials(earning_forecast.company.ticker)
        today = earning_forecast.date
        # print(today)
        yesterday = (today + timedelta(days=-1))
        # print(yesterday)
        today_str = str(today)
        yesterday_str = str(yesterday)
        data = yahoo_financials.get_historical_price_data(start_date='2015-01-01',
                                                          end_date=str(today + timedelta(days=1)),
                                                          time_interval='daily')
        # print(data)
        df = pd.DataFrame(data[earning_forecast.company.ticker]['prices'])
        df = df.drop('date', axis=1).set_index('formatted_date')
        # print(df.tail())
        try:
            today_price = int(100 * df.filter(items=[today_str], axis=0)['adjclose']) / 100
            # print(today_price)
            earning_forecast.today_price = today_price
        except Exception as ex:
            pass
            # print('ex 1')
            # print(ex)
        is_ok = False
        while not is_ok:
            try:
                yesterday_price = int(100 * df.filter(items=[yesterday_str], axis=0)['adjclose']) / 100
                # print(yesterday_price)
                earning_forecast.yesterday_price = yesterday_price
                is_ok = True
            except Exception as ex:
                yesterday = (yesterday + timedelta(days=-1))
                # print(yesterday)
                yesterday_str = str(yesterday)
                # print('ex 2')
                # print(ex)
        earning_forecast.save()
        dic = {'status': 'ok'}
        return dic

    # Should delete this function
    def get_earning_forecast_sp500_view(self):
        d = {}
        for t in XBRLSPEarningForecast.objects.all():
            if t.company.ticker not in d:
                d[t.company.ticker] = {}
            new_ticker = False
            if t.year not in d[t.company.ticker]:
                d[t.company.ticker][t.year] = {}
                new_ticker = True
            if t.quarter not in d[t.company.ticker][t.year]:
                d[t.company.ticker][t.year][t.quarter] = ['f', 'a', 'p', '-p']
            d[t.company.ticker][t.year][t.quarter][0] = str(t.forecast)
            d[t.company.ticker][t.year][t.quarter][1] = str(t.actual)
            d[t.company.ticker][t.year][t.quarter][2] = str(t.today_price)
            d[t.company.ticker][t.year][t.quarter][3] = str(t.yesterday_price)
            if new_ticker:
                d[t.company.ticker][t.year][t.quarter].append(str(t.next_release_date))

        # print(d)
        return {'status': 'ok', 'earning_forecast_sp500_view': d}

    def get_earning_forecast_sp500_view_main_detail(self, ticker):
        # print(ticker)
        d = {}
        for r in XBRLSPEarningForecast.objects.filter(company__ticker=ticker).order_by('-year', 'quarter').all():
            try:
                d[str(int(r.year) * 10 + int(r.quarter))] = [r.year, r.quarter, float(r.forecast), float(r.actual),
                                                             float(r.today_price), float(r.yesterday_price)]
            except Exception as ex:
                pass
                # print('error get_earning_forecast_sp500_view_main')
                # print(ex)
        # print(d)
        return {'status': 'ok', 'get_earning_forecast_sp500_view_main_detail': d}

    def get_earning_forecast_sp500_view_main(self, order_by):
        # print('get_earning_forecast_sp500_view_main')
        # print('order_by')
        # print(order_by)
        # print('order_by')

        d = {}
        for t in XBRLSPStatistics.objects.select_related("company__industry__main_sic").order_by(order_by).all():
            if t.company.ticker not in d:
                try:
                    d[str(t.company.ticker)] = {"cn": str(t.company.company_name), "nrd": str(t.next_release_date),
                                                "mapc": str(t.mean_abs_price_change),
                                                "maafc": str(t.mean_abs_actual_forecast_change),
                                                "maafcm": str(t.mean_abs_actual_forecast_change_money),
                                                "cafp": str(t.correlation_afp),
                                                "ud": str(t.updated),
                                                "sic": t.company.industry.sic_code,
                                                "msic": t.company.industry.main_sic.sic_code,
                                                "bfp": str(t.butterfly_price),
                                                "stp": str(t.straddle_price),
                                                "a": str(t.announcement_time)}
                except Exception as ex:
                    print('error get_earning_forecast_sp500_view_main')
                    print(ex)
        # print(d)
        return {'status': 'ok', 'earning_forecast_sp500_view_main': d}

    # # #
    #  Data processing
    def create_company_by_ticker(self, ticker=None):
        try:
            company_id = -1
            type = '10-k'
            #
            headers = {'User-Agent': 'amos@drbaranes.com'}
            base_url = "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={}&type={}"  # &dateb={}"
            url = base_url.format(ticker, type)
            #
            # print('-'*100)
            # print(url)
            # print('-'*100)
            #
            edgar_resp = requests.get(url, headers=headers, timeout=30)
            edgar_str = edgar_resp.text
            #
            cik0 = ''
            cik_re = re.compile(r'.*CIK=(\d{10}).*')
            results = cik_re.findall(edgar_str)
            if len(results):
                results[0] = int(re.sub('\.[0]*', '.', results[0]))
                cik0 = str(results[0])
            # print('-1'*10)
            # print(cik0)
            # print('-1'*10)
            #
            sic0 = ''
            cik_re = re.compile(r'.*SIC=(\d{4}).*')
            results = cik_re.findall(edgar_str)
            if len(results):
                results[0] = int(re.sub('\.[0]*', '.', results[0]))
                sic0 = int(results[0])

            # print('-1'*10)
            # print(sic0)
            # print('-1'*10)

            soup = BeautifulSoup(edgar_str, 'html.parser')
            company_name_ = soup.find('span', class_='companyName').text
            company_name_n = company_name_.index("CIK#")
            company_name_ = company_name_[0:company_name_n]
            # print(company_name_)

            table_tag = soup.find('table', class_='tableFile2')
            rows = table_tag.find_all('tr')
            dic_data = {}
            for row in rows:
                try:
                    cells = row.find_all('td')
                    if len(cells) > 3:
                        # print(cells[3].text)
                        for filing_year in range(self.xbrl_start_year, self.today_year + 1):
                            if str(filing_year) in cells[3].text:
                                dic_data[filing_year] = {'href': 'https://www.sec.gov' + cells[1].a['href']}
                except Exception as ex:
                    pass

            # print('len(dic_data)')
            # print(len(dic_data))

            if len(dic_data) > 0:
                industry_ = XBRLIndustryInfo.objects.get(sic_code=sic0)
                # print(industry_)
                # print(cik0)
                # print(sic0)
                # print(ticker)
                # print(company_name_)
                # print(company_name_[0].upper())

                company, created = XBRLCompanyInfo.objects.get_or_create(industry=industry_, ticker=ticker, cik=cik0,
                                                                         company_name=company_name_,
                                                                         company_letter=company_name_[0].upper())
                company_id = company.id

        except Exception as ex:
            pass
            # print('error 108: '+str(ex))
        return {'status': 'ok', 'id': company_id, 'company_name': company_name_}

    def get_companies_for_exchange(self, exchange, exchange_url):
        companies = pd.DataFrame(columns=['exchange', 'name', 'ticker', 'letter'])
        company_name = []
        company_ticker = []
        company_letter = []
        letters = string.ascii_uppercase
        for letter in letters:
            company_name, company_ticker, company_letter = self.get_companies_for_letter(
                letter, exchange_url + letter, company_name, company_ticker, company_letter)
        companies['name'] = company_name
        companies['ticker'] = company_ticker
        companies['exchange'] = exchange
        companies['letter'] = company_letter
        companies = companies[companies['ticker'] != '']
        return companies

    def get_companies_for_letter(self, letter, url, company_name, company_ticker, company_letter):
        page = requests.get(url, timeout=30)
        soup = BeautifulSoup(page.text, 'html.parser')
        odd_rows = soup.find_all('tr', attrs={'class': 'ts0'})
        even_rows = soup.find_all('tr', attrs={'class': 'ts1'})
        for r in odd_rows:
            cs = r.find_all('td')
            company_name.append(cs[0].text.strip())
            company_ticker.append(cs[1].text.strip())
            company_letter.append(letter)
        for r in even_rows:
            cs = r.find_all('td')
            company_name.append(cs[0].text.strip())
            company_ticker.append(cs[1].text.strip())
            company_letter.append(letter)
        return company_name, company_ticker, company_letter
    #

    # general purpose functions for testing
    def test(self):
        try:
            today_ = datetime.datetime.today()
            nday = today_.weekday()  # Monday = 0
            h = today_.hour
            m = today_.minute
            log_debug(str(nday)+" : "+str(h)+" : "+str(m))

            data = ()
            ssql = " insert into corporatevaluation_XBRLRealEquityPricesArchive(ticker,t,o,h,l,c,v) "
            ssql += "select ticker,t,o,h,l,c,v from corporatevaluation_XBRLRealEquityPrices"
            # print(ssql)

            count = SQL().exc_sql(ssql, data)
            # print(count)
            XBRLRealEquityPrices.truncate()

        except Exception as ex:
            print("Error200: "+str(ex))
            pass
        result = {'status': "ok"}
        return result

    def test1(self):
        XBRLRealEquityPrices.truncate()
        result = {'status': "ok"}
        return result

    def test2(self):
        XBRLRealEquityPricesArchive.truncate()
        result = {'status': "ok"}
        return result

#   --- Stage 1 get list of companies ---
    def set_sic_code(self):
        log_debug("Start set_sic_code")
        url = 'https://en.wikipedia.org/wiki/Standard_Industrial_Classification'
        headers = {'User-Agent': 'amos@drbaranes.com'}
        page = requests.get(url, headers=headers, timeout=30)
        soup = BeautifulSoup(page.text, 'html.parser')
        tables = soup.find_all('table')
        rows = tables[0].find_all('tr')
        z = 0
        for r in rows:
            if z == 0:
                z = 1
                continue
            cs = r.find_all('td')
            sic_code_ = int(cs[0].text.strip().split('-')[1])
            sic_description_ = cs[1].text.strip()
            XBRLMainIndustryInfo.objects.get_or_create(sic_code=sic_code_, sic_description=sic_description_)
        url = 'https://www.sec.gov/corpfin/division-of-corporation-finance-standard-industrial-classification-sic-code-list'
        page = requests.get(url, headers=headers, timeout=30)
        soup = BeautifulSoup(page.text, 'html.parser')
        rows = soup.find_all('tr')
        main_sics = XBRLMainIndustryInfo.objects.all()
        z = 0
        for r in rows:
            if z == 0:
                z = 1
                continue
            cs = r.find_all('td')
            # print(cs[0].text.strip() + ':' + cs[2].text.strip())

            sic_code_ = int(cs[0].text.strip())
            sic_description_ = cs[2].text.strip()
            main_sic_ = 1999
            for main_sic in main_sics:
                if sic_code_ < main_sic.sic_code:
                    main_sic_ = main_sic
                    break
            XBRLIndustryInfo.objects.get_or_create(sic_code=sic_code_, main_sic=main_sic_,
                                                   sic_description=sic_description_)
        log_debug("End set_sic_code")
        return {'status': 'ok'}

    def get_all_companies(self):
        log_debug("Start get_all_companies")
        exchanges = {
            'nyse': 'nyse/newyorkstockexchange',
            'nasdaq': 'nasdaq/nasdaq',
            'amex': 'amex/americanstockexchange'}

        # writer = pd.ExcelWriter(self.EXCEL_PATH+'/all_companies.xlsx', engine='xlsxwriter')
        n = 0
        companies = None
        for exchange in exchanges:
            # print(exchange)
            url = 'https://www.advfn.com/' + exchanges[exchange] + '.asp?companies='
            df = self.get_companies_for_exchange(exchange=exchange, exchange_url=url)
            if n == 0:
                companies = df
                n += 1
            else:
                frames = [companies, df]
                companies = pd.concat(frames)
            # df.to_excel(writer, sheet_name=exchange)
            # Close the Pandas Excel writer and output the Excel file.
        try:
            XBRLCompanyInfoInProcess.truncate()
        except Exception as ex:
            print("Error 107:  " + str(ex))

        # print('done collecting data on companies')

        try:
            # print(companies)
            for i, c in companies.iterrows():
                # print(c)
                # print(c['exchange'])
                # print(c['name'])

                # print(c['ticker'])

                # print(c['letter'])
                a, created = XBRLCompanyInfoInProcess.objects.get_or_create(exchange=c['exchange'],
                                                                            company_name=c['name'],
                                                                            ticker=c['ticker'],
                                                                            company_letter=c['letter'])
        except Exception as ex:
            print("Error 2:  " + str(ex))

        # print('done loading data to XBRLCompanyInfoInProcess')

        # self.companies.reset_index(drop=True, inplace=True)
        # self.companies.to_excel(writer, sheet_name='all')
        # writer.save()
        log_debug("End get_all_companies")
        dic = {'status': 'ok'}
        return dic

    def clean_data_for_all_companies(self):
        log_debug("Start clean_data_for_all_companies")
        companies_ = XBRLCompanyInfoInProcess.objects.all()
        for company in companies_:
            ticker = company.ticker
            type = '10-k'
            try:
                #
                headers = {'User-Agent': 'amos@drbaranes.com'}
                base_url = "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={}&type={}"  # &dateb={}"
                url = base_url.format(ticker, type)
                #
                # print('-'*100)
                # print(url)
                # print('-'*100)
                #
                edgar_resp = requests.get(url, headers=headers, timeout=30)
                edgar_str = edgar_resp.text
                #
                cik0 = ''
                cik_re = re.compile(r'.*CIK=(\d{10}).*')
                results = cik_re.findall(edgar_str)
                if len(results):
                    results[0] = int(re.sub('\.[0]*', '.', results[0]))
                    cik0 = str(results[0])
                # print('-1'*10)
                # print(cik0)
                # print('-1'*10)
                #
                sic0 = ''
                cik_re = re.compile(r'.*SIC=(\d{4}).*')
                results = cik_re.findall(edgar_str)
                if len(results):
                    results[0] = int(re.sub('\.[0]*', '.', results[0]))
                    sic0 = int(results[0])

                # print('-1'*10)
                # print(sic0)
                # print('-1'*10)

                company.cik = cik0
                company.sic = sic0
                company.save()
            except Exception as ex:
                # print('error 106: '+str(ex))
                company.is_error = True
                company.message = "Can not get CIK or SIC: " + str(ex)
                company.cik = ''
                company.sic = 0
                company.save()
                continue

            # print('-2'*10)
            #
            # Find the document links
            soup = BeautifulSoup(edgar_str, 'html.parser')
            table_tag = soup.find('table', class_='tableFile2')
            try:
                # print('-3'*10)
                rows = table_tag.find_all('tr')
                # print('-4'*10)

            except Exception as ex:
                # print('-5'*10)
                # print(str(ex))

                company.message = "there are no data rows: " + str(ex)
                # print('-51'*10)
                try:
                    company.save()
                except Exception as exc:
                    pass
                    # print('-52'*10)
                    # print(str(exc))
                # print('-6'*10)
                continue

            dic_data = {}
            for row in rows:
                try:
                    cells = row.find_all('td')
                    if len(cells) > 3:
                        # print(cells[3].text)
                        # for filing_year in range(2019, 2020):
                        for filing_year in range(self.xbrl_start_year, self.today_year + 1):
                            if str(filing_year) in cells[3].text:
                                dic_data[filing_year] = {'href': 'https://www.sec.gov' + cells[1].a['href']}
                except Exception as ex:
                    pass
                    # print(ex)

            # print('-8'*10)

            # need to add send message on no data
            if len(dic_data) == 0:
                company.message = "there are no data."
                company.save()
                continue

            # print(cik0)
            # print('--9--'*5)
            company.save()
            # print('--10--'*5)
        log_debug("End clean_data_for_all_companies")
        return {'status': 'ok'}

    def copy_processed_companies(self):
        log_debug("Start copy_processed_companies")
        companies_ = XBRLCompanyInfoInProcess.objects.filter(is_error=False).all()
        for c in companies_:
            try:
                # print(c.ticker)
                i_ = XBRLIndustryInfo.objects.get(sic_code=c.sic)
                XBRLCompanyInfo.objects.get_or_create(industry=i_, exchange=c.exchange, company_name=c.company_name,
                                                      ticker=c.ticker, company_letter=c.company_letter, cik=c.cik)
            except Exception as exc:
                print(str(exc))
        log_debug("End copy_processed_companies")
        dic = {'status': 'ok'}
        return dic

#   --- Stage 2 Regions, countries, ----
    def load_tax_rates_by_country_year(self):
        dic = {'status': 'ko'}
        log_debug("Start load_tax_rates_by_country_year.")
        url = "https://files.taxfoundation.org/20210125115215/io1980-2020-Corporate-Tax-Rates-Around-the-World.csv.xlsx"
        file = "world_taxes"
        self.download_excel_file(url, file)
        df = self.load_excel_data(file)

        XBRLCountryYearData.truncate()
        XBRLCountry.truncate()
        XBRLRegion.truncate()
        log_debug("tables XBRLCountryYearData, XBRLCountry, XBRLRegion were cleaned.")

        for i, r in df.iterrows():
            try:
                region, c = XBRLRegion.objects.get_or_create(name=r['continent'])
                # log_debug('created region: ' + str(r['continent']))
            except Exception as ex:
                log_debug("Error 1 creating region: " + str(r['continent']) + " " + str(ex))

            oecd = True if int(r['oecd']) > 0 else False
            eu27 = True if int(r['eu27']) > 0 else False
            gseven = True if int(r['gseven']) > 0 else False
            gtwenty = True if int(r['gtwenty']) > 0 else False
            brics = True if int(r['brics']) > 0 else False

            try:
                country, c = XBRLCountry.objects.get_or_create(region=region, name=r['country'], iso_2=r['iso_2'],
                                                               iso_3=r['iso_3'], oecd=oecd, gseven=gseven, eu27=eu27,
                                                               gtwenty=gtwenty, brics=brics)
                # log_debug("Created country: " + str(r['continent']) + " " + str(r['country']))
            except Exception as ex:
                log_debug("Error 1 get country: " + str(r['continent']) + " " + str(r['country']) + " " + str(ex))

            d, c = XBRLCountryYearData.objects.get_or_create(country=country, year=int(r['year']))

            if not pd.isna(r['rate']):
                d.tax_rate = round(r['rate'], 2)

            if not pd.isna(r['gdp']):
                d.gdp = round(r['gdp'], 5)

            try:
                d.save()
                log_debug("Done: " + str(r['continent']) + " " + str(r['country']) + " " + str(r['year']))
            except Exception as ex:
                log_debug("Error 2 save country: " + str(r['continent']) + " " + str(r['country']) + " " + str(
                    r['year']) + " " + str(ex))

        dic['status'] = 'ok'
        log_debug("End load_tax_rates_by_country_year.")
        return dic

    def load_country_premium(self, request):
        log_debug("Start load_country_premium.")
        # print('in object load_country_premium(request)')
        match = {'Czech Republic': 'Czechia',
                 'Moldova': 'Republic of Moldova',
                 'United Kingdom': 'United Kingdom of Great Britain and Northern Ireland',
                 'Jersey (States of)': 'Jersey',
                 'Guernsey (States of)': 'Guernsey',
                 'Bolivia': 'Bolivia (Plurinational State of)',
                 "Côte d'Ivoire": "Cote d'Ivoire",
                 'Democratic Republic of Congo': 'Democratic Republic of the Congo',
                 'Congo (Democratic Republic of)': 'Democratic Republic of the Congo',
                 'Korea': 'Republic of Korea',
                 'Bolivia(Plurinational State of)': 'Bolivia (Plurinational State of)',
                 'United Kingdom of Great Britain and NorthernIreland': 'United Kingdom of Great Britain and Northern Ireland'}
        file = "ctrypremJuly21"
        df = self.load_excel_data(file, sheet_name="Data1")
        # load country and regions
        name_list = [x for x in df['name'].unique() if str(x) != 'nan']
        for r in name_list:
            # print(r)
            # print(df.loc[df['name'] == r])
            # print("------")
            z = 0
            for i, c in df.loc[df['name'] == r].iterrows():
                if z == 0:
                    # print(c['name'])
                    try:
                        region, created = XBRLRegion.objects.get_or_create(name=c['name'])
                        region.full_name = c['Region']
                        if created:
                            region.updated_adamodar = True
                        region.save()
                        # log_debug("load_country_premium : updated " + str(c['name']))
                    except Exception as ex:
                        log_debug("Error load_country_premium 10: " + str(c['name']) + " " + str(ex))
                    z = 1
                try:
                    if c['Country'] in match:
                        s_country = match[c['Country']]
                    else:
                        s_country = c['Country']

                    if not XBRLCountry.objects.filter(name=s_country).all().count() > 0:
                        XBRLCountry.objects.create(region=region, name=s_country, updated_adamodar=True)
                        # print('created')
                        # print(c['Country'])
                    else:
                        country = XBRLCountry.objects.filter(name=s_country).all()[0]
                        country.region = region
                        country.updated_adamodar = True
                        country.save()
                        # print('updated')
                        # print(c['Country'])
                        # log_debug("load_country_premium : updated country " + str(c['name']) + " " + c['Country'])
                except Exception as ex:
                    log_debug("Error load_country_premium 100: " + str(ex))
        # print('--SP Moodys and tax rates for 2020')
        # log_debug('--SP Moodys and tax rates for 2020')
        for i, c in df.iterrows():
            s_country_ = 'Country1'
            if c[s_country_] in match:
                s_country = match[c[s_country_]]
            else:
                s_country = c[s_country_]
            try:
                if not s_country:
                    break
                country = XBRLCountry.objects.filter(name=s_country).all()[0]
                data, created = XBRLCountryYearData.objects.get_or_create(country=country, year=2020)
                if str(c['sp_rating_2020']) == 'nan':
                    s_sp = ''
                else:
                    s_sp = c['sp_rating_2020']
                data.sp_rating = s_sp
                # print(s_country + "                  " + str(c['Moodys_rating_2020']))

                if str(c['Moodys_rating_2020']) == 'nan':
                    s_moodys = ''
                else:
                    s_moodys = c['Moodys_rating_2020']
                data.moodys_rating = s_moodys

                data.tax_rate = c['tax_rate_2020']
                data.save()
                # print(data)
                # log_debug("data added for: " + s_country)
            except Exception as ex:
                print('Error 222: for ' + str(ex))
                log_debug("Error load_country_premium 101: " + s_country + " " + str(ex))
        # log_debug('-- End SP Moodys and tax rates for 2020')

        # print('--composite_risk_rating_7_21 --')
        # log_debug('--composite_risk_rating_7_21 --')

        for i, c in df.iterrows():
            s_country_ = 'Country2'
            if c[s_country_] in match:
                s_country = match[c[s_country_]]
            else:
                s_country = c[s_country_]
            # print(c)

            s_country = str(s_country)
            if s_country == 'nan':
                break
            try:
                country, created = XBRLCountry.objects.get_or_create(name=s_country)
                data, created = XBRLCountryYearData.objects.get_or_create(country=country, year=2020)
                data.composite_risk_rating = c['composite_risk_rating_7_21']
                data.save()
                # print(data)
                # log_debug("composite_risk_rating_7_21 data added for: " + s_country)
            except Exception as ex:
                print(ex)
                log_debug("Error load_country_premium 102: " + s_country + " " + str(ex))

        # print('-- CDS_07_01_20211 --')
        # log_debug('-- CDS_07_01_20211 --')
        for i, c in df.iterrows():
            s_country_ = 'Country3'
            if c[s_country_] in match:
                s_country = match[c[s_country_]]
            else:
                s_country = c[s_country_]
            # print(s_country)
            # print(c)
            s_country = str(s_country)
            if s_country == 'nan':
                break

            try:
                country = XBRLCountry.objects.filter(name=s_country).all()[0]
                data, created = XBRLCountryYearData.objects.get_or_create(country=country, year=2020)

                # need to consider this.  Since three country with missing data turned to 0.
                if str(c['CDS_01_01_2021']) == 'nan':
                    s_ = 0
                else:
                    s_ = 100 * c['CDS_01_01_2021']
                data.cds = s_
                data.save()
                # print(data.cds)
                # log_debug("CDS_07_01_20211 data added for: " + s_country)
            except Exception as ex:
                log_debug("Error load_country_premium 102: " + s_country + " " + str(ex))

        # print('-- SPMoodys 1--')
        # log_debug('-- SPMoodys 1--')
        XBRLSPMoodys.truncate()
        for i, c in df.iterrows():
            try:
                if str(c['sp_moodys_year']) != 'nan':
                    d, created = XBRLSPMoodys.objects.get_or_create(year=c['sp_moodys_year'], sp=c['SP'],
                                                                    moodys=c['Moodys'])
            except Exception as ex:
                pass

        # print('-- SPMoodys 2--')
        # log_debug('-- SPMoodys 2--')
        for i, c in df.iterrows():
            try:
                if str(c['Rating']) != 'nan':
                    d, created = XBRLSPMoodys.objects.get_or_create(year=int(c['rating_year']), moodys=c['Rating'])
                    dd = round(c['Default_Spread_1_1_2021'] / 100, 2)
                    d.default_spread = dd
                    d_from = round(100 * c['score_from'], 2) / 100
                    d_to = round(100 * c['score_to'], 2) / 100
                    d.score_from = d_from
                    d.score_to = d_to
                    d.save()
                    # log_debug("CDS_07_01_20211 data added for: " + str(c['rating_year']) + " " + str(c['Rating']))
            except Exception as ex:
                log_debug("Error load_country_premium 104: " + str(c['Rating']) + str(ex))

        today = datetime.date.today()
        today = datetime.date(today.year, today.month, today.day)
        today_5 = str(datetime.date(today.year - 5, today.month, today.day))
        today = str(today)
        # print(today)
        # print(today_5)
        s_baml = "https://fred.stlouisfed.org/graph/fredgraph.xls?bgcolor=%23e1e9f0&chart_type=line&drp=0&fo=open%20sans&graph_bgcolor=%23ffffff&height=450&mode=fred&recession_bars=on&txtcolor=%23444444&ts=12&tts=12&width=1168&nt=0&thu=0&trc=0&show_legend=yes&show_axis_titles=yes&show_tooltip=yes&id=BAMLEMPBPUBSICRPIEY&scale=left&cosd=" + today_5 + "&coed=" + today + "&line_color=%234572a7&link_values=false&line_style=solid&mark_type=none&mw=3&lw=2&ost=-99999&oet=99999&mma=0&fml=a&fq=Daily&fam=avg&fgst=lin&fgsnd=2020-02-01&line_index=1&transformation=lin&vintage_date=2021-09-17&revision_date=2021-09-17&nd=1998-12-31"
        s_bmi = "http://www.spglobal.com/spdji/en/idsexport/file.xls?hostIdentifier=48190c8c-42c4-46af-8d1a-0cd5db894797&redesignExport=true&languageId=1&selectedModule=PerformanceGraphView&selectedSubModule=Graph&yearFlag=fiveYearFlag&indexId=5457901"
        path_baml = self.download_excel_file(s_baml, "baml", ext='xls')
        path_bmi = self.download_excel_file(s_bmi, "bmi", ext='xls')
        try:
            wb_baml = xlrd.open_workbook(path_baml)
            wb_bmi = xlrd.open_workbook(path_bmi)
        except Exception as ex:
            print(ex)

        # log_debug('-- Downloaded baml bmi --')
        sh_bmi = wb_bmi.sheet_by_index(0)
        data_bmi = []
        z = 0
        for cur_row in range(0, sh_bmi.nrows):
            cell = sh_bmi.cell(cur_row, 0)
            # print(cell.value)
            try:
                if 'Effective date' in cell.value:
                    z = 1
                    continue
            except Exception as ex:
                pass
            if cell.value == '':
                z = 0
            if z == 1:
                data_bmi.append(sh_bmi.cell(cur_row, 1).value)
        # print('data_bmi')
        # print(data_bmi)
        data_bmi = [x2 / x1 - 1 for (x1, x2) in zip(data_bmi, data_bmi[1:])]
        std_bmi = statistics.pstdev(data_bmi) * (260 ** 0.5)
        # log_debug('-- Processed file bmi --')
        sh_baml = wb_baml.sheet_by_index(0)
        data_baml = []
        z = 0
        for cur_row in range(0, sh_baml.nrows):
            cell = sh_baml.cell(cur_row, 0)
            # print(cell.value)
            try:
                if 'observation' in cell.value:
                    z = 1
                    continue
            except Exception as ex:
                pass
            if cell.value == '':
                z = 0
            if z == 1:
                data_baml.append(sh_baml.cell(cur_row, 1).value)
        data_baml = [x for x in data_baml if x is not None]
        k1 = 0
        for i in range(len(data_baml)):
            if data_baml[i] == 0:
                data_baml[i] = k1
            k1 = data_baml[i]
        std_baml = statistics.pstdev(data_baml)
        mean_baml = statistics.mean(data_baml)
        cv = std_baml / mean_baml
        volatility_ratio = std_bmi / cv

        # log_debug('-- Processed file baml --')

        # ll = [volatility_ratio, std_bmi, cv, std_baml, mean_baml]
        # print(ll)
        project = Project.objects.filter(translations__language_code=get_language()).get(
            id=int(request.session['cv_project_id']))
        project.volatility_ratio = volatility_ratio
        project.save()
        dic = {'status': 'ok', 'volatility_ratio': volatility_ratio}
        log_debug("End load_country_premium.")
        return dic

    def update_region_risk_premium(self, request):
        project = Project.objects.filter(translations__language_code=get_language()).get(
            id=request.session['cv_project_id'])
        XBRLCountryYearData.project = project
        XBRLRegionYearData.project = project
        for r in XBRLRegion.objects.filter(updated_adamodar=True).all():
            # print('-------')
            # print(r)
            ltx = []
            lds = []
            lrp = []
            for c in r.countries.all():
                try:
                    cc = XBRLCountryYearData.objects.get(country=c, year=project.year)
                    # print(cc, cc.tax_rate, cc.rating_based_default_spread, cc.country_risk_premium_rating)
                    if cc.tax_rate is not None:
                        ltx.append(float(cc.tax_rate))
                    if cc.rating_based_default_spread is not None:
                        lds.append(float(cc.rating_based_default_spread))
                    if cc.country_risk_premium_rating is not None:
                        lrp.append(float(cc.country_risk_premium_rating))
                except Exception as ex:
                    pass
                    # print(c, ex)
            try:
                tx, ds, rp = round(100 * sum(ltx) / len(ltx)) / 100, round(100 * sum(lds) / len(lds)) / 100, round(
                    100 * sum(lrp) / len(lrp)) / 100
                # print(tx, ds, rp)
                rr, c = XBRLRegionYearData.objects.get_or_create(region=r, year=project.year)
                rr.tax_rate = tx
                rr.country_risk_premium_rating = rp
                rr.rating_based_default_spread = ds
                rr.save()
            except Exception as ex:
                log_debug("update_region_risk_premium: " + str(ex))
        dic = {'status': 'ok'}
        log_debug("End update_region_risk_premium.")
        return dic

    def update_etfs_companies(self, params):
        l = ["b", "c", "e", "f", "i", "k", "p", "re", "u", "v", "y"]
        # print("9085\n", l, "\n", "-"*50)
        for f in l:
            file = "etfs/index-holdings-xl"+f
            print(file)
            df = self.load_excel_data(file, sheet_name="Sheet1")
            df = df.reset_index()
            # print(df)
            try:
                f_ = "XL"+f.upper()
                # print(f_)
                etf_obj = ETFS.objects.get(symbol=f_)
                # print(etf_obj)
            except Exception as ex:
                print(ex)

            exc = []
            for i, r in df.iterrows():
                try:
                    # print(r['Symbol'])
                    XBRLCompanyInfo.objects.filter(ticker=r['Symbol']).update(etf=etf_obj)
                except Exception as ex:
                    exc.append(f+": "+r['Symbol'])
                    # print(ex)
        # print("9088\n", "Done", "\n", "-"*50)

        result = {"exc": exc, "status": "ok"}
        print("90901:  Result:\n", result, "\n", "-"*50)
        return result

#   --- Stage 3 - SP500 ---
#     loading excel file to XBRLHistoricalReturnsSP.
    def load_sp_returns(self):
        log_debug("Start load_sp_returns.")
        # https://github.com/7astro7/full_fred
        file = 'histretSP'
        df = self.load_excel_data(file, 'Data')
        df = df.sort_values(by=['year'])
        spi = 1
        bbbi = 1
        n = 0
        XBRLHistoricalReturnsSP.truncate()
        log_debug("XBRLHistoricalReturnsSP.truncate() done.")
        # print(df)

        for i, r in df.iterrows():
            try:
                year_data, c = XBRLHistoricalReturnsSP.objects.get_or_create(year=int(r['year']))
                if not pd.isna(r['AAA']):
                    year_data.aaa = r['AAA']
                if not pd.isna(r['BBB']):
                    year_data.bbb = r['BBB']
                if not pd.isna(r['TB3MS']):
                    year_data.tb3ms = r['TB3MS']
                if not pd.isna(r['TB10Y']):
                    year_data.tb10y = r['TB10Y']
                if not pd.isna(r['SP500']):
                    year_data.sp500 = r['SP500']
                if not pd.isna(r['DividendYield']):
                    year_data.dividend_yield = r['DividendYield']
                if not pd.isna(r['ReturnsOnRealEstate']):
                    year_data.return_on_real_estate = r['ReturnsOnRealEstate']
                if not pd.isna(r['HomePrices']):
                    year_data.home_prices = r['HomePrices']
                if not pd.isna(r['CPI']):
                    year_data.cpi = r['CPI']
            except Exception as ex:
                print('ex')
                print(ex)
                print('ex')
                log_debug('Error 1 create year_data: ' + str(r['year']) + " " + ex)

            try:
                year_data.save()
                if int(year_data.year) >= 1928:
                    spi = spi * (1 + year_data.return_on_sp500)
                    bbbi = bbbi * (1 + year_data.return_on_tbond)
                    n += 1
                    spi_ = spi ** (1 / n)
                    bbbi_ = bbbi ** (1 / n)
                    r = spi_ - bbbi_
                    year_data.risk_premium = round(10000 * r) / 10000
                year_data.save()
            except Exception as ex:
                log_debug('Error 2 year_data.save(): ' + str(r['year']) + " " + ex)
        dic = {'status': 'ok'}
        log_debug("End load_sp_returns.")
        return dic

    # not used
    def get_etfs(self, params):
        # print(55555)
        # print(params)
        # print(55555)
        from selenium import webdriver
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.common.by import By
        driver = webdriver.Chrome()
        url = "https://www.cnbc.com/sector-etfs/"
        driver.get(url)

        headers = {'User-Agent': 'amos@drbaranes.com'}
        edgar_resp = requests.get(url, headers=headers, timeout=30)
        edgar_str = edgar_resp.text
        # print(edgar_str)

        # Find the document links
        soup = BeautifulSoup(edgar_str, 'html.parser')
        table_tag = soup.find('table', class_='BasicTable-table')
        print(table_tag)
        try:
            rows = table_tag.find_all('tr')
            print(rows)
        except Exception as ex:
            return dic_company_info
        # # Obtain HTML for document page
        # dic_data = {}
        # for row in rows:
        #     try:
        #         cells = row.find_all('td')
        #         if len(cells) > 3:
        #             if cells[0].text.lower() != type_.lower():
        #                 continue
        #             # for filing_year in range(2019, 2020):
        #             for filing_year in range(self.xbrl_start_year, self.today_year + 1):
        #                 if str(filing_year) in cells[3].text:
        #                     # print(str(filing_year), cells[3].text, cells[1].a['href'], cells[0].text)
        #                     if filing_year not in dic_data and (cells[0].text == "10-K" or cells[0].text == "20-F"):
        #                         dic_data[filing_year] = {}
        #                         dic_data[filing_year]['href'] = 'https://www.sec.gov' + cells[1].a['href']
        #     except Exception as ex:
        #         pass
        #
        # dic_company_info['data'] = dic_data
        # # print(dic_company_info)
        # return dic_company_info
        result = {}
        return result


class FinancialAnalysis(object):
    def __init__(self):
        try:
            clear_log_debug()
        except Exception as ex:
            pass

    def update_chart_of_accounts(self, **kwargs):
        log_debug('start: update_chart_of_accounts')
        for account in XBRLValuationAccounts.objects.all():
            # print('-' * 50)
            # print(account.id, account.order, account.statement.id, account.statement.statement, account.sic,
            #       account.account, account.type, account.scale)
            try:
                a, c = XBRLDimAccount.objects.get_or_create(order=account.order, account=account.account,
                                                        statement_order=account.statement.order,
                                                        statement=account.statement.statement)
            except Exception as ex:
                print('Error 5432: '+str(ex))
        log_debug('End: update_chart_of_accounts')
        result = {'status': "ok"}
        return result

    def update_companies(self, **kwargs):
        # print("update_companies")
        for company in XBRLCompanyInfo.objects.all():
            # print(company)
            if company.is_active:
                try:
                    # print('-'*10)
                    # print(company)
                    # print(company.id, company.industry.main_sic.sic_code, company.industry.main_sic.sic_description,
                    #       company.industry.sic_code, company.industry.sic_description,
                    #       company.ticker, company.cik, company.company_name, company.exchange,
                    #       company.is_active, company.city, company.state, company.zip)

                    log_debug(company.ticker)
                    a, c = XBRLDimCompany.objects.get_or_create(id=company.id,
                                                                main_sic_code=company.industry.main_sic.sic_code,
                                                                main_sic_description=company.industry.main_sic.sic_description,
                                                                sic_code=company.industry.sic_code,
                                                                sic_description=company.industry.sic_description,
                                                                ticker=company.ticker, cik=company.cik,
                                                                company_name=company.company_name,
                                                                exchange=company.exchange,
                                                                city=company.city,
                                                                state=company.state,
                                                                zip=company.zip, is_active=company.is_active)
                except Exception as ex:
                    pass
        result = {'status': "ok"}
        # print(result)
        return result

    def update_time(self, **kwargs):
        print("update_time")
        up_to_year = datetime.datetime.now().year + 2

        for y in range(2012, up_to_year):
            for q in range(0, 5):
                ind = y * 10 + q
                log_debug(str(ind) + " " + str(y) + " " + str(q))
                a, c = XBRLDimTime.objects.get_or_create(id=ind, year=y, quarter=q)
        result = {'status': "ok"}
        # print(result)
        return result

    # https://webix-ui.medium.com/top-7-javascript-pivot-widgets-in-2019-2020-8d81d4042f51
    # https://github.com/nicolaskruchten/pivottable
    def update_fact_table(self, **kwargs):

        print(kwargs)

        log_debug("--update_chart_of_accounts--")
        # print("--update_chart_of_accounts--")
        ticker_ = kwargs['ticker'].upper()
        # log_debug(ticker_)
        # print("--update_chart_of_accounts--")
        # request = kwargs['request']
        # print(ticker_)
        company = XBRLCompanyInfo.objects.get(ticker=ticker_)
        log_debug("got company: " + ticker_)
        print("company", company)

        try:
            print("AAAA")
            print(company.id)
            company_, is_created = XBRLDimCompany.objects.get_or_create(id=company.id,
                                                                        ticker=ticker_,
                                                                        main_sic_code = company.industry.main_sic.sic_code,
                                                                        sic_code = company.industry.sic_code)

            print("is_created", is_created, company.id, company.company_name)
            # if is_created:
            company_.main_sic_description = company.industry.main_sic.sic_description
            company_.sic_description = company.industry.sic_description
            company_.cik = company.cik
            company_.company_name = company.company_name
            company_.is_active = True
            company_.save()

            # print("company_", company_)
        except Exception as ex:
            print("Error 9876: " + ticker_ + str(ex))
            log_debug("Error 9876: " + ticker_ + str(ex))

        log_debug("Start yearly data")
        print("Start yearly data")
        # print(company.financial_data['data'])
        for y in company.financial_data['data']:
            try:
                yd = company.financial_data['data'][y]
                # print("-"*30, "\n", yd, "\n", "-"*30)
                # print("-"*30, "\n", y, "\n", "-"*30)

                if int(yd['dei']['documentfiscalyearfocus']) < 2012:
                    continue
                yq = int(yd['dei']['documentfiscalyearfocus'] + "0")
                time_ = XBRLDimTime.objects.get(id=yq)
                log_debug("start update fact table for " + ticker_ + " year: " + str(y))
                for account in yd['year_data']:
                    # print(account)
                    account_ = XBRLDimAccount.objects.get(order=int(account))
                    amount_ = yd['year_data'][account]
                    # print('account_')
                    # print(account_)
                    # print('amount_')
                    # print(amount_)
                    try:
                        f, c = XBRLFactCompany.objects.get_or_create(company=company_, time=time_, account=account_)
                        # print(f)
                        f.amount = amount_
                        f.save()
                        # print('--saved--')
                    except Exception as ex:
                        log_debug(str(ex))
                        # print(str(account_) + "  " + str(ex))
                log_debug("End update fact table for " + ticker_ + " year: " + str(y))
                # print("End update fact table for " + ticker_ + " year: " + str(y))
            except Exception as ex:
                log_debug("Err 9123: "+str(ex))
                # print("Err 9123: "+str(ex))
                continue

        log_debug("Start quarterly data")
        print("Start quarterly data")
        try:
            for y in company.financial_dataq['data']:
                try:
                    for sq in company.financial_dataq['data'][y]:
                        yd = company.financial_dataq['data'][y][sq]
                        if int(yd['dei']['documentfiscalyearfocus']) < 2012:
                            continue
                        q = yd['dei']['documentfiscalperiodfocus'][1]
                        log_debug("start update fact table for ticker=" + ticker_ + " year=" + str(y) + " q=" + str(q))
                        yq = int(yd['dei']['documentfiscalyearfocus'] + str(q))
                        time_ = XBRLDimTime.objects.get(id=yq)
                        for account in yd['year_data']:
                            account_ = XBRLDimAccount.objects.get(order=int(account))
                            amount_ = yd['year_data'][account]
                            try:
                                f, c = XBRLFactCompany.objects.get_or_create(company=company_, time=time_, account=account_)
                                f.amount = amount_
                                f.save()
                            except Exception as ex:
                                log_debug(str(ex))
                        log_debug("End update fact table for ticker=" + ticker_ + " year=" + str(y) + " q=" + str(q))
                except Exception as ex:
                    log_debug("Err 6537: "+str(ex))
                    continue
        except Exception as exx:
            pass

        log_debug("End quarterly data")
        result = {'status': "ok"}
        # print(result)
        return result

    def get_data_for_ticker(self, **kwargs):
        print('get_data_for_ticker 1')
        ticker_ = kwargs['ticker'].upper()
        print('-' * 50)
        print(ticker_)
        qsv = XBRLFactCompany.objects.filter(company__ticker=ticker_, time__quarter__lte=0).all().values("company_id",
                                                                                                         "time_id",
                                                                                                         "account_id",
                                                                                                         "amount")

        # qsv = XBRLFactCompany.objects.filter(company__ticker=ticker_, time__quarter__gt=0).all().values("time_id", "account_id", "amount")
        print(qsv)

        df = pd.DataFrame(qsv)
        print(df)
        df = df.pivot(index='account_id', columns='time_id', values='amount')
        print(df)

        result = {'status': "ok"}
        # print(result)
        return result

    def get_pivot_data(self, dic):
        dic = eval(dic)
        # print('get_data_for_ticker 1')
        ticker_ = dic['ticker']
        # print(ticker_)
        sic_ = XBRLDimCompany.objects.get(ticker=ticker_).sic_code
        companies = XBRLDimCompany.objects.filter(sic_code=sic_).all()
        companies_dic = {}
        for c in companies:
            companies_dic[c.id] = c.company_name
        qsv = XBRLFactCompany.objects.filter(company__sic_code=sic_, time__quarter__lte=0).all().values("company_id", "time_id", "account_id", "amount")
        company_id = []
        time_id = []
        account_id = []
        amount = []
        for q in qsv:
            # print(q, q['time_id'], q['account_id'],q['amount'])
            company_id.append(q['company_id'])
            time_id.append(q['time_id'])
            account_id.append(q['account_id'])
            amount.append(float(q['amount']))

        data_ = {"company_id": company_id, "time_id": time_id, "account_id": account_id, "amount": amount, "companies_dic": companies_dic, "sic":sic_}
        # print(data_)
        result = {'status': "ok", "data": data_}

        # print(result)
        return result

    def get_market_portfolio(self, params):
        list_of_tickers = params["tickers"]
        # print('='*20)
        # print(list_of_tickers)
        # print('='*20)

        XP = data.DataReader(list_of_tickers, 'yahoo', start='2020/01/01', end='2022/08/03')
        # print('=XP'*3)
        # print(XP)

        X = XP['Adj Close'].pct_change()      # .apply(lambda x: np.log(1 + x))
        # print('=X'*3)
        # print(X)

        V = X.cov()
        # print('=V'*3)
        # print(V)

        xr = XP['Adj Close'].resample('Y').last().pct_change().mean()
        # print('xr')
        # print(xr)
        # print('xr')

        # dic = {"a": [1, 2, 3], "b": [3, 5, 1]}
        # print(dic)
        # df = pd.dataframe(dic)
        # df.apply.(lambda w : w*w)
        # xsd = X.std().apply(lambda x: x * np.sqrt(250))

        i_r = []  # Define an empty array for portfolio returns
        i_sd = []  # Define an empty array for portfolio volatility
        i_w = []  # Define an empty array for asset weights
        ns = len(X.columns)

        # print('ns')
        # print(ns)
        # print('ns')

        nb = 3000
        for i in range(nb):
            w = np.random.random(ns)
            w = w / np.sum(w)
            i_w.append(w)
            r = round(np.dot(w, xr), 3)
            i_r.append(r)
            # var = V.mul(w, axis=0).mul(w, axis=1).sum().sum()

            var = V.mul(w, axis=0)
            # print(var)
            var = var.mul(w, axis=1)
            # print(var)
            var = var.sum()
            # print(var)
            var = var.sum()
            # print(var)

            sd = np.sqrt(var)  # Daily standard deviation
            sd = round(sd * np.sqrt(250), 3)  # Annual standard deviation = volatility
            i_sd.append(sd)

        data_ = {"Returns": i_r, "Volatility": i_sd}

        for counter, symbol in enumerate(X.columns.tolist()):
            data_['w' + symbol] = [w[counter] for w in i_w]

        portfolios = pd.DataFrame(data_)

        # Finding the optimal portfolio
        rf = 0.01  # risk factor
        optimal_risky_port = portfolios.iloc[((portfolios['Returns'] - rf) / portfolios['Volatility']).idxmax()]

        result = {"data": {"r": i_r, "sd": i_sd, "w": optimal_risky_port.tolist()}}
        # print(result)
        return result

    def get_market_portfolio_np(self, params):
        list_of_tickers = params["tickers"]
        XP = data.DataReader(list_of_tickers, 'yahoo', start='2020/01/01', end='2022/08/03')
        X = XP['Adj Close'].pct_change().dropna().reset_index(drop=True)
        XN = X.to_numpy()
        # print('XN.T')
        # print(XN.T)
        V = np.cov(XN.T)
        # print(V)
        xr = XP['Adj Close'].resample('Y').last().pct_change().mean()
        i_r = []  # Define an empty array for portfolio returns
        i_sd = []  # Define an empty array for portfolio volatility
        i_w = []  # Define an empty array for asset weights
        ns = len(X.columns)
        nb = 10000
        for i in range(nb):
            w = np.random.random(ns)
            w = w / np.sum(w)
            i_w.append(w)
            r = round(np.dot(w, xr), 3)
            i_r.append(r)
            var = np.dot(w, np.dot(V, w))
            sd = np.sqrt(var)
            sd = round(sd * np.sqrt(250), 3)
            i_sd.append(sd)
        data_ = {"Returns": i_r, "Volatility": i_sd}

        for counter, symbol in enumerate(X.columns.tolist()):
            data_['w' + symbol] = [w[counter] for w in i_w]

        portfolios = pd.DataFrame(data_)

        # Finding the optimal portfolio
        rf = 0.01
        optimal_risky_port = portfolios.iloc[((portfolios['Returns'] - rf) / portfolios['Volatility']).idxmax()]

        result = {"data": {"r": i_r, "sd": i_sd, "w": optimal_risky_port.tolist()}}
        return result

#
class BaseCorporateValuationAlgo(object):
    def __init__(self, dic):  # to_data_path, target_field
        print("90050-01 BaseTrainingAlgo", dic, '\n', '-'*50)
        # super(BaseTrainingAlgo, self).__init__()
        # print("90050-02 BaseTrainingAlgo", dic, '\n', '-'*50)
        app_ = dic["app"]
        self.PROJECT_ROOT_DIR = os.path.join(settings.WEB_DIR, "data", dic["app"])
        # print(self.PROJECT_ROOT_DIR)
        # print('-'*50)
        os.makedirs(self.PROJECT_ROOT_DIR, exist_ok=True)
        self.TOPIC_ID = dic["topic_id"]  # "general"
        self.TO_DATA_PATH = os.path.join(self.PROJECT_ROOT_DIR, "datasets")
        os.makedirs(self.TO_DATA_PATH, exist_ok=True)
        self.TO_EXCEL = os.path.join(self.TO_DATA_PATH, "excel", self.TOPIC_ID)
        os.makedirs(self.TO_EXCEL, exist_ok=True)

        self.excel_dir = settings.MEDIA_ROOT + '/' + app_ + '/excel'
        os.makedirs(self.excel_dir, exist_ok=True)
        self.save_to_file = None
        self.second_time_save = ''

    def save_to_excel(self, df, folder, file_name=None):
        if file_name:
            self.save_to_file = os.path.join(self.excel_dir, file_name)
        wb2 = Workbook()
        wb2.save(self.save_to_file)
        wb2.close()
        wb2 = None
        with pd.ExcelWriter(self.save_to_file, engine='openpyxl', mode="a") as writer_o:
            try:
                df.to_excel(writer_o, sheet_name="Data")
            except Exception as ex:
                print("9006-3 " + str(ex))
            try:
                writer_o.save()
            except Exception as ex:
                print("9006-4 " + str(ex))
        wb = load_workbook(filename=self.save_to_file, read_only=False)
        del wb['Sheet']
        wb.save(self.save_to_file)
        wb.close()


class CorporateValuationDataProcessing(BaseDataProcessing, BaseCorporateValuationAlgo):
    def __init__(self, dic):
        super().__init__(dic)
        self.app = dic["app"]

    def download_companies_to_excel(self, dic):
        # print('    90033-100 dic\n', '-'*100, '\n', dic, '\n', '-'*100)
        app_ = dic["app"]
        etfwatchlist_symbol_ = dic["etfwatchlist_symbol"]
        file_name_ = dic['file_name']

        model_name_ = "xbrlcompanyinfo"
        model_xci = apps.get_model(app_label=app_, model_name=model_name_)
        qs = model_xci.objects.filter(etfwatchlist__symbol=etfwatchlist_symbol_).all()
        df = pd.DataFrame(list(qs.values('exchange', 'company_name', 'ticker', 'company_letter', 'cik', 'is_active')))
        try:
            self.save_to_excel(df, "Data", file_name=file_name_)
        except Exception as ex:
            print("Error 90876-5543"+ex)
        result = {"status": "ok"}
        return result

    def upload_companies_to_excel(self, dic):
        # print('    90033-100 dic\n', '-'*100, '\n', dic, '\n', '-'*100)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # print('90022-1 dic')
        # print(file_path)
        dic = dic["cube_dic"]
        # print('90022-1 dic', dic)

        model_name_ = "etfwatchlists"
        model_etfwl = apps.get_model(app_label=app_, model_name=model_name_)
        etfwl_obj, is_created = model_etfwl.objects.get_or_create(symbol="HighV")

        model_name_ = "xbrlcompanyinfo"
        model_xci = apps.get_model(app_label=app_, model_name=model_name_)
        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        # print(df)
        for index, row in df.iterrows():
            # print(str(row["ticker"]))
            o, is_created = model_xci.objects.get_or_create(ticker=str(row["ticker"]))
            o.etfwatchlist = etfwl_obj
            o.company_name = str(row["company_name"])
            o.company_letter = str(row["company_letter"])
            o.cik = str(row["cik"])
            o.is_active = bool(row["is_active"])
            o.financial_data = {"a":"a"}
            o.financial_dataq = {"a":"a"}
            o.save()
            # print("saved", str(row["ticker"]))
        result = {"status": "ok"}
        return result

    def data_transfer_to_process_fact(self, dic):
        # print('data_transfer_to_process_fact 90033-100 dic\n', '-'*100, '\n', dic, '\n', '-'*100)
        app_ = dic["app"]

        model_from = apps.get_model(app_label=app_, model_name="XBRLFactCompany")
        model_to = apps.get_model(app_label=app_, model_name="XBRLProcessedFactCompany")
        qs = model_from.objects.all()
        for q in qs:
            # print(q.company.id, q.time, q.account.order, q.amount)
            obj, is_created = model_to.objects.get_or_create(company_id=q.company.id, time=q.time, account=q.account.order)
            obj.amount=q.amount
            obj.save()

        result = {"status": "ok"}
        return result

    def create_new_group_accounts(self, dic):
        # print('data_transfer_to_process_fact 90044-100 dic\n', '-'*100, '\n', dic, '\n', '-'*100)
        app_ = dic["app"]
        model_from = apps.get_model(app_label=app_, model_name="XBRLProcessedFactCompany")
        model_to = apps.get_model(app_label=app_, model_name="XBRLAccountsGroupsFactCompany")

        #
        # Need to add special ratios like for the banking industry.
        #
        ll_aggregate_accounts = dic["aggregate_accounts"]
        dic_new_accounts = dic["new_accounts"]
        #
        qs = model_from.objects.filter(account__in=ll_aggregate_accounts).all()
        # print(qs)

        for q in qs:
            # print(q.company.id, q.time, q.account, q.amount)
            obj, is_created = model_to.objects.get_or_create(company_id=q.company.id, time=q.time, account=q.account)
            obj.amount=q.amount
            obj.save()

        cs = model_from.objects.values('company_id').distinct()
        df_cs=pd.DataFrame(list(cs))
        for index, row in df_cs.iterrows():
            c_id = int(row["company_id"])
            # print("-"*50,"\n",c_id,"\n","-"*50)
            csi = model_from.objects.filter(company_id=c_id).values('time_id').distinct()
            df_csi=pd.DataFrame(list(csi))
            for index_, row_ in df_csi.iterrows():
                for new_account in dic_new_accounts:
                    detail = dic_new_accounts[new_account]
                    new_account = int(new_account)
                    n = 0
                    try:
                        time_id = int(row_["time_id"])
                        # print(c_id, time_id)
                        add = detail["add"]
                        subtract = detail["subtract"]
                        for add_account in add:
                            obj_1 = model_from.objects.get(company_id=c_id, time_id=time_id, account=add_account)
                            n += obj_1.amount
                            # print("add_account\n", obj_1.amount, add_account, "\n", n)
                        for subtract_account in subtract:
                            # print("add_account, add_account")
                            obj_2 = model_from.objects.get(company_id=c_id, time_id=time_id, account=subtract_account)
                            n -= obj_2.amount
                            # print("subtract_account\n", obj_2.amount, subtract_account, "\n", n)
                    except Exception as ex:
                        pass
                        # print("Errror 560\n", ex)

                    try:
                        obj_to, is_created = model_to.objects.get_or_create(company_id=c_id, time_id=time_id, account=new_account)
                        # print("n\n", n)
                        obj_to.amount = n
                        obj_to.save()
                        # print("-"*50, "\n", obj_to, "\n", "-"*50)
                    except Exception as ex:
                        print("Errror 600\n", ex)

        result = {"status": "ok"}
        return result

    def create_ratios(self, dic):
        # print('create_ratios 90044-666-66 dic\n', '-'*100, '\n', dic, '\n', '-'*100)
        app_ = dic["app"]

        model_from = apps.get_model(app_label=app_, model_name="XBRLAccountsGroupsFactCompany")
        model_to = apps.get_model(app_label=app_, model_name="XBRLFactRatiosCompany")
        model_ratio = apps.get_model(app_label=app_, model_name="XBRLRatioDim")

        cs = model_from.objects.values('company_id').distinct()
        df_cs=pd.DataFrame(list(cs))

        qs_r = model_ratio.objects.all().values()
        df_r=pd.DataFrame(qs_r)
        # print(df_r)

        for index, row in df_cs.iterrows():
            c_id = int(row["company_id"])
            # print("-"*50,"\n",c_id,"\n","-"*50)
            csi = model_from.objects.filter(company_id=c_id).values('time_id').distinct()
            df_csi=pd.DataFrame(list(csi))
            for index_, row_ in df_csi.iterrows():
                time_id = int(row_["time_id"])
                # print("-"*50,"\n",time_id,"\n","-"*50)
                for index_r, row_r in df_r.iterrows():
                    try:
                        numerator = int(row_r["numerator"])
                        denominator = int(row_r["denominator"])
                        # print(numerator, denominator)
                        n = model_from.objects.get(company_id=c_id, time_id=time_id, account=numerator).amount
                        d = model_from.objects.get(company_id=c_id, time_id=time_id, account=denominator).amount
                        r = n/d
                        # print(n, d, r)
                        ratio_id = row_r.id
                        obj, is_created = model_to.objects.get_or_create(company_id=c_id, time_id=time_id, ratio_id = ratio_id)
                        obj.amount=r
                        obj.save()
                    except Exception as ex:
                        pass
                        # print("Error 202-202", ex)

        # qs = model_from.objects.all().values()
        #
        # df = pd.DataFrame(list(qs))
        # print(df)
        # print(qs)
        # for q in qs:
        #     print(q.company.id,
        #           q.time,
        #           q.account,
        #           q.amount)
        #     obj, is_created = model_to.objects.get_or_create(company_id=q.company.id, time=q.time, account=q.account.order)
        #     obj.amount=q.amount
        #     obj.save()

        result = {"status": "ok"}
        return result


class SimFin(BaseDataProcessing, BaseCorporateValuationAlgo):
    def __init__(self, dic):
        super().__init__(dic)

    def activate_simfin(self, dic):
        print("9010-100-20 input dic: \n", dic, "\n"+"-"*30)

        print(sf.__version__)
        simfin_key = '74f52806-6451-49a2-95f2-e63f952f6a1f'
        sf.set_api_key(simfin_key)
        sf.set_data_dir(self.TO_EXCEL)
        df = sf.load_income(variant='annual', market='us')
        # print(df.loc['MSFT', [REVENUE, NET_INCOME]])
        df_prices = sf.load_shareprices(market='us', variant='daily')
        # print(df_prices)
        print(df_prices.loc['MSFT', CLOSE]) # .plot(grid=True, figsize=(20,10), title='MSFT Close')

        dic = {'status': "ok"}
        return dic


class Option(object):
    def __init__(self):
        pass

    def bs_call(self, dic):
        S = float(dic["S"])
        T = int(dic["T"])
        K = float(dic["K"])
        r = float(dic["r"])  # Risk-free interest rate
        sigma = float(dic["sigma"])  # Volatility
        N = norm.cdf
        #
        d1 = (np.log(S / K) + (r + sigma ** 2 / 2) * T) / (sigma * np.sqrt(T))
        d2 = d1 - sigma * np.sqrt(T)
        return S * N(d1) - K * np.exp(-r * T) * N(d2)

    def bs_put(self, dic):
        S = float(dic["S"])
        T = int(dic["T"])
        K = float(dic["K"])
        r = float(dic["r"])  # Risk-free interest rate
        sigma = float(dic["sigma"])  # Volatility
        N = norm.cdf
        #
        d1 = (np.log(S / K) + (r + sigma ** 2 / 2) * T) / (sigma * np.sqrt(T))
        d2 = d1 - sigma * np.sqrt(T)
        return K * np.exp(-r * T) * N(-d2) - S * N(-d1)

    def bs_put_call(self, dic):
        print(dic)
        S = float(dic["S"])
        T = int(dic["T"])
        K = float(dic["K"])
        r = float(dic["r"])  # Risk-free interest rate
        sigma = float(dic["sigma"])  # Volatility
        N = norm.cdf
        #
        d1 = (np.log(S / K) + (r + sigma ** 2 / 2) * T) / (sigma * np.sqrt(T))
        d2 = d1 - sigma * np.sqrt(T)
        call_option_price = S * N(d1) - K * np.exp(-r * T) * N(d2)
        put_option_price = K * np.exp(-r * T) * N(-d2) - S * N(-d1)
        result = {"status": "ok", "data": {"call_option_price":call_option_price, "put_option_price":put_option_price}}
        print(result)
        return result

    def binomial_american_call_option(self, S, K, T, r, sigma, n):
        dt = T / n
        u = math.exp(sigma * math.sqrt(dt))
        d = 1 / u
        p = (math.exp(r * dt) - d) / (u - d)
        print("dt=", round(10000*dt)/10000, "u=", round(10000*u)/10000, "d=", round(10000*d)/10000, "p=", round(10000*p)/10000)

        # Initialize the option price at expiration (payoff)
        option_price_c = [max(0, S * (u ** (n - i)) * (d ** i) - K) for i in range(n + 1)]
        option_price_p = [max(0, K - S * (u ** (n - i)) * (d ** i)) for i in range(n + 1)]
        # print("\noption_price_c=", option_price_c, "\noption_price_p=", option_price_p, "\n")

        # Calculate option price at each step backward through the tree

        # for j in range(10 - 1, -1, -1):
        #     print("j=", j)

        for j in range(n - 1, -1, -1):
            for i in range(j + 1):
                # print(j, i, j-i)
                # print("j", j, "i", i, "CV=", round(100*(S * (u ** (j-i)) * (d ** i) - K))/100,
                #       "OV=", round(100*(math.exp(-r * dt) * (p * option_price_c[i] + (1 - p) * option_price_c[i + 1])))/100,
                #       "max", round(100*(max(S * (u ** (j-i)) * (d ** i) - K,
                #                       math.exp(-r * dt) * (p * option_price_c[i] + (1 - p) * option_price_[i + 1]))))/100)
                #
                # print("j", j, "i", i, "PV=", round(100*(K - S * (u ** (j-i)) * (d ** i)))/100,
                #       "OV=", round(100*(math.exp(-r * dt) * (p * option_price_p[i] + (1 - p) * option_price_p[i + 1])))/100,
                #       "max", round(100*(max(K - S * (u ** (j-i)) * (d ** i),
                #                       math.exp(-r * dt) * (p * option_price_p[i] + (1 - p) * option_price_p[i + 1]))))/100)

                option_price_c[i] = max(S * (u ** (j-i)) * (d ** i) - K,
                                      math.exp(-r * dt) * (p * option_price_c[i] + (1 - p) * option_price_c[i + 1]))

                option_price_p[i] = max(K - S * (u ** (j-i)) * (d ** i),
                                      math.exp(-r * dt) * (p * option_price_p[i] + (1 - p) * option_price_p[i + 1]))

                # print("j", j, "i", i, "\noption_price_c=", option_price_c, "\noption_price_p=", option_price_p, "\n")

        return round(100*option_price_c[0])/100, round(100*option_price_p[0])/100

    def put_call(self, dic):
        # print('90-90-90-11 data_transfer_to_process_fact 90055-300 dic\n', '-'*100, '\n', dic, '\n', '-'*100)
        app_ = dic["app"]

        # Example usage:
        S = float(dic["S"])  # Current stock price
        K = float(dic["K"])  # Strike price
        T = 1  # Time to expiration (in years)
        r = float(dic["r"])  # Risk-free interest rate
        sigma = float(dic["sigma"])  # Volatility
        n = int(dic["n"])  # Number of time steps

        call_option_price, put_option_price = self.binomial_american_call_option(S, K, T, r, sigma, n)
        # print("\n", f"The price of the American call option is: {call_option_price:.2f}")
        # print(f"The price of the American put option is: {put_option_price:.2f}")

        result = {"status": "ok", "data": {"call_option_price":call_option_price, "put_option_price":put_option_price}}
        return result

    def range_put_call(self, dic):
        print('90-90-90-222 range_put_call 90055-300 dic\n', '-'*100, '\n', dic, '\n', '-'*100)

        spread = float(dic["spread"])
        # T = int(dic["T"])
        app_ = dic['app']
        K = float(dic["K"])
        r_ = float(dic["r"])  # Risk-free interest rate
        sigma_ = float(dic["sigma"])  # Volatility
        n_ = int(dic["n"])
        inner_range = int(dic["inner_range"])
        range_ = int(dic["range"])
        range_bin = float(dic["range_bin"])
        S = int(dic["S"])  # Current stock price

        def binomial_price_probability(S, T, r, sigma, n):
            dt = T / n
            u = math.exp(sigma * math.sqrt(dt))
            d = 1 / u
            p = (math.exp(r * dt) - d) / (u - d)

            price = []
            dist = []
            for i in range(n+1):
                price.appande(S * (u ** (n - i)) * (d ** i))
                dist.appande(binom.pmf(i, n, p))

            return {"price":price, "dist":dist}

        def binomial_american_call_option(S, K, T, r, sigma, n):
            dt = T / n
            u = math.exp(sigma * math.sqrt(dt))
            d = 1 / u
            p = (math.exp(r * dt) - d) / (u - d)
            option_price_c = [max(0, S * (u ** (n - i)) * (d ** i) - K) for i in range(n + 1)]

            for j in range(n - 1, -1, -1):
                for i in range(j + 1):
                    option_price_c[i] = max(S * (u ** (j - i)) * (d ** i) - K,
                                            math.exp(-r * dt) * (
                                                        p * option_price_c[i] + (1 - p) * option_price_c[i + 1]))

            return round(100 * option_price_c[0]) / 100

        def binomial_american_put_option(S, K, T, r, sigma, n):
            dt = T / n
            u = math.exp(sigma * math.sqrt(dt))
            d = 1 / u
            p = (math.exp(r * dt) - d) / (u - d)

            option_price_p = [max(0, K - S * (u ** (n - i)) * (d ** i)) for i in range(n + 1)]

            for j in range(n - 1, -1, -1):
                for i in range(j + 1):

                    option_price_p[i] = max(K - S * (u ** (j - i)) * (d ** i),
                                            math.exp(-r * dt) * (
                                                        p * option_price_p[i] + (1 - p) * option_price_p[i + 1]))

            return round(100 * option_price_p[0]) / 100

        def sub_(l1, l2):
            return [round(100*(l1[i] - l2[i]))/100 for i in range(len(l1))]
        def add_(l1, l2):
            return [round(100*(l1[i] + l2[i]))/100 for i in range(len(l1))]

        def calc_(t, S, K, T, r, sigma, inner_range, spread, n):
            # print("1 of 5")
            # cy_ = [binomial_american_call_option(s, K-inner_range-spread, T, r, sigma, n) for s in ar]
            # print("2 of 5")
            # cy = [binomial_american_call_option(s, K-inner_range, T, r, sigma, n) for s in ar]
            # print("3 of 5")
            # py_ = [binomial_american_put_option(s, K+inner_range+spread, T, r, sigma, n) for s in ar]
            # print("4 of 5")
            # py = [binomial_american_put_option(s, K+inner_range, T, r, sigma, n) for s in ar]
            # #
            for s in ar:
                cy_ = binomial_american_call_option(s, K-inner_range-spread, T, r, sigma, n)
                cy = binomial_american_call_option(s, K - inner_range, T, r, sigma, n)
                py_ = binomial_american_put_option(s, K + inner_range + spread, T, r, sigma, n)
                py = binomial_american_put_option(s, K + inner_range, T, r, sigma, n)
                print(t, inner_range, spread, s, cy_, cy, py_, py)


            # csy = sub_(cy_, cy)
            # psy = sub_(py_, py)
            # print("5 of 5")
            # two_strategy = add_(psy, csy)
            #
            # result={"x": list(ar), "py_": py_, "py": py, "psy": psy, "cy":cy, "cy_":cy_, "csy":csy,
            #         "two_strategy": two_strategy}
            # # print(result)
            # df = pd.DataFrame.from_dict(result)
            # # print(df,"\n")
            #
            # print("5.1 of 5")
            #
            # print(df[df["x"]==S]["two_strategy"])
            # df["profit"] = round(100*(float(df[df["x"]==S]["two_strategy"])-df["two_strategy"]))/100
            # print(df)
            # return df

        nr = int(range_/range_bin)
        lr = list(range(nr, -nr, -1))
        ar = [S+lr[i]*range_bin for i in lr]
        ar.sort(key=float)
        # ar = range(S+range_, S - range_, -5)
        # print(list(ar))

        model_company = apps.get_model(app_label=app_, model_name="XBRLCompanyInfo")
        company_obj = model_company.objects.get(company_name='^GSPC')
        print(company_obj)
        model_ = apps.get_model(app_label=app_, model_name="FactSimulation")
        n_ = 390
        for t in range(1, n_+1, 1):
            print(t)
            T = t/n_
            r = T*r_
            sigma = sigma_ * math.sqrt(T)
            # if t in [1,n_]:
            print(t, T)
            # calc_(t, S, K, T, r, sigma, inner_range, spread, n_)
            for s in ar:
                cy_ = binomial_american_call_option(s, K - inner_range - spread, T, r, sigma, n_)
                cy = binomial_american_call_option(s, K - inner_range, T, r, sigma, n_)
                py_ = binomial_american_put_option(s, K + inner_range + spread, T, r, sigma, n_)
                py = binomial_american_put_option(s, K + inner_range, T, r, sigma, n_)
                print(t, inner_range, spread, s, cy_, cy, py_, py)
                obj, _ = model_.objects.get_or_create(company=company_obj, time=t, spread=spread,
                                                      inner_range=inner_range, stock_price=s)
                obj.ch = cy_
                obj.c = cy
                obj.ph = py_
                obj.p = py
                obj.save()

        t1 = 240
        t2 = 300
        T = (t2-t1)/n_
        sigma = sigma_ * math.sqrt(T)
        r = T*r_
        b = binomial_price_probability(S, T, r, sigma, n_)
        print(b)

        return {"status": "ok"}

    def get_two_spread_data(self, dic):
        print('80-80-80-333 get_two_spread_data 90055-300 dic\n', '-'*100, '\n', dic, '\n', '-'*100)
        app_ = dic["app"]
        model_ = apps.get_model(app_label=app_, model_name="FactSimulation")
        spread_=float(dic["spread"])
        inner_range_ = float(dic["inner_range"])
        S_=float(dic["S"])
        ticker_=str(dic["ticker"])

        sq = model_.objects.filter(company__ticker=ticker_, spread=spread_,
                                   inner_range=inner_range_).all().values('time', 'stock_price', 'c', 'ch', 'p', 'ph')
        df = pd.DataFrame(list(sq))

        df["cs"] = df["ch"] - df["c"]
        df["ps"] = df["ph"] - df["p"]
        df["cps"] = df["cs"] + df["ps"]
        # print(df)
        r = {}
        n_ = 390
        for i in range(n_, 0, -1):
            df_ = df[df["time"]==i]
            df_ = df_.drop('time', 1)
            # print(df_)
            cps_s = float(df_[df_["stock_price"]==S_]["cps"])
            df_["profit"] = cps_s
            df_ = df_.apply(pd.to_numeric)
            df_["profit"] = round(100*(df_["profit"] - df_["cps"]))/100
            # print(df_)
            k = n_-i+1
            # print(k)
            stock_price = []
            if i == 1:
                for index, row in df_.iterrows():
                    stock_price.append(row["stock_price"])
            r[k] = {"c":[], "ch":[], "p":[], "ph":[], "cs":[], "ps":[], "cps":[], "profit":[]}
            l_ = list(df_.columns.values)
            l_.remove('stock_price')
            for index, row in df_.iterrows():
                for c in l_:
                    # print(c, row[c])
                    r[k][c].append(row[c])

        # for k in r:
        #     if k == 385:
        #         print("\n", k)
        #         print("\n", r[k])

        # print(r)

        # print("\n")
        # print(stock_price)

        return {"status": "ok", "data": r, "stock_price":stock_price}


    def put_call_spread(self, dic):
        print('90-90-90-11 data_transfer_to_process_fact 90055-300 dic\n', '-'*100, '\n', dic, '\n', '-'*100)
        app_ = dic["app"]

        # Example usage:
        spread = float(dic["spread"])  # Current stock price
        S = float(dic["S"])  # Current stock price
        # cK = S-spread  # Strike price
        T = 1  # Time to expiration (in years)
        r = float(dic["r"])  # Risk-free interest rate
        sigma = float(dic["sigma"])  # Volatility
        gS = float(dic["gS"])  # change in stock price
        n = int(dic["n"])  # Number of time steps

        # call
        call_option_price_itm, put_option_price = self.binomial_american_call_option(S, S-spread, T, r, sigma, n)
        call_option_price_atm, put_option_price = self.binomial_american_call_option(S, S, T, r, sigma, n)
        spread_price_c_0 = call_option_price_itm-call_option_price_atm
        # up
        call_option_price_itm, put_option_price = self.binomial_american_call_option(S+gS, S-spread, T, r, sigma, n)
        call_option_price_atm, put_option_price = self.binomial_american_call_option(S+gS, S, T, r, sigma, n)
        spread_price_c_u = call_option_price_itm-call_option_price_atm
        # down
        call_option_price_itm, put_option_price = self.binomial_american_call_option(S-gS, S-spread, T, r, sigma, n)
        call_option_price_atm, put_option_price = self.binomial_american_call_option(S-gS, S, T, r, sigma, n)
        spread_price_c_d = call_option_price_itm-call_option_price_atm

        print("Call", "0=", round(100*spread_price_c_0)/100, "u=", round(100*spread_price_c_u)/100, "d=", round(100*spread_price_c_d)/100)

        # put
        call_option_price, put_option_price_itm = self.binomial_american_call_option(S, S + spread, T, r, sigma, n)
        call_option_price, put_option_price_atm = self.binomial_american_call_option(S, S, T, r, sigma, n)
        spread_price_p_0 = put_option_price_itm - put_option_price_atm
        # up
        call_option_price, put_option_price_itm = self.binomial_american_call_option(S + gS, S + spread, T, r, sigma, n)
        call_option_price, put_option_price_atm = self.binomial_american_call_option(S + gS, S, T, r, sigma, n)
        spread_price_p_u = put_option_price_itm - put_option_price_atm
        # down
        call_option_price, put_option_price_itm = self.binomial_american_call_option(S - gS, S + spread, T, r, sigma, n)
        call_option_price, put_option_price_atm = self.binomial_american_call_option(S - gS, S, T, r, sigma, n)
        spread_price_p_d = put_option_price_itm - put_option_price_atm

        print("Put ", "0=", round(100*spread_price_p_0)/100, "u=", round(100*spread_price_p_u)/100, "d=", round(100*spread_price_p_d)/100)

        #
        two_spread_price_0 = spread_price_c_0 + spread_price_p_0
        two_spread_price_u = spread_price_c_u + spread_price_p_u
        two_spread_price_d = spread_price_c_d + spread_price_p_d
        print(" call_put_0", round(100*two_spread_price_0)/100, "\n",
              "call_put_u", round(100*two_spread_price_u)/100, "\n",
              "call_put_d", round(100*two_spread_price_d)/100, "\n",
              "call_put_u_profit", round(100 * (two_spread_price_0 - two_spread_price_u)) / 100, "\n",
              "call_put_d_profit", round(100 * (two_spread_price_0 - two_spread_price_d)) / 100
              )

        result = {"status": "ok", "data": {"two_spread_price_0":round(100 * two_spread_price_0)/100,
                                           "two_spread_price_u":round(100 * two_spread_price_u)/100,
                                           "two_spread_price_d":round(100 * two_spread_price_d)/100,
                                           "two_spread_price_pu":round(100 * (two_spread_price_0-two_spread_price_u)) / 100,
                                           "two_spread_price_pd":round(100 * (two_spread_price_0-two_spread_price_d)) / 100}}
        return result
