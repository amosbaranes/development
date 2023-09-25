import warnings
import os
from pathlib import Path
#
from django.conf import settings
from django.contrib.auth.models import Group
import matplotlib as mpl
mpl.use('Agg')
from ..ml.basic_ml_objects import BaseDataProcessing, BasePotentialAlgo
from django.apps import apps
#
import pandas as pd
import numpy as np
import random
#
import time
from datetime import timedelta, date
import shutil
#
import string
# from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.contrib.auth import authenticate
#
from openpyxl import Workbook, load_workbook
#
from ...core.utils import log_debug, clear_log_debug

from .models import Soldiers, DoubleShoot as DoubleShootModel
from django.db.models.fields.related import ForeignKey
from django.http import JsonResponse
import requests
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models.fields import *


class TestManager(object):
    def __init__(self, dic=None):
        self.obj_dic = dic


class DoubleShoot(object):
    def __init__(self, dic=None):
        headers = {
          'accept': 'application/json','content-type': 'application/json',
          'origin': 'https://manage.double-shoot.com','referer': 'https://manage.double-shoot.com/FlexiCore/',
        }
        json_data = {'email': 'admin@synrgai.com','password': 'synrgaiadmin',}
        response = requests.post(
          'https://manage.double-shoot.com/FlexiCore/rest/authenticationNew/login', headers=headers, json=json_data,
        )
        if response.status_code == 200:
          result = response.json()
          self.authentication_key = result.get('authenticationKey')
        else:
          print("Request failed. Status code:", response.status_code)
        self.head = {
          'authority': 'manage.double-shoot.com', 'accept': 'application/json',
          'authenticationkey': self.authentication_key,
          'origin': 'https://manage.double-shoot.com','referer': 'https://manage.double-shoot.com/FlexiCore/',
          'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36',
        }
        self.head_url = 'https://manage.double-shoot.com/FlexiCore/rest/plugins/Member/getAllMembers'
        self.head_heat = {
          'accept': 'application/json',
          'authenticationkey': self.authentication_key,
          'content-type': 'application/json',
          'origin': 'https://manage.double-shoot.com','referer': 'https://manage.double-shoot.com/FlexiCore/',
        }
        self.head_heat_url = 'https://manage.double-shoot.com/FlexiCore/rest/plugins/Heat/getAllHeats'
        # self.head_heat_url = 'https://manage.double-shoot.com/FlexiCore/rest/plugins/Heat/getAllHeatsForMember'

    def get_soldier_data(self, dic=None):
        print(dic)
        ds_soldier_id = dic["ds_soldier_id"]
        function_name = dic["function_name"]
        # print("function name: ", function_name, "\nsoldier_id: ", soldier_id, "\n", "="*10)
        url_ = self.url_.format(function_name, ds_soldier_id)
        # print('\nurl_: ', url_, "\n", "="*10)
        return requests.get(url_, headers=self.head).json()

    # https://www.yellowduck.be/posts/outputting-django-queryset-json
    # print("need to get double shoot id")
    def update_or_get_soldier_data(self, dic=None):
        # print("-"*100, "\nupdate_or_get_soldier_data\n", dic, "\n", "-"*100)
        soldier_id = dic["soldier_id"]
        data = {
                'soldier_id': soldier_id,
                'double_shoot_id': "-1"
               }
        try:
            # print("90123-0 update_or_get_soldier_data 100\n", "solder_id: ", soldier_id, "\n", "-" * 100)
            soldier = Soldiers.objects.get(id=soldier_id)
            # print("90123-1 soldier obj: ", soldier)
            ds_row, is_created = DoubleShootModel.objects.get_or_create(soldier=soldier)
            data['double_shoot_id'] = ds_row.double_shoot_id
            # print(ds_row)
            # print("update_or_get_soldier_data 100-1", "\n", "-" * 100)
        except Exception as ex:
            print("90876-23: "+str(ex))
            data['double_shoot_id'] = "-100 error create record in ds model"
        # print('data 1')
        # print(data)
        # print('data 1')

        if not ds_row.is_pulled:
            try:
                # print(soldier_id)
                # print("Get data and update our system")
                function_name = dic["function_name"]
                ds_update_data = self.get_soldier_data({"ds_soldier_id": data['double_shoot_id'], "function_name": function_name})
                data["pulled_data"] = ds_update_data
                #
                # # # # update the DoubleShootModel with the pulled data.
                #
                ds_row.is_pulled = True
                ds_row.save()
            except Exception as ex:
                print("Error 9085-125 DoubleShoot update_or_get_soldier_data:\n"+ex)
                data['double_shoot_id'] = "-2"

        # update the data dictionary with data from our DoubleShootModel
        # print('data')
        # print(data)
        # print('data')

        return data

    def get_members(self, dic):
        # print("test_list\n", "-"*50, "\n", dic, "\n", "-"*50)
        app_ = dic["app"]
        json_data = {'pageSize': 1000,'currentPage': 0,}
        response = requests.post(self.head_url, headers=self.head, json=json_data,)
        if response.status_code == 200:
          response_json = response.json()
          members = response_json['list']
          model_ = apps.get_model(app_label=app_, model_name="doubleshootmembers")
          for member in members:
            # print("member\n", "-"*50, "\n", member, "\n", "-"*50)
            member_name = member.get('name')
            member_id = member.get('id')
            try:
              obj, is_created= model_.objects.get_or_create(ds_id=member_id)
              obj.ds_name = member_name
              obj.save()
            except Exception as ex:
              print("Error 22222222: "+str(ex))
        else:
          print("Request failed. Status code:", response.status_code)
        result = {"status": "ok"}
        # print(result)
        return result

    def process_data(self, list_,data,gun_type,sight_types,targets,ranges, shooting_exams, members, n__):
        for heat in list_["list"]:
            try:
                try:
                    maxDistance = 0
                    maxDistance = round(heat["maxDistance"]*100)/100
                except Exception as ex:
                    pass
                dateHeld = ''.join(heat["dateHeld"][:10].split("-"))
                try:
                    averageHitPointX = 0
                    averageHitPointX = round(heat["averageHitPointX"]*100)/100
                except Exception as ex:
                    pass
                try:
                    averageHitPointY = 0
                    averageHitPointY = round(heat["averageHitPointY"]*100)/100
                except Exception as ex:
                    pass

                numberOfHits = heat["numberOfHits"]
                shotsFired = heat["shotsFired"]
                try:
                    gunType = heat["gunType"]["name"]
                    gun_type[heat["gunType"]["json-id"]] = gunType
                except Exception as ex:
                    # print(gun_type, "\n", heat["gunType"])
                    gunType = gun_type[heat["gunType"]]
                try:
                    sightType = heat["sightType"]["name"]
                    sight_types[heat["sightType"]["json-id"]] = sightType
                except Exception as ex:
                    sightType = sight_types[heat["sightType"]]
                try:
                    target = heat["target"]["name"]
                    targets[heat["target"]["json-id"]] = target
                except Exception as ex:
                    target = targets[heat["target"]]
                # ??? do we need more info from target ?
                #         "widthInCm": 18.8,
                #         "heightInCm": 27.8,
                #         "bullXCmFromLeft": 9.4,
                #         "bullYCmFromTop": 9.4,
                try:
                    range_ = heat["range"]["name"]
                    ranges[heat["range"]["json-id"]] = range_
                except Exception as ex:
                    range_ = ranges[heat["range"]]
                sp = heat["shootingPractice"]
                try:
                    name_ = "none"
                    member = sp["member"]
                    print("memeber\n","-"*10,"\n",member,"\n","-"*10,"\n")
                    name_ = member["name"]
                    members[member["json-id"]] = name_
                    # print("exist\n", members)
                except Exception as ex:
                    # print("Error 1:\n","-"*10, "\n: "+str(ex), "\n","-"*10,"\n")
                    # print(sp)
                    print("error\n", member, "\n", members)
                    try:
                        name_ = members[member]
                    except Exception as ex:
                        print("error name\n", member, "\n", members)
                # print(name_)
                try:
                    score = -1
                    shooting_exam = 'none'
                    try:
                        seo = sp["shootingExamOccurrence"]
                        # print(seo)
                        score = seo["score"]
                        # print(score)
                        shooting_exam = seo["shootingExam"]["name"]
                        # print(shooting_exam)
                        # print(seo["shootingExam"]["json-id"])
                        shooting_exams[seo["shootingExam"]["json-id"]] = shooting_exam
                        # print(shooting_exams)
                        # print("heat shooting_exam\n", "=" * 10, "\n", heat, "\n", "=" * 10, "\n")
                    except Exception as ex:
                        shooting_exam = shooting_exams[seo["shootingExam"]]

                    print(name_, score, shooting_exam)
                except Exception as eex:
                    pass

                try:
                    if str(name_).isnumeric():
                        n__ += 1
                        if str(name_) not in data:
                            data[str(name_)] = {}
                        if gunType not in data[str(name_)]:
                            data[str(name_)][gunType] = {}
                        if range_ not in data[str(name_)][gunType]:
                            data[str(name_)][gunType][range_] = {}
                        if dateHeld not in data[str(name_)][gunType][range_]:
                            data[str(name_)][gunType][range_][dateHeld] = {}

                        print(n__, name_, gunType, range_, dateHeld, maxDistance, averageHitPointX, averageHitPointY,
                              numberOfHits, shotsFired, sightType, target, shooting_exam, score)
                        data[str(name_)][gunType][range_][dateHeld] = [maxDistance, averageHitPointX, averageHitPointY,
                                                                       numberOfHits, shotsFired, sightType, target,
                                                                       shooting_exam, score]
                except Exception as ex:
                    print("Error 11:\n", name_, str(ex))
                # print(1007)
            except Exception as ex:
                print("Error 2:\n","-"*10, "\n: "+str(ex), "\n","-"*10,"\n")

    def get_members_data(self, dic):
        print("get_members_data\n", "-"*50, "\n", dic, "\n", "-"*50)
        members_ = [] # 'F11jUGZlR7y3oqZc1uvuXw'
        page_size = 50
        data_ = {'fields':['maxDistance', 'averageHitPointX', 'averageHitPointY', 'numberOfHits', 'shotsFired', 'sightType', 'target'], 'data':{}}
        data = data_["data"]
        gun_type = {}
        sight_types = {}
        targets = {}
        ranges = {}
        shooting_exams = {}
        members = {}
        n__ = 0

        json_data_heat = {'pageSize': page_size,'currentPage': 0, 'memeberIds': members_}
        response = requests.post(self.head_heat_url, headers=self.head_heat, json=json_data_heat,)
        if response.status_code == 200:
            list_ = response.json()
            # print(list_)
            total_pages = list_['totalPages']
            print(total_pages)
            self.process_data(list_, data, gun_type, sight_types, targets, ranges, shooting_exams, members, n__)
            for i in range(1, total_pages):
                print(i)
                json_data_heat = {'pageSize': page_size,'currentPage': i, 'memeberIds': members_}
                response = requests.post(self.head_heat_url, headers=self.head_heat, json=json_data_heat,)
                if response.status_code == 200:
                    list_ = response.json()
                    self.process_data(list_, data, gun_type, sight_types, targets, ranges, shooting_exams, members, n__)
                else:
                    print("Request failed. Status code:", response.status_code)
        else:
            print("Request failed. Status code:", response.status_code)

        # list_ = {
        #     "list": [
        #     {
        #       "json-id": "c451d159-e242-4795-b3fd-51ce8520df8a",
        #       "id": "02qCEMBGSuSjSaha63RO4g",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-01-18T18:57:36.318+01:00",
        #       "updateDate": "2023-01-21T09:52:53.130+01:00",
        #       "dtype": "Heat",
        #       "tenant": {
        #         "json-id": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "id": "VYL49flzRKKroW7koxnt0w",
        #         "name": "SYNRGAI",
        #         "description": "",
        #         "softDelete": False,
        #         "creationDate": "2022-12-11T10:13:25.452+01:00",
        #         "updateDate": "2022-12-11T10:13:25.453+01:00",
        #         "dtype": "Tenant",
        #         "tenant": {
        #           "json-id": "999a4182-746c-44b4-8d15-19f69c70e536",
        #           "id": "jgV8M9d0Qd6owkPPFrbWIQ",
        #           "name": "defaulttenant",
        #           "description": "The default tenantToBaseClassPremissions all users belong to unless having special tenantToBaseClassPremissions",
        #           "softDelete": False,
        #           "creationDate": "2017-09-22T08:59:53.397+02:00",
        #           "updateDate": "2023-05-14T15:21:31.444+02:00",
        #           "dtype": "Tenant",
        #           "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #           "systemObject": False,
        #           "externalId": None,
        #           "icon": None,
        #           "noSQL": False,
        #           "javaType": "com.flexicore.model.Tenant",
        #           "json-type": "com.flexicore.model.Tenant"
        #         },
        #         "systemObject": False,
        #         "externalId": None,
        #         "icon": None,
        #         "noSQL": False,
        #         "javaType": "com.flexicore.model.Tenant",
        #         "json-type": "com.flexicore.model.Tenant"
        #       },
        #       "systemObject": False,
        #       "dateHeld": "2023-01-18T17:22:26.000+01:00",
        #       "externalId": "15d05c48-6264-4f5f-845a-cb84beb5220a",
        #       "score": None,
        #       "heatType": None,
        #       "maxDistance": 8.166504912493789,
        #       "averageHitPointX": 204.1301523844401,
        #       "averageHitPointY": 329.92139689127606,
        #       "ordinal": 1,
        #       "numberOfHits": 22,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 4,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": {
        #         "json-id": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #         "id": "-CgpgJOES0KYXEKWQmGUfQ",
        #         "name": "MZ-4P",
        #         "description": "MZ-4P",
        #         "softDelete": False,
        #         "creationDate": "2022-12-11T10:52:30.541+01:00",
        #         "updateDate": "2022-12-12T19:29:26.054+01:00",
        #         "dtype": "GunType",
        #         "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #         "systemObject": False,
        #         "horizontalClicksToCmRatioPerMeter": 0.04,
        #         "verticalClicksToCmRatioPerMeter": 0.06,
        #         "caliber": 5.56,
        #         "demo": False,
        #         "fileResource": None,
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.GunType",
        #         "json-type": "com.doubleshot.model.GunType"
        #       },
        #       "sightType": {
        #         "json-id": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #         "id": "1KMKPmfRT9uYXhQXQEnhiw",
        #         "name": "Iron Sight",
        #         "description": "Iron Sight",
        #         "softDelete": False,
        #         "creationDate": "2019-10-05T09:20:07.540+02:00",
        #         "updateDate": "2019-10-15T07:44:35.352+02:00",
        #         "dtype": "SightType",
        #         "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #         "systemObject": False,
        #         "horizontalClicksToCmRatioPerMeter": 0.028,
        #         "verticalClicksToCmRatioPerMeter": 0.036,
        #         "ironSight": True,
        #         "demo": True,
        #         "userCreated": False,
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.SightType",
        #         "json-type": "com.doubleshot.model.SightType"
        #       },
        #       "target": {
        #         "json-id": "2812e906-4642-44a0-8439-1ee6f5303a31",
        #         "id": "aNIM19TgScmIVyIptNQgmQ",
        #         "name": "DS Zeroing A4 Target",
        #         "description": "DS Zeroing Target",
        #         "softDelete": False,
        #         "creationDate": "2019-10-05T08:59:56.948+02:00",
        #         "updateDate": "2023-03-19T19:29:36.115+01:00",
        #         "dtype": "Target",
        #         "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #         "systemObject": False,
        #         "targetImage": {
        #           "json-id": "67047fd7-41c6-4b6a-a4d9-87e201bb032f",
        #           "id": "vH53-kYBQ4+AjuTh4mYWNA",
        #           "name": "Target-Civil-B.jpg",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-11T11:13:37.578+01:00",
        #           "updateDate": "2022-12-11T11:13:38.332+01:00",
        #           "md5": "64889e979e012df9ad192ae7e6572399",
        #           "offset": 646807,
        #           "actualFilename": "8993852c-3f93-429f-84c0-f6e0e3ef35aa.jpg",
        #           "originalFilename": "Target-Civil-B.jpg",
        #           "done": True,
        #           "dateTaken": None,
        #           "nonDownloadable": False,
        #           "keepUntil": None,
        #           "onlyFrom": None,
        #           "javaType": "com.wizzdi.flexicore.file.model.FileResource",
        #           "json-type": "com.wizzdi.flexicore.file.model.FileResource"
        #         },
        #         "noDefaultFiles": False,
        #         "ratioY": 297,
        #         "ratioX": 210,
        #         "settingResource": {
        #           "json-id": "5e57f229-e087-4d8c-a4cc-2ddea1add2a9",
        #           "id": "8Xp0a9GhRJW+PCZriH9wLg",
        #           "name": "dsZeroParams.txt",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": None,
        #           "updateDate": None,
        #           "md5": "85a2438b711ec60f4d386ba2b2d9dd6d",
        #           "offset": 703,
        #           "actualFilename": "90da6f73-eee6-495c-9a03-5d66a3bdc738.txt",
        #           "originalFilename": "dsZeroParams.txt",
        #           "done": False,
        #           "dateTaken": None,
        #           "nonDownloadable": False,
        #           "keepUntil": None,
        #           "onlyFrom": None,
        #           "javaType": "com.wizzdi.flexicore.file.model.FileResource",
        #           "json-type": "com.wizzdi.flexicore.file.model.FileResource"
        #         },
        #         "usingClicks": True,
        #         "usingScore": False,
        #         "continueWhenZeroed": True,
        #         "finishWhenNotZeroed": True,
        #         "hitsRadius": 10,
        #         "svmWhite": {
        #           "json-id": "13e80c7a-eb87-4253-9ebf-6b3f36346052",
        #           "id": "7C1OHejGS6KNIK7VQd3cXg",
        #           "name": "dm2015.svm",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": None,
        #           "updateDate": None,
        #           "md5": "2df7de6a6c59939d3a5f6c2eca0a4d22",
        #           "offset": 19893,
        #           "actualFilename": "1c56651e-ffdd-46f6-8037-fe5c0b5ea15e.svm",
        #           "originalFilename": "dm2015.svm",
        #           "done": True,
        #           "dateTaken": None,
        #           "nonDownloadable": False,
        #           "keepUntil": None,
        #           "onlyFrom": None,
        #           "javaType": "com.wizzdi.flexicore.file.model.FileResource",
        #           "json-type": "com.wizzdi.flexicore.file.model.FileResource"
        #         },
        #         "svmBlack": "13e80c7a-eb87-4253-9ebf-6b3f36346052",
        #         "mask": {
        #           "json-id": "e49c1308-ff6d-4ed3-9b51-108586674448",
        #           "id": "kvWQiARwSEG+fsc+50IqGg",
        #           "name": "dsZeroBin.png",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": None,
        #           "updateDate": None,
        #           "md5": "abf91e8f0b1a86954e38ef1c1e9feaad",
        #           "offset": 42477,
        #           "actualFilename": "8cf2a0f8-60ef-4fbc-bb73-d2cb73047df0.png",
        #           "originalFilename": "dsZeroBin.png",
        #           "done": True,
        #           "dateTaken": None,
        #           "nonDownloadable": False,
        #           "keepUntil": None,
        #           "onlyFrom": None,
        #           "javaType": "com.wizzdi.flexicore.file.model.FileResource",
        #           "json-type": "com.wizzdi.flexicore.file.model.FileResource"
        #         },
        #         "scoringCalculator": None,
        #         "widthInCm": 18.8,
        #         "heightInCm": 27.8,
        #         "bullXCmFromLeft": 9.4,
        #         "bullYCmFromTop": 9.4,
        #         "availableForDemo": True,
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.Target",
        #         "json-type": "com.doubleshot.model.Target"
        #       },
        #       "range": {
        #         "json-id": "c8c13311-22a0-4bc8-b212-3c5277849532",
        #         "id": "tiJ1oCtOQNCCbSrgyANKJQ",
        #         "name": "25m",
        #         "description": "25m",
        #         "softDelete": False,
        #         "creationDate": "2019-10-06T12:28:52.295+02:00",
        #         "updateDate": "2019-10-06T12:28:52.297+02:00",
        #         "dtype": "Range",
        #         "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #         "systemObject": False,
        #         "rangeInUnit": 25,
        #         "measureUnit": "METRIC",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.Range",
        #         "json-type": "com.doubleshot.model.Range"
        #       },
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "5cc11c60-30a4-41d3-b8ba-ef7ad474ed58",
        #         "id": "q2IIDZkHQHKR4s6428Vpqw",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-01-18T18:57:36.318+01:00",
        #         "updateDate": "2023-01-23T22:24:09.482+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-01-18T17:22:11.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 0,
        #         "shootingExamOccurrence": None,
        #         "externalId": "a4abb898-1d36-49f5-be16-3d04d866d4f7",
        #         "member": {
        #           "json-id": "f8a4d26d-aa3c-49e7-8ac2-e0ae05c58410",
        #           "id": "F11jUGZlR7y3oqZc1uvuXw",
        #           "name": "23155",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-28T15:53:14.481+01:00",
        #           "updateDate": "2022-12-28T15:53:14.661+01:00",
        #           "dtype": "Member",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "email": "23155@syn.com",
        #           "homeDir": "/home/flexicore/users/23155F11jUGZlR7y3oqZc1uvuXw",
        #           "surName": "",
        #           "disabled": False,
        #           "dateApproved": None,
        #           "uiConfiguration": None,
        #           "lastVerificationDate": None,
        #           "totpEnabled": False,
        #           "phoneNumber": "",
        #           "externalId": None,
        #           "verificationDate": None,
        #           "contactByPhone": False,
        #           "lastUsedDeviceIdentification": None,
        #           "productId": None,
        #           "tokenId": None,
        #           "packageName": None,
        #           "platform": None,
        #           "appPurchase": False,
        #           "image": None,
        #           "commandingUnit": None,
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.Member",
        #           "json-type": "com.doubleshot.model.Member"
        #         },
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "214258b1-adc5-49c6-93f0-890734f6191b",
        #       "id": "02xhSE2kR9SIRGw7832A9A",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2022-12-28T15:46:28.475+01:00",
        #       "updateDate": "2022-12-28T15:46:47.992+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2022-12-28T15:46:18.000+01:00",
        #       "externalId": "62c74075-a269-4dd8-83ec-48a05d900a5b",
        #       "score": None,
        #       "heatType": None,
        #       "maxDistance": 1.7533200649054344,
        #       "averageHitPointX": 213.66666666666666,
        #       "averageHitPointY": 133.66666666666666,
        #       "ordinal": 2,
        #       "numberOfHits": 8,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 4,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": {
        #         "json-id": "6fe0fd65-a56b-4223-a32d-1951cc152399",
        #         "id": "iv9GhivzRw+-o2zIA5Ld6g",
        #         "name": "Aimpoint COMPM5/M5S",
        #         "description": "Aimpoint COMPM5/M5S",
        #         "softDelete": False,
        #         "creationDate": "2019-10-05T09:14:37.120+02:00",
        #         "updateDate": "2019-10-05T09:14:37.122+02:00",
        #         "dtype": "SightType",
        #         "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #         "systemObject": False,
        #         "horizontalClicksToCmRatioPerMeter": 0.01,
        #         "verticalClicksToCmRatioPerMeter": 0.01,
        #         "ironSight": False,
        #         "demo": False,
        #         "userCreated": False,
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.SightType",
        #         "json-type": "com.doubleshot.model.SightType"
        #       },
        #       "target": "2812e906-4642-44a0-8439-1ee6f5303a31",
        #       "range": "c8c13311-22a0-4bc8-b212-3c5277849532",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "60624e85-e42e-47c7-aac4-18acf6fedb42",
        #         "id": "l6rQ1yFuQcqIsjdhoBTM9A",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2022-12-28T15:46:28.469+01:00",
        #         "updateDate": "2022-12-28T15:46:47.992+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2022-12-28T15:45:52.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 14,
        #         "shootingExamOccurrence": None,
        #         "externalId": "7fa3259c-c813-42f3-857c-6f4b36edcfb6",
        #         "member": {
        #           "json-id": "f08829c9-8122-4d3d-afa1-bc0172d65e3c",
        #           "id": "RAKKenh9SAGvs5xDNvECGA",
        #           "name": "SYN-RG-Ai_1/23",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-21T11:50:16.282+01:00",
        #           "updateDate": "2023-05-12T13:44:01.326+02:00",
        #           "dtype": "Member",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "email": "syn@jan23.com",
        #           "homeDir": "/home/flexicore/users/SYN-RG-Ai 1/23RAKKenh9SAGvs5xDNvECGA",
        #           "surName": None,
        #           "disabled": False,
        #           "dateApproved": None,
        #           "uiConfiguration": None,
        #           "lastVerificationDate": None,
        #           "totpEnabled": False,
        #           "phoneNumber": None,
        #           "externalId": None,
        #           "verificationDate": None,
        #           "contactByPhone": False,
        #           "lastUsedDeviceIdentification": "test",
        #           "productId": None,
        #           "tokenId": None,
        #           "packageName": None,
        #           "platform": None,
        #           "appPurchase": False,
        #           "image": {
        #             "json-id": "234d7b90-51c1-42c9-b8fb-51f36a692a62",
        #             "id": "ulKlvncdQc6Hj+uDcNoV5g",
        #             "name": "image_picker6265075270655949444.jpg",
        #             "description": None,
        #             "softDelete": False,
        #             "creationDate": "2022-12-28T15:38:57.184+01:00",
        #             "updateDate": "2022-12-28T15:38:57.409+01:00",
        #             "md5": "17afcf95c0f1bdd505ce21b85cd0a321",
        #             "offset": 34827,
        #             "actualFilename": "d581246a-e39c-4fc8-8f60-9c48cb637cab.jpg",
        #             "originalFilename": "image_picker6265075270655949444.jpg",
        #             "done": True,
        #             "dateTaken": None,
        #             "nonDownloadable": False,
        #             "keepUntil": None,
        #             "onlyFrom": None,
        #             "javaType": "com.wizzdi.flexicore.file.model.FileResource",
        #             "json-type": "com.wizzdi.flexicore.file.model.FileResource"
        #           },
        #           "commandingUnit": None,
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.Member",
        #           "json-type": "com.doubleshot.model.Member"
        #         },
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "fc90f05d-27eb-4c51-a127-7b842408eb5a",
        #       "id": "03gqglh0Svi6XEOtvU0CqA",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-01-23T22:46:31.560+01:00",
        #       "updateDate": "2023-01-23T22:46:31.562+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-01-23T15:13:47.000+01:00",
        #       "externalId": "54022cf2-87f7-49da-9ab3-ec1af5cc52dd",
        #       "score": None,
        #       "heatType": None,
        #       "maxDistance": 6.428616200240919,
        #       "averageHitPointX": 336,
        #       "averageHitPointY": 232.5,
        #       "ordinal": 1,
        #       "numberOfHits": 5,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 4,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "2812e906-4642-44a0-8439-1ee6f5303a31",
        #       "range": "c8c13311-22a0-4bc8-b212-3c5277849532",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "9c79b4ee-77ad-4989-af0b-5c738b14ef00",
        #         "id": "k7IrNZCySXObM5JzzIvZkg",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-01-23T22:46:31.560+01:00",
        #         "updateDate": "2023-01-25T14:50:22.144+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-01-23T15:13:02.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 0,
        #         "shootingExamOccurrence": None,
        #         "externalId": "bfebba8c-71d4-4880-a80d-ea54730d175f",
        #         "member": {
        #           "json-id": "f48a330b-5523-4fa9-9134-f5355dd216db",
        #           "id": "wPOehd2QQSG8OWzIpwrhcA",
        #           "name": "23343",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-28T16:02:50.501+01:00",
        #           "updateDate": "2022-12-28T16:03:01.472+01:00",
        #           "dtype": "Member",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "email": "23343@syn.com",
        #           "homeDir": "/home/flexicore/users/23343wPOehd2QQSG8OWzIpwrhcA",
        #           "surName": "",
        #           "disabled": False,
        #           "dateApproved": None,
        #           "uiConfiguration": None,
        #           "lastVerificationDate": None,
        #           "totpEnabled": False,
        #           "phoneNumber": "",
        #           "externalId": None,
        #           "verificationDate": None,
        #           "contactByPhone": False,
        #           "lastUsedDeviceIdentification": None,
        #           "productId": None,
        #           "tokenId": None,
        #           "packageName": None,
        #           "platform": None,
        #           "appPurchase": False,
        #           "image": None,
        #           "commandingUnit": None,
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.Member",
        #           "json-type": "com.doubleshot.model.Member"
        #         },
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "be906ef9-07e4-4001-8991-2dd826d5651b",
        #       "id": "04Orci0DQeiBtTKOHd6WsQ",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-01-17T18:27:31.156+01:00",
        #       "updateDate": "2023-01-21T09:32:42.608+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-01-17T11:16:40.000+01:00",
        #       "externalId": "8e0ff242-7f29-4026-b5b7-72fd01dc307a",
        #       "score": None,
        #       "heatType": None,
        #       "maxDistance": 4.463401533583999,
        #       "averageHitPointX": 189.66666666666666,
        #       "averageHitPointY": 68.66666666666667,
        #       "ordinal": 2,
        #       "numberOfHits": 16,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 4,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "2812e906-4642-44a0-8439-1ee6f5303a31",
        #       "range": "c8c13311-22a0-4bc8-b212-3c5277849532",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "63d12ddc-06db-49ae-b0e3-c4c3a4f39663",
        #         "id": "r9RuZfWaQsKpd5w76uzWeQ",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-01-17T18:27:31.156+01:00",
        #         "updateDate": "2023-01-21T09:37:25.911+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-01-17T10:56:03.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 0,
        #         "shootingExamOccurrence": None,
        #         "externalId": "a6b27293-d4e9-4dea-9ca2-7ca641e71559",
        #         "member": {
        #           "json-id": "d49d339e-1a3c-45e7-8b53-85174a326799",
        #           "id": "nXsNuVt8SXWIB8S0KyCsJA",
        #           "name": "23053",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-28T15:35:26.318+01:00",
        #           "updateDate": "2022-12-28T15:35:35.514+01:00",
        #           "dtype": "Member",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "email": "23053@syn.com",
        #           "homeDir": "/home/flexicore/users/23053nXsNuVt8SXWIB8S0KyCsJA",
        #           "surName": "",
        #           "disabled": False,
        #           "dateApproved": None,
        #           "uiConfiguration": None,
        #           "lastVerificationDate": None,
        #           "totpEnabled": False,
        #           "phoneNumber": "",
        #           "externalId": None,
        #           "verificationDate": None,
        #           "contactByPhone": False,
        #           "lastUsedDeviceIdentification": None,
        #           "productId": None,
        #           "tokenId": None,
        #           "packageName": None,
        #           "platform": None,
        #           "appPurchase": False,
        #           "image": None,
        #           "commandingUnit": None,
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.Member",
        #           "json-type": "com.doubleshot.model.Member"
        #         },
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "f613e393-bd63-4626-8a08-584806063661",
        #       "id": "05ql9APjRpWQQNvC03sAiw",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-05-12T13:27:04.905+02:00",
        #       "updateDate": "2023-05-12T13:27:04.905+02:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-05-11T18:28:13.000+02:00",
        #       "externalId": "e7999ea9-b645-404c-aeed-b5b01a65d242",
        #       "score": 115,
        #       "heatType": None,
        #       "maxDistance": 42.55834965080296,
        #       "averageHitPointX": 215.7826086956522,
        #       "averageHitPointY": 286.2173913043478,
        #       "ordinal": 1,
        #       "numberOfHits": 23,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 23,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "6fe0fd65-a56b-4223-a32d-1951cc152399",
        #       "target": {
        #         "json-id": "c82b4c9e-ee01-4e27-b6a6-6cdbabadd5e2",
        #         "id": "Q4lCRtUSRH+vW35zCyKchA",
        #         "name": "ABC - K",
        #         "description": "ABC with a different calculator ",
        #         "softDelete": False,
        #         "creationDate": "2023-02-07T09:04:57.754+01:00",
        #         "updateDate": "2023-02-14T14:50:17.206+01:00",
        #         "dtype": "Target",
        #         "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #         "systemObject": False,
        #         "targetImage": {
        #           "json-id": "fccbaced-9490-4c19-bca6-970cacda534c",
        #           "id": "ULQ6ul53Rh2+Edlh-bbRtQ",
        #           "name": "abc",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-02-07T09:03:54.084+01:00",
        #           "updateDate": "2023-02-07T09:03:54.084+01:00",
        #           "md5": "0f94f94a05571cb1dc91826e35a80422",
        #           "offset": 21377,
        #           "actualFilename": "31352974-8f80-4c5e-a7ca-de3cbe8a554f.",
        #           "originalFilename": "abc",
        #           "done": True,
        #           "dateTaken": None,
        #           "nonDownloadable": False,
        #           "keepUntil": None,
        #           "onlyFrom": None,
        #           "javaType": "com.wizzdi.flexicore.file.model.FileResource",
        #           "json-type": "com.wizzdi.flexicore.file.model.FileResource"
        #         },
        #         "noDefaultFiles": False,
        #         "ratioY": 0,
        #         "ratioX": 0,
        #         "settingResource": {
        #           "json-id": "c877fbaf-4226-4376-9eac-d57cb9347aea",
        #           "id": "lzBVlCj7Q0avoE-CSED6xg",
        #           "name": "abcParams.txt",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-02-07T09:04:05.072+01:00",
        #           "updateDate": "2023-02-07T09:04:05.072+01:00",
        #           "md5": "d97660ac31e7f9424350bd53625644cf",
        #           "offset": 883,
        #           "actualFilename": "24bf59c4-0e63-4a75-b7c4-a19071751a84.txt",
        #           "originalFilename": "abcParams.txt",
        #           "done": True,
        #           "dateTaken": None,
        #           "nonDownloadable": False,
        #           "keepUntil": None,
        #           "onlyFrom": None,
        #           "javaType": "com.wizzdi.flexicore.file.model.FileResource",
        #           "json-type": "com.wizzdi.flexicore.file.model.FileResource"
        #         },
        #         "usingClicks": False,
        #         "usingScore": True,
        #         "continueWhenZeroed": True,
        #         "finishWhenNotZeroed": True,
        #         "hitsRadius": 10,
        #         "svmWhite": {
        #           "json-id": "07541396-c6aa-497e-83de-a999e8af82e3",
        #           "id": "ZAo7vxIISn+-3TuB7lk32A",
        #           "name": "abcGrey.svm",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-02-07T09:04:23.244+01:00",
        #           "updateDate": "2023-02-07T09:12:12.211+01:00",
        #           "md5": "9da62a22076c11cd05cbcd8b4b4a8ee7",
        #           "offset": 15250,
        #           "actualFilename": "727cbe3d-6da8-4281-8a50-8bdffe4cacf4.svm",
        #           "originalFilename": "abcGrey.svm",
        #           "done": True,
        #           "dateTaken": None,
        #           "nonDownloadable": False,
        #           "keepUntil": None,
        #           "onlyFrom": None,
        #           "javaType": "com.wizzdi.flexicore.file.model.FileResource",
        #           "json-type": "com.wizzdi.flexicore.file.model.FileResource"
        #         },
        #         "svmBlack": "07541396-c6aa-497e-83de-a999e8af82e3",
        #         "mask": "07541396-c6aa-497e-83de-a999e8af82e3",
        #         "scoringCalculator": "ABCScoringCalculatorK",
        #         "widthInCm": 50,
        #         "heightInCm": 70,
        #         "bullXCmFromLeft": 25,
        #         "bullYCmFromTop": 35,
        #         "availableForDemo": False,
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.Target",
        #         "json-type": "com.doubleshot.model.Target"
        #       },
        #       "range": {
        #         "json-id": "b0f018d6-ee65-4c47-be41-6cc3b216a4d0",
        #         "id": "BxuLSssCRpyBCQZF3oVhfw",
        #         "name": "30m",
        #         "description": "30m",
        #         "softDelete": False,
        #         "creationDate": "2019-10-06T18:45:09.756+02:00",
        #         "updateDate": "2019-10-06T18:45:09.758+02:00",
        #         "dtype": "Range",
        #         "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #         "systemObject": False,
        #         "rangeInUnit": 30,
        #         "measureUnit": "METRIC",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.Range",
        #         "json-type": "com.doubleshot.model.Range"
        #       },
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "019d2e43-8ebb-40e2-8c99-3338206809c7",
        #         "id": "qL3HJanNTd61ej7jw+9AMA",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-05-12T13:27:04.905+02:00",
        #         "updateDate": "2023-05-12T13:27:04.905+02:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-05-11T18:07:44.000+02:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 29,
        #         "shootingExamOccurrence": {
        #           "json-id": "f05a5eb1-b248-48b4-8b1e-111118d21400",
        #           "id": "a5Q+fCqGQCepUISses8ecQ",
        #           "name": None,
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-05-12T13:27:04.605+02:00",
        #           "updateDate": "2023-05-12T13:27:04.905+02:00",
        #           "dtype": "ShootingExamOccurrence",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "score": 115,
        #           "shootingExam": {
        #             "json-id": "e8e772d0-db12-4ef6-ace4-fad1e077495e",
        #             "id": "1atJNzKcTvm4CGVGtEc0IA",
        #             "name": "1.Rifleman Summary Session",
        #             "description": "10+10 rounds, standing and kneeling on the ABC target at 30m",
        #             "softDelete": False,
        #             "creationDate": "2022-12-27T10:04:36.281+01:00",
        #             "updateDate": "2023-02-05T19:48:54.055+01:00",
        #             "dtype": "ShootingExam",
        #             "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #             "systemObject": False,
        #             "minimumScore": 0,
        #             "manual": None,
        #             "noSQL": False,
        #             "javaType": "com.doubleshot.model.ShootingExam",
        #             "json-type": "com.doubleshot.model.ShootingExam"
        #           },
        #           "member": {
        #             "json-id": "ef2d2436-0c6e-4a1d-b1c1-4e17743fb516",
        #             "id": "IfOYoSHkSwuJ5s+PZFOxyA",
        #             "name": "23120",
        #             "description": None,
        #             "softDelete": False,
        #             "creationDate": "2022-12-28T15:51:00.252+01:00",
        #             "updateDate": "2022-12-28T15:51:02.347+01:00",
        #             "dtype": "Member",
        #             "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #             "systemObject": False,
        #             "email": "23120@syn.com",
        #             "homeDir": "/home/flexicore/users/23120IfOYoSHkSwuJ5s+PZFOxyA",
        #             "surName": "",
        #             "disabled": False,
        #             "dateApproved": None,
        #             "uiConfiguration": None,
        #             "lastVerificationDate": None,
        #             "totpEnabled": False,
        #             "phoneNumber": "",
        #             "externalId": None,
        #             "verificationDate": None,
        #             "contactByPhone": False,
        #             "lastUsedDeviceIdentification": None,
        #             "productId": None,
        #             "tokenId": None,
        #             "packageName": None,
        #             "platform": None,
        #             "appPurchase": False,
        #             "image": None,
        #             "commandingUnit": None,
        #             "noSQL": False,
        #             "javaType": "com.doubleshot.model.Member",
        #             "json-type": "com.doubleshot.model.Member"
        #           },
        #           "externalId": "bb151d28-c230-4934-8d66-8d9e94e81248",
        #           "passed": True,
        #           "dateHeld": "2023-05-11T18:28:33.000+02:00",
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.ShootingExamOccurrence",
        #           "json-type": "com.doubleshot.model.ShootingExamOccurrence"
        #         },
        #         "externalId": "058cd96e-5137-4795-81b3-477b43f869dd",
        #         "member": "ef2d2436-0c6e-4a1d-b1c1-4e17743fb516",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "0246d550-1ba0-4d8c-b5bc-d99837ec90ee",
        #       "id": "06nSeri5RGyB-0DbID6J3g",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-02-09T15:16:36.747+01:00",
        #       "updateDate": "2023-02-09T15:16:38.645+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-02-09T09:52:21.000+01:00",
        #       "externalId": "a61cacfd-f452-4cf4-9894-03c79c5c6249",
        #       "score": 100,
        #       "heatType": None,
        #       "maxDistance": 44.54259363368889,
        #       "averageHitPointX": 136.30325241088866,
        #       "averageHitPointY": 239.9533504486084,
        #       "ordinal": 1,
        #       "numberOfHits": 33,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 20,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": {
        #         "json-id": "904e4057-a32c-458a-a18c-96dd760bd4b1",
        #         "id": "IPxowbZGQu2mI0FCHqvafg",
        #         "name": "9mm/0.40/0.45",
        #         "description": "Generic Gun 9mm",
        #         "softDelete": False,
        #         "creationDate": "2019-10-06T18:42:44.500+02:00",
        #         "updateDate": "2023-02-22T10:26:46.725+01:00",
        #         "dtype": "GunType",
        #         "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #         "systemObject": False,
        #         "horizontalClicksToCmRatioPerMeter": 0,
        #         "verticalClicksToCmRatioPerMeter": 0,
        #         "caliber": 9,
        #         "demo": True,
        #         "fileResource": None,
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.GunType",
        #         "json-type": "com.doubleshot.model.GunType"
        #       },
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "c82b4c9e-ee01-4e27-b6a6-6cdbabadd5e2",
        #       "range": {
        #         "json-id": "1e5fbaeb-0b41-45ed-9854-5ab6f0c86197",
        #         "id": "JGPpXvCYR+OBj1t4Ks+LPw",
        #         "name": "7m",
        #         "description": "7m",
        #         "softDelete": False,
        #         "creationDate": "2022-10-24T08:34:55.578+02:00",
        #         "updateDate": "2022-10-24T08:34:55.578+02:00",
        #         "dtype": "Range",
        #         "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #         "systemObject": False,
        #         "rangeInUnit": 7,
        #         "measureUnit": "METRIC",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.Range",
        #         "json-type": "com.doubleshot.model.Range"
        #       },
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "7f374302-4e68-49e2-8c73-8e499f54069e",
        #         "id": "xETFeaCkRXO7iVifj3R5uQ",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-02-09T15:16:36.745+01:00",
        #         "updateDate": "2023-02-09T15:16:38.645+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-02-09T09:45:53.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 121,
        #         "shootingExamOccurrence": {
        #           "json-id": "7b5ac2c4-9eac-442d-a32a-8e6876ae1bd9",
        #           "id": "2BlxeYFbTxSKslne9Pvhiw",
        #           "name": None,
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-02-09T15:16:36.664+01:00",
        #           "updateDate": "2023-02-09T15:16:38.645+01:00",
        #           "dtype": "ShootingExamOccurrence",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "score": 100,
        #           "shootingExam": {
        #             "json-id": "90b1bbab-66ab-456f-8236-1eecadb3a3b3",
        #             "id": "-G9Gu+fSTpGq1gdBpwBz-w",
        #             "name": "2.Pistol Level Test",
        #             "description": "10+10 rounds on the ABC target at 7m",
        #             "softDelete": False,
        #             "creationDate": "2022-12-28T11:15:53.103+01:00",
        #             "updateDate": "2023-02-05T19:49:18.479+01:00",
        #             "dtype": "ShootingExam",
        #             "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #             "systemObject": False,
        #             "minimumScore": 0,
        #             "manual": None,
        #             "noSQL": False,
        #             "javaType": "com.doubleshot.model.ShootingExam",
        #             "json-type": "com.doubleshot.model.ShootingExam"
        #           },
        #           "member": {
        #             "json-id": "b11ff8aa-329f-4a68-aad5-f6234381f211",
        #             "id": "NtWfGvsWRuaiTbuHNSscWw",
        #             "name": "23464",
        #             "description": None,
        #             "softDelete": False,
        #             "creationDate": "2022-12-28T16:09:40.439+01:00",
        #             "updateDate": "2022-12-28T16:09:48.300+01:00",
        #             "dtype": "Member",
        #             "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #             "systemObject": False,
        #             "email": "23464@syn.com",
        #             "homeDir": "/home/flexicore/users/23464NtWfGvsWRuaiTbuHNSscWw",
        #             "surName": "",
        #             "disabled": False,
        #             "dateApproved": None,
        #             "uiConfiguration": None,
        #             "lastVerificationDate": None,
        #             "totpEnabled": False,
        #             "phoneNumber": "",
        #             "externalId": None,
        #             "verificationDate": None,
        #             "contactByPhone": False,
        #             "lastUsedDeviceIdentification": None,
        #             "productId": None,
        #             "tokenId": None,
        #             "packageName": None,
        #             "platform": None,
        #             "appPurchase": False,
        #             "image": None,
        #             "commandingUnit": None,
        #             "noSQL": False,
        #             "javaType": "com.doubleshot.model.Member",
        #             "json-type": "com.doubleshot.model.Member"
        #           },
        #           "externalId": "58cfb0c5-57a4-43e7-8db6-16cb6306373c",
        #           "passed": True,
        #           "dateHeld": "2023-02-09T09:54:08.000+01:00",
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.ShootingExamOccurrence",
        #           "json-type": "com.doubleshot.model.ShootingExamOccurrence"
        #         },
        #         "externalId": "620f1913-b650-457c-9575-bf65b71d527d",
        #         "member": "b11ff8aa-329f-4a68-aad5-f6234381f211",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "cae0c90c-9a0f-4f19-b1ad-feef0c207ab2",
        #       "id": "07nSSjCrR+GUnT72Dv9TlA",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-01-25T15:02:48.424+01:00",
        #       "updateDate": "2023-01-25T15:02:48.424+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-01-25T10:47:19.000+01:00",
        #       "externalId": "4b3af335-a3e0-401a-acf4-0132a337f823",
        #       "score": 30,
        #       "heatType": None,
        #       "maxDistance": 23.662126912008567,
        #       "averageHitPointX": 176.7359381781684,
        #       "averageHitPointY": 244.02197096082898,
        #       "ordinal": 1,
        #       "numberOfHits": 18,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 10,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": {
        #         "json-id": "8cf8a8ce-6fb1-4c82-b70b-f7533c6da7bb",
        #         "id": "db-NXsNOQgmypvgdvfgpzA",
        #         "name": "White Demon",
        #         "description": "White Demon",
        #         "softDelete": False,
        #         "creationDate": "2021-02-21T11:06:45.165+01:00",
        #         "updateDate": "2022-12-11T11:01:53.913+01:00",
        #         "dtype": "Target",
        #         "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #         "systemObject": False,
        #         "targetImage": {
        #           "json-id": "81867cc5-7256-4fcb-99f1-3c55ab3c8771",
        #           "id": "NH17x3m3Tf+KilFCNz07DQ",
        #           "name": "%D7%93%D7%90%D7%91%D7%9C%20%D7%A9%D7%95%D7%98%2070_50.jpg",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-11T11:01:49.089+01:00",
        #           "updateDate": "2023-02-07T12:51:01.649+01:00",
        #           "md5": "009f49c55fd0d38f28dce8318b698ffd",
        #           "offset": 743838,
        #           "actualFilename": "69a5ea50-6b49-4d48-927a-93f9ee030f10.jpg",
        #           "originalFilename": "%D7%93%D7%90%D7%91%D7%9C%20%D7%A9%D7%95%D7%98%2070_50.jpg",
        #           "done": True,
        #           "dateTaken": None,
        #           "nonDownloadable": False,
        #           "keepUntil": None,
        #           "onlyFrom": None,
        #           "javaType": "com.wizzdi.flexicore.file.model.FileResource",
        #           "json-type": "com.wizzdi.flexicore.file.model.FileResource"
        #         },
        #         "noDefaultFiles": True,
        #         "ratioY": 297,
        #         "ratioX": 210,
        #         "settingResource": {
        #           "json-id": "f6472737-44b6-4f50-87c5-020b2f26a967",
        #           "id": "j5QWxBCNRSGcaeuVH2HgzA",
        #           "name": "whiteParams.txt",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-02-07T09:11:56.256+01:00",
        #           "updateDate": "2023-02-07T09:11:56.256+01:00",
        #           "md5": "992c094732b32a05d17d085df90d0d43",
        #           "offset": 834,
        #           "actualFilename": "742d84d7-1998-4049-aa83-d1d305fbec1d.txt",
        #           "originalFilename": "whiteParams.txt",
        #           "done": True,
        #           "dateTaken": None,
        #           "nonDownloadable": False,
        #           "keepUntil": None,
        #           "onlyFrom": None,
        #           "javaType": "com.wizzdi.flexicore.file.model.FileResource",
        #           "json-type": "com.wizzdi.flexicore.file.model.FileResource"
        #         },
        #         "usingClicks": False,
        #         "usingScore": True,
        #         "continueWhenZeroed": True,
        #         "finishWhenNotZeroed": True,
        #         "hitsRadius": 10,
        #         "svmWhite": "07541396-c6aa-497e-83de-a999e8af82e3",
        #         "svmBlack": "07541396-c6aa-497e-83de-a999e8af82e3",
        #         "mask": None,
        #         "scoringCalculator": "WhiteDemonScoringCalculator",
        #         "widthInCm": 50,
        #         "heightInCm": 70,
        #         "bullXCmFromLeft": 25,
        #         "bullYCmFromTop": 35,
        #         "availableForDemo": False,
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.Target",
        #         "json-type": "com.doubleshot.model.Target"
        #       },
        #       "range": {
        #         "json-id": "705c4218-ab35-42ec-8892-350ad9b6afdf",
        #         "id": "+k6ilTOeT6aBifvk53mqjw",
        #         "name": "100m",
        #         "description": "100m",
        #         "softDelete": False,
        #         "creationDate": "2019-10-26T12:45:57.656+02:00",
        #         "updateDate": "2019-10-26T12:45:57.658+02:00",
        #         "dtype": "Range",
        #         "tenant": {
        #           "json-id": "bdd6c397-96ef-4f5e-857c-6a22c83dca6b",
        #           "id": "eNmnX0ViSKaO3VHRE937wg",
        #           "name": "Testing",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2019-10-07T11:36:22.051+02:00",
        #           "updateDate": "2019-10-07T11:36:22.053+02:00",
        #           "dtype": "Tenant",
        #           "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #           "systemObject": False,
        #           "externalId": None,
        #           "icon": None,
        #           "noSQL": False,
        #           "javaType": "com.flexicore.model.Tenant",
        #           "json-type": "com.flexicore.model.Tenant"
        #         },
        #         "systemObject": False,
        #         "rangeInUnit": 100,
        #         "measureUnit": "METRIC",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.Range",
        #         "json-type": "com.doubleshot.model.Range"
        #       },
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "8e0099a4-99e5-4bcf-9c5d-aaf7ee0e0c35",
        #         "id": "bU6NFQluTO+C9ihINsScpA",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-01-25T15:02:48.422+01:00",
        #         "updateDate": "2023-01-25T15:02:48.424+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-01-25T10:33:40.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 63,
        #         "shootingExamOccurrence": None,
        #         "externalId": "a54ee435-4eee-42e0-917d-d09a4c135107",
        #         "member": {
        #           "json-id": "e48583d4-a9bc-4d37-bc8a-d378dfa694ea",
        #           "id": "B2dV-JX0QoWHNyDUk1K95w",
        #           "name": "23329",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-28T16:02:51.080+01:00",
        #           "updateDate": "2022-12-28T16:03:01.472+01:00",
        #           "dtype": "Member",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "email": "23329@syn.com",
        #           "homeDir": "/home/flexicore/users/23329B2dV-JX0QoWHNyDUk1K95w",
        #           "surName": "",
        #           "disabled": False,
        #           "dateApproved": None,
        #           "uiConfiguration": None,
        #           "lastVerificationDate": None,
        #           "totpEnabled": False,
        #           "phoneNumber": "",
        #           "externalId": None,
        #           "verificationDate": None,
        #           "contactByPhone": False,
        #           "lastUsedDeviceIdentification": None,
        #           "productId": None,
        #           "tokenId": None,
        #           "packageName": None,
        #           "platform": None,
        #           "appPurchase": False,
        #           "image": None,
        #           "commandingUnit": None,
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.Member",
        #           "json-type": "com.doubleshot.model.Member"
        #         },
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "4ea01ba4-b76e-4701-ae5e-db99ad63b88d",
        #       "id": "080ZSuzKShaMWmXYBLkjeg",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-01-23T22:28:26.633+01:00",
        #       "updateDate": "2023-01-23T22:28:26.633+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-01-23T18:07:10.000+01:00",
        #       "externalId": "9ef2ec3e-0418-48b9-adb1-095e0d014f0f",
        #       "score": None,
        #       "heatType": None,
        #       "maxDistance": 5.887080031632949,
        #       "averageHitPointX": 201.1845359802246,
        #       "averageHitPointY": 180.34675979614258,
        #       "ordinal": 1,
        #       "numberOfHits": 24,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 4,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "2812e906-4642-44a0-8439-1ee6f5303a31",
        #       "range": "c8c13311-22a0-4bc8-b212-3c5277849532",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "1be2da54-293d-48c5-a771-2d3d653d0f81",
        #         "id": "GkGzisd3RdaHw5BGPdO1Pg",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-01-23T22:28:26.632+01:00",
        #         "updateDate": "2023-01-23T22:28:26.633+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": True,
        #         "dateHeld": "2023-01-23T18:06:56.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 41,
        #         "shootingExamOccurrence": None,
        #         "externalId": "256250ae-8c70-41b0-9628-079b5d0aa516",
        #         "member": {
        #           "json-id": "84c2eae7-0f22-4ecb-9ca7-1fcec742e7da",
        #           "id": "XsE1BxKBRBq9fSrKgVVA1Q",
        #           "name": "23418",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-28T16:07:27.781+01:00",
        #           "updateDate": "2022-12-28T16:07:32.806+01:00",
        #           "dtype": "Member",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "email": "23418@syn.com",
        #           "homeDir": "/home/flexicore/users/23418XsE1BxKBRBq9fSrKgVVA1Q",
        #           "surName": "",
        #           "disabled": False,
        #           "dateApproved": None,
        #           "uiConfiguration": None,
        #           "lastVerificationDate": None,
        #           "totpEnabled": False,
        #           "phoneNumber": "",
        #           "externalId": None,
        #           "verificationDate": None,
        #           "contactByPhone": False,
        #           "lastUsedDeviceIdentification": None,
        #           "productId": None,
        #           "tokenId": None,
        #           "packageName": None,
        #           "platform": None,
        #           "appPurchase": False,
        #           "image": None,
        #           "commandingUnit": None,
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.Member",
        #           "json-type": "com.doubleshot.model.Member"
        #         },
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "604ed432-0605-4f69-ae3a-c34aea75dca5",
        #       "id": "09V9sYhsTvuiUqqWC-wdHw",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-02-09T14:49:02.196+01:00",
        #       "updateDate": "2023-02-09T14:49:02.196+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-02-09T12:31:40.000+01:00",
        #       "externalId": "21eb540f-f66d-48b7-9bf3-3583294cc5c9",
        #       "score": 100,
        #       "heatType": None,
        #       "maxDistance": 27.173631520280832,
        #       "averageHitPointX": 202.9047619047619,
        #       "averageHitPointY": 241,
        #       "ordinal": 1,
        #       "numberOfHits": 23,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 21,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "904e4057-a32c-458a-a18c-96dd760bd4b1",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "c82b4c9e-ee01-4e27-b6a6-6cdbabadd5e2",
        #       "range": "1e5fbaeb-0b41-45ed-9854-5ab6f0c86197",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "c2f22181-d6e2-4f96-8a69-ffae2f6afe05",
        #         "id": "D1SmmiaORiy1U10uk9hVEQ",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-02-09T14:49:02.196+01:00",
        #         "updateDate": "2023-02-09T14:49:02.196+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-02-09T12:22:16.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 21,
        #         "shootingExamOccurrence": {
        #           "json-id": "72893151-a9bf-4a8e-b963-466c15050399",
        #           "id": "uJgWLfReQRKnPOCDMBQa9w",
        #           "name": None,
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-02-09T14:49:02.137+01:00",
        #           "updateDate": "2023-02-09T14:49:02.196+01:00",
        #           "dtype": "ShootingExamOccurrence",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "score": 100,
        #           "shootingExam": "90b1bbab-66ab-456f-8236-1eecadb3a3b3",
        #           "member": {
        #             "json-id": "c5506a5d-d5f8-404d-9334-acd0e0981d49",
        #             "id": "kgQSdrw2QnKTP+4YVHRf7g",
        #             "name": "23255",
        #             "description": None,
        #             "softDelete": False,
        #             "creationDate": "2022-12-28T15:59:23.017+01:00",
        #             "updateDate": "2022-12-28T15:59:33.470+01:00",
        #             "dtype": "Member",
        #             "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #             "systemObject": False,
        #             "email": "23255@syn.com",
        #             "homeDir": "/home/flexicore/users/23255kgQSdrw2QnKTP+4YVHRf7g",
        #             "surName": "",
        #             "disabled": False,
        #             "dateApproved": None,
        #             "uiConfiguration": None,
        #             "lastVerificationDate": None,
        #             "totpEnabled": False,
        #             "phoneNumber": "",
        #             "externalId": None,
        #             "verificationDate": None,
        #             "contactByPhone": False,
        #             "lastUsedDeviceIdentification": None,
        #             "productId": None,
        #             "tokenId": None,
        #             "packageName": None,
        #             "platform": None,
        #             "appPurchase": False,
        #             "image": None,
        #             "commandingUnit": None,
        #             "noSQL": False,
        #             "javaType": "com.doubleshot.model.Member",
        #             "json-type": "com.doubleshot.model.Member"
        #           },
        #           "externalId": "49184c27-329b-4e8a-868c-72e312dd2b60",
        #           "passed": True,
        #           "dateHeld": "2023-02-09T12:32:01.000+01:00",
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.ShootingExamOccurrence",
        #           "json-type": "com.doubleshot.model.ShootingExamOccurrence"
        #         },
        #         "externalId": "b55a4774-de89-42f7-b7c8-e62429d5a990",
        #         "member": "c5506a5d-d5f8-404d-9334-acd0e0981d49",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "3a3909ba-b54a-4d56-a002-0ed0f0228413",
        #       "id": "0aofksc3T0moAv6S0IbdEA",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-01-25T15:13:22.597+01:00",
        #       "updateDate": "2023-01-25T15:13:22.598+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-01-25T12:32:48.000+01:00",
        #       "externalId": "880a7ce8-38cb-48f2-a8f5-c4ba9bb5c98f",
        #       "score": 30,
        #       "heatType": None,
        #       "maxDistance": 30.97225627463065,
        #       "averageHitPointX": 264.7490495954241,
        #       "averageHitPointY": 130.5759985787528,
        #       "ordinal": 1,
        #       "numberOfHits": 34,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 10,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "8cf8a8ce-6fb1-4c82-b70b-f7533c6da7bb",
        #       "range": "705c4218-ab35-42ec-8892-350ad9b6afdf",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "af1cf78d-94ac-4490-b7c9-86d7bd1c531f",
        #         "id": "AudRuwp9SD6e+nUQA3Fcrg",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-01-25T15:13:22.597+01:00",
        #         "updateDate": "2023-01-25T15:13:22.598+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-01-25T12:19:15.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 58,
        #         "shootingExamOccurrence": {
        #           "json-id": "d2ce2e9f-a702-4c5a-aa88-19a2ae9eb9f6",
        #           "id": "25zTTdj8RKCCYQvG-COZjg",
        #           "name": None,
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-01-25T15:13:22.434+01:00",
        #           "updateDate": "2023-01-25T15:13:22.598+01:00",
        #           "dtype": "ShootingExamOccurrence",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "score": 30,
        #           "shootingExam": {
        #             "json-id": "f72dabd8-41d7-4b37-aec3-bb613177cc1f",
        #             "id": "6lyu4oFVRPmHq6nef3KlHg",
        #             "name": "4.Accurate Lying Shooting",
        #             "description": "10 rounds on the White Demon target at 60m ",
        #             "softDelete": False,
        #             "creationDate": "2022-12-27T10:11:36.945+01:00",
        #             "updateDate": "2023-02-05T19:51:52.355+01:00",
        #             "dtype": "ShootingExam",
        #             "tenant": "999a4182-746c-44b4-8d15-19f69c70e536",
        #             "systemObject": False,
        #             "minimumScore": 0,
        #             "manual": None,
        #             "noSQL": False,
        #             "javaType": "com.doubleshot.model.ShootingExam",
        #             "json-type": "com.doubleshot.model.ShootingExam"
        #           },
        #           "member": {
        #             "json-id": "d60982c9-ac3b-49fe-8e4e-f8c728b88289",
        #             "id": "fXZNbSjuQSOvg6J2Lwcs4Q",
        #             "name": "23355",
        #             "description": None,
        #             "softDelete": False,
        #             "creationDate": "2022-12-28T16:04:59.545+01:00",
        #             "updateDate": "2022-12-28T16:05:07.126+01:00",
        #             "dtype": "Member",
        #             "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #             "systemObject": False,
        #             "email": "23355@syn.com",
        #             "homeDir": "/home/flexicore/users/23355fXZNbSjuQSOvg6J2Lwcs4Q",
        #             "surName": "",
        #             "disabled": False,
        #             "dateApproved": None,
        #             "uiConfiguration": None,
        #             "lastVerificationDate": None,
        #             "totpEnabled": False,
        #             "phoneNumber": "",
        #             "externalId": None,
        #             "verificationDate": None,
        #             "contactByPhone": False,
        #             "lastUsedDeviceIdentification": None,
        #             "productId": None,
        #             "tokenId": None,
        #             "packageName": None,
        #             "platform": None,
        #             "appPurchase": False,
        #             "image": None,
        #             "commandingUnit": None,
        #             "noSQL": False,
        #             "javaType": "com.doubleshot.model.Member",
        #             "json-type": "com.doubleshot.model.Member"
        #           },
        #           "externalId": "ceae5f8d-ee7d-4d41-a8bb-dc4eb4b6e290",
        #           "passed": True,
        #           "dateHeld": "2023-01-25T12:33:15.000+01:00",
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.ShootingExamOccurrence",
        #           "json-type": "com.doubleshot.model.ShootingExamOccurrence"
        #         },
        #         "externalId": "a8afcf0f-4170-42eb-8356-85d1013b8fb5",
        #         "member": "d60982c9-ac3b-49fe-8e4e-f8c728b88289",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "5ba6b361-b312-4f6c-b3bf-4a92d2eb4027",
        #       "id": "0BSPYfotTEqV8-n26BeODg",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-02-09T14:33:46.432+01:00",
        #       "updateDate": "2023-02-09T15:21:10.581+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-02-09T10:30:40.000+01:00",
        #       "externalId": "02f3e0ec-5d53-41aa-bc57-8f9913e961b9",
        #       "score": 65,
        #       "heatType": None,
        #       "maxDistance": 63.0090717674844,
        #       "averageHitPointX": 215.8387451171875,
        #       "averageHitPointY": 219.58570963541666,
        #       "ordinal": 1,
        #       "numberOfHits": 15,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 20,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "904e4057-a32c-458a-a18c-96dd760bd4b1",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "c82b4c9e-ee01-4e27-b6a6-6cdbabadd5e2",
        #       "range": "1e5fbaeb-0b41-45ed-9854-5ab6f0c86197",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "0bbac6a9-be77-41a9-9e5a-e9bcebabf37f",
        #         "id": "QflvOWNaQjW9jvlNGW-F3A",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-02-09T14:33:46.432+01:00",
        #         "updateDate": "2023-02-09T15:21:10.581+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-02-09T10:23:04.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 15,
        #         "shootingExamOccurrence": {
        #           "json-id": "ad38cb2d-2b78-4db6-84a7-54b1426265df",
        #           "id": "p5WYdNHHTdSzghJCW-IXtQ",
        #           "name": None,
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-02-09T14:33:46.371+01:00",
        #           "updateDate": "2023-02-09T15:21:10.581+01:00",
        #           "dtype": "ShootingExamOccurrence",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "score": 65,
        #           "shootingExam": "90b1bbab-66ab-456f-8236-1eecadb3a3b3",
        #           "member": {
        #             "json-id": "cf9b004d-e50a-49a5-ac40-87b9aabc70d7",
        #             "id": "OaOh4buARn+TV18naJn9wg",
        #             "name": "23116",
        #             "description": None,
        #             "softDelete": False,
        #             "creationDate": "2022-12-28T15:50:55.039+01:00",
        #             "updateDate": "2022-12-28T15:51:02.347+01:00",
        #             "dtype": "Member",
        #             "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #             "systemObject": False,
        #             "email": "23116@syn.com",
        #             "homeDir": "/home/flexicore/users/23116OaOh4buARn+TV18naJn9wg",
        #             "surName": "",
        #             "disabled": False,
        #             "dateApproved": None,
        #             "uiConfiguration": None,
        #             "lastVerificationDate": None,
        #             "totpEnabled": False,
        #             "phoneNumber": "",
        #             "externalId": None,
        #             "verificationDate": None,
        #             "contactByPhone": False,
        #             "lastUsedDeviceIdentification": None,
        #             "productId": None,
        #             "tokenId": None,
        #             "packageName": None,
        #             "platform": None,
        #             "appPurchase": False,
        #             "image": None,
        #             "commandingUnit": None,
        #             "noSQL": False,
        #             "javaType": "com.doubleshot.model.Member",
        #             "json-type": "com.doubleshot.model.Member"
        #           },
        #           "externalId": "00a680e0-0692-4eca-8d0b-f217d4b3f5dd",
        #           "passed": True,
        #           "dateHeld": "2023-02-09T10:31:13.000+01:00",
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.ShootingExamOccurrence",
        #           "json-type": "com.doubleshot.model.ShootingExamOccurrence"
        #         },
        #         "externalId": "475a5de0-d3fe-4947-8a4e-024bed3cc3ed",
        #         "member": "cf9b004d-e50a-49a5-ac40-87b9aabc70d7",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "37acc839-cb26-49f6-85bb-acef9c665e19",
        #       "id": "+0Cl9OrkQqCXPz72eR4FAg",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-01-17T18:31:34.455+01:00",
        #       "updateDate": "2023-01-21T09:58:38.884+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-01-17T12:25:30.000+01:00",
        #       "externalId": "f5f0c0aa-20ba-4184-ac5c-f9a95d694a1d",
        #       "score": None,
        #       "heatType": None,
        #       "maxDistance": 2.98970484389738,
        #       "averageHitPointX": 202.83618545532227,
        #       "averageHitPointY": 194.70619583129883,
        #       "ordinal": 5,
        #       "numberOfHits": 44,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 5,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "2812e906-4642-44a0-8439-1ee6f5303a31",
        #       "range": "c8c13311-22a0-4bc8-b212-3c5277849532",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "2925f405-ec96-4df8-90bf-9edfcc8a151e",
        #         "id": "EaCp7yUHRmK29EY3rbHbug",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-01-17T18:31:34.453+01:00",
        #         "updateDate": "2023-01-21T09:58:38.884+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": True,
        #         "dateHeld": "2023-01-17T10:28:34.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 106,
        #         "shootingExamOccurrence": None,
        #         "externalId": "baecbccf-8586-4f57-a6aa-221d75d9e87c",
        #         "member": {
        #           "json-id": "7b12b6bf-a361-4fdf-b0ad-99eac6648450",
        #           "id": "IB2ighxtQUGnEvG3Gc+UrA",
        #           "name": "23056",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-28T15:35:29.326+01:00",
        #           "updateDate": "2022-12-28T15:35:35.514+01:00",
        #           "dtype": "Member",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "email": "23056@syn.com",
        #           "homeDir": "/home/flexicore/users/23056IB2ighxtQUGnEvG3Gc+UrA",
        #           "surName": "",
        #           "disabled": False,
        #           "dateApproved": None,
        #           "uiConfiguration": None,
        #           "lastVerificationDate": None,
        #           "totpEnabled": False,
        #           "phoneNumber": "",
        #           "externalId": None,
        #           "verificationDate": None,
        #           "contactByPhone": False,
        #           "lastUsedDeviceIdentification": None,
        #           "productId": None,
        #           "tokenId": None,
        #           "packageName": None,
        #           "platform": None,
        #           "appPurchase": False,
        #           "image": None,
        #           "commandingUnit": None,
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.Member",
        #           "json-type": "com.doubleshot.model.Member"
        #         },
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "0e451fa2-6049-4435-855e-92356b874c4c",
        #       "id": "0DarGBbkSxS3nYw-wrsKKQ",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-01-23T22:52:34.562+01:00",
        #       "updateDate": "2023-01-23T22:52:34.563+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-01-23T16:26:37.000+01:00",
        #       "externalId": "86bdd570-1d93-4aa8-a5ea-50f7f3ceeb59",
        #       "score": None,
        #       "heatType": None,
        #       "maxDistance": 5.270343276296146,
        #       "averageHitPointX": 234.25,
        #       "averageHitPointY": 130.25,
        #       "ordinal": 1,
        #       "numberOfHits": 6,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 4,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "2812e906-4642-44a0-8439-1ee6f5303a31",
        #       "range": "c8c13311-22a0-4bc8-b212-3c5277849532",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "25455147-dddd-4650-b61c-a399a4830df6",
        #         "id": "5qsp4RBfQ4mTKyaHszPNuQ",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-01-23T22:52:34.562+01:00",
        #         "updateDate": "2023-01-23T22:52:34.563+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": True,
        #         "dateHeld": "2023-01-23T16:26:20.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 18,
        #         "shootingExamOccurrence": None,
        #         "externalId": "7a988092-fa95-4464-8bf0-9f30d7a0b96d",
        #         "member": {
        #           "json-id": "8f05a703-54c8-4175-bf0e-d760bb9a8966",
        #           "id": "nPLKN9aMTqeQ+cZs5BDUzw",
        #           "name": "23283",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-28T16:01:02.209+01:00",
        #           "updateDate": "2022-12-28T16:01:14.099+01:00",
        #           "dtype": "Member",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "email": "23283@syn.com",
        #           "homeDir": "/home/flexicore/users/23283nPLKN9aMTqeQ+cZs5BDUzw",
        #           "surName": "",
        #           "disabled": False,
        #           "dateApproved": None,
        #           "uiConfiguration": None,
        #           "lastVerificationDate": None,
        #           "totpEnabled": False,
        #           "phoneNumber": "",
        #           "externalId": None,
        #           "verificationDate": None,
        #           "contactByPhone": False,
        #           "lastUsedDeviceIdentification": None,
        #           "productId": None,
        #           "tokenId": None,
        #           "packageName": None,
        #           "platform": None,
        #           "appPurchase": False,
        #           "image": None,
        #           "commandingUnit": None,
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.Member",
        #           "json-type": "com.doubleshot.model.Member"
        #         },
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "ceac388a-2fd4-4227-96db-2fc0e878e171",
        #       "id": "0DKEVaFoRr6lPCp4AydTmA",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-01-23T22:42:42.477+01:00",
        #       "updateDate": "2023-01-23T22:42:42.478+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-01-23T15:12:10.000+01:00",
        #       "externalId": "d56d9921-01d0-4cf4-8aee-2124a981c79b",
        #       "score": None,
        #       "heatType": None,
        #       "maxDistance": 5.635421191002497,
        #       "averageHitPointX": 278.3333333333333,
        #       "averageHitPointY": 159.33333333333334,
        #       "ordinal": 1,
        #       "numberOfHits": 6,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 5,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "2812e906-4642-44a0-8439-1ee6f5303a31",
        #       "range": "c8c13311-22a0-4bc8-b212-3c5277849532",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "b27f5e90-2056-4b78-8c8b-7a96a9160f86",
        #         "id": "4p-y0EQlQ4q8F2sSqQDJNw",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-01-23T22:42:42.476+01:00",
        #         "updateDate": "2023-01-25T14:54:30.148+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-01-23T15:11:45.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 0,
        #         "shootingExamOccurrence": None,
        #         "externalId": "132b95e2-f6e6-4bb8-8b96-f4cbc908eddb",
        #         "member": {
        #           "json-id": "30037e7b-d7b0-4b4f-983e-04d5a220ec0e",
        #           "id": "ttckq457SeOlQPQbT0XgWw",
        #           "name": "﻿23312",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-28T16:02:51.388+01:00",
        #           "updateDate": "2022-12-28T16:03:01.472+01:00",
        #           "dtype": "Member",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "email": "23312@syn.com",
        #           "homeDir": "/home/flexicore/users/﻿23312ttckq457SeOlQPQbT0XgWw",
        #           "surName": "",
        #           "disabled": False,
        #           "dateApproved": None,
        #           "uiConfiguration": None,
        #           "lastVerificationDate": None,
        #           "totpEnabled": False,
        #           "phoneNumber": "",
        #           "externalId": None,
        #           "verificationDate": None,
        #           "contactByPhone": False,
        #           "lastUsedDeviceIdentification": None,
        #           "productId": None,
        #           "tokenId": None,
        #           "packageName": None,
        #           "platform": None,
        #           "appPurchase": False,
        #           "image": None,
        #           "commandingUnit": None,
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.Member",
        #           "json-type": "com.doubleshot.model.Member"
        #         },
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "c6ed8ff5-fc4d-45a4-9813-5e2630215ea6",
        #       "id": "0dO1xdCSSgmDfo-dRzzc0Q",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-05-11T19:08:18.245+02:00",
        #       "updateDate": "2023-05-11T19:08:18.246+02:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-05-11T18:15:42.000+02:00",
        #       "externalId": "86d06636-eb74-4471-b5d8-1d7a409679f5",
        #       "score": 75,
        #       "heatType": None,
        #       "maxDistance": 21.631371893617843,
        #       "averageHitPointX": 191.86666666666667,
        #       "averageHitPointY": 333.8,
        #       "ordinal": 1,
        #       "numberOfHits": 38,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 20,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "6fe0fd65-a56b-4223-a32d-1951cc152399",
        #       "target": "c82b4c9e-ee01-4e27-b6a6-6cdbabadd5e2",
        #       "range": "b0f018d6-ee65-4c47-be41-6cc3b216a4d0",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "06610973-d443-4c2d-8e00-3c9e89f21373",
        #         "id": "Etz0Lb-NR0C9OBFAgDK0KQ",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-05-11T19:08:18.244+02:00",
        #         "updateDate": "2023-05-11T19:08:18.246+02:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-05-11T18:04:29.000+02:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 62,
        #         "shootingExamOccurrence": {
        #           "json-id": "ed13d69f-8bdb-46c6-92fd-c5b79262512d",
        #           "id": "yT31lHg-RmyH+T3su1uxSA",
        #           "name": None,
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-05-11T19:08:18.056+02:00",
        #           "updateDate": "2023-05-11T19:08:18.246+02:00",
        #           "dtype": "ShootingExamOccurrence",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "score": 75,
        #           "shootingExam": "e8e772d0-db12-4ef6-ace4-fad1e077495e",
        #           "member": {
        #             "json-id": "3ea3710d-51a4-43ff-9138-5e2d3cc8c53c",
        #             "id": "ly4lkiuATLutp+FMZ+Lqwg",
        #             "name": "﻿23042",
        #             "description": None,
        #             "softDelete": False,
        #             "creationDate": "2022-12-28T15:35:28.310+01:00",
        #             "updateDate": "2022-12-28T15:35:35.514+01:00",
        #             "dtype": "Member",
        #             "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #             "systemObject": False,
        #             "email": "23042@syn.com",
        #             "homeDir": "/home/flexicore/users/﻿23042ly4lkiuATLutp+FMZ+Lqwg",
        #             "surName": "",
        #             "disabled": False,
        #             "dateApproved": None,
        #             "uiConfiguration": None,
        #             "lastVerificationDate": None,
        #             "totpEnabled": False,
        #             "phoneNumber": "",
        #             "externalId": None,
        #             "verificationDate": None,
        #             "contactByPhone": False,
        #             "lastUsedDeviceIdentification": None,
        #             "productId": None,
        #             "tokenId": None,
        #             "packageName": None,
        #             "platform": None,
        #             "appPurchase": False,
        #             "image": None,
        #             "commandingUnit": None,
        #             "noSQL": False,
        #             "javaType": "com.doubleshot.model.Member",
        #             "json-type": "com.doubleshot.model.Member"
        #           },
        #           "externalId": "cda2a56c-cb0b-4421-8ef8-d69e485187d9",
        #           "passed": True,
        #           "dateHeld": "2023-05-11T18:16:21.000+02:00",
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.ShootingExamOccurrence",
        #           "json-type": "com.doubleshot.model.ShootingExamOccurrence"
        #         },
        #         "externalId": "1e55066d-7276-4952-a976-96c847e0255b",
        #         "member": "3ea3710d-51a4-43ff-9138-5e2d3cc8c53c",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "56160a58-7933-4f3c-b293-a82e8714f295",
        #       "id": "0DP4whTdRIiTw2tF+ap8Lg",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-02-09T14:48:56.226+01:00",
        #       "updateDate": "2023-02-09T14:48:56.229+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-02-09T10:05:47.000+01:00",
        #       "externalId": "d4e1ec60-1fd4-4985-8e01-9d6d3ee19b3f",
        #       "score": 100,
        #       "heatType": None,
        #       "maxDistance": 58.066707726611334,
        #       "averageHitPointX": 114.79755424050724,
        #       "averageHitPointY": 224.5363567576689,
        #       "ordinal": 1,
        #       "numberOfHits": 35,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 34,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "904e4057-a32c-458a-a18c-96dd760bd4b1",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "c82b4c9e-ee01-4e27-b6a6-6cdbabadd5e2",
        #       "range": "1e5fbaeb-0b41-45ed-9854-5ab6f0c86197",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "5b237c35-f337-4d1e-ae07-7f74e786edb1",
        #         "id": "I-sHmwtNSlC9sI90P9ORuQ",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-02-09T14:48:56.226+01:00",
        #         "updateDate": "2023-02-09T14:48:56.229+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-02-09T09:55:53.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 51,
        #         "shootingExamOccurrence": {
        #           "json-id": "0400393e-ad94-4a9c-9701-b19f5980e786",
        #           "id": "dOKy4WxISRObS4m2pjPP0A",
        #           "name": None,
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-02-09T14:48:56.189+01:00",
        #           "updateDate": "2023-02-09T14:48:56.229+01:00",
        #           "dtype": "ShootingExamOccurrence",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "score": 100,
        #           "shootingExam": "90b1bbab-66ab-456f-8236-1eecadb3a3b3",
        #           "member": {
        #             "json-id": "2bd7a94a-d226-4194-b4b7-03e6b7853e5f",
        #             "id": "6H9ezTdQT1CMZkh0myutMw",
        #             "name": "23195",
        #             "description": None,
        #             "softDelete": False,
        #             "creationDate": "2022-12-28T15:57:23.327+01:00",
        #             "updateDate": "2022-12-28T15:57:28.147+01:00",
        #             "dtype": "Member",
        #             "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #             "systemObject": False,
        #             "email": "23195@syn.com",
        #             "homeDir": "/home/flexicore/users/231956H9ezTdQT1CMZkh0myutMw",
        #             "surName": "",
        #             "disabled": False,
        #             "dateApproved": None,
        #             "uiConfiguration": None,
        #             "lastVerificationDate": None,
        #             "totpEnabled": False,
        #             "phoneNumber": "",
        #             "externalId": None,
        #             "verificationDate": None,
        #             "contactByPhone": False,
        #             "lastUsedDeviceIdentification": None,
        #             "productId": None,
        #             "tokenId": None,
        #             "packageName": None,
        #             "platform": None,
        #             "appPurchase": False,
        #             "image": None,
        #             "commandingUnit": None,
        #             "noSQL": False,
        #             "javaType": "com.doubleshot.model.Member",
        #             "json-type": "com.doubleshot.model.Member"
        #           },
        #           "externalId": "e0129c71-17fd-4b5f-a739-57b86202fd6a",
        #           "passed": True,
        #           "dateHeld": "2023-02-09T10:06:57.000+01:00",
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.ShootingExamOccurrence",
        #           "json-type": "com.doubleshot.model.ShootingExamOccurrence"
        #         },
        #         "externalId": "d97e7bfb-c324-498a-b24e-da103a55e7e7",
        #         "member": "2bd7a94a-d226-4194-b4b7-03e6b7853e5f",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "417fe068-b52d-43b1-bb4e-6f5cced68db7",
        #       "id": "0DR7WlqmT1Slb273U2Jjpw",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-01-23T22:28:10.259+01:00",
        #       "updateDate": "2023-01-23T22:28:11.034+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-01-23T10:20:37.000+01:00",
        #       "externalId": "1176ed86-395a-42a1-8111-d0197c76874a",
        #       "score": None,
        #       "heatType": None,
        #       "maxDistance": 2.393982786626194,
        #       "averageHitPointX": 204.20059967041016,
        #       "averageHitPointY": 225.42171478271484,
        #       "ordinal": 1,
        #       "numberOfHits": 12,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 4,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "2812e906-4642-44a0-8439-1ee6f5303a31",
        #       "range": "c8c13311-22a0-4bc8-b212-3c5277849532",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "d267180f-1ea7-4560-bc83-64c5c10c5ce6",
        #         "id": "mBKOf3QtQbS8uAt2RQD6sA",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-01-23T22:28:10.259+01:00",
        #         "updateDate": "2023-01-23T22:28:11.034+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-01-23T10:20:09.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 18,
        #         "shootingExamOccurrence": None,
        #         "externalId": "fe838624-45c8-4f8e-88d1-7a0a95a6272b",
        #         "member": {
        #           "json-id": "39ca5e4f-eb38-4d85-87a0-1a0154683f21",
        #           "id": "EK4VTejJSPODS2Mmb-RCFQ",
        #           "name": "23518",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-28T16:11:00.201+01:00",
        #           "updateDate": "2022-12-28T16:11:01.762+01:00",
        #           "dtype": "Member",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "email": "23518@syn.com",
        #           "homeDir": "/home/flexicore/users/23518EK4VTejJSPODS2Mmb-RCFQ",
        #           "surName": "",
        #           "disabled": False,
        #           "dateApproved": None,
        #           "uiConfiguration": None,
        #           "lastVerificationDate": None,
        #           "totpEnabled": False,
        #           "phoneNumber": "",
        #           "externalId": None,
        #           "verificationDate": None,
        #           "contactByPhone": False,
        #           "lastUsedDeviceIdentification": None,
        #           "productId": None,
        #           "tokenId": None,
        #           "packageName": None,
        #           "platform": None,
        #           "appPurchase": False,
        #           "image": None,
        #           "commandingUnit": None,
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.Member",
        #           "json-type": "com.doubleshot.model.Member"
        #         },
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "e67e9e24-2b58-47d5-9e0e-db63eb3ec7f0",
        #       "id": "0DrC16P5QhyCaKtXdBY0mA",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-01-23T22:52:29.977+01:00",
        #       "updateDate": "2023-01-23T22:52:29.977+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-01-23T15:14:46.000+01:00",
        #       "externalId": "14099290-fd9b-4b23-8988-a908ac1ed077",
        #       "score": None,
        #       "heatType": None,
        #       "maxDistance": 6.843875065350787,
        #       "averageHitPointX": 295.0398864746094,
        #       "averageHitPointY": 170.2733123779297,
        #       "ordinal": 1,
        #       "numberOfHits": 14,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 5,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "2812e906-4642-44a0-8439-1ee6f5303a31",
        #       "range": "c8c13311-22a0-4bc8-b212-3c5277849532",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "d4d15933-e900-4d6a-babb-e993de052ed7",
        #         "id": "+VGKd04eQ9OoN3jpjmb0kA",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-01-23T22:52:29.976+01:00",
        #         "updateDate": "2023-01-25T14:45:42.835+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-01-23T15:14:29.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 0,
        #         "shootingExamOccurrence": None,
        #         "externalId": "c2d1ce13-4ad8-4e74-bf7d-f7f36e875a41",
        #         "member": {
        #           "json-id": "75d0abb3-88c9-4a87-80b4-d993da3d1416",
        #           "id": "JDAvgqbYTs++LxLn-JicVw",
        #           "name": "23334",
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2022-12-28T16:02:57.158+01:00",
        #           "updateDate": "2022-12-28T16:03:01.472+01:00",
        #           "dtype": "Member",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "email": "23334@syn.com",
        #           "homeDir": "/home/flexicore/users/23334JDAvgqbYTs++LxLn-JicVw",
        #           "surName": "",
        #           "disabled": False,
        #           "dateApproved": None,
        #           "uiConfiguration": None,
        #           "lastVerificationDate": None,
        #           "totpEnabled": False,
        #           "phoneNumber": "",
        #           "externalId": None,
        #           "verificationDate": None,
        #           "contactByPhone": False,
        #           "lastUsedDeviceIdentification": None,
        #           "productId": None,
        #           "tokenId": None,
        #           "packageName": None,
        #           "platform": None,
        #           "appPurchase": False,
        #           "image": None,
        #           "commandingUnit": None,
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.Member",
        #           "json-type": "com.doubleshot.model.Member"
        #         },
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "603ed10e-6838-4bfe-bd79-a11131c292e6",
        #       "id": "0ffZGZdCRKK0eCNtkGR5Tw",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-02-16T14:59:00.254+01:00",
        #       "updateDate": "2023-02-16T14:59:00.255+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-02-16T09:09:24.000+01:00",
        #       "externalId": "2c0c793e-6d7d-4eb4-a7d6-0fc15982e3d5",
        #       "score": 60,
        #       "heatType": None,
        #       "maxDistance": 58.659105218201205,
        #       "averageHitPointX": 200.79375584920248,
        #       "averageHitPointY": 230.64727656046549,
        #       "ordinal": 1,
        #       "numberOfHits": 12,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 20,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "c82b4c9e-ee01-4e27-b6a6-6cdbabadd5e2",
        #       "range": "b0f018d6-ee65-4c47-be41-6cc3b216a4d0",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "4f36216a-34cf-4f6b-8e41-1d63ea1ffa3a",
        #         "id": "jGh2WrHURVytZ2ZVsH2vyA",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-02-16T14:59:00.254+01:00",
        #         "updateDate": "2023-02-16T14:59:00.255+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-02-16T08:58:57.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 12,
        #         "shootingExamOccurrence": {
        #           "json-id": "8d7517e2-8442-486e-8684-42a40f29c70c",
        #           "id": "dO4DYpUPQJyYKjTASytAfQ",
        #           "name": None,
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-02-16T14:58:59.932+01:00",
        #           "updateDate": "2023-02-16T14:59:00.255+01:00",
        #           "dtype": "ShootingExamOccurrence",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "score": 60,
        #           "shootingExam": "e8e772d0-db12-4ef6-ace4-fad1e077495e",
        #           "member": {
        #             "json-id": "5263c144-2e1d-4627-b68e-26013584f5df",
        #             "id": "YEgl0uVATyK66vButcPKrg",
        #             "name": "23379",
        #             "description": None,
        #             "softDelete": False,
        #             "creationDate": "2022-12-28T16:05:05.676+01:00",
        #             "updateDate": "2022-12-28T16:05:07.126+01:00",
        #             "dtype": "Member",
        #             "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #             "systemObject": False,
        #             "email": "23379@syn.com",
        #             "homeDir": "/home/flexicore/users/23379YEgl0uVATyK66vButcPKrg",
        #             "surName": "",
        #             "disabled": False,
        #             "dateApproved": None,
        #             "uiConfiguration": None,
        #             "lastVerificationDate": None,
        #             "totpEnabled": False,
        #             "phoneNumber": "",
        #             "externalId": None,
        #             "verificationDate": None,
        #             "contactByPhone": False,
        #             "lastUsedDeviceIdentification": None,
        #             "productId": None,
        #             "tokenId": None,
        #             "packageName": None,
        #             "platform": None,
        #             "appPurchase": False,
        #             "image": None,
        #             "commandingUnit": None,
        #             "noSQL": False,
        #             "javaType": "com.doubleshot.model.Member",
        #             "json-type": "com.doubleshot.model.Member"
        #           },
        #           "externalId": "13f8c061-2766-4dc9-a3d3-f034597fbac9",
        #           "passed": True,
        #           "dateHeld": "2023-02-16T09:09:43.000+01:00",
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.ShootingExamOccurrence",
        #           "json-type": "com.doubleshot.model.ShootingExamOccurrence"
        #         },
        #         "externalId": "123d1a80-2cb2-4953-8a4c-fe5e95561c80",
        #         "member": "5263c144-2e1d-4627-b68e-26013584f5df",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     },
        #     {
        #       "json-id": "75095bc0-5d88-4df9-87c7-6f7e23df785f",
        #       "id": "0FiMNRE9RWugL66GiIKbvg",
        #       "name": None,
        #       "description": None,
        #       "softDelete": False,
        #       "creationDate": "2023-01-25T15:09:45.185+01:00",
        #       "updateDate": "2023-01-25T15:09:45.185+01:00",
        #       "dtype": "Heat",
        #       "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #       "systemObject": False,
        #       "dateHeld": "2023-01-25T11:54:46.000+01:00",
        #       "externalId": "a7409010-f3ef-4780-abd2-4b273a881836",
        #       "score": 30,
        #       "heatType": None,
        #       "maxDistance": 35.94955297400657,
        #       "averageHitPointX": 222.5462890625,
        #       "averageHitPointY": 151.7126678466797,
        #       "ordinal": 1,
        #       "numberOfHits": 35,
        #       "failed": False,
        #       "reverted": False,
        #       "shotsFired": 10,
        #       "hitsInScoringAreas": 0,
        #       "shootingExamOccurrence": None,
        #       "gunType": "e21e1fb2-38e8-410c-aa5e-96ec47b422ea",
        #       "sightType": "14ddef49-5300-47d6-94b2-b9f1abbfd232",
        #       "target": "8cf8a8ce-6fb1-4c82-b70b-f7533c6da7bb",
        #       "range": "705c4218-ab35-42ec-8892-350ad9b6afdf",
        #       "event": None,
        #       "shootingPractice": {
        #         "json-id": "c0b716e5-b4f5-4fb2-b88e-0fc3b0c18103",
        #         "id": "5xWupcuQQjeqn36OKkJCMQ",
        #         "name": None,
        #         "description": None,
        #         "softDelete": False,
        #         "creationDate": "2023-01-25T15:09:45.185+01:00",
        #         "updateDate": "2023-01-25T15:09:45.185+01:00",
        #         "dtype": "ShootingPractice",
        #         "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #         "systemObject": False,
        #         "locationName": "ultimate",
        #         "lat": 0,
        #         "lon": 0,
        #         "endedSuccessfully": False,
        #         "dateHeld": "2023-01-25T11:54:22.000+01:00",
        #         "training": False,
        #         "forReview": False,
        #         "reviewed": False,
        #         "errorReport": None,
        #         "validHits": 62,
        #         "shootingExamOccurrence": {
        #           "json-id": "5c34eaec-dc0c-4cbf-9acf-c95920bf9100",
        #           "id": "81LiE7tPTTqEQPmYia9PoA",
        #           "name": None,
        #           "description": None,
        #           "softDelete": False,
        #           "creationDate": "2023-01-25T15:09:45.142+01:00",
        #           "updateDate": "2023-01-25T15:09:45.185+01:00",
        #           "dtype": "ShootingExamOccurrence",
        #           "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #           "systemObject": False,
        #           "score": 30,
        #           "shootingExam": "f72dabd8-41d7-4b37-aec3-bb613177cc1f",
        #           "member": {
        #             "json-id": "bed6a5f6-0fd5-43d7-9f3f-79a1d7b4784c",
        #             "id": "uotroRCJRA2dQVR4TZQv0w",
        #             "name": "23476",
        #             "description": None,
        #             "softDelete": False,
        #             "creationDate": "2022-12-28T16:09:43.767+01:00",
        #             "updateDate": "2022-12-28T16:09:48.300+01:00",
        #             "dtype": "Member",
        #             "tenant": "e815e555-ae99-4f3c-a97a-b23e752b35b8",
        #             "systemObject": False,
        #             "email": "23476@syn.com",
        #             "homeDir": "/home/flexicore/users/23476uotroRCJRA2dQVR4TZQv0w",
        #             "surName": "",
        #             "disabled": False,
        #             "dateApproved": None,
        #             "uiConfiguration": None,
        #             "lastVerificationDate": None,
        #             "totpEnabled": False,
        #             "phoneNumber": "",
        #             "externalId": None,
        #             "verificationDate": None,
        #             "contactByPhone": False,
        #             "lastUsedDeviceIdentification": None,
        #             "productId": None,
        #             "tokenId": None,
        #             "packageName": None,
        #             "platform": None,
        #             "appPurchase": False,
        #             "image": None,
        #             "commandingUnit": None,
        #             "noSQL": False,
        #             "javaType": "com.doubleshot.model.Member",
        #             "json-type": "com.doubleshot.model.Member"
        #           },
        #           "externalId": "b36d9573-4daf-4a37-a447-9189ed00c9b3",
        #           "passed": True,
        #           "dateHeld": "2023-01-25T11:55:50.000+01:00",
        #           "noSQL": False,
        #           "javaType": "com.doubleshot.model.ShootingExamOccurrence",
        #           "json-type": "com.doubleshot.model.ShootingExamOccurrence"
        #         },
        #         "externalId": "09cf45fa-fc9b-4d42-a47f-23ffbd31c69a",
        #         "member": "bed6a5f6-0fd5-43d7-9f3f-79a1d7b4784c",
        #         "noSQL": False,
        #         "javaType": "com.doubleshot.model.ShootingPractice",
        #         "json-type": "com.doubleshot.model.ShootingPractice"
        #       },
        #       "noSQL": False,
        #       "javaType": "com.doubleshot.model.Heat",
        #       "json-type": "com.doubleshot.model.Heat"
        #     }
        #   ],
        #     "totalRecords": 3130,
        #     "totalPages": 157,
        #     "startPage": 0,
        #     "endPage": 156
        # }

        # print(list_)

        # for heat in list_["list"]:
        #     try:
        #         # print("heat\n","="*10,"\n",heat,"\n","="*10,"\n")
        #         try:
        #             maxDistance = 0
        #             maxDistance = round(heat["maxDistance"]*100)/100
        #         except Exception as ex:
        #             pass
        #         dateHeld = ''.join(heat["dateHeld"][:10].split("-"))
        #         try:
        #             averageHitPointX = 0
        #             averageHitPointX = round(heat["averageHitPointX"]*100)/100
        #         except Exception as ex:
        #             pass
        #         try:
        #             averageHitPointY = 0
        #             averageHitPointY = round(heat["averageHitPointY"]*100)/100
        #         except Exception as ex:
        #             pass
        #
        #         numberOfHits = heat["numberOfHits"]
        #         shotsFired = heat["shotsFired"]
        #         try:
        #             gunType = heat["gunType"]["name"]
        #             gun_type[heat["gunType"]["json-id"]] = gunType
        #         except Exception as ex:
        #             # print(gun_type, "\n", heat["gunType"])
        #             gunType = gun_type[heat["gunType"]]
        #         try:
        #             sightType = heat["sightType"]["name"]
        #             sight_types[heat["sightType"]["json-id"]] = sightType
        #         except Exception as ex:
        #             sightType = sight_types[heat["sightType"]]
        #         try:
        #             target = heat["target"]["name"]
        #             targets[heat["target"]["json-id"]] = target
        #         except Exception as ex:
        #             target = targets[heat["target"]]
        #         # ??? do we need more info from target ?
        #         #         "widthInCm": 18.8,
        #         #         "heightInCm": 27.8,
        #         #         "bullXCmFromLeft": 9.4,
        #         #         "bullYCmFromTop": 9.4,
        #         try:
        #             range_ = heat["range"]["name"]
        #             ranges[heat["range"]["json-id"]] = range_
        #         except Exception as ex:
        #             range_ = ranges[heat["range"]]
        #         sp = heat["shootingPractice"]
        #         try:
        #             member = sp["member"]
        #             # print("memeber\n","-"*10,"\n",member,"\n","-"*10,"\n")
        #             name_ = member["name"]
        #         except Exception as ex:
        #             # print("Error 1:\n","-"*10, "\n: "+str(ex), "\n","-"*10,"\n")
        #             # print(sp)
        #             name_ = member
        #         try:
        #             if str(name_).isnumeric():
        #                 n__ += 1
        #                 print(n__, name_, gunType, range_, dateHeld, maxDistance, averageHitPointX, averageHitPointY,
        #                       numberOfHits, shotsFired, sightType, target)
        #                 if str(name_) not in data:
        #                     data[str(name_)] = {}
        #                 if gunType not in data[str(name_)]:
        #                     data[str(name_)][gunType] = {}
        #                 if range_ not in data[str(name_)][gunType]:
        #                     data[str(name_)][gunType][range_] = {}
        #                 if dateHeld not in data[str(name_)][gunType][range_]:
        #                     data[str(name_)][gunType][range_][dateHeld] = {}
        #                 data[str(name_)][gunType][dateHeld] = [maxDistance, averageHitPointX, averageHitPointY, numberOfHits,
        #                                                        shotsFired, sightType, target]
        #
        #         except Exception as ex:
        #             print("Error 11:\n", name_, str(ex))
        #         # print(1007)
        #
        #     except Exception as ex:
        #         print("Error 2:\n","-"*10, "\n: "+str(ex), "\n","-"*10,"\n")

        print(data_)
        result = self.update_members_data(dic, data_)
        # result = {"status": "ok", "data": data_}
        # print(result)
        return result

    def update_members_data(self, dic, data_=None):
        print("update_members_data\n", "-"*50, "\n", dic, "\n", "-"*50)
        app_ = dic["app"]
        # data_ = {'fields':['maxDistance', 'averageHitPointX', 'averageHitPointY', 'numberOfHits', 'shotsFired', 'sightType', 'target'],
        #      'data':{'23155': {
        #          'MZ-4P': {'25m': {'20230118': [11.38, 194.09, 177.76, 34, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23343': {
        #          'MZ-4P': {'25m': {'20230123': [4.12, 190.24, 176.57, 20, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                            '20230511': [5.77, 153.87, 167.15, 13, 5, 'Aimpoint COMPM5/M5S',
        #                                         'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [38.12, 244.1, 236.5, 26, 10, 'Iron Sight', 'White Demon']}}},
        #      '23053': {
        #          'MZ-4P': {'25m': {'20230117': [6.68, 145.82, 228.51, 10, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                            '20230511': [2.51, 352.6, 170.2, 7, 5, 'Aimpoint COMPM5/M5S',
        #                                         'DS Zeroing A4 Target']},
        #                    '30m': {'20230216': [50.16, 190.0, 237.0, 12, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23329': {'MZ-4P': {'100m': {'20230125': [23.66, 176.74, 244.02, 18, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [4.08, 201.25, 177.25, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '30m': {'20230216': [36.42, 165.59, 173.88, 17, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23418': {
        #          'MZ-4P': {'25m': {'20230123': [5.89, 201.18, 180.35, 24, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                            '20230511': [2.32, 69.0, 30.0, 8, 4, 'Aimpoint COMPM5/M5S',
        #                                         'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [63.18, 199.38, 204.97, 47, 49, 'Iron Sight', 'ABC - K']}}},
        #      '23056': {
        #          'MZ-4P': {'25m': {'20230117': [8.61, 166.34, 198.76, 5, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23283': {'MZ-4P': {
        #          '25m': {'20230123': [3.71, 193.09, 182.46, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                  '20230202': [46.3, 200.19, 274.12, 53, 20, 'Iron Sight', 'ABC '],
        #                  '20230119': [4.72, 190.09, 184.54, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23518': {
        #         'MZ-4P': {'25m': {'20230123': [12.13, 217.67, 237.67, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23334': {
        #          'MZ-4P': {'25m': {'20230123': [3.81, 195.0, 206.44, 21, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [33.14, 142.3, 228.9, 20, 10, 'Iron Sight', 'White Demon']}}},
        #      '23105': {
        #          'MZ-4P': {'25m': {'20230118': [7.47, 218.36, 193.21, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23532': {
        #          'MZ-4P': {'25m': {'20230123': [2.14, 200.4, 174.38, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [37.42, 145.2, 235.1, 20, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23098': {'MZ-4P': {'100m': {'20230125': [41.52, 208.42, 166.59, 29, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {'20230118': [8.54, 210.4, 213.9, 13, 5, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230117': [8.74, 192.9, 208.57, 14, 4, 'Iron Sight',
        #                                               'DS Zeroing A4 Target']}}}, '23369': {'MZ-4P': {
        #         '25m': {'20230123': [5.09, 190.33, 183.92, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                 '20230511': [12.83, 221.5, 52.62, 8, 8, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target']}}},
        #      '23193': {'MZ-4P': {'25m': {'20230123': [0, 195.0, 100.0, 7, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23166': {
        #          'MZ-4P': {'25m': {'20230123': [10.48, 239.75, 289.25, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23396': {
        #          'MZ-4P': {'25m': {'20230123': [6.06, 189.75, 231.5, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23461': {
        #          'MZ-4P': {'25m': {'20230123': [0.93, 196.02, 202.19, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23200': {
        #          'MZ-4P': {'25m': {'20230123': [12.32, 190.44, 195.5, 8, 6, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23494': {
        #          'MZ-4P': {'25m': {'20230123': [2.82, 201.78, 211.93, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23447': {
        #          'MZ-4P': {'25m': {'20230123': [9.47, 282.0, 177.5, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23151': {
        #          'MZ-4P': {'25m': {'20230118': [5.73, 196.09, 202.58, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [21.86, 138.6, 180.6, 42, 10, 'Iron Sight', 'White Demon']}}},
        #      '23173': {
        #          'MZ-4P': {'25m': {'20230123': [4.53, 331.67, 178.33, 19, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23169': {
        #          'MZ-4P': {'25m': {'20230123': [4.32, 175.01, 144.11, 22, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23355': {
        #          'MZ-4P': {'25m': {'20230123': [31.24, 211.67, 192.33, 4, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23220': {'MZ-4P': {'25m': {'20230123': [6.04, 233.0, 60.2, 6, 5, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #                '9mm/0.40/0.45': {'7m': {'20230209': [26.06, 205.67, 248.55, 25, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23466': {
        #          'MZ-4P': {'25m': {'20230123': [18.6, 195.97, 149.68, 11, 10, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23435': {
        #          'MZ-4P': {'25m': {'20230123': [6.67, 187.87, 183.24, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23152': {
        #          'MZ-4P': {'25m': {'20230118': [10.65, 199.12, 208.31, 33, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23108': {
        #          'MZ-4P': {'25m': {'20230118': [5.52, 208.59, 344.66, 10, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23448': {
        #          'MZ-4P': {'25m': {'20230123': [13.17, 160.64, 87.61, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23045': {
        #          'MZ-4P': {'25m': {'20230117': [11.8, 192.56, 201.96, 24, 5, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [56.2, 152.33, 249.39, 18, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23086': {
        #          'MZ-4P': {'30m': {'20230511': [29.67, 209.1, 291.15, 46, 20, 'Aimpoint COMPM5/M5S', 'ABC - K']}}},
        #      '23174': {'MZ-4P': {'25m': {'20230123': [4.38, 226.67, 114.33, 4, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [4.26, 258.02, 190.39, 11, 5, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']},
        #                          '30m': {'20230511': [31.3, 158.0, 310.42, 19, 20, 'Aimpoint COMPM5/M5S', 'ABC - K']}}},
        #      '23314': {
        #          'MZ-4P': {'25m': {'20230123': [8.0, 219.25, 180.0, 9, 8, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23165': {'MZ-4P': {
        #          '25m': {'20230123': [5.48, 177.07, 120.87, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                  '20230511': [10.02, 231.48, 147.35, 10, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target']}}},
        #      '23406': {'MZ-4P': {'25m': {'20230123': [17.2, 318.4, 202.6, 6, 5, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [0, 0.0, 0.0, 0, 4, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']}}}, '23017': {
        #         'MZ-4P': {'30m': {'20230216': [30.81, 221.96, 282.33, 28, 20, 'Iron Sight', 'ABC - K']}, '25m': {
        #             '20230511': [4.75, 299.33, 218.33, 4, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #             '20230116': [14.06, 179.0, 250.8, 6, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23516': {'9mm/0.40/0.45': {'7m': {'20230209': [53.38, 176.5, 206.0, 20, 20, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {
        #                    '25m': {'20230123': [1.6, 212.62, 230.51, 7, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23408': {'MZ-4P': {'25m': {'20230123': [0, 45.0, 21.0, 4, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23508': {
        #          'MZ-4P': {'25m': {'20230123': [2.96, 195.34, 152.42, 29, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23123': {
        #          'MZ-4P': {'25m': {'20230118': [26.42, 146.78, 245.37, 12, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23260': {
        #          'MZ-4P': {'25m': {'20230123': [14.09, 178.5, 207.5, 5, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23313': {
        #          'MZ-4P': {'25m': {'20230123': [5.19, 293.33, 242.67, 5, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23315': {'MZ-4P': {'100m': {'20230125': [33.17, 205.29, 151.29, 7, 10, 'Iron Sight', 'White Demon']}}},
        #      '23006': {'MZ-4P': {
        #          '25m': {'20230116': [5.89, 179.12, 199.85, 9, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230511': [8.16, 280.41, 148.51, 5, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target']}}},
        #      '23128': {'MZ-4P': {'100m': {'20230125': [19.66, 189.2, 139.1, 32, 10, 'Iron Sight', 'White Demon']}}},
        #      '23332': {
        #          'MZ-4P': {'25m': {'20230123': [12.25, 301.41, 204.33, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230216': [44.78, 189.39, 258.0, 30, 20, 'Iron Sight', 'ABC - K']},
        #                    '100m': {'20230125': [42.66, 269.78, 252.33, 18, 10, 'Iron Sight', 'White Demon']}}},
        #      '23473': {
        #          'MZ-4P': {'25m': {'20230123': [6.2, 189.21, 214.36, 17, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23445': {'MZ-4P': {
        #          '25m': {'20230123': [12.75, 237.92, 221.5, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                  '20230511': [5.73, 368.46, 211.05, 16, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target']}}},
        #      '23232': {
        #          'MZ-4P': {'25m': {'20230123': [4.13, 239.5, 194.0, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23208': {'MZ-4P': {'25m': {'20230123': [7.6, 127.25, 148.0, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #                '9mm/0.40/0.45': {'7m': {'20230209': [43.88, 149.88, 281.86, 20, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23330': {
        #          'MZ-4P': {'25m': {'20230123': [9.15, 207.12, 203.55, 22, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230216': [48.89, 162.35, 226.6, 23, 20, 'Iron Sight', 'ABC - K']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [34.12, 149.55, 297.68, 26, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23392': {
        #          'MZ-4P': {'25m': {'20230123': [7.09, 218.75, 184.0, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23388': {
        #          'MZ-4P': {'25m': {'20230123': [8.69, 283.5, 95.0, 10, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23281': {'MZ-4P': {
        #          '25m': {'20230119': [7.36, 186.63, 196.48, 21, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                  '20230202': [50.21, 202.06, 244.18, 28, 20, 'Iron Sight', 'ABC ']}}}, '23380': {
        #         'MZ-4P': {'25m': {'20230123': [5.15, 199.23, 182.6, 14, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23449': {
        #          'MZ-4P': {'25m': {'20230123': [14.68, 196.21, 270.72, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23295': {
        #          'MZ-4P': {'25m': {'20230119': [4.75, 267.33, 171.0, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [10.72, 221.46, 213.14, 37, 10, 'Iron Sight', 'White Demon']},
        #                    '30m': {'20230202': [43.77, 215.93, 310.64, 16, 20, 'Iron Sight', 'ABC '],
        #                            '20230216': [39.08, 228.11, 273.11, 18, 20, 'Iron Sight', 'ABC - K']}}}, '23290': {
        #         'MZ-4P': {'30m': {'20230202': [34.17, 130.0, 331.33, 6, 20, 'Iron Sight', 'ABC ']},
        #                   '25m': {'20230119': [4.63, 84.33, 224.0, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23033': {
        #          'MZ-4P': {'25m': {'20230116': [12.78, 293.25, 248.0, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23236': {
        #          'MZ-4P': {'25m': {'20230123': [9.46, 283.25, 172.0, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23082': {'MZ-4P': {
        #          '25m': {'20230511': [4.22, 201.95, 193.49, 27, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230117': [8.08, 133.06, 327.51, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #          '100m': {'20230125': [8.52, 200.73, 201.59, 13, 10, 'Iron Sight', 'White Demon']}}}, '23237': {
        #         'MZ-4P': {'25m': {'20230123': [16.02, 153.4, 170.0, 23, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23256': {
        #          'MZ-4P': {'25m': {'20230123': [10.78, 209.0, 231.4, 17, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23286': {'MZ-4P': {'25m': {'20230119': [4.62, 193.53, 202.64, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230202': [37.96, 200.33, 281.83, 34, 20, 'Iron Sight', 'ABC ']}}},
        #      '23060': {
        #          'MZ-4P': {'25m': {'20230117': [11.86, 222.74, 194.31, 22, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23310': {
        #          'MZ-4P': {'25m': {'20230119': [7.46, 178.38, 204.88, 27, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230202': [49.19, 210.71, 258.24, 17, 20, 'Iron Sight', 'ABC ']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [22.68, 212.5, 188.67, 53, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23190': {
        #          'MZ-4P': {'25m': {'20230123': [22.96, 127.83, 231.14, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [47.07, 153.3, 249.55, 28, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23333': {'MZ-4P': {'100m': {'20230125': [21.25, 146.71, 178.8, 21, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {'20230511': [4.56, 199.42, 209.0, 15, 5, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target'],
        #                                  '20230123': [7.48, 170.5, 305.5, 24, 4, 'Iron Sight',
        #                                               'DS Zeroing A4 Target']}}}, '23325': {
        #         'MZ-4P': {'25m': {'20230123': [11.68, 218.25, 258.75, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23342': {'MZ-4P': {'100m': {'20230125': [25.09, 254.5, 250.2, 22, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [5.46, 203.34, 196.03, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23026': {
        #          'MZ-4P': {'25m': {'20230116': [1.75, 223.63, 195.3, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23528': {
        #          'MZ-4P': {'25m': {'20230123': [5.14, 198.76, 258.63, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [48.81, 201.5, 218.95, 17, 51, 'Iron Sight', 'ABC - K']}}},
        #      '23509': {'9mm/0.40/0.45': {'7m': {'20230209': [51.82, 198.27, 239.01, 24, 42, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {
        #                    '25m': {'20230123': [20.6, 227.02, 157.26, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {
        #                        '20230125': [11.71, 180.44, 207.68, 25, 10, 'Aimpoint COMPM5/M5S', 'White Demon']}}},
        #      '23514': {
        #          'MZ-4P': {'25m': {'20230123': [9.7, 180.14, 80.46, 19, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23372': {
        #          'MZ-4P': {'25m': {'20230123': [6.69, 196.55, 206.18, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23095': {'MZ-4P': {'100m': {'20230125': [40.99, 215.87, 273.97, 18, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230117': [6.06, 212.19, 226.64, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23234': {
        #          'MZ-4P': {'25m': {'20230123': [5.09, 277.25, 176.0, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23129': {
        #          'MZ-4P': {'25m': {'20230118': [9.75, 203.42, 203.54, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [18.87, 192.78, 221.56, 26, 10, 'Iron Sight', 'White Demon']}}},
        #      '23013': {
        #          'MZ-4P': {'25m': {'20230116': [14.47, 205.23, 207.61, 15, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23218': {
        #          'MZ-4P': {'25m': {'20230123': [17.33, 278.33, 230.67, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23067': {'9mm/0.40/0.45': {'7m': {'20230209': [44.88, 175.85, 252.85, 41, 34, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {
        #                    '25m': {'20230117': [8.92, 75.74, 115.02, 19, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23117': {'MZ-4P': {
        #          '25m': {'20230511': [7.66, 240.4, 182.8, 5, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230118': [15.69, 210.8, 199.27, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23254': {
        #         'MZ-4P': {'25m': {'20230123': [3.97, 196.87, 202.5, 12, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23002': {'MZ-4P': {'25m': {'20230117': [12.42, 175.0, 135.14, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230118': [7.54, 242.42, 220.36, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230115': [3.04, 201.0, 201.25, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230122': [2.07, 108.5, 281.5, 4, 5, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']}}}, '23399': {
        #         'MZ-4P': {'25m': {'20230123': [7.3, 82.47, 310.09, 22, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23240': {'MZ-4P': {
        #          '25m': {'20230511': [5.3, 201.82, 193.85, 20, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230123': [11.49, 144.67, 42.33, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23089': {
        #         'MZ-4P': {
        #             '25m': {'20230511': [13.04, 201.65, 200.78, 32, 6, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                     '20230117': [3.77, 206.02, 196.46, 13, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #             '100m': {'20230125': [65.06, 161.95, 234.05, 20, 10, 'Iron Sight', 'White Demon']}}}, '23362': {
        #         'MZ-4P': {'25m': {'20230123': [4.4, 192.17, 199.18, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23458': {
        #          'MZ-4P': {'25m': {'20230123': [16.46, 271.3, 105.18, 6, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23510': {'MZ-4P': {'30m': {'20230216': [36.56, 216.64, 253.0, 24, 20, 'Iron Sight', 'ABC - K']}, '25m': {
        #          '20230123': [2.08, 204.67, 126.67, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #                '9mm/0.40/0.45': {'7m': {'20230209': [40.74, 89.62, 221.31, 16, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23273': {'MZ-4P': {'30m': {'20230202': [33.82, 199.29, 319.88, 17, 20, 'Iron Sight', 'ABC ']}, '25m': {
        #          '20230511': [1.66, 198.09, 192.26, 15, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #          '20230119': [2.23, 205.87, 200.64, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23104': {
        #         'MZ-4P': {'25m': {'20230118': [14.5, 196.75, 214.25, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23107': {
        #          'MZ-4P': {'25m': {'20230118': [11.47, 205.36, 205.88, 14, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23214': {
        #          'MZ-4P': {'25m': {'20230123': [9.11, 204.26, 213.9, 28, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23450': {'MZ-4P': {'25m': {'20230123': [0, 0.0, 0.0, 0, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '100m': {'20230125': [18.1, 206.67, 269.57, 25, 10, 'Iron Sight', 'White Demon']},
        #                          '30m': {'20230216': [34.9, 224.1, 256.4, 20, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23527': {
        #          'MZ-4P': {'25m': {'20230123': [26.27, 169.0, 234.4, 10, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23488': {'9mm/0.40/0.45': {'7m': {'20230209': [46.13, 130.28, 264.74, 39, 20, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {'100m': {'20230125': [0, 239.34, 244.11, 25, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [9.64, 237.13, 234.87, 15, 8, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23215': {'MZ-4P': {'30m': {'20230216': [34.32, 156.07, 251.0, 40, 20, 'Iron Sight', 'ABC - K']},
        #                          '100m': {'20230125': [15.49, 189.44, 209.78, 28, 10, 'Iron Sight', 'White Demon']}},
        #                '9mm/0.40/0.45': {'7m': {'20230209': [39.4, 215.15, 227.85, 20, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23350': {'MZ-4P': {'25m': {'20230123': [7.02, 210.5, 72.75, 18, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [8.95, 166.0, 239.0, 8, 5, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']},
        #                          '30m': {'20230511': [67.77, 229.38, 297.19, 16, 20, 'Aimpoint COMPM5/M5S', 'ABC - K']},
        #                          '100m': {'20230125': [11.2, 185.58, 208.54, 34, 10, 'Iron Sight', 'White Demon']}}},
        #      '23440': {
        #          'MZ-4P': {'25m': {'20230123': [9.5, 98.67, 287.67, 4, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23115': {
        #          'MZ-4P': {'25m': {'20230118': [5.6, 200.32, 207.18, 39, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [32.22, 250.82, 117.1, 18, 10, 'Iron Sight', 'White Demon']}}},
        #      '23305': {
        #          'MZ-4P': {'25m': {'20230119': [5.32, 127.78, 328.7, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230202': [50.74, 164.77, 257.92, 13, 20, 'Iron Sight', 'ABC ']}}}, '23120': {
        #         'MZ-4P': {'25m': {'20230118': [12.92, 210.87, 152.76, 24, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23417': {
        #          'MZ-4P': {'25m': {'20230123': [8.55, 262.67, 258.67, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23278': {'MZ-4P': {
        #          '25m': {'20230511': [4.18, 199.0, 176.02, 18, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230123': [6.2, 197.06, 189.61, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #          '30m': {'20230202': [33.59, 163.43, 284.33, 21, 20, 'Iron Sight', 'ABC ']}}}, '23031': {
        #         'MZ-4P': {'25m': {'20230116': [7.37, 163.33, 199.33, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                   '30m': {'20230216': [37.42, 286.25, 264.5, 48, 20, 'Iron Sight', 'ABC - K']}}}, '23496': {
        #         'MZ-4P': {'100m': {'20230125': [8.4, 193.52, 167.48, 37, 10, 'Iron Sight', 'White Demon']},
        #                   '25m': {'20230123': [13.42, 188.31, 203.88, 19, 8, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23087': {'MZ-4P': {'100m': {'20230125': [35.75, 178.53, 279.8, 22, 13, 'Iron Sight', 'White Demon']},
        #                          '25m': {'20230118': [10.09, 182.26, 173.53, 14, 4, 'Iron Sight',
        #                                               'DS Zeroing A4 Target']}}}, '23213': {'MZ-4P': {
        #         '25m': {'20230511': [4.6, 202.71, 192.95, 22, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                 '20230123': [7.67, 149.32, 297.33, 24, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #         '30m': {'20230511': [17.38, 216.28, 318.45, 20, 20, 'Aimpoint COMPM5/M5S', 'ABC - K']}}}, '23511': {
        #         'MZ-4P': {'25m': {'20230123': [3.81, 203.79, 215.3, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23061': {
        #          'MZ-4P': {'25m': {'20230117': [3.75, 210.0, 156.0, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23131': {
        #          'MZ-4P': {'25m': {'20230118': [6.97, 229.62, 100.98, 10, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [54.62, 158.37, 260.63, 21, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23287': {'MZ-4P': {'30m': {'20230202': [60.4, 206.21, 269.0, 14, 20, 'Iron Sight', 'ABC ']}, '25m': {
        #          '20230119': [16.48, 190.02, 156.37, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #          '20230202': [41.18, 157.06, 281.56, 34, 20, 'Iron Sight', 'ABC '],
        #          '20230511': [3.38, 175.97, 178.26, 10, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target']}}}, '23298': {
        #         'MZ-4P': {'25m': {'20230123': [12.44, 139.33, 59.33, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                           '20230202': [34.5, 263.67, 293.83, 19, 20, 'Iron Sight', 'ABC ']}}}, '23498': {
        #         'MZ-4P': {'100m': {'20230125': [22.47, 219.75, 275.03, 38, 10, 'Iron Sight', 'White Demon']},
        #                   '25m': {'20230123': [14.05, 208.25, 235.25, 21, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23336': {
        #          'MZ-4P': {'25m': {'20230123': [6.52, 169.15, 311.59, 14, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [25.74, 163.3, 240.9, 29, 10, 'Iron Sight', 'White Demon']}}},
        #      '23292': {
        #          'MZ-4P': {'25m': {'20230119': [12.34, 209.8, 191.92, 28, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23504': {
        #          'MZ-4P': {'25m': {'20230123': [6.63, 214.76, 193.6, 37, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23132': {
        #          'MZ-4P': {'25m': {'20230118': [13.92, 134.96, 156.38, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23452': {'MZ-4P': {'25m': {'20230123': [1.13, 307.0, 97.67, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [4.89, 285.68, 122.24, 16, 6, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']}}}, '23322': {
        #         'MZ-4P': {'25m': {'20230123': [6.17, 216.4, 172.4, 5, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                   '100m': {'20230125': [20.72, 212.69, 83.79, 10, 10, 'Iron Sight', 'White Demon']}}},
        #      '23505': {
        #          'MZ-4P': {'25m': {'20230123': [8.61, 192.68, 213.37, 27, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23047': {'MZ-4P': {'25m': {'20230117': [8.45, 148.64, 228.94, 9, 5, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [3.53, 258.33, 73.0, 4, 4, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']}}}, '23085': {
        #         'MZ-4P': {'25m': {'20230117': [2.43, 88.5, 164.17, 7, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                   '100m': {'20230125': [20.09, 196.1, 284.79, 18, 10, 'Iron Sight', 'White Demon']}}},
        #      '23258': {
        #          'MZ-4P': {'25m': {'20230123': [0.92, 108.85, 104.72, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23277': {
        #          'MZ-4P': {'25m': {'20230119': [7.94, 252.33, 109.67, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230202': [29.65, 225.53, 309.43, 9, 20, 'Iron Sight', 'ABC ']}}}, '23326': {
        #         'MZ-4P': {'100m': {'20230125': [22.72, 196.32, 212.27, 20, 10, 'Iron Sight', 'White Demon']},
        #                   '25m': {'20230123': [4.65, 199.29, 210.74, 23, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23272': {'MZ-4P': {'25m': {'20230119': [2.14, 199.38, 208.44, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [5.09, 266.25, 66.5, 17, 5, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']},
        #                          '30m': {'20230202': [13.16, 120.83, 329.83, 6, 20, 'Iron Sight', 'ABC ']}}},
        #      '23080': {'9mm/0.40/0.45': {'7m': {'20230209': [32.1, 169.22, 266.5, 18, 20, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {'25m': {'20230117': [7.78, 134.56, 315.64, 9, 5, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [16.42, 238.69, 224.23, 13, 13, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']}}}, '23463': {
        #         'MZ-4P': {'25m': {'20230123': [11.61, 224.8, 186.2, 10, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23243': {'MZ-4P': {'100m': {'20230125': [9.38, 159.67, 29.11, 32, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [5.77, 195.88, 179.62, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23050': {
        #          'MZ-4P': {'25m': {'20230117': [3.2, 202.67, 193.6, 22, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23046': {'MZ-4P': {'25m': {'20230122': [2.06, 108.25, 280.75, 5, 5, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230117': [5.52, 92.78, 340.46, 9, 4, 'Iron Sight',
        #                                               'DS Zeroing A4 Target']}},
        #                '9mm/0.40/0.45': {'7m': {'20230209': [35.28, 220.8, 220.05, 20, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23191': {
        #          'MZ-4P': {'25m': {'20230123': [4.39, 219.66, 203.38, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23429': {
        #          'MZ-4P': {'25m': {'20230123': [6.21, 221.06, 198.42, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230216': [39.64, 145.22, 247.29, 21, 20, 'Iron Sight', 'ABC - K']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [46.96, 160.25, 271.2, 41, 42, 'Iron Sight', 'ABC - K']}}},
        #      '23513': {'MZ-4P': {'100m': {'20230125': [25.32, 131.54, 105.12, 15, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {'20230123': [15.23, 216.83, 204.15, 19, 5, 'Iron Sight',
        #                                               'DS Zeroing A4 Target']}}}, '23032': {
        #         'MZ-4P': {'25m': {'20230116': [1.02, 198.81, 195.92, 33, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23225': {'MZ-4P': {'25m': {'20230123': [0, 0.0, 0.0, 1, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23339': {'MZ-4P': {'100m': {'20230125': [42.66, 192.42, 134.22, 24, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [4.58, 193.13, 203.98, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23282': {'MZ-4P': {'25m': {'20230202': [47.3, 146.47, 269.35, 32, 20, 'Iron Sight', 'ABC '],
        #                                  '20230119': [8.26, 184.77, 195.32, 21, 4, 'Iron Sight',
        #                                               'DS Zeroing A4 Target']}}}, '23130': {
        #         'MZ-4P': {'25m': {'20230118': [7.3, 204.8, 194.82, 22, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23323': {
        #          'MZ-4P': {'25m': {'20230123': [5.44, 319.67, 199.33, 4, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [14.82, 219.4, 220.7, 11, 10, 'Iron Sight', 'White Demon']}}},
        #      '23304': {'MZ-4P': {'30m': {'20230202': [41.53, 140.2, 306.8, 10, 20, 'Iron Sight', 'ABC ']}, '25m': {
        #          '20230119': [3.85, 198.67, 187.45, 32, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23523': {
        #         'MZ-4P': {'25m': {'20230123': [2.1, 203.39, 208.14, 22, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23341': {
        #          'MZ-4P': {'25m': {'20230123': [5.88, 202.84, 186.83, 21, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230216': [49.71, 243.62, 220.19, 16, 20, 'Iron Sight', 'ABC - K']},
        #                    '100m': {'20230125': [25.3, 263.33, 206.22, 19, 10, 'Iron Sight', 'White Demon']}}},
        #      '23247': {
        #          'MZ-4P': {'25m': {'20230123': [9.37, 109.5, 214.0, 11, 6, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23296': {'MZ-4P': {
        #          '25m': {'20230511': [10.94, 124.67, 86.33, 3, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230119': [10.02, 239.33, 128.0, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                  '20230202': [44.47, 202.27, 311.33, 29, 20, 'Iron Sight', 'ABC ']}},
        #                '9mm/0.40/0.45': {'7m': {'20230209': [51.74, 208.84, 196.48, 62, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23168': {
        #          'MZ-4P': {'25m': {'20230123': [14.47, 266.79, 157.96, 10, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23410': {
        #          'MZ-4P': {'25m': {'20230123': [6.06, 193.55, 171.44, 23, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23048': {
        #          'MZ-4P': {'25m': {'20230117': [15.46, 224.19, 209.21, 14, 5, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [33.68, 179.53, 282.31, 20, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23420': {
        #          'MZ-4P': {'25m': {'20230123': [3.73, 130.67, 301.33, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23052': {
        #          'MZ-4P': {'25m': {'20230117': [5.12, 205.31, 105.16, 9, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [24.33, 207.78, 194.25, 21, 10, 'Iron Sight', 'White Demon']}}},
        #      '23184': {
        #          'MZ-4P': {'25m': {'20230123': [4.28, 170.82, 164.7, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [37.08, 149.46, 254.96, 25, 10, 'Iron Sight', 'White Demon']}}},
        #      '23487': {
        #          'MZ-4P': {'25m': {'20230123': [10.76, 170.9, 175.01, 59, 12, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [19.04, 296.58, 167.45, 19, 10, 'Iron Sight', 'White Demon']},
        #                    '30m': {'20230216': [33.91, 127.79, 249.37, 40, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23426': {'9mm/0.40/0.45': {'7m': {'20230209': [39.64, 159.47, 261.84, 19, 20, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {
        #                    '25m': {'20230123': [12.09, 213.67, 110.67, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230216': [36.51, 246.15, 97.28, 20, 20, 'Iron Sight', 'ABC - K']},
        #                    '100m': {'20230125': [27.11, 242.11, 133.78, 18, 10, 'Iron Sight', 'White Demon']}}},
        #      '23266': {
        #          'MZ-4P': {'25m': {'20230119': [7.79, 201.02, 194.0, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230202': [37.89, 203.67, 223.2, 15, 20, 'Iron Sight', 'ABC ']}}}, '23167': {
        #         'MZ-4P': {'25m': {'20230123': [4.35, 203.6, 124.0, 19, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23172': {'MZ-4P': {'25m': {'20230123': [0, 142.28, 369.81, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [7.44, 173.81, 191.46, 25, 5, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']}, '30m': {
        #          '20230511': [25.8, 237.61, 303.67, 18, 20, 'Aimpoint COMPM5/M5S', 'ABC - K']}}}, '23034': {
        #         'MZ-4P': {'25m': {'20230116': [10.91, 137.11, 199.01, 10, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23156': {
        #          'MZ-4P': {'25m': {'20230118': [3.62, 203.65, 197.39, 38, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23534': {
        #          'MZ-4P': {'25m': {'20230123': [1.58, 194.39, 201.42, 8, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [24.23, 140.0, 239.8, 19, 10, 'Iron Sight', 'White Demon']}}},
        #      '23090': {'MZ-4P': {
        #          '25m': {'20230117': [17.86, 240.91, 169.84, 9, 5, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                  '20230118': [4.42, 200.54, 200.9, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #          '100m': {'20230125': [19.29, 269.19, 239.87, 9, 10, 'Iron Sight', 'White Demon']}}}, '23457': {
        #         'MZ-4P': {'25m': {'20230123': [4.7, 220.55, 126.44, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23291': {'MZ-4P': {
        #          '25m': {'20230119': [3.18, 177.24, 203.57, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                  '20230511': [12.6, 245.2, 184.91, 15, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target']},
        #          '30m': {'20230202': [30.32, 199.8, 315.07, 15, 20, 'Iron Sight', 'ABC ']}}}, '23197': {
        #         'MZ-4P': {'25m': {'20230123': [7.44, 199.41, 188.44, 21, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23027': {'MZ-4P': {'25m': {'20230115': [2.15, 108.0, 280.25, 4, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '100m': {'20230125': [23.25, 106.33, 189.0, 33, 10, 'Iron Sight', 'White Demon']}}},
        #      '23203': {'MZ-4P': {'30m': {'20230511': [26.67, 231.0, 293.45, 20, 20, 'Aimpoint COMPM5/M5S', 'ABC - K']},
        #                          '25m': {'20230123': [1.48, 171.81, 192.5, 33, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [1.51, 282.71, 91.09, 11, 4, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']}}}, '23222': {
        #         'MZ-4P': {'25m': {'20230123': [10.99, 299.0, 222.5, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23096': {'MZ-4P': {'100m': {'20230125': [42.21, 250.12, 172.46, 26, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {'20230117': [11.17, 216.55, 208.84, 12, 5, 'Iron Sight',
        #                                               'DS Zeroing A4 Target']}}}, '23072': {'MZ-4P': {
        #         '25m': {'20230117': [9.57, 362.19, 153.46, 25, 5, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                 '20230511': [5.63, 285.73, 99.53, 14, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target']},
        #         '30m': {'20230511': [66.29, 166.5, 325.35, 20, 20, 'Aimpoint COMPM5/M5S', 'ABC - K']}}}, '23073': {
        #         'MZ-4P': {'25m': {'20230117': [21.93, 337.19, 146.3, 25, 5, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #         '9mm/0.40/0.45': {'7m': {'20230209': [40.18, 199.55, 179.1, 20, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23221': {
        #          'MZ-4P': {'25m': {'20230123': [5.14, 57.33, 297.67, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23319': {
        #          'MZ-4P': {'25m': {'20230123': [4.89, 218.04, 246.42, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23383': {
        #          'MZ-4P': {'25m': {'20230123': [2.66, 209.51, 299.08, 21, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23097': {'MZ-4P': {'100m': {'20230125': [0, 171.74, 36.35, 16, 10, 'Iron Sight', 'White Demon']}, '25m': {
        #          '20230118': [15.08, 202.41, 205.8, 10, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23112': {
        #         'MZ-4P': {'25m': {'20230118': [3.81, 173.74, 205.76, 20, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                           '20230511': [2.63, 171.5, 229.86, 7, 5, 'Aimpoint COMPM5/M5S',
        #                                        'DS Zeroing A4 Target']}}}, '23035': {
        #         'MZ-4P': {'25m': {'20230116': [7.47, 199.93, 289.25, 8, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                   '30m': {'20230216': [60.81, 178.48, 273.36, 62, 21, 'Iron Sight', 'ABC - K']}}}, '23446': {
        #         'MZ-4P': {'25m': {'20230123': [15.14, 152.36, 275.73, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23025': {'MZ-4P': {'25m': {'20230116': [6.82, 244.5, 54.0, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [5.11, 150.25, 57.5, 4, 4, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']},
        #                          '30m': {'20230216': [69.02, 169.32, 264.44, 35, 20, 'Iron Sight', 'ABC - K']},
        #                          '100m': {'20230125': [11.95, 201.8, 201.8, 22, 10, 'Iron Sight', 'White Demon']}}},
        #      '23271': {
        #          'MZ-4P': {'30m': {'20230202': [44.74, 198.41, 208.18, 18, 20, 'Vector Rifle scope 1-8X', 'ABC ']},
        #                    '25m': {'20230119': [6.7, 244.22, 166.89, 10, 9, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23423': {'MZ-4P': {'25m': {'20230123': [7.05, 230.0, 208.33, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '30m': {'20230216': [32.15, 217.93, 311.0, 14, 20, 'Iron Sight', 'ABC - K']},
        #                          '100m': {'20230125': [25.95, 237.56, 261.93, 19, 10, 'Iron Sight', 'White Demon']}}},
        #      '23219': {
        #          'MZ-4P': {'25m': {'20230123': [7.7, 171.01, 247.06, 27, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23071': {
        #          'MZ-4P': {'25m': {'20230117': [6.94, 205.3, 177.6, 36, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23535': {
        #          'MZ-4P': {'25m': {'20230123': [1.72, 203.86, 171.38, 21, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23455': {'MZ-4P': {'100m': {'20230125': [14.82, 205.56, 237.0, 9, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [3.33, 202.94, 199.2, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23010': {'MZ-4P': {'25m': {'20230116': [4.81, 206.01, 179.13, 8, 5, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230115': [3.05, 201.0, 201.5, 9, 5, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #                '9mm/0.40/0.45': {'7m': {'20230209': [52.54, 189.69, 261.19, 16, 16, 'Iron Sight', 'ABC - K']}}},
        #      '23294': {'MZ-4P': {'30m': {'20230202': [28.14, 163.75, 336.0, 11, 20, 'Iron Sight', 'ABC ']}, '25m': {
        #          '20230119': [7.18, 194.92, 192.53, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23054': {
        #         'MZ-4P': {'25m': {'20230511': [3.61, 219.5, 192.0, 3, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                           '20230117': [6.82, 187.5, 181.25, 9, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23328': {'MZ-4P': {'100m': {'20230125': [16.89, 173.04, 234.24, 22, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [1.21, 180.0, 213.5, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23246': {
        #          'MZ-4P': {'25m': {'20230123': [7.12, 204.12, 191.58, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23181': {'MZ-4P': {'30m': {'20230511': [62.94, 209.54, 192.54, 33, 26, 'Aimpoint COMPM5/M5S', 'ABC - K']},
        #                          '25m': {'20230511': [4.21, 191.38, 213.74, 20, 5, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']}}}, '23486': {
        #         'MZ-4P': {'25m': {'20230123': [3.33, 271.79, 181.43, 21, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23187': {'MZ-4P': {'100m': {'20230125': [12.66, 204.58, 224.57, 25, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [11.22, 166.5, 169.75, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23531': {'MZ-4P': {
        #          '25m': {'20230511': [15.51, 242.7, 104.63, 29, 8, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230123': [3.09, 196.56, 202.38, 15, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23540': {'MZ-4P': {'25m': {'20230123': [0, 0.0, 0.0, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23044': {'MZ-4P': {'25m': {'20230117': [2.08, 201.31, 183.0, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [6.79, 288.0, 162.0, 3, 4, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']}, '30m': {
        #          '20230511': [61.48, 139.65, 278.3, 20, 20, 'Aimpoint COMPM5/M5S', 'ABC - K']}}}, '23479': {
        #         'MZ-4P': {'25m': {'20230123': [12.39, 214.62, 142.53, 18, 17, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23111': {'MZ-4P': {
        #          '25m': {'20230511': [7.8, 152.83, 242.1, 13, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230118': [4.64, 172.08, 21.94, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23077': {
        #         'MZ-4P': {'25m': {'20230117': [8.82, 21.14, 80.18, 8, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23491': {'MZ-4P': {'100m': {'20230125': [0, 251.0, 197.0, 32, 10, 'Iron Sight', 'White Demon']}, '25m': {
        #          '20230123': [16.46, 208.71, 172.9, 15, 9, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23230': {
        #         'MZ-4P': {'25m': {'20230123': [5.73, 149.88, 305.92, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23374': {
        #          'MZ-4P': {'25m': {'20230123': [14.78, 187.12, 278.83, 15, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23100': {'MZ-4P': {
        #          '25m': {'20230118': [10.85, 94.74, 221.47, 18, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                  '20230511': [6.69, 325.06, 93.98, 6, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target']},
        #          '100m': {'20230125': [3.69, 172.42, 230.34, 20, 10, 'Iron Sight', 'White Demon']}}}, '23544': {
        #         'MZ-4P': {
        #             '25m': {'20230511': [4.24, 296.33, 156.33, 22, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                     '20230123': [2.19, 199.11, 163.33, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23317': {
        #          'MZ-4P': {'25m': {'20230123': [11.35, 194.4, 178.8, 14, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23118': {
        #          'MZ-4P': {'25m': {'20230118': [8.92, 108.53, 304.08, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23407': {
        #          'MZ-4P': {'25m': {'20230123': [16.92, 266.2, 275.38, 23, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23255': {
        #          'MZ-4P': {'25m': {'20230123': [8.41, 118.89, 189.05, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [51.98, 141.67, 213.78, 9, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23069': {
        #          'MZ-4P': {'25m': {'20230117': [17.74, 209.3, 211.72, 25, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23062': {
        #          'MZ-4P': {'25m': {'20230117': [2.31, 162.89, 80.97, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23340': {'MZ-4P': {'30m': {'20230216': [45.17, 137.89, 216.63, 52, 20, 'Iron Sight', 'ABC - K']}, '25m': {
        #          '20230123': [5.45, 169.33, 180.33, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '100m': {'20230125': [33.61, 180.83, 230.89, 22, 10, 'Iron Sight', 'White Demon']}}},
        #      '23284': {'9mm/0.40/0.45': {'7m': {'20230209': [26.96, 197.73, 278.51, 20, 20, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {'25m': {'20230202': [44.6, 165.12, 292.44, 48, 20, 'Iron Sight', 'ABC '],
        #                                  '20230119': [5.96, 203.38, 196.03, 21, 4, 'Iron Sight',
        #                                               'DS Zeroing A4 Target']}}}, '23400': {
        #         'MZ-4P': {'25m': {'20230123': [8.23, 206.25, 101.5, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23059': {'MZ-4P': {
        #          '25m': {'20230511': [19.66, 283.0, 164.86, 7, 7, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230117': [9.42, 207.79, 185.85, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23443': {
        #         'MZ-4P': {'25m': {'20230123': [5.01, 210.93, 185.03, 21, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23413': {'MZ-4P': {'100m': {'20230125': [11.92, 211.15, 195.62, 30, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [4.06, 197.98, 211.25, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23394': {
        #          'MZ-4P': {'25m': {'20230123': [5.47, 197.09, 205.14, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23041': {
        #          'MZ-4P': {'25m': {'20230116': [2.14, 195.33, 187.25, 19, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23020': {
        #          'MZ-4P': {'25m': {'20230116': [13.67, 174.75, 237.25, 12, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23520': {'MZ-4P': {'30m': {'20230216': [62.63, 230.59, 228.99, 23, 20, 'Iron Sight', 'ABC - K']},
        #                          '25m': {'20230123': [0, 0.0, 0.0, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23039': {
        #          'MZ-4P': {'25m': {'20230116': [5.01, 175.25, 98.75, 18, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [57.06, 137.26, 264.63, 19, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23309': {'MZ-4P': {'30m': {'20230202': [10.51, 278.0, 315.0, 2, 20, 'Iron Sight', 'ABC ']}, '25m': {
        #          '20230119': [14.49, 149.34, 185.22, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23386': {
        #         'MZ-4P': {'25m': {'20230123': [7.41, 200.26, 213.02, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23344': {'MZ-4P': {'100m': {'20230125': [21.56, 170.56, 240.56, 29, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [4.28, 190.5, 141.75, 10, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23175': {
        #          'MZ-4P': {'25m': {'20230123': [2.37, 272.33, 135.0, 28, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [33.54, 223.2, 224.6, 10, 10, 'Iron Sight', 'White Demon']}}},
        #      '23401': {'MZ-4P': {
        #          '25m': {'20221109': [7.67, 137.6, 182.4, 18, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230123': [0, 0.0, 0.0, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23549': {
        #         'MZ-4P': {'25m': {'20230123': [8.46, 139.67, 138.5, 25, 6, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                   '100m': {'20230125': [40.99, 206.57, 124.43, 17, 14, 'Iron Sight', 'White Demon']}}},
        #      '23127': {'MZ-4P': {'25m': {'20230118': [3.63, 93.64, 65.75, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '100m': {'20230125': [31.46, 209.64, 249.09, 41, 10, 'Iron Sight', 'White Demon']}}},
        #      '23192': {
        #          'MZ-4P': {'25m': {'20230123': [8.9, 141.72, 189.09, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23226': {
        #          'MZ-4P': {'25m': {'20230123': [7.62, 144.91, 319.84, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23004': {
        #          'MZ-4P': {'25m': {'20230116': [2.58, 156.83, 376.38, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'25m': {'20230209': [28.08, 157.47, 289.42, 19, 19, 'Iron Sight', 'ABC - K']}}},
        #      '23249': {
        #          'MZ-4P': {'25m': {'20230123': [14.6, 150.54, 217.04, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [28.53, 152.1, 306.9, 10, 10, 'Iron Sight', 'White Demon']}}},
        #      '23502': {
        #          'MZ-4P': {'25m': {'20230123': [7.93, 96.11, 237.18, 23, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23091': {'MZ-4P': {'25m': {'20230118': [2.23, 284.88, 60.29, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '100m': {'20230125': [27.63, 101.51, 266.57, 25, 10, 'Iron Sight', 'White Demon']}}},
        #      '23212': {
        #          'MZ-4P': {'25m': {'20230123': [22.3, 191.5, 212.5, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23411': {
        #          'MZ-4P': {'25m': {'20230123': [5.11, 207.52, 203.81, 42, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230216': [66.91, 210.27, 273.6, 15, 20, 'Iron Sight', 'ABC - K']}}}, '23070': {
        #         'MZ-4P': {'25m': {'20230117': [3.68, 186.69, 202.77, 27, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23145': {'MZ-4P': {
        #          '25m': {'20230511': [11.92, 187.6, 144.58, 22, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230118': [7.39, 194.66, 205.81, 29, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23431': {
        #         'MZ-4P': {'25m': {'20230123': [12.85, 184.15, 197.9, 25, 7, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23299': {
        #          'MZ-4P': {'25m': {'20230119': [4.87, 169.5, 182.63, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230202': [22.98, 102.67, 77.33, 3, 20, 'Iron Sight', 'ABC ']}}}, '23126': {
        #         'MZ-4P': {'100m': {'20230125': [17.49, 242.6, 131.93, 37, 10, 'Iron Sight', 'White Demon']},
        #                   '25m': {'20230118': [4.48, 203.73, 199.52, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23063': {
        #          'MZ-4P': {'25m': {'20230117': [13.2, 199.91, 171.34, 26, 6, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [33.16, 194.55, 215.91, 26, 11, 'Iron Sight', 'White Demon']}}},
        #      '23196': {
        #          'MZ-4P': {'25m': {'20230123': [8.01, 264.33, 349.33, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23288': {'MZ-4P': {'25m': {'20230119': [7.13, 168.5, 52.0, 23, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230202': [47.41, 263.29, 264.43, 24, 20, 'Iron Sight', 'ABC ']}}},
        #      '23524': {
        #          'MZ-4P': {'25m': {'20230123': [3.38, 213.33, 186.67, 10, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23137': {'MZ-4P': {'100m': {'20230125': [21.82, 209.82, 167.36, 45, 11, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230118': [9.16, 208.55, 197.52, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23382': {'MZ-4P': {'100m': {'20230125': [11.99, 208.75, 223.35, 32, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [8.61, 236.75, 91.0, 7, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23088': {'MZ-4P': {'100m': {'20230125': [53.63, 257.28, 169.1, 19, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230117': [10.54, 193.2, 188.77, 11, 6, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23016': {
        #          'MZ-4P': {'25m': {'20230116': [12.45, 173.4, 112.2, 17, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230216': [57.58, 230.0, 255.38, 22, 20, 'Iron Sight', 'ABC - K']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [34.96, 162.5, 302.5, 16, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23261': {
        #          'MZ-4P': {'25m': {'20230123': [6.32, 264.75, 313.5, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23146': {
        #          'MZ-4P': {'25m': {'20230118': [5.38, 202.85, 197.54, 36, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [42.08, 130.06, 268.31, 16, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23142': {
        #          'MZ-4P': {'25m': {'20230118': [8.33, 112.8, 261.61, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23158': {
        #          'MZ-4P': {'25m': {'20230118': [5.15, 187.5, 150.8, 24, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23444': {
        #          'MZ-4P': {'25m': {'20230123': [8.79, 155.36, 150.46, 7, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [33.91, 196.13, 153.6, 18, 10, 'Iron Sight', 'White Demon']}}},
        #      '23404': {'MZ-4P': {
        #          '25m': {'20230123': [16.75, 152.73, 173.87, 20, 15, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                  '20230511': [4.69, 209.73, 215.24, 4, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target']}}},
        #      '23064': {'MZ-4P': {'25m': {'20230117': [8.75, 306.15, 143.87, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [5.03, 286.66, 200.91, 13, 4, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']},
        #                          '30m': {'20230511': [44.72, 229.84, 311.97, 25, 20, 'Aimpoint COMPM5/M5S', 'ABC - K']},
        #                          '100m': {'20230125': [17.25, 242.65, 214.22, 23, 10, 'Iron Sight', 'White Demon']}}},
        #      '23427': {
        #          'MZ-4P': {'25m': {'20230123': [3.77, 190.42, 200.82, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23484': {'9mm/0.40/0.45': {'7m': {'20230209': [24.86, 204.04, 261.99, 55, 20, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {'100m': {'20230125': [9.63, 190.17, 209.93, 34, 10, 'Iron Sight', 'White Demon']}}},
        #      '23550': {'MZ-4P': {'25m': {'20230123': [0, 0.0, 0.0, 10, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23257': {
        #          'MZ-4P': {'25m': {'20230123': [8.38, 246.3, 226.12, 20, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23241': {
        #          'MZ-4P': {'25m': {'20230123': [14.13, 151.8, 243.4, 31, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23485': {
        #          'MZ-4P': {'25m': {'20230123': [16.46, 196.38, 232.99, 25, 13, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230216': [51.78, 185.94, 187.13, 23, 20, 'Iron Sight', 'ABC - K']}}}, '23279': {
        #         'MZ-4P': {'25m': {'20230119': [8.48, 121.5, 97.5, 10, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23414': {
        #          'MZ-4P': {'25m': {'20230123': [4.81, 89.0, 246.67, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23517': {'MZ-4P': {'30m': {'20230216': [32.4, 199.42, 293.53, 19, 20, 'Aimpoint COMPM5/M5S', 'ABC - K']},
        #                          '25m': {
        #                              '20230123': [0.49, 193.16, 197.86, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23478': {
        #          'MZ-4P': {'25m': {'20230123': [10.68, 226.8, 195.8, 16, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23384': {'MZ-4P': {'30m': {'20230216': [62.91, 161.3, 225.39, 19, 20, 'Iron Sight', 'ABC - K']},
        #                          '100m': {'20230125': [0, 194.34, 224.57, 1, 10, 'Iron Sight', 'White Demon']}, '25m': {
        #              '20230123': [3.39, 154.5, 134.0, 13, 5, 'Vector Rifle scope 1-8X', 'DS Zeroing A4 Target']}}},
        #      '23337': {'MZ-4P': {
        #          '25m': {'20230511': [1.97, 196.46, 190.27, 15, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230123': [3.57, 182.21, 181.74, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #          '30m': {'20230511': [25.35, 191.15, 332.7, 21, 20, 'Aimpoint COMPM5/M5S', 'ABC - K']},
        #          '100m': {'20230125': [22.22, 192.32, 141.18, 31, 10, 'Iron Sight', 'White Demon']}}}, '23405': {
        #         'MZ-4P': {'25m': {'20230123': [1.09, 175.96, 233.56, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                   '30m': {'20230216': [38.53, 228.24, 117.9, 20, 20, 'Iron Sight', 'ABC - K']}}}, '23345': {
        #         'MZ-4P': {'25m': {'20230123': [13.55, 208.14, 210.77, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                   '100m': {'20230125': [27.18, 202.0, 269.33, 26, 10, 'Iron Sight', 'White Demon']}}},
        #      '23389': {
        #          'MZ-4P': {'25m': {'20230123': [19.3, 225.33, 268.0, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23231': {
        #          'MZ-4P': {'25m': {'20230123': [4.65, 208.6, 241.4, 15, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23153': {
        #          'MZ-4P': {'25m': {'20230118': [7.68, 206.5, 207.39, 24, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23481': {'MZ-4P': {'100m': {'20230125': [21.35, 179.43, 189.54, 36, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [2.45, 187.38, 195.13, 24, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23198': {
        #          'MZ-4P': {'25m': {'20230123': [4.8, 135.0, 180.5, 20, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23119': {
        #          'MZ-4P': {'25m': {'20230118': [6.27, 205.8, 176.93, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [27.31, 247.2, 194.2, 28, 10, 'Iron Sight', 'White Demon']}}},
        #      '23434': {
        #          'MZ-4P': {'25m': {'20230123': [18.59, 292.0, 190.25, 22, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23148': {
        #          'MZ-4P': {'25m': {'20230118': [7.24, 200.47, 208.79, 35, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23233': {
        #          'MZ-4P': {'25m': {'20230123': [5.65, 215.6, 179.0, 20, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23179': {'MZ-4P': {'25m': {'20230123': [7.41, 94.5, 240.0, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23276': {'MZ-4P': {'30m': {'20230202': [42.22, 216.97, 228.15, 15, 20, 'Iron Sight', 'ABC ']}, '25m': {
        #          '20230511': [9.03, 57.8, 139.2, 5, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #          '20230119': [3.05, 106.0, 261.33, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23023': {
        #         'MZ-4P': {'25m': {'20230116': [15.94, 169.0, 255.5, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23357': {
        #          'MZ-4P': {'25m': {'20230123': [8.29, 220.5, 183.75, 7, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23270': {
        #          'MZ-4P': {'25m': {'20230119': [11.51, 226.22, 158.0, 10, 9, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230202': [40.84, 227.54, 290.0, 13, 20, 'Iron Sight', 'ABC ']}}}, '23195': {
        #         'MZ-4P': {'25m': {'20230123': [9.34, 209.5, 188.5, 23, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23081': {
        #          'MZ-4P': {'25m': {'20230117': [5.96, 263.97, 243.2, 13, 5, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [26.61, 174.37, 261.42, 19, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23361': {
        #          'MZ-4P': {'25m': {'20230123': [2.39, 199.01, 200.75, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23210': {
        #          'MZ-4P': {'25m': {'20230123': [5.81, 186.0, 179.67, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23223': {
        #          'MZ-4P': {'25m': {'20230123': [4.78, 253.42, 205.62, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23425': {'MZ-4P': {'25m': {'20230123': [6.3, 138.67, 202.0, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '30m': {'20230216': [35.27, 187.73, 281.48, 20, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23311': {'MZ-4P': {'25m': {'20230119': [3.07, 37.56, 183.48, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '30m': {'20230202': [25.42, 95.33, 301.44, 9, 20, 'Iron Sight', 'ABC ']}},
        #                '9mm/0.40/0.45': {'7m': {'20230209': [34.02, 130.09, 289.78, 35, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23235': {'MZ-4P': {'25m': {'20230123': [0, 204.0, 135.0, 22, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23346': {
        #          'MZ-4P': {'25m': {'20230123': [10.71, 238.5, 150.0, 16, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230216': [39.09, 150.91, 262.11, 29, 20, 'Iron Sight', 'ABC - K']}}}, '23307': {
        #         'MZ-4P': {'30m': {'20230202': [58.98, 180.36, 185.5, 14, 20, 'Iron Sight', 'ABC ']},
        #                   '25m': {'20230119': [6.08, 200.81, 204.28, 28, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #         '9mm/0.40/0.45': {'7m': {'20230209': [32.88, 211.18, 212.82, 60, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23280': {'MZ-4P': {'25m': {'20230202': [40.84, 187.0, 248.59, 35, 22, 'Iron Sight', 'ABC ']},
        #                          '100m': {'20230125': [11.16, 197.4, 198.36, 27, 10, 'Iron Sight', 'White Demon']}}},
        #      '23015': {'MZ-4P': {'25m': {'20230116': [6.03, 177.85, 171.6, 9, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '30m': {'20230216': [40.73, 212.66, 209.42, 29, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23007': {'MZ-4P': {'30m': {'20230216': [30.44, 224.43, 301.82, 36, 20, 'Iron Sight', 'ABC - K']},
        #                          '25m': {'20230116': [8.81, 215.98, 85.36, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [1.43, 182.0, 180.5, 3, 3, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']}}}, '23188': {
        #         'MZ-4P': {'25m': {'20230123': [8.76, 176.88, 238.54, 23, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23366': {
        #          'MZ-4P': {'25m': {'20230123': [7.74, 204.73, 216.74, 29, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23049': {
        #          'MZ-4P': {'25m': {'20230117': [22.18, 197.72, 195.89, 27, 8, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23019': {
        #          'MZ-4P': {'25m': {'20230116': [7.62, 194.41, 81.74, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23416': {
        #          'MZ-4P': {'25m': {'20230123': [2.58, 201.41, 230.24, 9, 5, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'25m': {'20230209': [48.66, 167.89, 214.61, 18, 18, 'Iron Sight', 'ABC - K']}}},
        #      '23421': {
        #          'MZ-4P': {'25m': {'20230123': [6.49, 195.83, 151.33, 14, 6, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23207': {
        #          'MZ-4P': {'25m': {'20230123': [9.16, 207.11, 210.38, 26, 7, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23170': {
        #          'MZ-4P': {'25m': {'20230123': [4.59, 152.75, 129.5, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [68.44, 132.11, 209.38, 22, 21, 'Iron Sight', 'ABC - K']}}},
        #      '23274': {
        #          'MZ-4P': {'25m': {'20230119': [10.23, 109.75, 95.75, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [22.29, 199.44, 196.89, 9, 10, 'Iron Sight', 'White Demon']},
        #                    '30m': {'20230202': [43.7, 238.13, 251.67, 15, 20, 'Iron Sight', 'ABC ']}}}, '23051': {
        #         'MZ-4P': {'25m': {'20230117': [7.69, 176.25, 245.5, 29, 5, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #         '9mm/0.40/0.45': {'25m': {'20230209': [42.18, 165.08, 203.99, 19, 23, 'Iron Sight', 'ABC - K']}}},
        #      '23194': {
        #          'MZ-4P': {'25m': {'20230123': [4.31, 252.25, 278.0, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23102': {'MZ-4P': {'100m': {'20230125': [24.83, 197.78, 77.89, 9, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230118': [6.89, 196.22, 205.68, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23251': {
        #          'MZ-4P': {'25m': {'20230123': [7.22, 198.52, 191.08, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23475': {
        #          'MZ-4P': {'25m': {'20230123': [5.2, 273.33, 69.67, 18, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23359': {
        #          'MZ-4P': {'25m': {'20230123': [8.47, 228.0, 108.25, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [70.29, 136.95, 268.08, 34, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23043': {'MZ-4P': {'25m': {'20230117': [9.44, 173.88, 190.02, 7, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [2.96, 236.0, 61.0, 4, 5, 'Aimpoint COMPM5/M5S',
        #                                               'DS Zeroing A4 Target']}}}, '23285': {'MZ-4P': {
        #         '25m': {'20230202': [40.1, 132.12, 313.0, 29, 20, 'Iron Sight', 'ABC '],
        #                 '20230119': [4.44, 191.9, 210.07, 20, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23297': {
        #         'MZ-4P': {'25m': {'20230119': [3.42, 205.29, 201.72, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                           '20230202': [35.5, 248.75, 286.0, 29, 20, 'Iron Sight', 'ABC ']}}}, '23003': {
        #         'MZ-4P': {'25m': {'20230116': [5.03, 336.75, 294.0, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23460': {'MZ-4P': {'25m': {'20230123': [4.49, 93.5, 358.0, 3, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23347': {
        #          'MZ-4P': {'25m': {'20230123': [4.55, 204.09, 210.16, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23437': {'MZ-4P': {
        #          '25m': {'20230125': [0, 0.0, 0.0, 30, 5, 'Vector Rifle scope 1-8X', 'DS Zeroing A4 Target']}}},
        #      '23318': {'9mm/0.40/0.45': {'25m': {'20230209': [69.14, 158.44, 212.67, 8, 9, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {'100m': {'20230125': [19.87, 202.22, 163.03, 12, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [6.94, 291.0, 86.75, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23495': {
        #          'MZ-4P': {'25m': {'20230123': [3.36, 205.98, 202.02, 24, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23205': {
        #          'MZ-4P': {'25m': {'20230123': [2.81, 114.67, 161.67, 23, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23356': {'MZ-4P': {'100m': {'20230125': [26.64, 211.02, 238.95, 29, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {'20230123': [0, -1.0, -1.0, 2, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23135': {
        #          'MZ-4P': {'25m': {'20230118': [3.18, 145.27, 246.85, 10, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23300': {
        #          'MZ-4P': {'25m': {'20230119': [7.82, 189.75, 200.25, 19, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230202': [27.59, 220.12, 169.44, 16, 20, 'Iron Sight', 'ABC ']}}},
        #      '23227': {'MZ-4P': {'25m': {'20230123': [5.72, 327.0, 49.5, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23539': {
        #          'MZ-4P': {'25m': {'20230123': [2.34, 205.69, 187.08, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23497': {
        #          'MZ-4P': {'25m': {'20230123': [14.67, 202.1, 225.15, 13, 6, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23430': {
        #          'MZ-4P': {'25m': {'20230123': [16.55, 189.0, 161.0, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23141': {
        #          'MZ-4P': {'25m': {'20230118': [8.35, 281.79, 54.6, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23433': {
        #          'MZ-4P': {'25m': {'20230123': [4.68, 186.0, 285.0, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23134': {
        #          'MZ-4P': {'25m': {'20230118': [4.54, 185.9, 197.43, 34, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23259': {
        #          'MZ-4P': {'25m': {'20230123': [10.88, 221.61, 141.58, 19, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23250': {'MZ-4P': {'25m': {'20230123': [0, 258.0, 236.0, 19, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23367': {
        #          'MZ-4P': {'25m': {'20230123': [17.13, 199.25, 142.25, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23136': {'MZ-4P': {'25m': {'20230118': [6.59, 179.9, 201.9, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '100m': {'20230125': [44.28, 194.78, 220.22, 33, 10, 'Iron Sight', 'White Demon']}}},
        #      '23177': {
        #          'MZ-4P': {'25m': {'20230123': [12.09, 191.2, 180.4, 6, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23029': {
        #          'MZ-4P': {'25m': {'20230116': [5.01, 192.42, 140.8, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23066': {'9mm/0.40/0.45': {'7m': {'20230209': [21.94, 206.12, 281.76, 25, 25, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {
        #                    '25m': {'20230117': [18.14, 202.31, 187.79, 18, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23338': {'MZ-4P': {'100m': {'20230125': [6.79, 192.78, 201.94, 23, 10, 'Iron Sight', 'White Demon']},
        #                          '30m': {'20230216': [48.2, 179.77, 143.01, 16, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23109': {'MZ-4P': {
        #          '25m': {'20230511': [13.79, 81.86, 39.71, 10, 7, 'Vector Rifle scope 1-8X', 'DS Zeroing A4 Target'],
        #                  '20230118': [6.92, 180.61, 194.22, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #          '100m': {'20230125': [32.7, 208.39, 247.2, 29, 10, 'Iron Sight', 'White Demon']}}}, '23238': {
        #         'MZ-4P': {'25m': {'20230123': [12.81, 243.44, 326.59, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23239': {
        #          'MZ-4P': {'25m': {'20230123': [11.75, 225.65, 189.35, 24, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23125': {'MZ-4P': {'25m': {'20230118': [1.5, 205.06, 86.17, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '100m': {'20230125': [30.34, 222.67, 237.67, 37, 10, 'Iron Sight', 'White Demon']}}},
        #      '23530': {
        #          'MZ-4P': {'25m': {'20230123': [3.48, 203.74, 208.77, 20, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23065': {
        #          'MZ-4P': {'25m': {'20230117': [10.54, 227.25, 187.25, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23121': {
        #          'MZ-4P': {'25m': {'20230118': [3.66, 205.64, 198.72, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23262': {
        #          'MZ-4P': {'25m': {'20230123': [3.62, 195.78, 203.61, 28, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23036': {
        #          'MZ-4P': {'25m': {'20230116': [8.16, 226.1, 178.27, 18, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23201': {
        #          'MZ-4P': {'25m': {'20230123': [7.83, 283.25, 230.0, 19, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23078': {
        #          'MZ-4P': {'25m': {'20230117': [8.38, 354.5, 149.25, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23521': {'9mm/0.40/0.45': {'7m': {'20230209': [43.08, 105.47, 232.88, 17, 20, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {'25m': {'20230123': [0, 0.0, 0.0, 8, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23327': {
        #          'MZ-4P': {'25m': {'20230123': [10.89, 160.33, 122.99, 12, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23351': {
        #          'MZ-4P': {'25m': {'20230123': [7.44, 150.92, 242.55, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [0, 242.24, 221.06, 21, 10, 'Iron Sight', 'White Demon']}}}, '23436': {
        #         'MZ-4P': {'7m': {'20230209': [31.35, 186.37, 159.6, 41, 20, 'Aimpoint COMPM5/M5S', 'ABC - K']},
        #                   '25m': {'20230123': [8.51, 184.01, 202.34, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23267': {'MZ-4P': {'25m': {'20230119': [0, 0.0, 0.0, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '30m': {'20230202': [35.74, 180.0, 258.03, 14, 20, 'Iron Sight', 'ABC ']}}}, '23419': {
        #         'MZ-4P': {'100m': {'20230125': [27.31, 239.6, 140.3, 20, 10, 'Vector Rifle scope 1-8X', 'White Demon']},
        #                   '25m': {'20230123': [9.22, 187.74, 185.72, 20, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23150': {
        #          'MZ-4P': {'25m': {'20230118': [5.93, 200.68, 187.26, 24, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23354': {
        #          'MZ-4P': {'25m': {'20230123': [8.02, 217.4, 140.0, 11, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23269': {
        #          'MZ-4P': {'25m': {'20230119': [3.48, 222.33, 155.33, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '30m': {'20230202': [46.39, 143.65, 206.0, 17, 20, 'Iron Sight', 'ABC ']}}}, '23275': {
        #         'MZ-4P': {'100m': {'20230125': [19.67, 193.5, 210.1, 10, 10, 'Iron Sight', 'White Demon']},
        #                   '30m': {'20230202': [54.54, 176.9, 249.31, 19, 20, 'Iron Sight', 'ABC ']},
        #                   '25m': {'20230119': [6.99, 170.5, 187.75, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23368': {'MZ-4P': {
        #          '25m': {'20230511': [7.47, 201.99, 185.45, 8, 4, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230123': [0.28, 192.0, 277.5, 24, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23490': {
        #         'MZ-4P': {'25m': {'20230123': [14.57, 147.98, 218.62, 17, 9, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                   '100m': {'20230125': [35.46, 184.84, 202.26, 37, 10, 'Iron Sight', 'White Demon']}}},
        #      '23182': {
        #          'MZ-4P': {'25m': {'20230123': [8.08, 181.14, 193.86, 9, 7, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23224': {'MZ-4P': {
        #          '25m': {'20230123': [7.96, 195.77, 202.13, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                  '20230511': [3.17, 189.13, 185.67, 15, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target']}}},
        #      '23353': {'MZ-4P': {
        #          '25m': {'20230511': [7.41, 115.25, 161.25, 10, 5, 'Aimpoint COMPM5/M5S', 'DS Zeroing A4 Target'],
        #                  '20230123': [3.2, 201.06, 247.97, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23364': {
        #         'MZ-4P': {'25m': {'20230123': [2.65, 177.6, 213.83, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23268': {'MZ-4P': {'25m': {'20230119': [6.13, 149.0, 268.67, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '30m': {'20230202': [48.16, 197.91, 267.55, 11, 20, 'Iron Sight', 'ABC ']}},
        #                '9mm/0.40/0.45': {'7m': {'20230209': [65.45, 149.05, 217.18, 59, 40, 'Iron Sight', 'ABC - K']}}},
        #      '23265': {'MZ-4P': {'30m': {'20230202': [44.32, 193.29, 286.62, 22, 21, 'Iron Sight', 'ABC ']}}},
        #      '23533': {'MZ-4P': {'30m': {'20230216': [43.95, 211.4, 259.25, 20, 20, 'Iron Sight', 'ABC - K']}, '25m': {
        #          '20230123': [1.3, 201.25, 214.19, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23370': {
        #         'MZ-4P': {'25m': {'20230123': [1.38, 205.23, 141.8, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23160': {
        #          'MZ-4P': {'25m': {'20230118': [11.4, 180.66, 192.11, 25, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23424': {'MZ-4P': {'30m': {'20230216': [31.21, 175.62, 281.54, 20, 20, 'Iron Sight', 'ABC - K']}, '25m': {
        #          '20230123': [3.67, 221.98, 223.04, 22, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23005': {
        #         'MZ-4P': {'25m': {'20230116': [5.75, 217.27, 250.63, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23543': {'MZ-4P': {'25m': {'20230123': [0, 0.0, 0.0, 23, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23140': {
        #          'MZ-4P': {'25m': {'20230118': [13.14, 200.14, 180.14, 38, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23139': {
        #          'MZ-4P': {'25m': {'20230118': [3.07, 111.15, 136.62, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23358': {
        #          'MZ-4P': {'25m': {'20230123': [3.16, 199.4, 205.92, 30, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [8.61, 197.15, 202.36, 22, 10, 'Iron Sight', 'White Demon']}}},
        #      '23456': {'MZ-4P': {'100m': {'20230125': [18.03, 208.19, 225.36, 20, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [2.94, 190.63, 198.13, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '30m': {'20230216': [59.14, 209.75, 172.0, 20, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23467': {
        #          'MZ-4P': {'25m': {'20230123': [9.87, 159.5, 116.75, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23011': {
        #          'MZ-4P': {'25m': {'20230116': [8.55, 163.31, 37.63, 7, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23308': {
        #          'MZ-4P': {'25m': {'20230119': [5.41, 263.85, 257.91, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23293': {
        #          'MZ-4P': {'25m': {'20230119': [3.87, 209.4, 175.63, 19, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #          '9mm/0.40/0.45': {'7m': {'20230209': [0, 283.0, 221.0, 19, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23522': {'MZ-4P': {'25m': {'20230123': [0, 0.0, 0.0, 8, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23289': {'MZ-4P': {'30m': {'20230202': [43.15, 165.81, 287.06, 16, 20, 'Iron Sight', 'ABC ']}, '25m': {
        #          '20230119': [7.12, 172.75, 357.5, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23503': {
        #         'MZ-4P': {'25m': {'20230123': [1.71, 205.29, 162.66, 22, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23464': {
        #          'MZ-4P': {'25m': {'20230123': [8.79, 206.05, 185.21, 15, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23157': {'9mm/0.40/0.45': {'7m': {'20230209': [47.95, 144.05, 256.16, 19, 20, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {
        #                    '25m': {'20230118': [9.99, 229.18, 279.11, 19, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23183': {
        #          'MZ-4P': {'25m': {'20230123': [7.18, 64.75, 192.5, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23390': {
        #          'MZ-4P': {'25m': {'20230123': [4.06, 161.5, 42.0, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23133': {
        #          'MZ-4P': {'25m': {'20230118': [3.73, 189.96, 176.85, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23331': {
        #          'MZ-4P': {'25m': {'20230123': [6.91, 197.52, 210.52, 8, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23335': {'MZ-4P': {'25m': {'20230123': [5.81, 201.64, 173.9, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '100m': {'20230125': [45.35, 226.6, 304.2, 17, 10, 'Iron Sight', 'White Demon']}}},
        #      '23324': {'MZ-4P': {'25m': {'20230123': [7.07, 226.0, 87.25, 10, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '100m': {'20230125': [29.7, 226.75, 282.38, 19, 10, 'Iron Sight', 'White Demon']}}},
        #      '23348': {'9mm/0.40/0.45': {'7m': {'20230209': [56.62, 161.36, 255.93, 20, 20, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {'100m': {'20230125': [19.45, 201.2, 226.9, 39, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [8.12, 211.92, 160.55, 19, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '30m': {'20230216': [42.23, 202.18, 239.12, 23, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23529': {'MZ-4P': {'25m': {'20230123': [3.99, 216.86, 232.4, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                                  '20230511': [7.56, 85.84, 200.67, 14, 5, 'Iron Sight',
        #                                               'DS Zeroing A4 Target']}}}, '23537': {
        #         'MZ-4P': {'25m': {'20230123': [1.05, 201.55, 170.18, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23301': {'MZ-4P': {'25m': {'20230202': [41.87, 164.06, 179.03, 33, 20, 'Iron Sight', 'ABC '],
        #                                  '20230119': [9.16, 323.47, 173.44, 13, 4, 'Iron Sight',
        #                                               'DS Zeroing A4 Target']}},
        #                '9mm/0.40/0.45': {'7m': {'20230209': [33.51, 189.55, 260.36, 96, 30, 'Iron Sight', 'ABC - K']}}},
        #      '23538': {
        #          'MZ-4P': {'25m': {'20230123': [12.37, 130.59, 317.09, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23306': {'MZ-4P': {'25m': {'20230119': [8.38, 164.9, 321.7, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '30m': {'20230202': [6.97, 194.0, 337.0, 2, 20, 'Iron Sight', 'ABC ']}}}, '23547': {
        #         'MZ-4P': {'25m': {'20230123': [1.52, 172.74, 214.53, 4, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23476': {
        #          'MZ-4P': {'25m': {'20230123': [13.44, 206.63, 242.06, 26, 8, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23507': {
        #          'MZ-4P': {'25m': {'20230123': [6.49, 209.27, 211.66, 26, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23428': {
        #          'MZ-4P': {'25m': {'20230123': [3.68, 205.56, 205.17, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23409': {
        #          'MZ-4P': {'25m': {'20230123': [12.33, 194.59, 211.32, 21, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23395': {
        #          'MZ-4P': {'25m': {'20230123': [2.67, 278.6, 100.02, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23101': {
        #          'MZ-4P': {'25m': {'20230118': [4.41, 168.33, 116.03, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23412': {'9mm/0.40/0.45': {'7m': {'20230209': [35.38, 168.2, 146.7, 31, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23536': {
        #          'MZ-4P': {'25m': {'20230123': [12.39, 166.0, 140.25, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23321': {'MZ-4P': {'100m': {'20230125': [55.88, 226.78, 174.63, 11, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {'20230123': [5.98, 125.0, 87.0, 4, 4, 'Iron Sight', 'DS Zeroing A4 Target']}},
        #                '9mm/0.40/0.45': {'7m': {'20230209': [33.89, 217.35, 225.55, 14, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23365': {
        #          'MZ-4P': {'25m': {'20230123': [7.86, 191.65, 178.45, 15, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23349': {'MZ-4P': {'100m': {'20230125': [13.43, 185.11, 211.11, 30, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [4.72, 183.55, 173.1, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23546': {
        #          'MZ-4P': {'25m': {'20230123': [3.11, 205.16, 208.28, 18, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23161': {
        #          'MZ-4P': {'25m': {'20230118': [15.55, 150.07, 83.19, 27, 5, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '60m': {'20230125': [6.33, 222.05, 225.73, 13, 10, 'Iron Sight', 'White Demon']}}},
        #      '23526': {'MZ-4P': {'25m': {'20230123': [5.16, 204.25, 173.5, 7, 4, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                          '100m': {'20230125': [49.65, 203.0, 202.8, 17, 10, 'Iron Sight', 'White Demon']}}},
        #      '23492': {
        #          'MZ-4P': {'25m': {'20230123': [17.16, 150.66, 271.66, 21, 7, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [0, 176.02, 28.45, 34, 10, 'Iron Sight', 'White Demon']}}}, '23402': {
        #         'MZ-4P': {'25m': {'20230123': [2.66, 199.99, 240.5, 35, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23512': {
        #          'MZ-4P': {'25m': {'20230123': [2.0, 209.92, 210.35, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23055': {
        #          'MZ-4P': {'25m': {'20230117': [8.85, 162.01, 161.38, 10, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23217': {
        #          'MZ-4P': {'25m': {'20230123': [13.13, 255.0, 222.75, 4, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23515': {
        #          'MZ-4P': {'25m': {'20230123': [9.93, 126.5, 117.5, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23114': {
        #          'MZ-4P': {'25m': {'20230118': [8.55, 195.43, 172.32, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23462': {
        #          'MZ-4P': {'25m': {'20230123': [5.16, 303.75, 114.0, 6, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23178': {'MZ-4P': {'100m': {'20230125': [38.6, 203.48, 161.58, 26, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [7.95, 211.0, 199.75, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23076': {'MZ-4P': {'30m': {'20230216': [37.13, 150.38, 301.19, 16, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23147': {
        #          'MZ-4P': {'25m': {'20230118': [8.77, 192.11, 205.92, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23092': {
        #          'MZ-4P': {'25m': {'20230118': [13.91, 189.38, 192.78, 12, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23432': {
        #          'MZ-4P': {'25m': {'20230123': [5.42, 189.34, 172.49, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23159': {
        #          'MZ-4P': {'25m': {'20230118': [3.8, 199.06, 171.37, 25, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23209': {
        #          'MZ-4P': {'25m': {'20230123': [7.0, 192.09, 194.77, 32, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23084': {'MZ-4P': {
        #          '25m': {'20230118': [24.89, 190.85, 177.22, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target'],
        #                  '20230117': [13.34, 87.89, 249.53, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}, '23154': {
        #         'MZ-4P': {'25m': {'20230118': [5.74, 200.57, 190.51, 23, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23149': {
        #          'MZ-4P': {'25m': {'20230118': [2.25, 36.53, 95.46, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23525': {
        #          'MZ-4P': {'25m': {'20230123': [2.5, 207.21, 209.6, 31, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23018': {
        #          'MZ-4P': {'25m': {'20230116': [5.81, 179.27, 199.55, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23472': {
        #          'MZ-4P': {'25m': {'20230123': [3.65, 195.76, 225.6, 21, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23008': {
        #          'MZ-4P': {'25m': {'20230116': [8.26, 180.38, 208.17, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23252': {
        #          'MZ-4P': {'25m': {'20230123': [3.24, 194.39, 188.53, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23122': {
        #          'MZ-4P': {'25m': {'20230118': [3.93, 196.6, 182.43, 22, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23422': {
        #          'MZ-4P': {'25m': {'20230123': [5.41, 202.45, 206.21, 24, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23074': {
        #          'MZ-4P': {'25m': {'20230117': [16.43, 221.51, 181.15, 28, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23040': {
        #          'MZ-4P': {'25m': {'20230116': [3.38, 227.15, 299.72, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23459': {'MZ-4P': {'25m': {'20230123': [0, 0.0, 0.0, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23176': {
        #          'MZ-4P': {'25m': {'20230123': [3.62, 199.98, 218.15, 10, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23371': {
        #          'MZ-4P': {'25m': {'20230123': [16.19, 202.57, 212.52, 11, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23028': {
        #          'MZ-4P': {'25m': {'20230116': [4.11, 280.63, 175.42, 8, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23471': {'MZ-4P': {'25m': {'20230123': [4.93, 47.25, 8.5, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23489': {
        #          'MZ-4P': {'25m': {'20230123': [13.85, 233.43, 181.05, 22, 8, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [52.32, 226.4, 201.33, 30, 10, 'Iron Sight', 'White Demon']}}},
        #      '23199': {
        #          'MZ-4P': {'25m': {'20230123': [6.95, 228.75, 191.75, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23373': {
        #          'MZ-4P': {'25m': {'20230123': [11.63, 201.05, 141.92, 12, 6, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23377': {
        #          'MZ-4P': {'25m': {'20230123': [5.67, 174.75, 281.04, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23057': {
        #          'MZ-4P': {'25m': {'20230117': [9.43, 212.21, 184.63, 13, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23470': {
        #          'MZ-4P': {'25m': {'20230123': [5.9, 151.0, 113.67, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23037': {'9mm/0.40/0.45': {'7m': {'20230209': [25.86, 173.0, 340.44, 16, 20, 'Iron Sight', 'ABC - K']}},
        #                'MZ-4P': {
        #                    '25m': {'20230116': [14.87, 208.57, 186.07, 24, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23103': {
        #          'MZ-4P': {'25m': {'20230118': [6.89, 107.21, 58.41, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23079': {
        #          'MZ-4P': {'25m': {'20230117': [1.43, 236.75, 116.75, 10, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23381': {
        #          'MZ-4P': {'25m': {'20230123': [4.8, 118.87, 191.81, 13, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23320': {'MZ-4P': {'100m': {'20230125': [16.92, 210.7, 230.4, 10, 10, 'Iron Sight', 'White Demon']},
        #                          '25m': {
        #                              '20230123': [3.14, 186.75, 208.5, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23548': {'MZ-4P': {'25m': {'20230123': [0, 33.0, 137.0, 2, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23499': {
        #          'MZ-4P': {'25m': {'20230123': [6.91, 165.67, 229.33, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23113': {
        #          'MZ-4P': {'25m': {'20230118': [11.74, 192.17, 208.29, 24, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23469': {'MZ-4P': {'25m': {'20230123': [4.8, 157.0, 143.2, 7, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23303': {'MZ-4P': {'30m': {'20230202': [38.0, 195.75, 195.5, 4, 20, 'Iron Sight', 'ABC ']}}}, '23439': {
        #         'MZ-4P': {'25m': {'20230123': [2.34, 188.17, 140.06, 18, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23204': {'MZ-4P': {'25m': {'20230123': [5.8, 129.5, 274.0, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23075': {
        #          'MZ-4P': {'25m': {'20230117': [14.09, 199.0, 186.41, 24, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23545': {
        #          'MZ-4P': {'25m': {'20230123': [7.9, 213.04, 197.92, 27, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23379': {'MZ-4P': {'25m': {'20230123': [9.0, 199.0, 205.0, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23012': {
        #          'MZ-4P': {'25m': {'20230116': [2.18, 161.24, 347.81, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23483': {
        #          'MZ-4P': {'25m': {'20230123': [6.4, 209.41, 236.62, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23393': {
        #          'MZ-4P': {'25m': {'20230123': [5.78, 178.0, 174.0, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23143': {
        #          'MZ-4P': {'25m': {'20230118': [5.77, 168.95, 190.89, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23378': {
        #          'MZ-4P': {'25m': {'20230123': [17.55, 184.0, 189.38, 17, 8, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23397': {
        #          'MZ-4P': {'25m': {'20230123': [1.0, 157.93, 273.77, 8, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23360': {
        #          'MZ-4P': {'25m': {'20230123': [6.34, 212.75, 199.75, 14, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23453': {
        #          'MZ-4P': {'25m': {'20230123': [2.74, 192.82, 133.54, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23542': {
        #          'MZ-4P': {'25m': {'20230123': [2.14, 204.86, 176.59, 19, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23474': {
        #          'MZ-4P': {'25m': {'20230123': [5.85, 222.9, 207.23, 26, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23202': {
        #          'MZ-4P': {'25m': {'20230123': [6.15, 203.23, 172.32, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23480': {
        #          'MZ-4P': {'25m': {'20230123': [10.98, 198.5, 176.0, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23171': {'9mm/0.40/0.45': {'7m': {'20230209': [24.37, 164.83, 178.62, 11, 20, 'Iron Sight', 'ABC - K']}}},
        #      '23144': {
        #          'MZ-4P': {'25m': {'20230118': [4.15, 174.47, 202.12, 17, 6, 'Iron Sight', 'DS Zeroing A4 Target']},
        #                    '100m': {'20230125': [51.64, 121.55, 165.23, 37, 11, 'Iron Sight', 'White Demon']}}},
        #      '23099': {
        #          'MZ-4P': {'25m': {'20230117': [12.12, 181.39, 214.13, 12, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23106': {
        #          'MZ-4P': {'25m': {'20230118': [17.62, 210.0, 201.45, 11, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23116': {
        #          'MZ-4P': {'25m': {'20230118': [12.53, 197.82, 187.49, 34, 7, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23244': {
        #          'MZ-4P': {'25m': {'20230123': [6.91, 200.9, 195.52, 16, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23138': {
        #          'MZ-4P': {'25m': {'20230118': [4.84, 183.05, 194.51, 21, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23442': {
        #          'MZ-4P': {'25m': {'20230123': [3.69, 196.54, 219.32, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23206': {
        #          'MZ-4P': {'25m': {'20230123': [5.37, 202.22, 205.01, 17, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23398': {
        #          'MZ-4P': {'25m': {'20230123': [7.24, 206.75, 174.25, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23385': {
        #          'MZ-4P': {'25m': {'20230123': [3.42, 223.14, 139.99, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23189': {'MZ-4P': {'25m': {'20230123': [5.5, 192.0, 207.0, 3, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23454': {'MZ-4P': {'100m': {'20230125': [15.57, 179.6, 223.1, 10, 10, 'Iron Sight', 'White Demon']}}},
        #      '23391': {
        #          'MZ-4P': {'25m': {'20230123': [5.85, 158.06, 250.31, 13, 7, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23014': {
        #          'MZ-4P': {'25m': {'20230116': [5.28, 176.81, 142.1, 10, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23438': {
        #          'MZ-4P': {'25m': {'20230123': [3.63, 200.71, 200.36, 15, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23375': {
        #          'MZ-4P': {'25m': {'20230123': [5.53, 162.5, 133.0, 5, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23468': {
        #          'MZ-4P': {'25m': {'20230123': [7.62, 240.09, 85.03, 9, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23441': {
        #          'MZ-4P': {'25m': {'20230123': [3.85, 201.64, 205.82, 21, 5, 'Iron Sight', 'DS Zeroing A4 Target']}}},
        #      '23162': {
        #          'MZ-4P': {'25m': {'20230118': [9.02, 197.45, 206.53, 20, 4, 'Iron Sight', 'DS Zeroing A4 Target']}}}}
        # }
        data = data_["data"]
        model_ = apps.get_model(app_label=app_, model_name="doubleshootmembers")
        for k in data:
            print(k, "\n", data[k])
            obj = model_.objects.get(ds_name=k)
            obj.data = data[k]
            obj.save()
        result = {"status": "ok"}
        # print(result)
        return result

    def test(self, dic):
        print("test\n", "-"*50, "\n", dic, "\n", "-"*50)
        json_data = {'pageSize': 1,'currentPage': 0,}
        response = requests.post(self.head_url, headers=self.head, json=json_data, )
        if response.status_code == 200:
          response_json = response.json()
          members = response_json['list']
          members_id = []
          for member in members:
            member_id = member.get('id')
            members_id.append(member_id)
          print(members_id)
          json_data_heat = {'pageSize': 1, 'currentPage': 0, 'memberIds': members_id}
          response = requests.post(self.head_heat_url, headers=self.head_heat, json=json_data_heat, )
          if response.status_code == 200:
              list_ = response.json()
              print(list_)
          else:
              print("Request failed-heats. Status code:", response.status_code)
        else:
          print("Request failed. Status code:", response.status_code)
        result = {"status": "ok"}
        # print(result)
        return result


class BaseTrainingAlgo(object):
    def __init__(self, dic):  # to_data_path, target_field
        # print("90050-01 BaseTrainingAlgo", dic, '\n', '-'*50)
        # super(BaseTrainingAlgo, self).__init__()
        # print("90050-02 BaseTrainingAlgo", dic, '\n', '-'*50)
        app_ = dic["app"]
        self.excel_dir = settings.MEDIA_ROOT + '/'+app_+'/excel'
        os.makedirs(self.excel_dir, exist_ok=True)
        self.save_to_file = None
        self.second_time_save = ''

    def save_to_excel(self, df, folder, file_name=None):
        if file_name:
            self.save_to_file = os.path.join(self.excel_dir, file_name)
        # print(self.save_to_file)
        try:
            # create a Path object with the path to the file
            path = Path(self.save_to_file)
            if not path.is_file():
                wb2 = Workbook()
                wb2.save(self.save_to_file)
                wb2.close()
            else:
                wb = load_workbook(filename=self.save_to_file, read_only=False)
                sheet_names = wb.sheetnames
                for f in sheet_names:
                    if f == folder:
                        wb.remove(wb[folder])
                        wb.save(self.save_to_file)
                        break
            df2 = df.copy()
            total, used, free = shutil.disk_usage("/")
        except Exception as ee:
            print("90555-52 Error objects save_to_excel: "+str(ee))
        nnn = 0
        try:
            with pd.ExcelWriter(self.save_to_file, engine='openpyxl', mode='a') as writer_:
                df2.to_excel(writer_, sheet_name=folder)
                writer_.save()
                time.sleep(5)
            if self.second_time_save != '':
                print("save ok:", self.second_time_save)
            self.second_time_save = ''
            nnn = 1
            # print("OK")
        except Exception as ee:
            print("90555-12 Error objects save_to_excel: "+str(ee))
            time.sleep(5)
            self.save_to_excel(df2, folder)
            self.second_time_save = self.save_to_file
            nnn = 1
        finally:
            if nnn == 0:
                print(self.save_to_file + ' 55 finally -' + str(nnn) + ' - ' + folder)
                time.sleep(5)
                print(self.save_to_file + ' 551 finally -' + str(nnn) + ' - ' + folder)
                self.save_to_excel(df2, folder)
                self.second_time_save = self.save_to_file


# a = {"112": {"data": {
#                 "5516": {"data": {"5517": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5518": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5519": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5520": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5521": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1}
#                                   },
#                          "title": "ALPHA", "wet_day": 0, "raid_day": 0, "wet_night": 0, "raid_night": 0},
#                 "5522": {"data": {"5523": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5524": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5525": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5532": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5539": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5540": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5541": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1}
#                                   },
#                          "title": "DELTA", "wet_day": 0, "raid_day": 0, "wet_night": 0, "raid_night": 0},
#                 "5526": {"data": {"5527": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5528": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5529": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5530": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5531": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1}
#                                   },
#                          "title": "BRAVO", "wet_day": 0, "raid_day": 0, "wet_night": 0, "raid_night": 0},
#                 "5533": {"data": {"5534": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5535": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5536": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5537": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1},
#                                   "5538": {"wet_day": 1, "raid_day": 1, "wet_night": 1, "raid_night": 1}
#                                   },
#                          "title": "CHARLIE", "wet_day": 0, "raid_day": 0, "wet_night": 0, "raid_night": 0}
#                 },
#              "title": "battalion 112", "wet_day": 0, "raid_day": 0, "wet_night": 0, "raid_night": 0}
#      }
#
# b = {"112": {"data": {
#                 "5516": {"data": {"5517": {"is_equiped": 1},
#                                   "5518": {"is_equiped": 1},
#                                   "5519": {"is_equiped": 1},
#                                   "5520": {"is_equiped": 1},
#                                   "5521": {"is_equiped": 1}
#                                   },
#                          "is_equiped": 0},
#                 "5522": {"data": {"5523": {"is_equiped": 1},
#                                   "5524": {"is_equiped": 1},
#                                   "5525": {"is_equiped": 1},
#                                   "5532": {"is_equiped": 1},
#                                   "5539": {"is_equiped": 1},
#                                   "5540": {"is_equiped": 1},
#                                   "5541": {"is_equiped": 1}},
#                          "is_equiped": 0},
#                 "5526": {"data": {"5527": {"is_equiped": 1},
#                                   "5528": {"is_equiped": 1},
#                                   "5529": {"is_equiped": 1},
#                                   "5530": {"is_equiped": 1},
#                                   "5531": {"is_equiped": 1}},
#                          "is_equiped": 1},
#                 "5533": {"data": {"5534": {"is_equiped": 1},
#                                   "5535": {"is_equiped": 1},
#                                   "5536": {"is_equiped": 1},
#                                   "5537": {"is_equiped": 1},
#                                   "5538": {"is_equiped": 1}},
#                          "is_equiped": 1}}, "is_equiped": 0}
#      }

class TrainingDataProcessing(BaseDataProcessing, BaseTrainingAlgo):
    def __init__(self, dic):
        super().__init__(dic)

        print("="*100, "\n", dic, "\n", "="*100)

        self.df_positions = pd.DataFrame.from_dict({"id" :[1,2,3,4,5,0],
                                          "position_name": ["Captain","Officer","Soldier","Colonel","Sous Officer","Other"]})

        self.df_instructor_positions = pd.DataFrame.from_dict({"id" :[1,2,3,4,5,6,7,8],
                                          "position_name": ['מפקד גדוד', 'מפקד פלוגה', 'מספר 2 – צ', 'מוביל צוות', 'מספר 2',
                                                            'אימון גופני', 'מתגבר', 'קמ”ג']})
        self.inventory_with_pn = ["mz4psn", "ramonsn", "mz10", "mz15", "negev"]

    # To be deleted
    def process_period_1(self, app_, n, nav_tables_, result, qid=None, battalion=None):
        nn = len(nav_tables_)
        if n < nn:
            model_ = apps.get_model(app_label=app_, model_name=nav_tables_[n]+"s")
            p_key_field_name = model_._meta.pk.name
            if qid:
                model_name = nav_tables_[n-1]
                model_p = apps.get_model(app_label=app_, model_name=model_name+"s").objects.get(id=qid)
                qs = eval("model_.objects.filter("+model_name+"=model_p).all()")
            else:
                qs = model_.objects.filter(id=battalion).all()
            n += 1
            if nav_tables_[n-1] == "soldier":
                model_periods = apps.get_model(app_label=app_, model_name="periods")
                period_obj, is_created = model_periods.objects.get_or_create(battalion__id=battalion, period_number=1)
                model_soldiers = apps.get_model(app_label=app_, model_name="soldiers")
                model_unit_soldiers = apps.get_model(app_label=app_, model_name="unitsoldiers")
                df = pd.DataFrame(list(qs.values('user_id', 'platoon_id')))
                # print(df)
                for index, row in df.iterrows():
                    soldier_obj = model_soldiers.objects.get(user_id=row["user_id"])
                    u_obj, is_created = model_unit_soldiers.objects.get_or_create(period=period_obj, soldier=soldier_obj)
                    u_obj.unit_number = row["platoon_id"]
                    u_obj.save()
                return result
            for q in qs:
                if p_key_field_name=="user":
                    qid_ = eval("q."+p_key_field_name+".id")
                else:
                    qid_ = eval("q."+p_key_field_name)
                if n < nn-1:
                    n_ = self.get_next_number({"app": app_})
                else:
                    n_ = qid_
                result[n_]={"title":str(q), "data":{}}
                self.process_period_1(app_, n, nav_tables_, result[n_]["data"], qid=qid_, battalion=battalion)
        return result

    # To be deleted
    def get_units_structure_battallion_1_Period_1(self, dic):
        # print("\n", "-"*50, '\n90035-11 dic\n', dic, "\n", "-"*50)
        app_ = dic["app"]
        battalion_ = dic["battalion"]
        period_ = dic["period"]
        #
        groups_ = dic["groups"]
        nav_tables_ = groups_["nav_tables"]
        result = {"title":"title-top", "data":{}}

        result = self.process_period_1(app_, 0, nav_tables_, result["data"], battalion=battalion_)
        model_periods = apps.get_model(app_label=app_, model_name="periods")
        period_obj, is_created = model_periods.objects.get_or_create(battalion__id=battalion_, period_number=period_)
        period_obj.structure = result
        period_obj.period_name = "Battalion: " + str(battalion_) +" Period: " + str(period_)
        period_obj.save()

        # print("-"*100, "\n", result, "\n", "-"*100)
        result = {"status": "ok", "result":result}
        return result

    def process_weeks_1_5(self, app_, n, nav_tables_, result, qid=None, battalion=None):
        nn = len(nav_tables_)
        if n < nn:
            title = nav_tables_[n]
            # print(title)
            # result[nav_tables_[n]]={}
            model_ = apps.get_model(app_label=app_, model_name=nav_tables_[n]+"s")
            p_key_field_name = model_._meta.pk.name
            if qid:
                model_name = nav_tables_[n-1]
                # print(title, p_key_field_name, model_name)
                model_p = apps.get_model(app_label=app_, model_name=model_name+"s").objects.get(id=qid)
                qs = eval("model_.objects.filter("+model_name+"=model_p).all()")
            else:
                qs = model_.objects.filter(id=battalion).all()
            n += 1
            for q in qs:
                if p_key_field_name=="user":
                    qid_ = eval("q."+p_key_field_name+".id")
                    # print("q." + p_key_field_name, model_, qid_, q.userid)
                else:
                    qid_ = eval("q."+p_key_field_name)
                result[qid_]={"title":str(q), "model":nav_tables_[n-1], "data":{}}
                self.process_weeks_1_5(app_, n, nav_tables_, result[qid_]["data"], qid=qid_, battalion=battalion)
        return result

    def process_weeks_6_10(self, app_, weeks, weeks_dic, result, qid=None):
        model_ = apps.get_model(app_label=app_, model_name="unitesoldiers")
        objs = model_.objects.filter(weeks = weeks).all()

        print(weeks_dic)

        return weeks_dic

    def get_units_structure(self, dic):
        # print("\n", "-"*50, '\n90035-12 dic\n', dic, "\n", "-"*50)
        app_ = dic["app"]
        weeks_ = dic["weeks"]
        battalion_ = dic["battalion"]
        #
        groups_ = dic["groups"]
        nav_tables_ = groups_["nav_tables"]
        result = {"title":"title-top", "data":{}}
        if weeks_ == "1-5":
            result = self.process_weeks_1_5(app_, 0, nav_tables_, result["data"], battalion=battalion_)

        # print("-"*100, "\n", result, "\n", "-"*100)
        result = {"status": "ok", "result":result}
        return result

    def get_units_structure_new(self, dic):
        print("\n", "-"*50, '\n90035-1 dic\n', dic, "\n", "-"*50)
        app_ = dic["app"]
        battalion_ = int(dic["battalion"])
        # soldiers
        try:
            soldiers = {"id":[], "userid":[], "first_name":[], "last_name":[], "name":[]}
            model_soldiers = apps.get_model(app_label=app_, model_name="soldiers")
            qs = model_soldiers.objects.filter(battalion__id=battalion_).all()
            # print(qs)

            df_s = pd.DataFrame(list(qs.values("user_id", "userid", "first_name", "last_name")))
            # print(df_s)
            for index, row in df_s.iterrows():
                soldiers["id"].append(str(row["user_id"]))
                soldiers["userid"].append(str(row["userid"]))
                soldiers["first_name"].append(str(row["first_name"]))
                soldiers["last_name"].append(str(row["last_name"]))
                soldiers["name"].append(str(row["first_name"])+" "+str(row["last_name"]))
        except Exception as ex:
            print("Error 1: "+str(ex))

        # unitsoldiers
        try:
            unitsoldiers = {}
            model_unitsoldiers = apps.get_model(app_label=app_, model_name="unitsoldiers")
            qs = model_unitsoldiers.objects.filter(period__battalion__id=battalion_).all()
            df_us = pd.DataFrame(list(qs.values()))
            # print(df_us)
            period_id_ = []
            unit_number_ = []
            for index, row in df_us.iterrows():
                if int(row["period_id"]) not in period_id_:
                    # print(int(row["period_id"]))
                    unitsoldiers[int(row["period_id"])] = {}
                    period_id_.append(int(row["period_id"]))
                if int(row["unit_number"]) not in unit_number_:
                    unitsoldiers[int(row["period_id"])][int(row["unit_number"])] = []
                    unit_number_.append(int(row["unit_number"]))
                unitsoldiers[int(row["period_id"])][int(row["unit_number"])].append(int(row["soldier_id"]))
            # print(n, n1)

            # print(period_id_, "\n", len(unit_number_))
            # print(unitsoldiers)
        except Exception as ex:
            print("Error 2: "+str(ex))

        # structure periods
        try:
            structure = {}
            period = {}
            model_periods = apps.get_model(app_label=app_, model_name="periods")
            period_objs = model_periods.objects.filter(battalion__id=battalion_).all()
            # print(period_objs)
            for period_obj in period_objs:
                structure[period_obj.id] = period_obj.structure
                period[period_obj.id] = {"n_limit": period_obj.n_limit}
        except Exception as ex:
            print("Error 3: "+str(ex))
        # print("-"*100, "\n", structure, "\n", "-"*100)

        result = {"status": "ok", "soldiers": soldiers, "structure":structure, "unitsoldiers":unitsoldiers,
                  "period":period}
        # print(result)
        return result

    def set_instructors(self, dic):
        # print('90088-17 dic', dic)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # print("-"*100, "\n", file_path, "\n", "-"*100)
        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        # print(df)
        # print("-"*100)
        model_instructors = apps.get_model(app_label=app_, model_name="instructors")
        model_companys = apps.get_model(app_label=app_, model_name="companys")
        model_platoons = apps.get_model(app_label=app_, model_name="platoons")
        model_battalions = apps.get_model(app_label=app_, model_name="battalions")
        my_group, is_created = Group.objects.get_or_create(name='t_simple_user')

        try:
            u = User.objects.get(first_name="admin", last_name="admin")
            print(u)
            instructor_obj, is_created = model_instructors.objects.get_or_create(user=u)
            instructor_obj.first_name = "admin"
            instructor_obj.last_name = "admin"
            # instructor_obj.position = 0
            instructor_obj.save()
        except Exception as ex:
            print("9001-01 Error " + str(ex))

        for index, row in df.iterrows():
            company_name_ = str(row["company_name"]).upper()
            first_name_ = str(row["first_name"])
            last_name_ = str(row["last_name"])
            username_ = str(row["username"])
            password_ = str(row["password"])
            position_name_ = str(row["position_name"])
            position_ = int(self.df_instructor_positions[self.df_instructor_positions["position_name"]==position_name_]["id"])
            # print("9088-1", company_name_, first_name_, " ", last_name_, position_name_)

            try:
                u = User.objects.get(username=username_)
                count = u.delete()
                # print("B count\n", count, "\n")
            except Exception as ex:
                # pass
                print("9055-55 Error " + str(ex))
            try:
                u = User.objects.create_user(username=username_, email=username_+'@gmail.com',
                                             password=password_)
                # print(u.password)
                u.first_name = first_name_
                u.last_name = last_name_
                u.save()
                my_group.user_set.add(u)
                my_group.save()
            except Exception as ex:
                print("9000-00 Error " + str(ex))
            try:
                instructor_obj, is_created = model_instructors.objects.get_or_create(user=u)
                instructor_obj.first_name = first_name_
                instructor_obj.last_name = last_name_
                instructor_obj.position = position_
                instructor_obj.save()
            except Exception as ex:
                print("9001-01 Error " + str(ex))

            print("Record 9088-3   ", company_name_, instructor_obj, "  =", is_created)

            try:
                if company_name_ != "Battalion".upper():
                    # print(company_name_)
                    company_obj = model_companys.objects.get(company_name=company_name_)
                    if str(row["platoon_number"]).lower() != "nan":
                        platoon_number = float(str(row["platoon_number"]))
                        platoon_obj = model_platoons.objects.get(platoon_number=platoon_number, company=company_obj)
                        platoon_obj.instructor.add(instructor_obj)
                        if position_ == 2:
                            company_obj.instructor.add(instructor_obj)
                        platoon_obj.save()
                    else:
                        # print("no platoon")
                        company_obj.instructor.add(instructor_obj)
                        company_obj.save()
                else:
                    try:
                        battalion_obj, is_created = model_battalions.objects.get_or_create(battalion_number=1)
                    except Exception as ex:
                        print("9011-1111 Error " + str(ex))
                    try:
                        battalion_obj.instructor.add(instructor_obj)
                    except Exception as ex:
                        print("9011-1122 Error " + str(ex))
                    try:
                        battalion_obj.save()
                    except Exception as ex:
                        print("9011-1133 Error " + str(ex))
            except Exception as ex:
                print("9022-22 Error " + str(ex))

        print("Done")
        result = {"status": "ok"}
        return result

    # v2 --
    def set_soldiers_and_org_structure(self, dic):
        print('90022-7 dic', dic)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # print("-"*100, "\n", file_path, "\n", "-"*100)
        battalion_number_ = int(self.uploaded_filename.split(".")[0])
        model_name_ = "battalions"
        model_periods = apps.get_model(app_label=app_, model_name="periods")
        model_battalions = apps.get_model(app_label=app_, model_name=model_name_)
        model_unit_soldiers = apps.get_model(app_label=app_, model_name="unitsoldiers")
        model_soldiers = apps.get_model(app_label=app_, model_name="soldiers")
        model_soldierqualificationfact = apps.get_model(app_label=app_, model_name="soldierqualificationfact")
        # model_name_ = "companys"
        # model_companys = apps.get_model(app_label=app_, model_name=model_name_)
        # model_name_ = "platoons"
        # model_platoons = apps.get_model(app_label=app_, model_name=model_name_)
        battalion_obj, is_created = model_battalions.objects.get_or_create(battalion_number=battalion_number_)
        if is_created:
            battalion_obj.battalion_name = "Battalion " + str(battalion_number_)
            battalion_obj.save()
        period_obj, is_created = model_periods.objects.get_or_create(battalion=battalion_obj, period_number=1)
        if is_created:
            period_obj.period_name = "Battalion: "+ str(battalion_number_) + " Period " + str(1)
            period_obj.save()
        n__ = self.get_next_number({"app": app_})
        units_dic = {n__: {'title': battalion_obj.battalion_name, 'data': {}}}
        units_dic_b = units_dic[n__]["data"]

        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        # print(df)
        # print("-"*100)
        ll_c = []
        ll_p = []
        for index, row in df.iterrows():
            print(index)
            # if index < 545:
            #     continue
            # print(row, "\n", row["company"])
            company_name_ = str(row["company"]).upper()
            company_number = int(row["company_number"])
            if company_number not in ll_c:
                ll_c.append(company_number)
                n__ = self.get_next_number({"app": app_})
                units_dic_b[n__] = {}
                units_dic_c = units_dic_b[n__]
                units_dic_c['title'] = company_name_
                units_dic_c['data'] = {}
            platoon_number = int(row["platoon"])
            platoon_name_ = company_name_ + " " +str(platoon_number)
            if platoon_name_ not in ll_p:
                ll_p.append(platoon_name_)
                n__ = self.get_next_number({"app": app_})
                units_dic_c['data'][n__] = {}
                units_dic_p = units_dic_c['data'][n__]
                units_dic_p['title'] = platoon_name_
                units_dic_p['data'] = {}
            username_ = "u"+str(row["dshn"])
            u_sername_ = "U"+str(row["dshn"])
            userid = str(row["dshn"])
            uniqueid = str(row["dshn"])
            try:
                u = User.objects.get(username=username_)
                count = u.delete()
                # print("B count\n", count, "\n")
            except Exception as ex:
                pass
                # print("9055-55 Error " + str(ex))
            full_name = string.capwords(str(row["full_name"]))
            nnf = full_name.find(" ")
            first_name = full_name[:nnf]
            last_name = full_name[nnf+1:]
            #
            my_group, is_created = Group.objects.get_or_create(name='t_simple_user')
            # print(my_group)
            # Users
            try:
                u = User.objects.create_user(username=username_, email=username_+'@gmail.com', password=u_sername_+'#')
                # print(u.password)
                u.first_name = first_name
                u.last_name = last_name
                u.save()
                # print(u)
                my_group.user_set.add(u)
                my_group.save()
            except Exception as ex:
                print("9011-11-1 Error " + str(ex))
            # Soldiers
            try:
                soldier_obj, is_created = model_soldiers.objects.get_or_create(user=u)
                soldier_obj.battalion=battalion_obj
                soldier_obj.first_name = first_name
                soldier_obj.last_name = last_name
                soldier_obj.userid = userid
                soldier_obj.uniqueid = uniqueid
                soldier_obj.save()
            except Exception as ex:
                print("9011-22-1 Error " + str(ex))

            try:
                u_obj, is_created = model_unit_soldiers.objects.get_or_create(period=period_obj, soldier=soldier_obj)
                u_obj.unit_number = n__
                u_obj.save()
                # print("saved: "+ str(n__))
            except Exception as ex:
                print("error 200: ", full_name)

            self.set_basic_soldier_data(row, soldier_obj, battalion_obj, model_soldierqualificationfact)

        # -----
        # print(units_dic)
        # print("\n", "Done")

        period_obj.structure=units_dic
        period_obj.save()
        result = {"status": "ok"}
        return result

    def update_soldiers_info(self, dic):
        print('90022-2 dic', dic)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # print("-"*100, "\n", file_path, "\n", "-"*100)
        #
        model_inventorys = apps.get_model(app_label=app_, model_name="inventorys")
        model_inventoryfact = apps.get_model(app_label=app_, model_name="inventoryfact")
        #
        # print(self.uploaded_filename)
        s = self.uploaded_filename.split("_")
        # print(s)
        battalion_number_ = int(s[0])
        period_number_ = int(s[1])
        # print(battalion_number_, period_number_)
        #
        model_periods = apps.get_model(app_label=app_, model_name="periods")
        model_battalions = apps.get_model(app_label=app_, model_name="battalions")
        model_unit_soldiers = apps.get_model(app_label=app_, model_name="unitsoldiers")
        model_soldiers = apps.get_model(app_label=app_, model_name="soldiers")
        model_soldierqualificationfact = apps.get_model(app_label=app_, model_name="soldierqualificationfact")
        battalion_obj = model_battalions.objects.get(battalion_number=battalion_number_)
        period_obj = model_periods.objects.get(battalion=battalion_obj, period_number=period_number_)
        units_dic = period_obj.structure
        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        # print(df, "\n", "-"*100)

        columns = df.columns[13:]
        # print("columns\n", columns)

        try:
            objs = model_soldierqualificationfact.objects.filter(soldier__battalion=battalion_obj, skill=2).all()
            objs.delete()
            objs = model_soldierqualificationfact.objects.filter(soldier__battalion=battalion_obj, skill=3).all()
            objs.delete()
            objs = model_unit_soldiers.objects.filter(period=period_obj).all()
            objs.delete()
            objs = model_inventoryfact.objects.filter(soldier__battalion=battalion_obj).all()
            objs.delete()
        except Exception as ex:
            print(ex)

        for index, row in df.iterrows():
            # ------
            company_name_ = str(row["company"]).upper()
            userid = str(row["dshn"])
            username_ = "u"+userid
            print(index, userid)

            platoon_number = row["platoon"]
            platoon_name_ = company_name_ + " " +str(platoon_number)
            uniqueid = str(row["dshn"])
            full_name = string.capwords(str(row["full_name"]))
            nnf = full_name.find(" ")
            first_name = full_name[:nnf]
            last_name = full_name[nnf+1:]
            status = row["status"]

            # User
            try:
                u = User.objects.get(username=username_)
                u.first_name = first_name
                u.last_name = last_name
                u.save()
            except Exception as ex:
                print("9011-11-2 Error ", username_, str(ex))
                try:
                    u_username_ = "U"+userid
                    my_group, is_created = Group.objects.get_or_create(name='t_simple_user')
                    u = User.objects.create_user(username=username_, email=username_ + '@gmail.com',
                                                 password=u_username_ + '#')
                    # print(u.password)
                    u.first_name = first_name
                    u.last_name = last_name
                    u.save()
                    # print(u)
                    my_group.user_set.add(u)
                    my_group.save()
                except Exception as ex:
                    print("9011-11-3 Error " + str(ex))

            # Soldiers
            try:
                soldier_obj, is_created = model_soldiers.objects.get_or_create(user=u)
                soldier_obj.battalion = battalion_obj
                soldier_obj.userid = userid
                soldier_obj.uniqueid = uniqueid
                soldier_obj.save()
            except Exception as ex:
                print("9011-22-1 Error " + str(ex))

            if status == -1:
                try:
                    print("the following soldier was deleted")
                    print("-"*100, "\n", status, soldier_obj, "Delete Soldier", "\n", "-"*100)
                    soldier_obj.delete()
                except Exception as ex:
                    pass
                continue
            if status == 1:
                print("the following soldier was replaced by another soldier")
                print("soldier_obj", soldier_obj, "\n", "-"*100)
                try:
                    # print("soldier_obj.user_id=", soldier_obj.user_id)
                    objs = model_soldiers_for_events.objects.filter(soldier_number=soldier_obj.user_id).all()
                    count = objs.delete()
                    # print(count)
                except Exception as ex:
                    pass

            self.set_basic_soldier_data(row, soldier_obj, battalion_obj, model_soldierqualificationfact)

            self.set_unit_soldier(soldier_obj, units_dic, platoon_name_, period_obj, model_unit_soldiers)

            self.set_soldier_inventory(row, columns, soldier_obj, model_inventorys, model_inventoryfact,
                                       battalion_obj, model_soldierqualificationfact)

        result = {"status": "ok"}
        print(result)
        return result

    def set_units_equipment(self, dic):
        print('90022-2 dic\n', dic)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # print("-"*100, "\n", file_path, "\n", "-"*100)
        #
        model_inventorys = apps.get_model(app_label=app_, model_name="inventorys")
        model_inventoryunitfact = apps.get_model(app_label=app_, model_name="inventoryunitfact")
        #
        # print(self.uploaded_filename)
        s = self.uploaded_filename.split("_")
        # print(s)
        battalion_number_ = int(s[0])
        period_number_ = int(s[1])
        print(battalion_number_, period_number_)
        #
        model_periods = apps.get_model(app_label=app_, model_name="periods")
        model_battalions = apps.get_model(app_label=app_, model_name="battalions")

        battalion_obj = model_battalions.objects.get(battalion_number=battalion_number_)
        period_obj = model_periods.objects.get(battalion=battalion_obj, period_number=period_number_)
        def get_list_of_units(ll_, dic_):
            for k in dic_:
                # print(k, dic_[k]["title"])
                ll_.append(k)
                get_list_of_units(ll_, dic_[k]["data"])
            return ll

        units_dic = period_obj.structure
        # print(units_dic)

        ll = []
        get_list_of_units(ll, units_dic)
        # print("="*50, "\n", ll, "\n", "="*50)
        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        print(df, "\n", "-"*100)
        columns = df.columns[3:]
        # print("columns\n", columns)

        try:
            objs = model_inventoryunitfact.objects.filter(unit__in=ll).all()
            objs.delete()
        except Exception as ex:
            print(ex)

        dic = {}
        for index, row in df.iterrows():
            # ------
            battalion_ = str(row["battalion"]).upper()
            battalion_title = "Battalion " + battalion_
            battalion_unit_number = self.get_unit_number(units_dic, battalion_title)
            company_ = str(row["company"]).upper()
            platoon_ = str(row["platoon"]).upper()
            # print("\nrecord A=\n", battalion_, company_, platoon_)
            unit_type_ = 0
            if platoon_ != "NAN":
                company_title = company_
                company_unit_number = self.get_unit_number(units_dic, company_title)
                platoon_title = company_ + " " + platoon_
                title = platoon_title
                platoon_unit_number = self.get_unit_number(units_dic, platoon_title)
                # print(battalion_unit_number, company_title, company_unit_number, platoon_title, platoon_unit_number, dic)
                unit_type_ = 3
                if platoon_unit_number not in dic[battalion_unit_number]["data"][company_unit_number]["data"]:
                    dic[battalion_unit_number]["data"][company_unit_number]["data"][platoon_unit_number] = {}
                platoon_unit__ = dic[battalion_unit_number]["data"][company_unit_number]["data"][platoon_unit_number]
                company_unit__ = dic[battalion_unit_number]["data"][company_unit_number]
                unit__ = platoon_unit__
                unit_number = platoon_unit_number
            elif company_ != "NAN":
                company_title = company_
                title = company_title
                company_unit_number = self.get_unit_number(units_dic, company_title)
                unit_type_ = 2
                if company_unit_number not in dic[battalion_unit_number]["data"]:
                    dic[battalion_unit_number]["data"][company_unit_number] = {"data":{}}
                platoon_unit__ = None
                company_unit__ = dic[battalion_unit_number]["data"][company_unit_number]
                unit__ = company_unit__
                unit_number = company_unit_number
            else:
                unit_type_ = 1
                # print(battalion_unit_number)
                if battalion_unit_number not in dic:
                    dic[battalion_unit_number] = {"data":{}}
                platoon_unit__ = None
                company_unit__ = None
                unit__ = dic[battalion_unit_number]
                unit_number = battalion_unit_number
                title = battalion_title
            battalion_unit__ = dic[battalion_unit_number]

            # print("\nrecord B=\nbattalion_unit__\n", battalion_unit__, "\ncompany_unit__\n", company_unit__, "\nplatoon_unit__\n", platoon_unit__, "\nunit__=\n", unit__)

            is_equiped = 1
            for k in columns:
                # print("\n", "-"*20, "\n", k, str(row[k]), "\n", "-"*20)
                try:
                    v = str(row[k])
                    # print("1","k=", k, " v=", "="+v+"=")
                    if v == "" or v == "nan":
                        continue
                    # print("2","k=", k, " v=", "="+v+"=")
                    kl = k.lower()
                    if kl in self.inventory_with_pn:
                        v = 1
                        # print("3","k=", k, " v=", "="+str(v)+"=")
                    else:
                        # print("1","k=", k, " v=", "="+str(v)+"=")
                        v = int(float(v))
                    # print("4\n", "k=", k, " v=", "="+str(v)+"=","\n", "-"*100)
                except Exception as ex:
                    print("9011-77-77-1 Error " + str(ex))
                try:
                    inventory_obj = model_inventorys.objects.get(item_name=kl)
                except Exception as ex:
                    print("9011-55-1 Error ", k, str(ex))
                try:
                    f_obj, is_created = model_inventoryunitfact.objects.get_or_create(inventory=inventory_obj,
                                                                                      unit=unit_number)
                    f_obj.value = v
                    f_obj.save()
                    unit_critical = int(inventory_obj.unit_critical)
                    q_ = inventory_obj.qty_per_soldier
                    # print("AA == ", kl, "v=", v, "q_=", q_, "unit_critical=", unit_critical, "unit_type_=", unit_type_)
                    if unit_critical == 0 or unit_critical != unit_type_:
                        continue
                    if kl == "drone_t1" and title.lower()=="delta":
                        q_ -= 1
                    if kl == "drone_t2" and unit_type_ == 2 and title.lower()=="delta":
                        q_ = 0
                    elif kl == "drone_t2" and unit_type_ == 1:
                        q_ = 1
                    if v < q_:
                        is_equiped = 0
                except Exception as ex:
                    print("9011-55-2 Error \n", k, v, str(ex))
            unit__["is_equiped"] = is_equiped
            print(title, unit_type_, is_equiped)

        general_data_model = apps.get_model(app_label="core", model_name="generaldata")
        group_ = s[0] + "_" + str(s[1])
        data_name_ = "unit_equipment"
        obj, is_created = general_data_model.objects.get_or_create(app=app_, group=group_, data_name=data_name_)
        obj.data_json = dic
        obj.save()

        #     try:
        #         is_equiped = 1
        #         inventory_objs = model_inventorys.objects.all()
        #         for q in inventory_objs:
        #             k = q.item_name
        #             try:
        #                 v = str(row[k])
        #                 if v == "" or v == "nan":
        #                     v = 0
        #                 else:
        #                     if k in self.inventory_with_pn:
        #                         v = 1
        #                     else:
        #                         v = int(float(v))
        #             except Exception as ex:
        #                 pass
        #                 # print("9011-77-77-2 Error " + str(ex))
        #             qty_per_unit = 1 #  maby we need to add a column for qty_per_unit in the table q.qty_per_soldier
        #             unit_critical = q.unit_critical
        #             if unit_critical == 0 or unit_critical != unit_type_:
        #                 continue
        #             if v < qty_per_unit:
        #                 is_equiped = 0
        #                 break
        #
        #         if is_equiped == 1:
        #             unit__["is_equiped"] = is_equiped
        #
        #
            # except Exception as ex:
            #     print("9011-55-2 Error \n", k, v, soldier_obj, str(ex))
        #
        #
        #
        #
        #
        #             k = self.get_unit_number(units_dic, company)
        #             if k not in battalion_dic["data"]:
        #                 battalion_dic["data"][k]={"title": company, "wet_day":0, "wet_night":0, "raid_day":0, "raid_night":0, "data":{}}
        #                 company_dic = battalion_dic["data"][k]
        #                 company_dic["wet_day"] = wet_day
        #                 company_dic["wet_night"] = wet_night
        #                 company_dic["raid_day"] = raid_day
        #                 company_dic["raid_night"] = raid_night
        #
        #                 battalion_dic["wet_day"] *= wet_day
        #                 battalion_dic["wet_night"] *= wet_night
        #                 battalion_dic["raid_day"] *= raid_day
        #                 battalion_dic["raid_night"] *= raid_night
        #
        #         else:
        #             continue
        #         company_dic = battalion_dic["data"][k]
        #         if platoon != "nan":
        #             platoon_name_ = company + " " + platoon
        #             p = self.get_unit_number(units_dic, platoon_name_)
        #             print(p, platoon_name_)
        #             if p not in company_dic["data"]:
        #                 company_dic["data"][p]={}
        #                 platoon_dic = company_dic["data"][p]
        #             platoon_dic["wet_day"] = wet_day
        #             platoon_dic["wet_night"] = wet_night
        #             platoon_dic["raid_day"] = raid_day
        #             platoon_dic["raid_night"] = raid_night
        #
        #             company_dic["wet_day"] *= wet_day
        #             company_dic["wet_night"] *= wet_night
        #             company_dic["raid_day"] *= raid_day
        #             company_dic["raid_night"] *= raid_night
        #
        #             dic[battalion_number]["wet_day"] *= wet_day
        #             dic[battalion_number]["wet_night"] *= wet_night
        #             dic[battalion_number]["raid_day"] *= raid_day
        #             dic[battalion_number]["raid_night"] *= raid_night
        #         else:
        #             continue
        #
        #     general_data_model = apps.get_model(app_label="core", model_name="generaldata")
        #     group_ = s[0]+"_"+str(s[1])
        #     data_name_ = "unit_qualification"
        #     print(group_, "\n", data_name_, "\n", dic)
        #     obj, is_created = general_data_model.objects.get_or_create(app=app_, group=group_, data_name=data_name_)
        #     obj.data_json = dic
        #     obj.save()

        print(group_, "\n", data_name_, "\n", dic)

        result = {"status": "ok"}
        # print(result)
        return result

    def upload_inventory_list(self, dic):
        print('90055-1 upload_inventory_list dic\n', dic)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # print("-"*100, "\n", file_path, "\n", "-"*100)
        model_inventorys = apps.get_model(app_label=app_, model_name="inventorys")
        model_inventorycategorys = apps.get_model(app_label=app_, model_name="inventorycategorys")

        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        # print(df)
        # print("-"*100)
        shoes_size = ["40", "41", "42", "43", "44", "45"]
        shoes_size_ll = ["running_shoes", "boots"]
        clothes_size = ["XL", "L", "M", "S"]
        clothes_size_ll = ["sport_shirt", "sport_short", "working_uniform", "tactical_uniform"]

        for index, row in df.iterrows():
            # print(index)
            category_obj, is_created = model_inventorycategorys.objects.get_or_create(category_name=str(row["category_name"]))
            inventory_obj, is_created = model_inventorys.objects.get_or_create(item_name=str(row["item_name"]))
            inventory_obj.inventorycategory = category_obj
            inventory_obj.item_number = int(row["item_number"])
            inventory_obj.pn = str(row["pn"])
            inventory_obj.description = str(row["description"])
            inventory_obj.critical = str(row["critical"])
            inventory_obj.unit_critical = str(row["unit_critical"])
            inventory_obj.qty_per_soldier = int(row["qty_per_soldier"])
            inventory_obj.save()
        # -----
        # print("\n", "Done")
        # -----
        result = {"status": "ok"}
        print(result)
        return result

    def set_new_structure(self, dic):
        print('90088-19 dic', dic)
        clear_log_debug()
        log_debug("Start 1")
        app_ = dic["app"]
        try:
            ll = eval(dic["sheet_name"])
            # print("battalion_id, period_number", battalion_name, battalion_id, period_number, n_limit)

            log_debug("Start 2")
            file_path = self.upload_file(dic)["file_path"]
            log_debug(file_path)
            print(file_path)

            s = self.uploaded_filename.split("_")
            battalion_number_ = int(s[0])
            period_number = int(s[2])
            battalion_name = "Battalion " + s[0]
            n_limit = int(ll[0])
            model_battalion = apps.get_model(app_label=app_, model_name="battalions")
            battalion_obj = model_battalion.objects.get(battalion_number=battalion_number_)
            battalion_id = battalion_obj.id

            units_dic = {battalion_number_: {'title': battalion_name, 'data': {}}}
            # print("s", s, "battalion_number_", battalion_number_, "period_number", period_number, "battalion_name", battalion_name, "n_limit", n_limit, "battalion_id", battalion_id)
            # print(file_path)
            df = pd.read_excel(file_path, sheet_name="Data", header=0)
            # print(df)

            # print("1 columns\n", df.columns)
            log_debug("Start 3")
            columns = df.columns[13:]
            print("2 columns\n", columns)

            try:
                model_inventorys = apps.get_model(app_label=app_, model_name="inventorys")
                model_inventoryfact = apps.get_model(app_label=app_, model_name="inventoryfact")
                model_inventoryfact.objects.filter(soldier__battalion=battalion_obj).all().delete()
            except Exception as ex:
                print(ex)
            print("Start 4")

            model_periods = apps.get_model(app_label=app_, model_name="periods")
            period_obj, is_created = model_periods.objects.get_or_create(battalion=battalion_obj, period_number=period_number)
            print("Start 5")
            if not is_created:
                # by deleting the period all records in model_unit_soldiers
                period_obj.delete()
                period_obj, is_created = model_periods.objects.get_or_create(battalion=battalion_obj,
                                                                             period_number=period_number)

            print("Start 6")
        except Exception as ex:
            print(ex)

        model_unit_soldiers = apps.get_model(app_label=app_, model_name="unitsoldiers")
        model_soldiers_for_events = apps.get_model(app_label=app_, model_name="soldiersforevents")
        model_soldiers = apps.get_model(app_label=app_, model_name="soldiers")
        model_soldierqualificationfact = apps.get_model(app_label=app_, model_name="soldierqualificationfact")

        # company
        units_dic[battalion_number_] = {"title": battalion_name, "data": {}}
        data_ = units_dic[battalion_number_]["data"]
        companies_ = []
        companies_numbers = []
        platoons_ = []
        platoons_numbers = []
        n__ = 0
        print("Start 10")
        for index, row in df.iterrows():
            n__+= 1
            company = str(row["company"])
            platoon = str(row["platoon"]).strip().upper()
            status = row["status"]
            #
            dshn = str(row["dshn"])
            print(index, dshn)
            try:
                soldier_obj = model_soldiers.objects.get(userid=str(dshn))
            except Exception as ex:
                continue
            if status == -1:
                print("the following soldier was deleted")
                print("-"*100, "\n", status, dshn, "Delete Soldier", "\n", "-"*100)
                try:
                    soldier_obj.delete()
                except Exception as ex:
                    pass
                continue
            if status == 1:
                print("the following soldier was replaced by another soldier")
                print("soldier_obj", soldier_obj, "\n", "-"*100)
                try:
                    # print("soldier_obj.user_id=", soldier_obj.user_id)
                    objs = model_soldiers_for_events.objects.filter(soldier_number=soldier_obj.user_id).all()
                    count = objs.delete()
                    # print(count)
                except Exception as ex:
                    pass

            if company not in ["ALPHA", "BRAVO", "CHARLIE", "DELTA", "WOLF", "SWORD", "VICTORY"]:
                continue
            #
            if company not in companies_:
                companies_.append(company)
                number_c = self.get_next_number({"app": app_})
                companies_numbers.append(number_c)
                data_[number_c] = {"title": company, "data": {}}
            else:
                number_c = companies_numbers[companies_.index(company)]
            data_c = data_[number_c]["data"]
            platoon_name_ = company + " " + platoon
            if platoon_name_ not in platoons_:
                # platoon
                platoons_.append(platoon_name_)
                number_p = self.get_next_number({"app": app_})
                platoons_numbers.append(number_p)
                data_c[number_p] = {"title": platoon_name_, "data": {}}
            else:
                number_p = platoons_numbers[platoons_.index(platoon_name_)]
            data_p = data_c[number_p]["data"]
            #
            self.set_basic_soldier_data(row, soldier_obj, battalion_obj, model_soldierqualificationfact)

            self.set_unit_soldier(soldier_obj, units_dic, platoon_name_, period_obj, model_unit_soldiers)

            self.set_soldier_inventory(row, columns, soldier_obj, model_inventorys, model_inventoryfact,
                                       battalion_obj, model_soldierqualificationfact)

        period_obj.structure = units_dic
        period_obj.period_name = "Battalion: " + str(battalion_id) + " Period: " + str(period_number)
        period_obj.n_limit = n_limit
        period_obj.save()
        # -----
        result = {"status": "ok", "units_dic":units_dic}
        # print(result)
        return result

    # Assisting functions
    def get_unit_number(self, units_dic_, title):
            h = -1
            for k_ in units_dic_:
                if title.upper() == units_dic_[k_]["title"].upper():
                    return int(k_)
                else:
                    h = self.get_unit_number(units_dic_[k_]["data"], title)
                if h != -1:
                    break
            return h

    def set_basic_soldier_data(self, row, soldier_obj, battalion_obj, model_soldierqualificationfact):

        full_name = string.capwords(str(row["full_name"]))
        nn__ = full_name.find(" ")
        first_name = full_name[:nn__].rstrip().lstrip()
        last_name = full_name[nn__+1:].rstrip().lstrip()
        # log_debug("index " + str(index))
        dshn = str(row["dshn"])
        s = dshn + " " + first_name + " " + last_name
        # print(s)
        # log_debug(s)
        clothes_size = str(row["clothes_size"]).strip()
        shoes_size = str(row["shoes_size"]).strip()
        mz4psn = str(row["mz4psn"]).strip()
        ramonsn = str(row["ramonsn"]).strip()
        mz10 = str(row["mz10"]).strip()
        #
        rank = str(row["rank"]).strip()
        function = str(row["function"]).strip()
        profession = int(row["profession"])
        professional_qualified = int(row["professional_qualified"])
        try:
            soldier_obj.first_name = first_name
            soldier_obj.last_name = last_name
            soldier_obj.save()
        except Exception as ex:
            print("Error 200-200-12 basic Soldier: ", ex)

        try:
            soldier_obj.rank = rank
            soldier_obj.save()
        except Exception as ex:
            print("9011-20-1 Error ", rank, str(ex))
        try:
            soldier_obj.function = function
            soldier_obj.save()
        except Exception as ex:
            print("9011-20-1 Error ", function, str(ex))
        try:
            soldier_obj.profession = profession
            soldier_obj.save()
        except Exception as ex:
            print("9011-20-1 Error ", profession, str(ex))
        try:
            soldier_obj.professional_qualified = professional_qualified
            soldier_obj.save()
        except Exception as ex:
            print("9011-20-1 Error ", professional_qualified, str(ex))

        try:
            soldier_obj.clothes_size = clothes_size
            soldier_obj.save()
        except Exception as ex:
            print("9011-22-2 Error ", clothes_size, str(ex))
        try:
            soldier_obj.shoes_size = shoes_size
            soldier_obj.save()
        except Exception as ex:
            print("9011-22-3 Error ", shoes_size, str(ex))
        try:
            soldier_obj.mz4psn = mz4psn
            soldier_obj.save()
        except Exception as ex:
            print("9011-22-4 Error ", mz4psn, str(ex))
        try:
            soldier_obj.ramonsn = ramonsn
            soldier_obj.save()
        except Exception as ex:
            print("9011-22-5 Error ", mz4psn, str(ex))
        try:
            soldier_obj.mz10 = mz10
            soldier_obj.save()
        except Exception as ex:
            print("9011-22-6 Error ", mz10, str(ex))

        # professional
        try:
            if professional_qualified != 0:
                o, is_created = model_soldierqualificationfact.objects.get_or_create(
                    training_web=battalion_obj.training_web,
                    soldier=soldier_obj, skill=2)
                o.value = professional_qualified
                o.save()
        except Exception as ex:
            print("9011-22-7 Error ", mz10, str(ex))

    def set_unit_soldier(self, soldier_obj, units_dic, platoon_name_, period_obj, model_unit_soldiers):
        try:
            u_obj, is_created = model_unit_soldiers.objects.get_or_create(period=period_obj, soldier=soldier_obj)
            k = self.get_unit_number(units_dic, platoon_name_)
            # print(u_obj.unit_number, k, platoon_name_)
            u_obj.unit_number = k
            u_obj.save()
            # print("saved: "+ str(n__))
        except Exception as ex:
            print("error 200: ", full_name)

    def set_soldier_inventory(self, row, columns, soldier_obj, model_inventorys, model_inventoryfact,
                              battalion_obj, model_soldierqualificationfact):
        # print("set_soldier_inventory")
        # Update inventory
        for k in columns:
            # print("\n", "-"*20, "\n", k, "\n", soldier_obj.userid, "\n", "-"*20)
            try:
                v = str(row[k])
                # print("1","k=", k, " v=", "="+v+"=")
                if v == "" or v == "nan":
                    continue
                # print("2","k=", k, " v=", "="+v+"=")
                kl = k.lower()
                if kl in self.inventory_with_pn:
                    v = 1
                    # print("3","k=", k, " v=", "="+str(v)+"=")
                else:
                    # print("1","k=", k, " v=", "="+str(v)+"=")
                    v = int(float(v))
                # print("4", soldier_obj.userid, "\n", "k=", k, " v=", "="+str(v)+"=","\n", "-"*100)
            except Exception as ex:
                print("9011-77-77-10 Error \n", k, v, soldier_obj, str(ex))
            try:
                inventory_obj = model_inventorys.objects.get(item_name=k)
            except Exception as ex:
                print("9011-55-1 Error \n", k, v, soldier_obj, str(ex))
            try:
                f_obj, is_created = model_inventoryfact.objects.get_or_create(inventory=inventory_obj,
                                                                              soldier=soldier_obj)
                f_obj.value = v
                f_obj.save()
                # print("f_obj", f_obj)
            except Exception as ex:
                print("9011-55-20 Error \n", k, v, soldier_obj, str(ex))

        try:
            is_equiped = 1
            profession = soldier_obj.profession
            inventory_objs = model_inventorys.objects.all()
            for q in inventory_objs:
                k = q.item_name
                try:
                    v = str(row[k])
                    if v == "" or v == "nan":
                        v=0
                    else:
                        if k in self.inventory_with_pn:
                            v = 1
                        else:
                            v = int(float(v))
                except Exception as ex:
                    pass
                    # print("9011-77-77-2 Error " + str(ex))
                qty_per_soldier = q.qty_per_soldier
                critical = q.critical
                if critical == 0 or critical != profession:
                    continue
                if v < qty_per_soldier:
                    is_equiped = 0
                    break
            if is_equiped == 1:
                o, is_created = model_soldierqualificationfact.objects.get_or_create(
                    training_web=battalion_obj.training_web,
                    soldier=soldier_obj, skill=3)
                o.value=1
                o.save()
        except Exception as ex:
            print("9011-55-21 Error \n", k, v, soldier_obj, str(ex))

    # --------------------

    def set_units_qualification(self, dic):
        print('90022-1 dic', dic)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        print("-"*100, "\n", file_path, "\n", "-"*100)
        s = self.uploaded_filename.split("_")
        # print("ssssssssss=", s)
        battalion_number_ = int(s[0])
        period_number_ = int(s[1])
        # print(battalion_number_, period_number_)
        #
        model_periods = apps.get_model(app_label=app_, model_name="periods")
        model_battalions = apps.get_model(app_label=app_, model_name="battalions")
        battalion_obj = model_battalions.objects.get(battalion_number=battalion_number_)
        period_obj = model_periods.objects.get(battalion=battalion_obj, period_number=period_number_)
        units_dic = period_obj.structure

        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        # print(df)
        print("-"*100)

        dic = {}
        for index, row in df.iterrows():
            print(index)
            battalion_number = int(row["battalion"])
            company = str(row["company"])
            # print(battalion_number, company)
            platoon= str(row["platoon"])
            wet_day= int(row["wet_day"])
            wet_night= int(row["wet_night"])
            raid_day= int(row["raid_day"])
            raid_night = int(row["raid_night"])
            if battalion_number not in dic:
                dic[battalion_number]={"title": "battalion " + str(battalion_number), "wet_day":0, "wet_night":0, "raid_day":0, "raid_night":0, "data":{}}
                dic[battalion_number]["wet_day"] = wet_day
                dic[battalion_number]["wet_night"] = wet_night
                dic[battalion_number]["raid_day"] = raid_day
                dic[battalion_number]["raid_night"] = raid_night
            battalion_dic = dic[battalion_number]
            print(company, wet_day, wet_night, raid_day, raid_night)
            if company != "nan":
                k = self.get_unit_number(units_dic, company)
                if k not in battalion_dic["data"]:
                    battalion_dic["data"][k]={"title": company, "wet_day":0, "wet_night":0, "raid_day":0, "raid_night":0, "data":{}}
                    company_dic = battalion_dic["data"][k]
                    company_dic["wet_day"] = wet_day
                    company_dic["wet_night"] = wet_night
                    company_dic["raid_day"] = raid_day
                    company_dic["raid_night"] = raid_night

                    battalion_dic["wet_day"] *= wet_day
                    battalion_dic["wet_night"] *= wet_night
                    battalion_dic["raid_day"] *= raid_day
                    battalion_dic["raid_night"] *= raid_night

            else:
                continue
            company_dic = battalion_dic["data"][k]
            if platoon != "nan":
                platoon_name_ = company + " " + platoon
                p = self.get_unit_number(units_dic, platoon_name_)
                print(p, platoon_name_)
                if p not in company_dic["data"]:
                    company_dic["data"][p]={}
                    platoon_dic = company_dic["data"][p]
                platoon_dic["wet_day"] = wet_day
                platoon_dic["wet_night"] = wet_night
                platoon_dic["raid_day"] = raid_day
                platoon_dic["raid_night"] = raid_night

                company_dic["wet_day"] *= wet_day
                company_dic["wet_night"] *= wet_night
                company_dic["raid_day"] *= raid_day
                company_dic["raid_night"] *= raid_night

                dic[battalion_number]["wet_day"] *= wet_day
                dic[battalion_number]["wet_night"] *= wet_night
                dic[battalion_number]["raid_day"] *= raid_day
                dic[battalion_number]["raid_night"] *= raid_night
            else:
                continue

        general_data_model = apps.get_model(app_label="core", model_name="generaldata")
        group_ = s[0]+"_"+str(s[1])
        data_name_ = "unit_qualification"
        print(group_, "\n", data_name_, "\n", dic)
        obj, is_created = general_data_model.objects.get_or_create(app=app_, group=group_, data_name=data_name_)
        obj.data_json = dic
        obj.save()

        # print(units_dic)
        # print("-"*100)
        # print(dic)
        # print("-"*100)
        # -----
        result = {"status": "ok"}
        return result


    # --------------------
    def compliance_data(self, dic):
        # try:
        #     print('\n 90100-160 get_compliance_data dic ', "\n", dic, "\n", "-" * 100)
        # except Exception as ex:
        #     print(str(ex))
        #     # pass
        app_ = dic['app']
        week_model_ = dic["week_model"]
        day_model_ = dic["day_model"]
        company_obj_id_ = int(dic['company_obj_id'])
        week_start_day_ = int(dic["week_start_day"])

        time_unit_ = dic["time_unit"]
        week_conclusion_ = dic["week_conclusion"]

        model_battalion = apps.get_model(app_label=app_, model_name="battalions")
        model_week = apps.get_model(app_label=app_, model_name=week_model_)
        model_day = apps.get_model(app_label=app_, model_name=day_model_)
        model_time_dim = apps.get_model(app_label=app_, model_name="timedim")
        unit_ = int(dic["unit"])
        battalion_ = int(dic["battalion"])
        time_dim_ = int(dic["time_dim"])
        # print(company_obj_id_, battalion_, week_start_day_, unit_)
        battalion_obj = model_battalion.objects.get(id=battalion_)
        obj, is_created = model_week.objects.get_or_create(training_web__id=company_obj_id_,
                                                           battalion = battalion_obj,
                                                           week_start_day=week_start_day_,
                                                           unit=unit_)
        # print("-"*50, "\nbattalion\n", obj, "\n", "-"*50)
        dic = {"week_unit_id": obj.id}
        time_dim_obj = model_time_dim.objects.get(id=time_dim_)
        obj_day, is_created = model_day.objects.get_or_create(complianceweek=obj, time_dim=time_dim_obj)
        if is_created:
            obj_day.time_unit = {"last_number": 0, "times": {}, "conclusion":""}
            obj_day.save()

        if time_unit_ == "null":
            time_unit_ = obj_day.time_unit
            week_conclusion_ = obj.conclusion
        else:
            obj_day.time_unit = time_unit_
            obj_day.save()
            obj.conclusion = week_conclusion_
            obj.save()

        dic["time_unit"] = time_unit_
        dic["week_conclusion"] = week_conclusion_
        # print(dic)
        return dic

    # Upload tests
    def upload_tests(self, dic):
        print('90022-3 dic', dic)
        cube_dic=dic["cube_dic"]
        user_id = dic["user_id"]
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # print("-"*100, "\n", file_path, "\n", "-"*100)
        file_name = self.uploaded_filename.split(".")[0]
        ll_ = file_name.split("_")
        battalion_number = int(ll_[0])
        period_ = int(ll_[1])
        timedim_ = ll_[4]
        model_instructors =  apps.get_model(app_label=app_, model_name="instructors")
        model_timedim =  apps.get_model(app_label=app_, model_name="timedim")
        model_battalions = apps.get_model(app_label=app_, model_name="battalions")
        model_periods = apps.get_model(app_label=app_, model_name="periods")
        model_test_events = apps.get_model(app_label=app_, model_name="testevents")
        model_grades_for_events = apps.get_model(app_label=app_, model_name="gradesforevents")

        b_obj = model_battalions.objects.get(battalion_number=battalion_number)
        p_obj = model_periods.objects.get(battalion=b_obj, period_number=period_)
        i_obj = model_instructors.objects.get(user__id=user_id)
        t_obj = model_timedim.objects.get(id=timedim_)

        model_soldiers = apps.get_model(app_label=app_, model_name="soldiers")
        model_testsforevents =  apps.get_model(app_label=app_, model_name="testsforevents")
        model_soldiersforevents =  apps.get_model(app_label=app_, model_name="soldiersforevents")

        # print(p_obj, i_obj, t_obj)

        adj_ = ll_[2]
        company_ = ll_[3]
        try:
            test_events_obj, is_created = model_test_events.objects.get_or_create(instructor=i_obj,
                                                                                  period=p_obj,
                                                                                  time_dim=t_obj,
                                                                                  test_event_name=adj_+" - "+company_)

            test_events_obj.units_description = adj_ + " - " + company_
            test_events_obj.save()
            # print("\n", "="*50, "\n", test_events_obj,"\n", "="*50, "\n")
        except Exception as ex:
            print("20-20-20-20 test_events_obj\n", ex)
        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        # print(df)
        # print(df.columns)
        tests = []
        tests_names = []
        for c in df.columns:
            try:
                test_number_ = int(c.split("_")[0])
                tests.append(test_number_)
                tests_names.append(c)
            except Exception as ex:
                pass
        print("-"*100)
        n_ = len(tests_names)
        result = {"Errors": {}}
        for index, row in df.iterrows():
            # print(index, "\n", row, "\n", row["full_name"])
            full_name = string.capwords(str(row["full_name"]))
            nnf = full_name.find(" ")
            first_name = full_name[:nnf]
            last_name = full_name[nnf+1:]
            userid = str(row["userid"])
            try:
                s_obj = model_soldiers.objects.get(battalion=b_obj, first_name=first_name, last_name=last_name)
            except Exception as ex:
                print("Error 202020-20: soldier FN=", first_name, "LN=", last_name, "not in db")
                if "202020-20" not in result["Errors"]:
                    result["Errors"]["202020-20"] = {}
                result["Errors"]["202020-20"][first_name + " " + last_name] = " not in the database"
                try:
                    s_obj = model_soldiers.objects.get(userid=userid)
                except Exception as ex:
                    if "202020-21" not in result["Errors"]:
                        result["Errors"]["202020-21"] = {}
                        result["Errors"]["202020-21"][userid] = "soldier with userid=" + userid +" is not in the database"

            try:
                soldiersforevents_obj, is_created_e = model_soldiersforevents.objects.get_or_create(testevent=test_events_obj,
                                                                                                    soldier_number=s_obj.user_id)
                soldiersforevents_obj.value = 1
                soldiersforevents_obj.save()
                # print(s_obj.id, "saved", soldiersforevents_obj)
            except Exception as ex:
                print("Error 202020-20-11: soldier")

            for j in range(n_):
                v_ = -1
                try:
                    v_ = round(100*row[tests_names[j]])/100
                except Exception as ex:
                    pass
                t_ = tests[j]
                # print(v_, t_)
                try:
                    testsforevents_obj, is_created_e = model_testsforevents.objects.get_or_create(testevent=test_events_obj,
                                                                                                  test_number=t_)
                    testsforevents_obj.value = 1
                    testsforevents_obj.save()
                except Exception as ex:
                    print("Error 202020-30: ", t_, v_)
                try:
                    model_grades_for_events_obj, is_created = model_grades_for_events.objects.get_or_create(testevent=test_events_obj,
                                                                                                            soldiersforevent=soldiersforevents_obj,
                                                                                                            testsforevent=testsforevents_obj)
                    model_grades_for_events_obj.value = v_
                    model_grades_for_events_obj.save()
                except Exception as ex:
                    print("Error 202020-40: soldier ", t_, v_, ex)

        return result

    # Update Qualification
    def update_qualification_fact(self, dic):
        # print('90033-1 dic', dic)
        app_ = dic["app"]
        data = dic["data"]
        skill = data["skill"]
        data = data["data"]
        first_soldier_id = list(data.keys())[0]
        company_obj_id_ = int(dic['company_obj_id'])
        # print(company_obj_id_, battalion_)
        model_soldier = apps.get_model(app_label=app_, model_name="soldiers")
        model_sqf = apps.get_model(app_label=app_, model_name="soldierqualificationfact")
        try:
            battalion_obj = model_soldier.objects.get(user_id = first_soldier_id).battalion
            objs = model_sqf.objects.filter(soldier__battalion=battalion_obj, skill=skill).all()
            print(objs.count())
            objs.delete()
        except Exception as ex:
            print("Error 90-80-22:", ex)

        try:
            for k in data:
                # print(k)
                soldier_obj = model_soldier.objects.get(user__id=int(k))
                sqf_obj, is_created = model_sqf.objects.get_or_create(training_web__id=company_obj_id_,
                                                                      soldier=soldier_obj, skill=skill)
                sqf_obj.value = data[k]
                sqf_obj.save()
        except Exception as ex:
            print("Error 90-80-33:", ex)

        # -----
        result = {"status": "ok"}
        # print(result)
        return result

    # To Be Deleted
    def get_soldier_report(self, dic):
        # print('90033-1 dic', dic)
        app_ = dic["app"]

        soldier_number_ = dic["soldier_number"]
        model_soldier = apps.get_model(app_label=app_, model_name="soldiers")
        model_tests_for_events = apps.get_model(app_label=app_, model_name="testsforevents")
        model_sgf = apps.get_model(app_label=app_, model_name="gradesforevents")
        model_if = apps.get_model(app_label=app_, model_name="inventoryfact")
        # model_sqf = apps.get_model(app_label=app_, model_name="soldierqualificationfact")

        soldier_obj = model_soldier.objects.get(userid=soldier_number_)
        dic = {"tests": {}, "inventory":{"idx":[], "id":[], "value":[]}, "equipment": {}}
        try:
            sgi_objs = model_if.objects.filter(soldier=soldier_obj).order_by("inventory").all()
            df = pd.DataFrame(list(sgi_objs.values()))
            # print(df)
            for index, row in df.iterrows():
                dic["inventory"]["idx"].append(index)
                dic["inventory"]["id"].append(int(row["inventory_id"]))
                dic["inventory"]["value"].append(round(float(row["value"])*100)/100)
        except Exception as ex:
            print("Error 90-90-90", ex)

        try:
            sgf_objs = model_sgf.objects.filter(soldiersforevent__soldier_number=soldier_obj.user_id).all()

            df = pd.DataFrame(list(sgf_objs.values()))
            # print(df)
            for index, row in df.iterrows():
                t_obj = model_tests_for_events.objects.get(id=row["testsforevent_id"])
                # print(t_obj.test_number)
                # print(t_obj.testevent.time_dim.id)
                if t_obj.test_number not in dic["tests"]:
                    dic["tests"][t_obj.test_number] = {"idx": [], "test_date": [], "value": []}
                dic["tests"][t_obj.test_number]["idx"].append(index)
                dic["tests"][t_obj.test_number]["test_date"].append(t_obj.testevent.time_dim.id)
                dic["tests"][t_obj.test_number]["value"].append(round(float(row["value"])*100)/100)

            # sqf_obj, is_created = model_sqf.objects.get(soldier=soldier_obj, skill=3)
        except Exception as ex:
            print("Error 90-90-91", ex)

        result = {"status": "ok", "result":dic}
        # print(result)

        return result

    #  -- To be deleted --
    def set_soldiers(self, dic):
        # print('90022-4 dic', dic)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # print("-"*100, "\n", file_path, "\n", "-"*100)
        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        # print(df)
        # print("-"*100)

        model_name_ = "instructors"
        model_instructors = apps.get_model(app_label=app_, model_name=model_name_)

        try:
            u = User.objects.get(username="dinstructor")
            count = u.delete()
            # print("B count\n", count, "\n")
        except Exception as ex:
            pass
            # print("9055-55 Error " + str(ex))

        my_group, is_created = Group.objects.get_or_create(name='t_simple_user')
        try:
            u = User.objects.create_user(username="dinstructor", email='dinstructor@gmail.com', password='DINSTRUCTOR123#')
            # print(u.password)
            u.first_name = "dinstructor"
            u.last_name = "dinstructor"
            u.save()
            # print(u)
            my_group.user_set.add(u)
            my_group.save()
        except Exception as ex:
            print("9000-00 Error " + str(ex))

        try:
            instructor_obj, is_created = model_instructors.objects.get_or_create(first_name='default_instructor1',
                                                                                 user=u)
        except Exception as ex:
            print("9001-01 Error " + str(ex))
        # print("-"*100, "\n", instructor_obj, "  =", is_created, "\n", "-"*100)

        model_name_ = "brigades"
        model_brigades = apps.get_model(app_label=app_, model_name=model_name_)
        brigade_obj, is_created = model_brigades.objects.get_or_create(brigade_name='brigade 1')
        # print("-"*100, "\n", brigade_obj, "  =", is_created, "\n", "-"*100)

        model_name_ = "battalions"
        model_battalions = apps.get_model(app_label=app_, model_name=model_name_)
        battalion_obj, is_created = model_battalions.objects.get_or_create(battalion_name='battalion 1', brigade=brigade_obj)
        battalion_obj.instructor.add(instructor_obj)
        battalion_obj.save()
        # print("-"*100, "\n", battalion_obj, "  ", is_created, "\n", "-"*100)

        model_name_ = "courses"
        model_courses = apps.get_model(app_label=app_, model_name=model_name_)
        course_obj, is_created = model_courses.objects.get_or_create(course_name='course 1')
        course_obj.instructor.add(instructor_obj)

        model_name_ = "companys"
        model_companys = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = "platoons"
        model_platoons = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = "squads"
        model_squads = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = "soldiers"
        model_soldier = apps.get_model(app_label=app_, model_name=model_name_)

        # model_user = get_user_model()
        ll = []
        ll_ = []
        ll1 = []
        ll1_ = []
        llf = []
        llf_ = []
        llfn = []
        llfn_ = []
        n__ = 0
        for index, row in df.iterrows():
            if n__ > 10000:
                break
            n__ += 1
            # print(row, "\n", row["COMPANY"])
            company_name_ = str(row["COMPANY"]).upper()
            company_number = int(row["COMPANYID"])
            username_ = str(23*10000+company_number*1000+int(row["SN"]))
            company_obj, is_created = model_companys.objects.get_or_create(company_name=company_name_, battalion=battalion_obj)
            if is_created:
                # print("-"*100, "\n", company_obj, "  ", is_created, "\n", "-"*100)
                company_obj.company_number = company_number
                company_obj.instructor.add(instructor_obj)
                company_obj.save()
            # str(int(row["PLATOON"]))
            platoon_name_n = int(row["PLATOON"])
            platoon_name_ = company_name_ + " " +str(platoon_name_n)
            platoon_obj, is_created = model_platoons.objects.get_or_create(platoon_name=platoon_name_, company=company_obj)
            if is_created:
                platoon_obj.platoon_number = platoon_name_n
                platoon_obj.instructor.add(instructor_obj)
                platoon_obj.save()
                # print("-"*100, "\n", platoon_obj, "  ", is_created, "\n", "-"*100)

            squad_name_ = '1'
            squad_obj, is_created = model_squads.objects.get_or_create(squad_name=squad_name_, platoon=platoon_obj)
            squad_obj.save()
            # print("-"*100, "\n", squad_obj, "  ", is_created, "\n", "-"*100)
            full_name = string.capwords(str(row["FULLNAME"]))
            nn__ = full_name.find(" ")
            first_name = full_name[:nn__]
            last_name = full_name[nn__+1:]
            mz4psn = str(row["MZ4PSN"])
            ramonsn = str(row["RAMONSN"])
            shoes_size = float(row["shoes_size"])
            uniform_size = str(row["uniform_size"])
            sport_size = str(row["sport_size"])
            if mz4psn not in ll:
                ll.append(mz4psn)
            else:
                ll_.append(mz4psn)
                print("mz4psn= ", mz4psn)
            if ramonsn not in ll1:
                ll1.append(ramonsn)
            else:
                ll1_.append(ramonsn)
                print("ramonsn= ", ramonsn)

            # if first_name not in llf:
            #     llf.append(first_name)
            # else:
            #     llf_.append(first_name)
            #     print("first_name= ", first_name)
            # if full_name not in llfn:
            #     llfn.append(full_name)
            # else:
            #     llfn_.append(full_name)
            #     print("full_namen= ", full_name)

            position = str(row["POSITION"]).upper()

            if position == "NAN":
                position = ""
            #  "A \n", "-"*100, "\n",
            print("full name", full_name, " first name=", first_name,"username_",username_,
                  " last name=", last_name, " position=", position,
                  " mz4psn", mz4psn, " ramonsn=", ramonsn,
                  shoes_size, uniform_size, sport_size)

            try:
                u = User.objects.get(username=username_)
                count = u.delete()
                # print("B count\n", count, "\n")
            except Exception as ex:
                pass
                # print("9055-55 Error " + str(ex))

            try:
                u = User.objects.create_user(username=username_, email=username_+'@gmail.com', password=company_name_+username_+'#')
                # print(u.password)
                u.first_name = first_name
                u.last_name = last_name
                u.save()
                # print(u)
                my_group.user_set.add(u)
                my_group.save()
            except Exception as ex:
                print("9011-11 Error " + str(ex))

            try:
                soldier, is_created = model_soldier.objects.get_or_create(user=u, platoon=platoon_obj)
                soldier.first_name = first_name
                soldier.last_name = last_name
                soldier.userid = username_
                try:
                    p_ = position.lower().strip()
                    if p_ == "":
                        p_ = "Other"
                    elif p_ == "officier":
                        p_ = "Officer"
                    elif p_ == "soldeir":
                        p_ = "Soldier"
                    elif p_ == "sous officier":
                        p_ = "Sous Officer"
                    else:
                        p_ = p_.title()
                    position = int(self.df_positions[self.df_positions["position_name"]==p_]["id"])
                except Exception as ex:
                    print("90888-66 Training objects set_soldiers Error: " + p_ +str(ex))
                soldier.position = position
                soldier.mz4psn = mz4psn
                soldier.ramonsn = ramonsn
                #
                soldier.shoes_size = shoes_size
                soldier.uniform_size = uniform_size
                soldier.sport_size = sport_size
                soldier.save()
                course_obj.course_soldiers.add(soldier)
                course_obj.save()
            except Exception as ex:
                print("9033-53 Error " + str(ex))

        # print(llf, "\n\n", llf_)
        # print("1="*10, "\n\n")
        # print(llfn, "\n\n", llfn_)
        # print("2="*10, "\n\n")
        #
        print(ll, "\n\n", ll1)
        print("="*100, "\n\n")
        print(ll_, "\n\n", ll1_)
        print("="*100, "\n\n")
        #
        # ukkk = authenticate(username='testsold', password='Amos122#')
        # print(ukkk)
        # print(ukkk.email)
        # try:
        #     print(u.user_training)
        # except Exception as ex:
        #     print("9044-54 Error " + str(ex))

        # print("-"*100, "\n", my_group, "  ", is_created, "\n", "-"*100)
        # print('90022-3 dic')

        # my_group.user_set.add(your_user)

        result = {"status": "ok"}
        return result

    def update_soldiers(self, dic):
        print('90033-1 dic', dic)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        print("-" * 100, "\n", file_path, "\n", "-" * 100)
        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        print(df)
        print("-" * 100)

        # model_name_ = "companys"
        # model_companys = apps.get_model(app_label=app_, model_name=model_name_)
        # model_name_ = "platoons"
        # model_platoons = apps.get_model(app_label=app_, model_name=model_name_)
        # model_name_ = "soldiers"
        # model_soldier = apps.get_model(app_label=app_, model_name=model_name_)
        # for index, row in df.iterrows():
        #     soldier_obj = model_soldier.objects.get(ramonsn=str(row["RAMONSN"]))
        #     soldier_obj.mz4psn = str(row["MZ4PSN"])
        #     platoon_obj = model_platoons.objects.get(company__company_name=str(row["COMPANY"]),
        #                                                  platoon_name=str(row["PLATOON"]))
        #     soldier_obj.platoon = platoon_obj
        #
        #     soldier_obj.save()

        result = {"status": "ok"}
        return result

    def delete_soldiers(self, dic):
        print('90044-1 dic', dic)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        print("-" * 100, "\n", file_path, "\n", "-" * 100)
        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        print(df)
        print("-" * 100)

        result = {"status": "ok"}
        return result

    def de_activate_soldiers(self, dic):
        print('90044-1 dic', dic)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        print("-" * 100, "\n", file_path, "\n", "-" * 100)
        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        print(df)
        print("-" * 100)

        result = {"status": "ok"}
        return result

    def create_dates_in_time_dim(self, dic):
        # print('90099-1 dic', dic)
        app_ = dic["app"]
        model_name = dic["model_name"]
        model_ = apps.get_model(app_label=app_, model_name=model_name)
        start_date = date.today() + timedelta(days=-90) # Monday = 0
        for i in range(1000):
            t = start_date + timedelta(days=i)
            y = t.year
            m = t.month
            d = t.day
            id_ = y * 10000 + m * 100 + d
            # print(y, m, d, id_)
            obj, is_create = model_.objects.get_or_create(id=id_)
            obj.year = y
            obj.month = m
            obj.day = d
            obj.save()
        result = {"status": "ok"}
        return result

    def daily_run(self, dic):
        print(dic)
        app_ = dic["app"]
        model_name_ = "soldiers"
        model_soldier = apps.get_model(app_label=app_, model_name=model_name_)
        qs = model_soldier.objects.all()
        df = pd.DataFrame(list(qs.values()))
        print(df.columns)
        df = pd.DataFrame(list(qs.values("is_confirmed","first_name","last_name","image",
                                         "userid","mz4psn","ramonsn","address","city","state","zip","country",
                                         "email","phone","birthday","num_of_children","marital_status",
                                         "shoes_size", "clothes_size","height","weight","blood_type","position","rank",
                                         "profession","sub_profession","medical_condition")))

        self.save_to_excel(df, "SoldiersList", file_name="daily_run.xlsx")

        result = {"status": "ok"}
        return result

    def excel_gun_list(self, dic):
        print('90033-100 dic\n', '-'*100, '\n', dic, '\n', '-'*100)
        app_ = dic["app"]
        model_name_ = "soldiers"
        model_soldier = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = "inventorys"
        model_inventorys = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = "inventorycategorys"
        model_inventorycategorys = apps.get_model(app_label=app_, model_name=model_name_)

        qs = model_inventorys.objects.all()
        df_i = pd.DataFrame(list(qs.values('id', 'inventorycategory_id', 'inventory_number')))
        qs = model_inventorycategorys.objects.all()
        df_ic = pd.DataFrame(list(qs.values('id', 'category_name')))
        dfm = df_i.merge(df_ic, left_on='inventorycategory_id', right_on='id')
        dfm = dfm.drop(["inventorycategory_id", 'id_y'], axis=1)
        dfm.columns = ['id', 'inventory_number', 'category_name']

        qs = model_soldier.objects.all()
        df_s = pd.DataFrame(list(qs.values('user_id', 'userid', 'first_name', 'last_name', 'gun_mz4psn_id', 'gun_ramonsn_id')))
        df_s['userid'] = df_s['userid'] + ":" + df_s['first_name'] + " " + df_s['last_name']
        df_s = df_s.drop(['first_name', 'last_name'], axis=1)
        dfm = df_s.merge(dfm, how='outer', left_on='gun_mz4psn_id', right_on='id')
        #
        dfm = df_s.merge(dfm, how='outer', left_on='gun_ramonsn_id', right_on='id')
        #
        dfm = dfm.drop(['id', 'gun_mz4psn_id_x', 'gun_mz4psn_id_y', 'gun_ramonsn_id_x', 'gun_ramonsn_id_y'], axis=1)
        dfm['userid'] = dfm.apply(lambda row: row["userid_y"] if np.isnan(row["user_id_x"]) else row["userid_x"], axis=1)
        dfm = dfm.drop(['user_id_x', 'user_id_y', 'userid_x', 'userid_y'], axis=1)
        # print(dfm)
        # print("5="*20)
        df = dfm.pivot_table(values='inventory_number', index='userid', columns=['category_name']
                             , aggfunc='sum'
                             )
        print(df)
        self.save_to_excel(df, "GunsList", file_name="guns_list.xlsx")

        result = {"status": "ok"}
        return result

    def set_guns(self, dic):
        print('90088-11 dic', dic)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # print("-"*100, "\n", file_path, "\n", "-"*100)
        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        print(df)
        print("-"*100)
        model_inventorys = apps.get_model(app_label=app_, model_name="inventorys")
        model_inventorycategorys = apps.get_model(app_label=app_, model_name="inventorycategorys")
        columns = df.columns
        print(columns)
        for c in columns:
            print(c)
            inventory_category_obj, is_create = model_inventorycategorys.objects.get_or_create(category_name=c)
            inventory_category_obj.save()
            for index, row in df.iterrows():
                gun_number = str(row[c]).upper()
                print(gun_number)
                inventory_obj, is_created = model_inventorys.objects.get_or_create(inventorycategory=inventory_category_obj,
                                                                                   inventory_number=gun_number)
                inventory_obj.save()
        print("-"*50, "\n\nDone")

        result = {"status": "ok"}
        return result

    # need to delete
    def set_new_structure_old(self, dic):
        print('90088-16 dic', dic)
        app_ = dic["app"]
        battalion_name = dic["cube_dic"]["fact"]["model"]
        battalion_id = int(dic["cube_dic"]["fact"]["field_name"])
        period_number = int(dic["cube_dic"]["dimensions"]["time_dim"]["field_name"])

        units_dic = {battalion_id: {'title': battalion_name, 'data': {}}}
        file_path = self.upload_file(dic)["file_path"]

        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        print(df)

        return

        sheet_name_ = eval(dic["sheet_name"])
        model_periods = apps.get_model(app_label=app_, model_name="periods")

        period_obj, is_created = model_periods.objects.get_or_create(battalion__id=battalion_id, period_number=period_number)
        model_unit_soldiers = apps.get_model(app_label=app_, model_name="unitsoldiers")

        # model_unit_soldiers.truncate()

        model_soldiers = apps.get_model(app_label=app_, model_name="soldiers")
        missing_soldiers = []
        multiple_soldiers = []
        for sheet in sheet_name_:
            number = self.get_next_number({"app": app_})
            # company
            units_dic[battalion_id]["data"][number] = {"title": sheet, "data": {}}
            data_ = units_dic[battalion_id]["data"][number]["data"]
            df = pd.read_excel(file_path, sheet_name=sheet, header=0)
            # print(df)
            platoons_ = []
            n__ = 0
            for index, row in df.iterrows():
                n__+= 1
                # print("=====", n__, "=====")
                platoon = str(row["Unit"])
                if platoon not in platoons_:
                    # platoon
                    platoons_.append(platoon)
                    number_ = self.get_next_number({"app": app_})
                    data_[number_] = {"title": platoon, "data": {}}
                    sections_ = []
                section = str(row["Sub"])
                if section not in sections_:
                    sections_.append(section)
                    # print(platoon, section)
                    # print(platoons_, sections_)
                    number__ = self.get_next_number({"app": app_})
                    data_[number_]["data"][number__] = {"title": section, "data": {}}

                full_name = string.capwords(str(row["FULLNAME"]))
                if full_name == "Nan":
                    continue
                nn__ = full_name.find(" ")
                first_name = full_name[:nn__].rstrip().lstrip()
                last_name = full_name[nn__+1:].rstrip().lstrip()

                try:
                    soldier_obj = model_soldiers.objects.filter(first_name=first_name).all()
                    count = soldier_obj.count()
                    if count == 0:
                        soldier_obj = model_soldiers.objects.filter(last_name=last_name).all()
                        count = soldier_obj.count()
                        if count == 0:
                            missing_soldiers.append(platoon+": " + section + ": " + full_name)
                            print(n__, " AAA (count == 0)=", platoon+": " + section + ": " + full_name)
                            continue
                    elif count > 1:
                        soldier_obj = model_soldiers.objects.filter(first_name=first_name, last_name=last_name).all()
                        count = soldier_obj.count()
                        if count > 1:
                            multiple_soldiers.append("2 "+full_name)
                            print(n__, "BBB(count >1)=", platoon+": " + section + ": " + full_name)
                            continue
                        elif count == 0:
                            print(n__, "DDD (count == 1 w LN count=0)=", platoon+": " + section + ": " + full_name)
                            multiple_soldiers.append("2 "+full_name)
                            continue
                    if soldier_obj.count() == 1:
                        try:
                            u_obj, is_created = model_unit_soldiers.objects.get_or_create(period=period_obj, soldier=soldier_obj[0])
                            u_obj.unit_number = number__
                            u_obj.save()
                            # print("saved: "+ str(n__))
                        except Exception as ex:
                            print("error 200: ", full_name)
                    else:
                        print(str(n__)+" HHH Not saved: (count="+str(soldier_obj.count())+")= ", platoon+": " + section + ": " + full_name)
                except Exception as ex:
                    print("Missin soldier in DB: platoon=" + platoon + ": section=" + section + ": first_name=" + first_name + "= last_name=" + last_name + "=\n" + str(ex))
                    # missing_soldiers.append(full_name)

        period_obj.structure = units_dic
        period_obj.period_name = "Battalion: " + str(battalion_id) + " Period: " + str(period_number)
        period_obj.save()
        print('\n missing_soldiers')
        print(missing_soldiers)
        print('\n multiple_soldiers')
        print(multiple_soldiers)
        print(units_dic)

        print("-"*50, "\n\nDone")

    def update_gun_list(self, dic):
        print('900100-1 dic', dic)
        clear_log_debug()
        app_ = dic["app"]
        model_soldiers = apps.get_model(app_label=app_, model_name="soldiers")
        model_inventorys = apps.get_model(app_label=app_, model_name="inventorys")
        qs = model_soldiers.objects.all()
        for q in qs:
            try:
                print(q.mz4psn)
                log_debug(str(q.mz4psn))
                im_obj = model_inventorys.objects.get(inventory_number=q.mz4psn)
            except Exception as ex:
                log_debug("9095-100 Error updating guns: get q.mz4psn " + str(q.mz4psn))

            try:
                q.gun_mz4psn = im_obj
                q.save()
            except Exception as ex:
                # pass
                log_debug("9095-100 Error updating guns: set 1 and save " + str(q.mz4psn))

            try:
                ir_obj = model_inventorys.objects.get(inventory_number=q.ramonsn)
            except Exception as ex:
                log_debug("9095-100 Error updating guns: get ramonsn" + " " + str(q.ramonsn))

            try:
                q.gun_ramonsn = ir_obj
                q.save()
            except Exception as ex:
                log_debug("9095-100 Error updating guns: save ramonsn"  + " " + str(q.ramonsn))

        result = {"status": "ok"}
        return result

    def get_variable_data(self, dic):
        # print('90065-11 dic', dic, "\n", "-"*50)
        app_ = dic["app"]
        record_id = dic["tests_dic"]["record_id"]
        group_list = dic["group_dic"]
        #
        parent_test_table = dic["tests_dic"]["parent_table"]
        parent_test_model_ = apps.get_model(app_label=app_, model_name=parent_test_table)
        parent_test_obj = parent_test_model_.objects.get(id=record_id)
        test_table = dic["tests_dic"]["table"]
        test_model_ = apps.get_model(app_label=app_, model_name=test_table)
        #
        tests_objs = test_model_.objects.filter(testsvariable=parent_test_obj).all()
        # print(tests_objs)
        test_list = []
        test_dic = {}
        up_value = parent_test_obj.up_value
        up_color = parent_test_obj.up_color
        down_value = parent_test_obj.down_value
        down_color = parent_test_obj.down_color
        other_color = parent_test_obj.other_color
        var_dic = {"up_value":up_value, "up_color":up_color, "down_value":down_value,
                   "down_color":down_color, "other_color":other_color}
        for q in tests_objs:
            test_list.append(str(q.test_number))
            test_dic[str(q.test_number)] = {"id": q.id , "value": round(float(q.value),2)}
        # print('test_dic')
        # print(test_dic)

        grades_model_ = apps.get_model(app_label=app_, model_name="gradesforevents")
        q_gs = grades_model_.objects.filter(soldiersforevent__soldier_number__in=group_list,
                                            testsforevent__test_number__in=test_list).order_by("soldiersforevent__soldier_number",
                                                                                               "testevent__id").all()

        result={"soldier_number":[], "event_id":[], "event_date":[], "test_number":[], "value":[]}
        for q in q_gs:
            # print(q.testevent.test_event_name)
            # print(q.testevent.id)
            result["event_id"].append(q.testevent.id)
            result["event_date"].append(q.testevent.time_dim.id)
            result["soldier_number"].append(q.soldiersforevent.soldier_number)
            result["test_number"].append(q.testsforevent.test_number)
            result["value"].append(round(float(q.value),2))
        result = {"status": "ok", "result":result, "test_dic":test_dic, "var_dic": var_dic}
        # print(result)
        return result

    def check_soldiers(self, dic):
        print('90033-1 dic', dic)
        app_ = dic["app"]
        sheet = dic["sheet_name"]
        file_path = self.upload_file(dic)["file_path"]
        print("-" * 100, "\n", file_path, "\n", "-" * 100)
        df = pd.read_excel(file_path, sheet_name=sheet, header=0)
        print(df)
        print("-" * 100)
        model_name_ = "soldiers"
        model_soldier = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = "soldierqualificationfact"
        model_sqf = apps.get_model(app_label=app_, model_name=model_name_)
        for index, row in df.iterrows():
            first_name_ = string.capwords(str(row["first_name"])).strip()
            # print("first_name", first_name_)
            last_name_ = string.capwords(str(row["last_name"])).strip()
            # print("last_name", last_name)
            try:
                soldier_obj = model_soldier.objects.get(first_name=first_name_, last_name=last_name_)
                # print("name_in_db====", soldier_obj.first_name + "   " + soldier_obj.last_name)
                sqf_obj, is_created = model_sqf.objects.get_or_create(soldier=soldier_obj, skill=2)
                sqf_obj.value=100
                sqf_obj.save()
            except Exception as ex:
                pass
                # print("name_not_in_db===", "="+first_name_ + "=" + last_name_ + "=")

        result = {"status": "ok"}
        return result


    # def get_periods_of_battalion(self, dic):
    #     # print('90065-11 dic', dic)
    #     app_ = dic["app"]
    #     battalion_id = int(dic["battalion_id"])
    #     periods_model_ = apps.get_model(app_label=app_, model_name="periods")
    #     period_objs = periods_model_.objects.filter(battalion__id=battalion_id).all()
    #
    #     periods = []
    #     test_dic = {}
    #     up_value = parent_test_obj.up_value
    #     up_color = parent_test_obj.up_color
    #     down_value = parent_test_obj.down_value
    #     down_color = parent_test_obj.down_color
    #     other_color = parent_test_obj.other_color
    #     var_dic = {"up_value":up_value, "up_color":up_color, "down_value":down_value,
    #                "down_color":down_color, "other_color":other_color}
    #     for q in tests_objs:
    #         test_list.append(str(q.test_number))
    #         test_dic[str(q.test_number)] = {"id": q.id , "value": round(float(q.value),2)}
    #     # print('test_dic')
    #     # print(test_dic)
    #
    #     grades_model_ = apps.get_model(app_label=app_, model_name="gradesforevents")
    #     q_gs = grades_model_.objects.filter(soldiersforevent__soldier_number__in=group_list,
    #                                         testsforevent__test_number__in=test_list).order_by("soldiersforevent__soldier_number").all()
    #     result={"soldier_number":[], "test_number":[], "value":[]}
    #     for q in q_gs:
    #         result["soldier_number"].append(q.soldiersforevent.soldier_number)
    #         result["test_number"].append(q.testsforevent.test_number)
    #         result["value"].append(round(float(q.value),2))
    #
    #     result = {"status": "ok", "result":result, "test_dic":test_dic, "var_dic": var_dic}
    #     # print(result)
    #     return result

# double_shoot = DoubleShoot()
# # function = "getSolderData"
# dic = {"function_name": "getMember", "soldier_id": "RoxASKgvRGaR90ZuAnc3Gw"}
# r = double_shoot.get_soldier_data(dic)

