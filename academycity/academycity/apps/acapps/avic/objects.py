import warnings
import os
from django.conf import settings
import matplotlib as mpl
from bs4 import BeautifulSoup
mpl.use('Agg')
import matplotlib.pyplot as plt

import numpy as np
from openpyxl import Workbook, load_workbook

from sklearn import linear_model, neighbors
from sklearn import preprocessing
from sklearn import pipeline
import tarfile
import zipfile
from six.moves import urllib
import hashlib
from sklearn.model_selection import train_test_split, StratifiedShuffleSplit
import matplotlib.image as mpimg
from sklearn.metrics import mean_squared_error, mean_absolute_error
from sklearn.model_selection import cross_val_score
from scipy import stats
import joblib

"""
 to_data_path_ is the place datasets are kept
 topic_id name of the chapter to store images
"""
import pandas as pd
import numpy as np
from ..ml.basic_ml_objects import BaseDataProcessing, BasePotentialAlgo
from django.apps import apps


class AvicAlgo(object):
    def __init__(self, dic):  # to_data_path, target_field
        # print("90004-000 AvibAlgo", dic, '\n', '-'*50)
        try:
            super(AvicAlgo, self).__init__()
        except Exception as ex:
            print("Error 90004-010 AvibDataProcessing:\n"+str(ex), "\n", '-'*50)
        # print("90004-020 AvibAlgo", dic, '\n', '-'*50)


class AvicDataProcessing(BaseDataProcessing, BasePotentialAlgo, AvicAlgo):
    def __init__(self, dic):
        super().__init__(dic)

    def check_country(self, cc):
        if cc.lower().find("west")>-1:
            cc = "Germany"
        if cc.lower().find("yemen")>-1:
            cc = "Yemen"
        if cc.lower().find("papua")>-1:
            cc = "Papua New Guinea"
        if cc == "Ethiopia  and  Eritrea":
            cc = "Ethiopia"
        if cc == "Ethiopia and Eritrea":
            cc = "Ethiopia"
        if cc == "Eritrea and Ethiopia":
            cc = "Ethiopia"
        if cc == "Viet Nam":
            cc = "Vietnam"
        elif cc == "Guinea-bissau":
            cc = "Guinea Bissau"
        elif cc == "Upper Volta":
            cc = "Burkina Faso"
        elif cc == "Kampuchea.Dem.":
            cc = "Cambodia"
        elif cc == "Kampuchea, Dem":
            cc = "Cambodia"
        elif cc == "Kampuchea, Dem.":
            cc = "Cambodia"
        elif cc == "Kampuchea":
            cc = "Cambodia"
        elif cc == "Guinea-Bissau":
            cc = "Guinea Bissau"
        elif cc == "Congo' People's Rep.":
            cc = "Democratic Republic of the Congo (DRC)"
        elif cc == "Congo, People's Rep":
            cc = "Democratic Republic of the Congo (DRC)"
        elif cc == "Congo, People's Rep.":
            cc = "Democratic Republic of the Congo (DRC)"
        elif cc == "Congo,People's rep.":
            cc = "Democratic Republic of the Congo (DRC)"
        elif cc == "Congo, Dem. Rep.":
            cc = "Democratic Republic of the Congo (DRC)"
        elif cc == "Zaire":
            cc = "Democratic Republic of the Congo (DRC)"
        elif cc == "Zaire (Congo Kinshasa)":
            cc = "Democratic Republic of the Congo (DRC)"
        elif cc == "Congo":
            cc = "Democratic Republic of the Congo (DRC)"
        elif cc == "DR Congo":
            cc = "Democratic Republic of the Congo (DRC)"
        elif cc == "DRC":
            cc = "Democratic Republic of the Congo (DRC)"
        elif cc == "Democratic Republic Of The Congo":
            cc = "Democratic Republic of the Congo (DRC)"
        elif cc == "Congo, Dem.Rep.":
            cc = "Democratic Republic of the Congo (DRC)"
        elif cc == "Congo 'Brazzaville'":
            cc = "Republic of the Congo"
        elif cc == "Republic of the Congo":
            cc = "Republic of the Congo"
        elif cc == "Congo, Rep.":
            cc = "Republic of the Congo"
        elif cc == "Russian Federation":
            cc = "Russia"
        elif cc == "Total Former USSR":
            cc = "Russia"
        elif cc == "USSR":
            cc = "Russia"
        elif cc == "Côte d’Ivoire":
            cc = "Cote d'Ivoire"
        elif cc == "Ivory Coast":
            cc = "Cote d'Ivoire"
        elif cc == "Eswatini (Swaziland)":
            cc = "Eswatini"
        elif cc == "Swaziland":
            cc = "Eswatini"
        elif cc == "Slovak Republic":
            cc = "Slovakia"
        elif cc.lower().find("burma")>-1:
            cc = "Myanmar"
        elif cc == "Czech Republic":
            cc = "Czechia"
        elif cc == "United States Of America":
            cc = "United States"
        elif cc == "Cabo Verde":
            cc = "Cape Verde"
        elif cc == "United Republic Of Tanzania":
            cc = "Tanzania"
        elif cc == "Laos":
            cc = "Lao People's Democratic Republic"
        elif cc == "Lao":
            cc = "Lao People's Democratic Republic"
        elif cc == "Republic Of Moldova":
            cc = "Moldova"
        elif cc == "Republic Of North Macedonia":
            cc = "Macedonia"
        elif cc == "North Macedonia":
            cc = "Macedonia"
        elif cc == "Macedonia, FYR":
            cc = "Macedonia"
        elif cc == "EU (27)":
            cc = "EU27"
        elif cc == "Venezuela, RB":
            cc = "Venezuela"
        elif cc == "turkiye":
            cc = "turkey"
        elif cc == "Egypt, Arab Rep.":
            cc = "Egypt"
        elif cc == "Egypt":
            cc = "Egypt"
        elif cc == "China (Mainland)":
            cc = "China"
        elif cc == "Hong Kong, China":
            cc = "China-Hong Kong"
        elif cc == "Hong Kong SAR, China":
            cc = "China-Hong Kong"
        elif cc == "Hong Kong SAR":
            cc = "China-Hong Kong"
        elif cc == "Hong Kong (China)":
            cc = "China-Hong Kong"
        elif cc == "Hong Kong":
            cc = "China-Hong Kong"
        elif cc == "USA":
            cc = "United States"
        elif cc == "Slovak Republic":
            cc = "Slovakia"
        elif cc == "Macau":
            cc = "China-Macau"
        elif cc == "Cabo Verde":
            cc = "Cape Verde"
        elif cc == "United Republic Of Tanzania":
            cc = "Tanzania"
        elif cc == "Republic Of Moldova":
            cc = "Moldova"
        elif cc == "Gambia, The":
            cc = "Gambia"
        elif cc == "Haïti":
            cc = "Haiti"
        elif cc == "Indonesia (including Timor until 1999)":
            cc = "Indonesia"
        elif cc == "Iran, Islamic Rep.":
            cc = "Iran"
        elif cc == "Iran,Islamic Rep.":
            cc = "Iran"
        elif cc == "Kyrgyz Republiuc":
            cc = "Kyrgyz Republic"
        elif cc == "Kyrgyzstan":
            cc = "Kyrgyz Republic"
        elif cc == "Lao PDR":
            cc = "Lao People's Democratic Republic"
        elif cc == "Slovak Republic":
            cc = "Slovakia"
        elif cc == "Syrian Arab Republic":
            cc = "Syria"
        elif cc == "Syrian Arab Rep.":
            cc = "Syria"
        elif cc == "Turkiye":
            cc = "Turkey"
        elif cc == "Venezuela, RB":
            cc = "Venezuela"
        elif cc == "Comoro Islands":
            cc = "Comoros"
        elif cc == "Côte d'Ivoire":
            cc = "Cote d'Ivoire"
        elif cc == "Central African Rep.":
            cc = "Central African Republic"
        elif cc == "Dominican Rep.":
            cc = "Dominican Republic"
        elif cc == "Germany, Fed Rep.":
            cc = "Germany"
        elif cc == "Germany, Fed. Rep.":
            cc = "Germany"
        elif cc == "Germany Fed. Rep.":
            cc = "Germany"
        elif cc == "German Dem. Rep.":
            cc = "German Democratic Republic"
        elif cc == "Germany (West)":
            cc = "Germany"
        elif cc == "Korea, Dem. People's Rep.":
            cc = "North Korea"
        elif cc == "Korea, Dem. People's Rep":
            cc = "North Korea"
        elif cc == "Korea, Rep. of":
            cc = "Korea Rep."
        elif cc == "Korea":
            cc = "Korea Rep."
        elif cc == "Republic Of Korea":
            cc = "Korea, Rep."
        elif cc == "South Korea":
            cc = "Korea Rep."
        elif cc == "Korea, Rep.":
            cc = "Korea Rep."
        elif cc == "Yemen Arab Rep.":
            cc = "Yemen"
        elif cc == "Yemen, Arab Rep.":
            cc = "Yemen"
        elif cc == "El Salvodor":
            cc = "El Salvador"
        return cc

    def remove_country(self, cc):
        c = -1
        if cc in ["Cape Verde","Luxembourg",
                  "Serbia/Montenegro/Kosovo", "Bosnia and Herzegovina", "Serbia and Montenegro", "Central Arab Rep."]:
            c = 1
        return c

    # _1  I adjusted this function to work after uploading Eli data
    def load_wbfile_to_db(self, dic):
        print("90121-5: \n", dic, "="*50)
        app_ = dic["app"]
        try:
            file_path = self.upload_file(dic)["file_path"]
            # print('90022-1 dic')
            dic = dic["cube_dic"]
            # print('90022-1 dic', dic)
            df = pd.read_excel(file_path, sheet_name="Data", header=0)
            # print(df)
            # df = df.reset_index()  # make sure indexes pair with number of rows
            # print(df)

            model_min_max_cut = apps.get_model(app_label=app_, model_name="minmaxcut")
            model_name_ = dic["dimensions"]["time_dim"]["model"]
            model_time_dim = apps.get_model(app_label=app_, model_name=model_name_)
            model_name_ = dic["dimensions"]["country_dim"]["model"]
            model_country_dim = apps.get_model(app_label=app_, model_name=model_name_)
            model_name_ = "measuregroupdim"
            model_group_measure_dim = apps.get_model(app_label=app_, model_name=model_name_)
            model_name_ = dic["dimensions"]["measure_dim"]["model"]
            model_measure_dim = apps.get_model(app_label=app_, model_name=model_name_)
            model_name_ = dic["fact"]["model"]
            model_fact = apps.get_model(app_label=app_, model_name=model_name_)
        except Exception as ex:
            print("Error 201-201-201", ex)

        min_cut = []
        max_cut = []
        #
        model_country_group_dim = apps.get_model(app_label=app_, model_name="CountryGroupDim")
        country_group_obj, is_created = model_country_group_dim.objects.get_or_create(group_name="wb")
        #
        for index, row in df.iterrows():
            # print(row["Country Name"], row["Country Code"], row["Series Name"], row["Series Code"])
            # Country
            # print(row)

            # print(row["Country Name"], c, row["Series Name"])
            try:
                if row["Series Name"] == "Population, total":
                    measure_name = "TotalPop"
                    group_measure_name = "General"
                    description = "Population, total"
                elif row["Series Name"] == "High-technology exports (current US$)":
                    measure_name = "HighTech"
                    group_measure_name = "Technology"
                    description = "High-technology exports (current US$)"
                elif row["Series Name"] == "ICT service exports (BoP, current US$)":
                    measure_name = "ExportICT"
                    group_measure_name = "ExportICT"
                    description = "ICT service exports (BoP, current US$)"
                    # print(measure_name)
                elif row["Series Name"] == "Scientific and technical journal articles":
                    measure_name = "SciTechJA"
                    group_measure_name = "SciTechJA"
                    description = "Scientific and technical journal articles"
                elif row["Series Name"] == "GDP per capita (current US$)":
                    measure_name = "GDPPCCUS$"
                    group_measure_name = "GDP"
                    description = "GDP per capita (current US$)"
                elif row["Series Name"] == "GDP per capita (constant 2015 US$)":
                    measure_name = "GDPPC2015$"
                    group_measure_name = "GDP"
                    description = "GDP per capita (constant 2015 US$)"
                elif row["Series Name"] == "GDP per capita, PPP (current international $)":
                    measure_name = "GDPPCINT$"
                    group_measure_name = "GDP"
                    description = "GDP per capita, PPP (current international $)"
                elif row["Series Name"] == "GDP per capita, PPP (constant 2017 international $)":
                    measure_name = "GDPPC2017INT$"
                    group_measure_name = "GDP"
                    description = "GDP per capita, PPP (constant 2017 international $)"
                elif row["Series Name"] == "GNI per capita, Atlas method (current US$)":
                    measure_name = "GNIPCAINT$"
                    group_measure_name = "GDP"
                    description = "GNI per capita, Atlas method (current US$)"
                elif row["Series Name"] == "GNI per capita (constant 2015 US$)":
                    measure_name = "GNIPC2015$"
                    group_measure_name = "GDP"
                    description = "GNI per capita (constant 2015 US$)"
                elif row["Series Name"] == "GNI per capita, PPP (current international $)":
                    measure_name = "GNIPCPPINT$"
                    group_measure_name = "GDP"
                    description = "GNI per capita, PPP (current international $)"
                elif row["Series Name"] == "GNI per capita, PPP (constant 2017 international $)":
                    measure_name = "GNIPC2017INT$"
                    group_measure_name = "GDP"
                    description = "GNI per capita, PPP (constant 2017 international $)"
                elif row["Series Name"] == "Military expenditure (current USD)":
                    measure_name = "MExpCUS$"
                    group_measure_name = "Military"
                    description = "Military expenditure (current USD)"
                elif row["Series Name"] == "Armed forces personnel, total":
                    measure_name = "ArmedFPT"
                    group_measure_name = "Military"
                    description = "Armed forces personnel, total"
                elif row["Series Name"] == "Researchers in R&D (per million people)":
                    measure_name = "ResearchersPMP"
                    group_measure_name = "RandD"
                    description = "Researchers in R&D (per million people)"
                elif row["Series Name"] == "Technicians in R&D (per million people)":
                    measure_name = "TechniciansPMP"
                    group_measure_name = "RandD"
                    description = "Technicians in R&D (per million people)"
                elif row["Series Name"] == "Exports of goods and services (constant 2015 US$)":
                    measure_name = "GSC2015US$"
                    group_measure_name = "Exports"
                    description = "Exports of goods and services (constant 2015 US$)"
                elif row["Series Name"] == "Exports of goods and services (current US$)":
                    measure_name = "GSCUS$"
                    group_measure_name = "Exports"
                    description = "Exports of goods and services (current US$)"
                elif row["Series Name"] == "Exports of goods, services and primary income (BoP, current US$)":
                    measure_name = "GSPIBOPCUS$"
                    group_measure_name = "Exports"
                    description = "Exports of goods, services and primary income (BoP, current US$)"
                elif row["Series Name"] == "Merchandise exports (current US$)":
                    measure_name = "MerCUS$"
                    group_measure_name = "Exports"
                    description = "Merchandise exports (current US$)"
                elif row["Series Name"] == "Exports of goods and services (BoP, current US$)":
                    measure_name = "GSBOPCUS$"
                    group_measure_name = "Exports"
                    description = "Exports of goods and services (BoP, current US$)"
                elif row["Series Name"] == "Industry (including construction), value added (current US$)":
                    measure_name = "IndCUS$"
                    group_measure_name = "Industry"
                    description = "Industry (including construction), value added (current US$)"
                elif row["Series Name"] == "Industry (including construction), value added (constant 2015 US$)":
                    measure_name = "IndC2015$"
                    group_measure_name = "Industry"
                    description = "Industry (including construction), value added (constant 2015 US$)"
                elif row["Series Name"] == "Natural gas rents (% of GDP)":
                    measure_name = "GasPerGDP"
                    group_measure_name = "NaturalRes"
                    description = "Natural gas rents (% of GDP)"
                elif row["Series Name"] == "Total natural resources rents (% of GDP)":
                    measure_name = "TNRPerGDP"
                    group_measure_name = "NaturalRes"
                    description = "Total natural resources rents (% of GDP)"
            except Exception as ex:
                pass
            #
            try:
                gm, is_created = model_group_measure_dim.objects.get_or_create(group_name=group_measure_name)
                mm = row["Series Code"]
                m, is_created = model_measure_dim.objects.get_or_create(measure_code=mm, measure_group_dim=gm)
                m.measure_name = measure_name
                m.description = description
                m.save()
            except Exception as ex:
                print("90986-1 Error measure:"+str(ex))
            country_name = str(row["Country Name"]).strip()
            is_remove = self.remove_country(country_name)
            if is_remove == 1:
                continue
            # print("="*50)
            # print(country_name)
            country_name = self.check_country(country_name)
            # print(country_name)

            for j in range(4, len(df.columns)):
                jm = j-4
                k = df.columns[j]
                s = k.split(" ")
                # print(s[0])

                yy = int(s[0])
                t, is_created = model_time_dim.objects.get_or_create(id=yy)
                if is_created:
                    # t.year = yy
                    s = 't.' + dic["dimensions"]["time_dim"]["field_name"] + ' = yy'
                    # print(s)
                    exec(s)
                    t.save()
                try:
                    s_ = str(row[0]).lower()
                    # print("-"*20, "\n", jm, yy, s_)
                    if s_ == "min_cut" or s_ == "max_cut":
                        if s_ == "min_cut":
                            try:
                                min_cut.append(float("{:.2f}".format(row[k])))
                            except Exception as ex:
                                print("Error 2001", jm, ex)
                        if s_ == "max_cut":
                            try:
                                max_cut.append(float("{:.2f}".format(row[k])))
                            except Exception as ex:
                                print("Error 2002", jm, ex)
                        try:
                            min_cut_ = min_cut[jm]
                            max_cut_ = max_cut[jm]
                            min_max_cut_obj, is_created = model_min_max_cut.objects.get_or_create(time_dim=t, measure_dim=m)
                            min_max_cut_obj.min = min_cut_
                            min_max_cut_obj.max = max_cut_
                            min_max_cut_obj.save()
                        except Exception as ex:
                            pass
                    else:
                        try:
                            c = model_country_dim.objects.get(country_name=country_name)
                        except Exception as ex:
                            continue
                        if float("{:.2f}".format(row[k])) > 0 or float("{:.2f}".format(row[k])) < 0:
                            try:
                                a, is_created = model_fact.objects.get_or_create(time_dim=t, country_dim=c, measure_dim=m)
                                # if is_created:
                                # a.amount = float(row[k])
                                s = 'a.' + dic["fact"]["field_name"] + ' = ' + str(float("{:.2f}".format((row[k]))))
                                # print(s)
                                exec(s)
                                a.save()
                            except Exception as ex:
                                print("90986-201 Error measure:"+str(ex))
                except Exception as ex:
                    pass
        result = {"status": "ok"}
        return result

    def load_eli_file_to_db(self, dic):
        try:
            # print("90121-1: \n", "="*50, "\n", dic, "\n", "="*50)
            app_ = dic["app"]
            file_path = self.upload_file(dic)["file_path"]
            # print(file_path)
            # print('90121-2 dic')
            dic = dic["cube_dic"]
            # print('90121-3 dic', dic)

            model_name_ = dic["dimensions"]["time_dim"]["model"]
            model_time_dim = apps.get_model(app_label=app_, model_name=model_name_)
            #
            model_country_group_dim = apps.get_model(app_label=app_, model_name="CountryGroupDim")
            model_name_ = dic["dimensions"]["country_dim"]["model"]
            model_country_dim = apps.get_model(app_label=app_, model_name=model_name_)
            #
            model_name_ = dic["dimensions"]["measure_dim"]["model"]
            model_measure_dim = apps.get_model(app_label=app_, model_name=model_name_)
            model_measure_group_dim = apps.get_model(app_label=app_, model_name="measuregroupdim")
            #
            model_name_ = dic["fact"]["model"]
            model_fact = apps.get_model(app_label=app_, model_name=model_name_)
            #
            model_min_max_cut = apps.get_model(app_label=app_, model_name="minmaxcut")
            #
            country_group_obj, is_created = model_country_group_dim.objects.get_or_create(group_name="eli")
            #
            year = int(self.uploaded_filename.split(".")[0])
            year_obj, is_created = model_time_dim.objects.get_or_create(id=year)
            if is_created:
                s = 'year_obj.' + dic["dimensions"]["time_dim"]["field_name"]+' = year'
                exec(s)
                year_obj.save()
            wb = load_workbook(filename=file_path, read_only=False)
            sheet_names = wb.sheetnames
        except Exception as ex:
            pass
        for f in sheet_names:
            try:
                ws = wb[f]
                f = self.clean_name(f)
                group_obj, is_created = model_measure_group_dim.objects.get_or_create(group_name=f)
                if is_created:
                    group_obj.group_name = f
                    group_obj.save()
            except Exception as ex:
                pass
            data = ws.values
            columns = next(data)[0:]   # Get the first line in file as a header line
            # print(columns)
            # Create a DataFrame based on the second and subsequent lines of data
            df = pd.DataFrame(data, columns=columns)
            df = df.reset_index()  # make sure indexes pair with number of rows
            # print(df)
            min_cut = []
            max_cut = []
            for j in range(0, len(columns)):
                min_cut.append(None)
                max_cut.append(None)
            for index, row in df.iterrows():
                if row[1] is not None and str(row[1]) != "None" and str(row[1]) != "":
                    n_ = 0
                    country_name = str(row[1]).strip()
                    is_remove = self.remove_country(country_name)
                    if is_remove == 1:
                        continue

                    # print("=" * 50)
                    # print(country_name)
                    country_name = self.check_country(country_name)
                    # print(country_name)

                    for j in range(1, len(columns)):
                        if str(columns[j]) != "None":
                            n_ += 1
                            f_ = f + str(n_)
                            try:
                                measure_obj, is_created = model_measure_dim.objects.get_or_create(measure_group_dim=group_obj,
                                                                                                  measure_name=f_)
                                if is_created:
                                    measure_obj.measure_code = f_
                                    measure_obj.save()
                            except Exception as ex:
                                pass

                            s_ = str(row[1]).lower()
                            if s_ == "min_cut" or s_ == "max_cut":
                                # print(columns[j])
                                try:
                                    if s_ == "min_cut":
                                        # print("Min_Cut", row[columns[j]])
                                        min_cut[j] = float(row[columns[j]])
                                    if s_ == "max_cut":
                                        # print("Max_Cut", row[columns[j]])
                                        max_cut[j] = float(row[columns[j]])
                                    if min_cut[j] is not None and max_cut[j] is not None:
                                        # print(f + str(n_), columns[j], "=", min_cut[j], max_cut[j])
                                        try:
                                            min_max_cut_obj, is_created = model_min_max_cut.objects.get_or_create(time_dim=year_obj,
                                                                                                                  measure_dim=measure_obj)
                                            if is_created:
                                                min_max_cut_obj.min = min_cut[j]
                                                min_max_cut_obj.max = max_cut[j]
                                                min_max_cut_obj.save()
                                        except Exception as ex:
                                            pass
                                except Exception as ex:
                                    print(ex)
                            else:
                                try:
                                    # print(country_name)
                                    country_dim_obj, is_created = model_country_dim.objects.get_or_create(country_name=country_name)
                                    if is_created:
                                        country_dim_obj.country_code = country_name
                                        country_dim_obj.country_group_dim = country_group_obj
                                        country_dim_obj.save()
                                except Exception as ex:
                                    pass
                                try:
                                    v_ = float(str(row[columns[j]]))
                                    if v_ is not None and str(v_) != "nan":
                                        # print(row[columns[j]], float(str(row[columns[j]])))
                                        fact_obj, is_created = model_fact.objects.get_or_create(time_dim=year_obj,
                                                                                                country_dim=country_dim_obj,
                                                                                                measure_dim=measure_obj)
                                        fact_obj.amount = v_
                                        fact_obj.save()
                                except Exception as ex:
                                    pass
                                    # print("Error 9055-33: \ncountry_name", country_name, "\nv=", "="+v_+"=", "=\n", str(columns[j]), "\n", f_, str(ex))
        wb.close()

        result = {"status": "ok"}
        print(result)

        return result

