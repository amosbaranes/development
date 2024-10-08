import warnings
import os
from django.conf import settings
import matplotlib as mpl
mpl.use('Agg')

import sys
import shutil
import math
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
import pickle
from concurrent.futures import ThreadPoolExecutor as T
from openpyxl import Workbook, load_workbook
import time


import pandas as pd
import numpy as np
from ..ml.basic_ml_objects import BaseDataProcessing, BasePotentialAlgo
from django.apps import apps


class PotentialDataProcessing(BaseDataProcessing, BasePotentialAlgo):
    def __init__(self, dic):
        # print("90001-00 PotentialDataProcessing", dic, '\n', '-'*50)
        super(PotentialDataProcessing, self).__init__(dic)
        # print("90002-00 PotentialDataProcessing", dic, '\n', '-'*50)
        mdg_fn = "id"
        # try:
        #     mdg_fn = dic["measure_group_dim_field_name"]
        # except Exception as ex:
        #     print("Exc90002-1" + str(ex))
        # mdg_fv = dic["measure_group_dim_field_value"]
        # td_fv = dic["time_dim_value"]
        # # notice in the measure group, we should include index for every category
        # s = 'sq = Fact.objects.filter(measure_dim__measure_group_dim__'+mdg_fn+'='+mdg_fv+', time_dim__id=td_fv)'
        # exec(s)
        # df = pd.DataFrame(sq.values('country_dim', 'measure_dim', 'amount'))
        # print(df)

    def load_file_to_db(self, dic):
        print("90121-1: \n", dic, "="*50)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # print('90121-  dic')
        dic = dic["cube_dic"]
        # print('90121-3 dic', dic)
        model_name_ = dic["dimensions"]["time_dim"]["model"]
        model_time_dim = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = dic["dimensions"]["country_dim"]["model"]
        model_country_dim = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = dic["dimensions"]["measure_dim"]["model"]
        model_measure_dim = apps.get_model(app_label=app_, model_name=model_name_)
        model_measure_group_dim = apps.get_model(app_label=app_, model_name="measuregroupdim")
        model_min_max_cut = apps.get_model(app_label=app_, model_name="minmaxcut")
        model_name_ = dic["fact"]["model"]
        model_fact = apps.get_model(app_label=app_, model_name=model_name_)

        year = int(self.uploaded_filename.split(".")[0])
        try:
            year_obj, is_created = model_time_dim.objects.get_or_create(id=year)
            if is_created:
                s = 'year_obj.' + dic["dimensions"]["time_dim"]["field_name"]+' = year'
                exec(s)
                year_obj.save()
        except Exception as ex:
            pass

        # files = []
        wb = load_workbook(filename=file_path, read_only=False)
        sheet_names = wb.sheetnames
        for f in sheet_names:
            ws = wb[f]
            f = self.clean_name(f)
            try:
                group_obj, is_created = model_measure_group_dim.objects.get_or_create(group_name=f)
                if is_created:
                    group_obj.group_name = f
                    group_obj.save()
            except Exception as ex:
                pass
            # files.append(f)
            data = ws.values
            # Get the first line in file as a header line
            columns = next(data)[0:]
            # print(columns)
            # Create a DataFrame based on the second and subsequent lines of data
            df = pd.DataFrame(data, columns=columns)
            df = df.reset_index()  # make sure indexes pair with number of rows
            # print(df)
            min_cut = []
            max_cut = []
            for j in range(0, len(columns)):
                min_cut.append(None)
                max_cut.append(None)
            for index, row in df.iterrows():
                if row[1] is not None and str(row[1]) != "None" and str(row[1]) != "":
                    n_ = 0
                    for j in range(1, len(columns)):
                        if str(columns[j]) != "None":
                            n_ += 1
                            f_ = f + str(n_)
                            try:
                                measure_obj, is_created = model_measure_dim.objects.get_or_create(measure_group_dim=group_obj,
                                                                                                  measure_name=f_)
                                if is_created:
                                    measure_obj.measure_code = f_
                                    measure_obj.save()
                            except Exception as ex:
                                pass

                            if row[1] == "Min_Cut" or row[1] == "Max_Cut":
                                # print(columns[j])
                                try:
                                    if row[1] == "Min_Cut":
                                        # print("Min_Cut", row[columns[j]])
                                        min_cut[j] = float(row[columns[j]])
                                    if row[1] == "Max_Cut":
                                        # print("Max_Cut", row[columns[j]])
                                        max_cut[j] = float(row[columns[j]])
                                    if min_cut[j] is not None and max_cut[j] is not None:
                                        # print(f + str(n_), columns[j], "=", min_cut[j], max_cut[j])

                                        try:
                                            min_max_cut_obj, is_created = model_min_max_cut.objects.get_or_create(time_dim=year_obj,
                                                                                                                  measure_dim=measure_obj)
                                            if is_created:
                                                min_max_cut_obj.min = min_cut[j]
                                                min_max_cut_obj.max = max_cut[j]
                                                min_max_cut_obj.save()
                                        except Exception as ex:
                                            pass

                                except Exception as ex:
                                    print(ex)
                            else:
                                try:
                                    country_name = str(row[1]).strip()
                                    country_dim_obj, is_created = model_country_dim.objects.get_or_create(country_name=country_name)
                                    if is_created:
                                        country_dim_obj.country_code = row[1]
                                        country_dim_obj.save()
                                except Exception as ex:
                                    pass
                                try:
                                    v_ = float(str(row[columns[j]]))
                                    if v_ is not None and str(v_) != "nan":
                                        # print(row[columns[j]], float(str(row[columns[j]])))
                                        fact_obj, is_created = model_fact.objects.get_or_create(time_dim=year_obj,
                                                                                                country_dim=country_dim_obj,
                                                                                                measure_dim=measure_obj)
                                        fact_obj.amount = v_
                                        fact_obj.save()
                                except Exception as ex:
                                    print("Error 9055-33: "+str(ex))
        wb.close()

        result = {"status": "ok"}
        return result

