import warnings
import os
from django.conf import settings
import matplotlib as mpl
from bs4 import BeautifulSoup
mpl.use('Agg')
import matplotlib.pyplot as plt

import numpy as np
from openpyxl import Workbook, load_workbook

from sklearn import linear_model, neighbors
from sklearn import preprocessing
from sklearn import pipeline
import tarfile
import zipfile
from six.moves import urllib
import hashlib
from sklearn.model_selection import train_test_split, StratifiedShuffleSplit
import matplotlib.image as mpimg
from sklearn.metrics import mean_squared_error, mean_absolute_error
from sklearn.model_selection import cross_val_score
from scipy import stats
import joblib

"""
 to_data_path_ is the place datasets are kept
 topic_id name of the chapter to store images
"""
import pandas as pd
import numpy as np
from ..ml.basic_ml_objects import BaseDataProcessing, BasePotentialAlgo
from django.apps import apps


class AoAlgo(object):
    def __init__(self, dic):  # to_data_path, target_field
        # print("90004-000 AoAlgo", dic, '\n', '-'*50)
        try:
            super(AoAlgo, self).__init__()
        except Exception as ex:
            print("Error 90004-010 AoDataProcessing:\n"+str(ex), "\n", '-'*50)
        # print("90004-020 AoAlgo", dic, '\n', '-'*50)

class AoDataProcessing(BaseDataProcessing, BasePotentialAlgo, AoAlgo):
    def __init__(self, dic):
        super().__init__(dic)

    def load_wbfile_to_db(self, dic):
        print("90121-5: \n", dic, "="*50)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # print('90022-1 dic')
        dic = dic["cube_dic"]
        # print('90022-1 dic', dic)
        df = pd.read_excel(file_path, sheet_name="Data", header=0)
        print(df)
        model_name_ = dic["dimensions"]["time_dim"]["model"]
        model_time_dim = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = dic["dimensions"]["country_dim"]["model"]
        model_country_dim = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = "measuregroupdim"
        model_group_measure_dim = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = dic["dimensions"]["measure_dim"]["model"]
        model_measure_dim = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = dic["fact"]["model"]
        model_fact = apps.get_model(app_label=app_, model_name=model_name_)

        for index, row in df.iterrows():
            print(row["Country Name"], row["Country Code"], row["Series Name"], row["Series Code"])
            # Country
            try:
                c, is_created = model_country_dim.objects.get_or_create(country_code=row["Country Code"])
                if is_created:
                    cc = self.check_country_el(row["Country Name"])
                    s = 'c.' + dic["dimensions"]["country_dim"]["field_name"] + ' = "' + cc + '"'
                    exec(s)
                    c.save()
            except Exception as ex:
                print("90987-1 Error measure:"+str(ex))
            #
            if row["Series Name"] == "Population, total":
                measure_name = "TotalPop"
                group_measure_name = "General"
                description = "Population, total"
            elif row["Series Name"] == "High-technology exports (current US$)":
                measure_name = "HighTech"
                group_measure_name = "Technology"
                description = "High-technology exports (current US$)"
            elif row["Series Name"] == "ICT service exports (BoP, current US$)":
                measure_name = "ICT"
                group_measure_name = "Technology"
                description = "ICT service exports (BoP, current US$)"
            elif row["Series Name"] == "GDP per capita (current US$)":
                measure_name = "GDPPCCUS$"
                group_measure_name = "GDP"
                description = "GDP per capita (current US$)"
            elif row["Series Name"] == "GDP per capita (constant 2015 US$)":
                measure_name = "GDPPC2015$"
                group_measure_name = "GDP"
                description = "GDP per capita (constant 2015 US$)"
            elif row["Series Name"] == "GDP per capita, PPP (current international $)":
                measure_name = "GDPPCINT$"
                group_measure_name = "GDP"
                description = "GDP per capita, PPP (current international $)"
            elif row["Series Name"] == "GDP per capita, PPP (constant 2017 international $)":
                measure_name = "GDPPC2017INT$"
                group_measure_name = "GDP"
                description = "GDP per capita, PPP (constant 2017 international $)"
            elif row["Series Name"] == "GNI per capita, Atlas method (current US$)":
                measure_name = "GNIPCAINT$"
                group_measure_name = "GDP"
                description = "GNI per capita, Atlas method (current US$)"
            elif row["Series Name"] == "GNI per capita (constant 2015 US$)":
                measure_name = "GNIPC2015$"
                group_measure_name = "GDP"
                description = "GNI per capita (constant 2015 US$)"
            elif row["Series Name"] == "GNI per capita, PPP (current international $)":
                measure_name = "GNIPCPPINT$"
                group_measure_name = "GDP"
                description = "GNI per capita, PPP (current international $)"
            elif row["Series Name"] == "GNI per capita, PPP (constant 2017 international $)":
                measure_name = "GNIPC2017INT$"
                group_measure_name = "GDP"
                description = "GNI per capita, PPP (constant 2017 international $)"
            elif row["Series Name"] == "Military expenditure (current USD)":
                measure_name = "MExpCUS$"
                group_measure_name = "Military"
                description = "Military expenditure (current USD)"
            elif row["Series Name"] == "Armed forces personnel, total":
                measure_name = "ArmedFPT"
                group_measure_name = "Military"
                description = "Armed forces personnel, total"
            elif row["Series Name"] == "Researchers in R&D (per million people)":
                measure_name = "ResearchersPMP"
                group_measure_name = "RandD"
                description = "Researchers in R&D (per million people)"
            elif row["Series Name"] == "Technicians in R&D (per million people)":
                measure_name = "TechniciansPMP"
                group_measure_name = "RandD"
                description = "Technicians in R&D (per million people)"
            elif row["Series Name"] == "Exports of goods and services (constant 2015 US$)":
                measure_name = "GSC2015US$"
                group_measure_name = "Exports"
                description = "Exports of goods and services (constant 2015 US$)"
            elif row["Series Name"] == "Exports of goods and services (current US$)":
                measure_name = "GSCUS$"
                group_measure_name = "Exports"
                description = "Exports of goods and services (current US$)"
            elif row["Series Name"] == "Exports of goods, services and primary income (BoP, current US$)":
                measure_name = "GSPIBOPCUS$"
                group_measure_name = "Exports"
                description = "Exports of goods, services and primary income (BoP, current US$)"
            elif row["Series Name"] == "Merchandise exports (current US$)":
                measure_name = "MerCUS$"
                group_measure_name = "Exports"
                description = "Merchandise exports (current US$)"
            elif row["Series Name"] == "Exports of goods and services (BoP, current US$)":
                measure_name = "GSBOPCUS$"
                group_measure_name = "Exports"
                description = "Exports of goods and services (BoP, current US$)"
            elif row["Series Name"] == "Industry (including construction), value added (current US$)":
                measure_name = "IndCUS$"
                group_measure_name = "Industry"
                description = "Industry (including construction), value added (current US$)"
            elif row["Series Name"] == "Industry (including construction), value added (constant 2015 US$)":
                measure_name = "IndC2015$"
                group_measure_name = "Industry"
                description = "Industry (including construction), value added (constant 2015 US$)"
            elif row["Series Name"] == "Scientific and technical journal articles":
                measure_name = "SciTechJA"
                group_measure_name = "Science"
                description = "Scientific and technical journal articles"
            elif row["Series Name"] == "Natural gas rents (% of GDP)":
                measure_name = "GasPerGDP"
                group_measure_name = "NaturalRes"
                description = "Natural gas rents (% of GDP)"
            elif row["Series Name"] == "Total natural resources rents (% of GDP)":
                measure_name = "TNRPerGDP"
                group_measure_name = "NaturalRes"
                description = "Total natural resources rents (% of GDP)"
            print(group_measure_name, measure_nam, description)
            #
            try:
                gm, is_created = model_group_measure_dim.objects.get_or_create(group_name=group_measure_name)
                mm = row["Series Code"]
                m, is_created = model_measure_dim.objects.get_or_create(measure_code=mm, measure_group_dim=gm)
                if is_created:
                    s = 'm.' + dic["dimensions"]["measure_dim"]["field_name"] + ' = "' + measure_name + '"'
                    s = 'm.description = "' + description + '"'
                    # print(s)
                    exec(s)
                    m.save()
            except Exception as ex:
                print("90986-1 Error measure:"+str(ex))

            for j in range(4, len(df.columns)):
                k = df.columns[j]
                try:
                    if float("{:.2f}".format(row[k])) > 0 or float("{:.2f}".format(row[k])) < 0:
                        s = k.split(" ")
                        # print(s[0])
                        yy = int(s[0])
                        t, is_created = model_time_dim.objects.get_or_create(id=yy)
                        if is_created:
                            # t.year = yy
                            s = 't.' + dic["dimensions"]["time_dim"]["field_name"] + ' = yy'
                            # print(s)
                            exec(s)
                            t.save()

                        a, is_created = model_fact.objects.get_or_create(time_dim=t, country_dim=c, measure_dim=m)
                        if is_created:
                            # a.amount = float(row[k])
                            s = 'a.' + dic["fact"]["field_name"] + ' = ' + str(float("{:.2f}".format((row[k]))))
                            # print(s)
                            exec(s)
                            a.save()
                except Exception as ex:
                    pass
                    # print(k, row[k], "9065-55 Error: "+str(ex))

        result = {"status": "ok"}
        return result

    def load_eli_file_to_db(self, dic):
        # print("90121-1 load_eli_file_to_db: \n", "="*50, "\n", dic, "\n", "="*50)
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # print(file_path)

        dic = dic["cube_dic"]
        print('90121-3 dic', dic)

        model_name_ = dic["dimensions"]["time_dim"]["model"]
        model_time_dim = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = dic["dimensions"]["country_dim"]["model"]
        model_country_group_dim = apps.get_model(app_label=self.app, model_name="CountryGroupDim")
        model_country_dim = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = dic["dimensions"]["measure_dim"]["model"]
        model_measure_dim = apps.get_model(app_label=app_, model_name=model_name_)

        measure_group_dim_ = dic["dimensions"]["measure_group_dim"]["model"]
        model_measure_group_dim = apps.get_model(app_label=app_, model_name=measure_group_dim_)
        #
        country_group_obj, is_created = model_country_group_dim.objects.get_or_create(group_name="eli")

        model_min_max_cut = apps.get_model(app_label=app_, model_name="minmaxcut")

        model_name_ = dic["fact"]["model"]
        model_fact = apps.get_model(app_label=app_, model_name=model_name_)


        year = int(self.uploaded_filename.split(".")[0])
        # print(year)
        try:
            year_obj, is_created = model_time_dim.objects.get_or_create(id=year)
            if is_created:
                s = 'year_obj.' + dic["dimensions"]["time_dim"]["field_name"]+' = year'
                # print(s)
                exec(s)
                year_obj.save()
        except Exception as ex:
            pass

        # files = []
        wb = load_workbook(filename=file_path, read_only=False)
        sheet_names = wb.sheetnames
        # print(sheet_names)

        for f in sheet_names:
            print("Start", f)
            ws = wb[f]
            f = self.clean_name(f)
            try:
                group_obj, is_created = model_measure_group_dim.objects.get_or_create(group_name=f)
                if is_created:
                    group_obj.group_name = f
                    group_obj.save()
            except Exception as ex:
                pass

            # print("A>\n", f)

            # files.append(f)
            data = ws.values
            # Get the first line in file as a header line
            columns = next(data)[0:]
            # print(columns)
            # Create a DataFrame based on the second and subsequent lines of data
            df = pd.DataFrame(data, columns=columns)
            df = df.reset_index()  # make sure indexes pair with number of rows
            # if f == "Export":
            #     print(df)
            #     print(columns)

            # print("B>\n", f, "\n", df)

            min_cut = []
            max_cut = []
            for j in range(0, len(columns)):
                min_cut.append(None)
                max_cut.append(None)

            # if f == "Export":
            #     print("6", f)

            for index, row in df.iterrows():
                # if f == "Export":
                #     print(row)

                if row[1] is not None and str(row[1]) != "None" and str(row[1]) != "":
                    n_ = 0
                    for j in range(1, len(columns)):
                        if str(columns[j]) != "None":
                            n_ += 1
                            f_ = f + str(n_)
                            try:
                                measure_obj, is_created = model_measure_dim.objects.get_or_create(measure_group_dim=group_obj,
                                                                                                  measure_name=f_)
                                if is_created:
                                    measure_obj.measure_code = f_
                                    measure_obj.save()
                            except Exception as ex:
                                pass

                            s_ = str(row[1]).lower()
                            if s_ == "min_cut" or s_ == "max_cut":
                                # print(columns[j])
                                try:
                                    if s_ == "min_cut":
                                        # print("Min_Cut", row[columns[j]])
                                        min_cut[j] = float(row[columns[j]])
                                    if s_ == "max_cut":
                                        # print("Max_Cut", row[columns[j]])
                                        max_cut[j] = float(row[columns[j]])
                                    if min_cut[j] is not None and max_cut[j] is not None:
                                        # print(f + str(n_), columns[j], "=", min_cut[j], max_cut[j])
                                        try:
                                            min_max_cut_obj, is_created = model_min_max_cut.objects.get_or_create(time_dim=year_obj,
                                                                                                                  measure_dim=measure_obj)
                                            if is_created:
                                                min_max_cut_obj.min = min_cut[j]
                                                min_max_cut_obj.max = max_cut[j]
                                                min_max_cut_obj.save()
                                        except Exception as ex:
                                            pass

                                except Exception as ex:
                                    print("Error 9044-44: "+str(ex))
                            else:
                                try:
                                    country_name = str(row[1]).strip()
                                    country_name = self.check_country_el(country_name)
                                    # print(country_name)

                                    country_dim_obj, is_created = model_country_dim.objects.get_or_create(country_name=country_name)
                                    # print(country_dim_obj)
                                    if is_created:
                                        country_dim_obj.country_code = country_name
                                        country_dim_obj.country_group_dim = country_group_obj
                                        country_dim_obj.save()
                                except Exception as ex:
                                    print("Error 9044-44: "+str(ex))

                                try:
                                    v_ = str(row[columns[j]]).strip()
                                except Exception as ex:
                                    print("Error 9055-22: "+str(ex), columns[j], country_name, "="+v_+"=")
                                try:
                                    if v_ != "" and v_ != "nan" and v_ is not None and v_ != "None":
                                        v_ = float(v_)
                                        # print(row[columns[j]], float(str(row[columns[j]])))
                                        fact_obj, is_created = model_fact.objects.get_or_create(time_dim=year_obj,
                                                                                                country_dim=country_dim_obj,
                                                                                                measure_dim=measure_obj)
                                        fact_obj.amount = v_
                                        fact_obj.save()
                                except Exception as ex:
                                    print("Error 9055-33: "+str(ex), columns[j], country_name, "="+v_+"=")
            print("Done", f)
        wb.close()
        #
        result = {"status": "ok"}
        print(result)

        return result
