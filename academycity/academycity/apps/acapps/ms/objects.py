import warnings
import os
from django.conf import settings
from ..ml.basic_ml_objects import BaseDataProcessing, BasePotentialAlgo
import matplotlib as mpl
from django.apps import apps

import sys
import shutil
import pandas as pd
import numpy as np
import math
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
import pickle
from concurrent.futures import ThreadPoolExecutor as T
from statistics import mean, median
import copy

from openpyxl import Workbook, load_workbook

from collections import OrderedDict
from operator import getitem
from sklearn.decomposition import PCA

mpl.use('Agg')


class MSAlgo(object):
    def __init__(self, dic):  # to_data_path, target_field
        # print("90004-000-1 MSAlgo", dic, '\n', '-'*50)
        super(MSAlgo, self).__init__()
        # print("90004-000-  MSAlgo", dic, '\n', '-'*50)
        # -----
        self.to_save_normalize = []
        self.to_save_similarity = []
        self.to_save_similarity_by_index = []
        self.to_save_score = []


class MSDataProcessing(BaseDataProcessing, BasePotentialAlgo, MSAlgo):
    def __init__(self, dic):
        super().__init__(dic)
        app_ = dic["app"]
        self.Debug = apps.get_model(app_label=app_, model_name="debug")

    def log_debug(self, value):
        self.Debug.objects.create(value=value)

    def clear_log_debug(self):
        self.Debug.truncate()

    # def get_general_data(self, dic):
    #     # print("DataProcessing get_general_data 9012:\n")
    #     # print(dic)
    #     app_ = dic["app"]
    #
    #     # dic = {"app": "avi",
    #     #        "dimensions": {"time_dim": {"model": "TimeDim", "field_name": "year"},
    #     #                       "country_dim": {"model": "CountryDim", "field_name": "country_name"} }}
    #
    #     result = {}
    #     for k in dic["dimensions"]:
    #         dic_ = dic["dimensions"][k]
    #         s = k + ' = {}'
    #         try:
    #             exec(s)
    #             # print(eval(k))
    #             model_name_ = dic["dimensions"][k]["model"]
    #             # print(model_name_)
    #             model_ = apps.get_model(app_label=app_, model_name=model_name_)
    #             # print(model_)
    #             for r in model_.objects.all():
    #                 # print(r)
    #                 s = k + '["'+str(r.id)+'"] = r.' + dic["dimensions"][k]["field_name"]
    #                 # print(s)
    #                 exec(s)
    #             # print(eval(k))
    #         except Exception as ex:
    #             print("err 1000: " + str(ex))
    #
    #         # print('result[k] = ' + k)
    #         exec('result[k] = ' + k)
    #     # print(result)
    #     return result

    # aa = {"1": {"entity": [6, 7, 8, 14, 22, 23, 26, 29, 121, 124, 125, 132, 138, 140, 141, 143, 144, 275, 286, 287, 293, 377, 379, 389, 393, 394, 402, 404, 405, 420, 425, 426, 427, 429, 433, 437, 440, 448, 451, 453, 455, 456, 521],
    #            "centroid": 102.00792790697675,
    #            "entity_value": [94.8761, 100.213, 101.36, 96.5844, 104.179, 100.938, 99.9258, 94.534, 105.97, 103.62, 107.245, 107.458, 96.649, 96.4508, 96.8074, 96.5696, 99.7425, 103.478, 108.431, 107.416, 96.2503, 107.129, 108.6, 101.453, 104.551, 99.9243, 101.714, 101.277, 104.21, 103.059, 104.247, 96.9172, 107.689, 104.112, 101.338, 108.126, 106.549, 101.887, 104.459, 96.4507, 101.083, 103.018, 99.8498]},
    #       "2": {"entity": [25, 34, 35, 36, 38, 42, 45, 47, 49, 52, 71, 72, 73, 75, 77, 78, 85, 87, 88, 89, 91, 94, 99, 100, 101, 103, 107, 109, 110, 111, 112, 122, 127, 130, 134, 146, 147, 148, 150, 152, 155, 157, 172, 176, 179, 189, 190, 218, 219, 220, 221, 222, 223, 225, 234, 247, 253, 277, 289, 301, 310, 326, 329, 330, 332, 336, 341, 344, 353, 370, 373, 376, 382, 386, 391, 398, 411, 413, 417, 422, 430, 441, 445, 450, 460, 461, 463, 466, 467, 479, 486, 496, 499, 503, 506, 510, 520, 523, 524, 526, 527, 530, 533, 541, 544, 547, 550, 558, 562, 565, 566, 568, 569],
    #             "centroid": 129.2145575221239,
    #             "entity_value": [124.876, 134.408, 130.742, 122.931, 131.153, 126.913, 136.156, 128.931, 125.4, 132.59, 134.699, 129.072, 127.707, 130.516, 129.858, 135.772, 134.377, 134.834, 124.833, 133.143, 134.971, 133.228, 132.057, 129.984, 132.651, 135.098, 125.006, 127.951, 122.566, 128.018, 128.271, 129.737, 127.675, 127.442, 122.608, 133.366, 125.565, 123.537, 123.486, 134.543, 127.129, 128.967, 131.203, 131.649, 132.549, 133.487, 130.848, 126.913, 124.261, 130.064, 122.866, 124.315, 126.119, 126.738, 129.627, 128.826, 129.845, 125.252, 127.787, 134.942, 135.753, 130.582, 123.785, 131.767, 127.294, 134.678, 131.356, 130.465, 132.712, 135.263, 125.042, 126.744, 130.998, 128.689, 130.911, 127.96, 134.526, 125.477, 133.323, 123.63, 129.54, 130.181, 123.227, 126.051, 122.506, 128.998, 125.922, 134.873, 132.762, 131.458, 130.019, 134.315, 126.313, 130.068, 133.463, 126.663, 125.805, 124.561, 123.53, 128.393, 129.71, 124.097, 127.843, 124.338, 124.709, 132.822, 133.171, 134.471, 124.755, 136.205, 123.964, 128.798, 126.701]},
    #       "3": {"entity": [1, 4, 5, 9, 10, 11, 12, 13, 15, 16, 17, 18, 19, 113, 137, 139, 142, 145, 375, 381, 388, 407, 428, 449, 454],
    #             "centroid": 87.015004,
    #             "entity_value": [90.8986, 83.492, 87.5978, 88.2278, 87.0293, 92.721, 85.7844, 89.8529, 89.6005, 84.0518, 71.4408, 77.0146, 82.1335, 93.5224, 94.0468, 90.0493, 74.8268, 90.8752, 85.921, 88.2654, 91.2218, 92.7555, 88.8332, 82.3627, 92.85]},
    #       "4": {"entity": [20, 30, 39, 43, 46, 54, 58, 74, 76, 81, 84, 90, 93, 95, 97, 98, 116, 151, 153, 162, 164, 165, 166, 174, 175, 177, 178, 183, 185, 187, 188, 191, 207, 228, 229, 230, 231, 232, 233, 245, 246, 248, 250, 252, 271, 279, 284, 297, 302, 303, 304, 312, 321, 325, 327, 331, 335, 337, 339, 343, 364, 378, 383, 387, 395, 399, 409, 410, 415, 416, 432, 438, 439, 442, 447, 452, 457, 458, 459, 464, 469, 482, 483, 484, 487, 492, 497, 501, 502, 508, 511, 512, 513, 515, 516, 517, 529, 531, 532, 534, 535, 536, 537, 539, 546, 549, 553, 554, 555, 556, 560, 561, 563, 564, 567], "centroid": 143.38422608695652, "entity_value": [136.98, 140.893, 138.841, 145.739, 136.307, 141.19, 144.75, 151.513, 142.754, 144.923, 142.552, 142.241, 142.183, 137.671, 138.355, 142.763, 143.402, 136.342, 141.022, 148.964, 148.658, 149.832, 151.32, 142.752, 146.149, 143.184, 150.441, 147.363, 146.884, 145.332, 140.236, 138.499, 150.595, 144.899, 144.824, 137.592, 139.557, 148.499, 144.697, 136.607, 146.588, 150.624, 141.601, 146.46, 146.61, 147.194, 142.632, 147.765, 140.674, 147.327, 140.269, 142.27, 140.513, 144.192, 138.5, 147.604, 139.245, 137.527, 137.619, 143.18, 148.0, 141.545, 139.72, 148.229, 149.32, 147.63, 150.243, 136.774, 144.067, 145.586, 147.94, 140.632, 145.042, 146.719, 138.954, 148.967, 140.81, 147.999, 139.802, 139.374, 146.974, 142.049, 137.235, 141.647, 146.466, 145.421, 140.778, 146.654, 146.402, 144.77, 141.783, 144.375, 142.848, 145.151, 136.654, 139.8, 149.388, 141.67, 139.933, 136.746, 142.543, 146.551, 138.95, 138.434, 137.725, 146.284, 148.85, 144.043, 138.217, 141.179, 138.713, 147.058, 142.25, 150.598, 137.995]}, "5": {"entity": [2, 3, 21, 24, 27, 28, 33, 37, 40, 48, 50, 82, 105, 106, 108, 114, 115, 117, 118, 119, 120, 123, 126, 128, 129, 131, 133, 135, 136, 149, 154, 156, 158, 159, 160, 217, 224, 227, 249, 268, 276, 278, 288, 290, 291, 292, 294, 295, 296, 311, 338, 340, 342, 390, 392, 397, 403, 406, 412, 419, 421, 423, 424, 431, 434, 435, 443, 444, 446, 462, 470, 518, 519, 522, 525, 528, 543, 545, 552, 557, 559], "centroid": 115.78974074074074, "entity_value": [116.279, 117.357, 112.661, 119.153, 117.267, 117.368, 121.486, 118.137, 115.37, 111.265, 116.879, 117.138, 120.149, 111.851, 120.351, 117.748, 116.91, 120.041, 113.952, 112.608, 117.365, 118.21, 109.982, 119.746, 117.912, 111.578, 120.762, 117.945, 115.609, 118.186, 119.752, 112.918, 113.986, 112.7, 114.721, 110.222, 114.15, 118.212, 120.242, 120.014, 112.523, 115.967, 118.196, 112.711, 117.266, 122.263, 113.139, 116.307, 116.084, 117.113, 114.31, 112.576, 116.668, 110.464, 113.054, 114.797, 116.545, 112.165, 122.084, 113.586, 114.302, 120.731, 114.194, 121.34, 114.357, 114.769, 108.934, 114.167, 113.171, 113.551, 112.522, 115.712, 114.596, 109.759, 117.858, 118.192, 110.738, 110.55, 119.935, 120.663, 112.928]}, "6": {"entity": [31, 41, 51, 53, 55, 56, 57, 59, 60, 61, 62, 64, 66, 67, 69, 79, 80, 96, 102, 104, 161, 163, 168, 170, 173, 184, 186, 214, 226, 262, 263, 264, 265, 266, 267, 274, 300, 306, 308, 309, 313, 316, 320, 322, 333, 334, 347, 349, 354, 356, 358, 363, 372, 380, 385, 396, 400, 401, 408, 414, 436, 468, 471, 472, 474, 475, 480, 481, 485, 493, 505, 507, 514, 538, 540, 548, 551], "centroid": 160.34306493506494, "entity_value": [157.497, 158.597, 152.736, 158.608, 158.703, 160.269, 167.414, 165.165, 156.363, 151.9, 152.431, 158.694, 164.858, 165.919, 160.461, 167.053, 157.308, 157.766, 156.083, 166.504, 160.235, 157.668, 160.777, 162.594, 156.454, 163.912, 164.754, 158.126, 153.831, 158.74, 159.769, 152.694, 156.836, 166.654, 156.316, 158.094, 166.826, 165.429, 152.589, 162.37, 164.01, 166.385, 164.716, 167.986, 154.492, 153.558, 167.172, 163.171, 160.232, 154.686, 165.591, 166.291, 154.284, 157.984, 161.453, 155.733, 152.274, 152.589, 156.409, 164.667, 165.713, 165.928, 153.108, 152.051, 167.136, 165.257, 158.922, 157.573, 166.887, 165.311, 166.602, 158.362, 153.534, 166.416, 160.901, 167.173, 162.862]}, "7": {"entity": [32, 63, 68, 70, 83, 86, 92, 169, 171, 180, 181, 182, 194, 196, 197, 199, 202, 209, 212, 213, 216, 259, 260, 261, 269, 270, 272, 273, 280, 281, 298, 305, 318, 319, 323, 328, 352, 357, 359, 360, 374, 384, 465, 476, 478, 489, 490, 494, 498, 504, 509, 542], "centroid": 177.0655, "entity_value": [179.11, 178.366, 183.752, 169.574, 171.56, 171.915, 181.057, 173.677, 179.613, 168.811, 178.938, 183.118, 181.493, 177.328, 183.217, 183.723, 170.281, 179.777, 171.116, 182.225, 170.146, 183.367, 182.877, 177.333, 177.038, 169.788, 172.523, 179.596, 182.885, 178.825, 182.895, 184.099, 169.625, 177.047, 170.561, 181.037, 173.328, 182.186, 173.019, 171.128, 175.944, 169.097, 176.585, 185.064, 176.857, 182.609, 184.01, 169.902, 171.577, 172.796, 179.185, 175.826]}, "8": {"entity": [44, 193, 235, 236, 240, 242, 244], "centroid": 245.706, "entity_value": [239.956, 251.246, 248.779, 245.051, 246.424, 251.258, 237.228]}, "9": {"entity": [65, 198, 203, 206, 208, 215, 282, 299, 307, 314, 350, 365, 371, 418, 473, 477, 491, 500], "centroid": 193.3478888888889, "entity_value": [201.076, 195.067, 201.082, 186.665, 189.395, 185.873, 190.579, 195.875, 201.206, 193.533, 198.218, 187.407, 191.224, 188.162, 187.508, 193.392, 194.306, 199.694]}, "10": {"entity": [192, 195, 201, 210, 211, 238, 241, 254, 257, 283, 285, 351, 355, 362], "centroid": 227.93814285714285, "entity_value": [228.119, 220.784, 220.839, 221.889, 225.997, 233.195, 230.609, 232.178, 228.11, 230.487, 232.984, 232.038, 228.576, 225.329]}, "11": {"entity": [200, 361, 366, 367, 369], "centroid": 261.4186, "entity_value": [264.783, 259.063, 255.836, 259.67, 267.741]}, "12": {"entity": [237, 243, 345, 368], "centroid": 285.03774999999996, "entity_value": [284.895, 281.998, 287.082, 286.176]}, "13": {"entity": [167, 204, 205, 239, 251, 255, 256, 258, 315, 317, 324, 348, 488, 495], "centroid": 209.51364285714286, "entity_value": [203.084, 218.276, 206.348, 212.639, 208.477, 208.55, 212.342, 216.742, 203.51, 210.168, 213.71, 208.507, 205.225, 205.613]}, "14": {"entity": [346], "centroid": 299.025, "entity_value": [299.025]}}

    # Data upload
    def upload_personal_info_to_db(self, dic):
        # print("90121-1-3: \n", dic, "\n", "="*50)
        self.clear_log_debug()
        self.log_debug("=== upload_personal_info_to_db 100 ===")
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # file_path = "/home/amos/projects/development/academycity/data/ms/datasets/excel/raw_data/RAW DATA (607).xlsx"
        # print('file_path')
        # print(file_path)
        # print('file_path')
        # print('90121-  dic')
        dic = dic["cube_dic"]
        # print('90121-3 dic', dic)

        # dep variables uploaded to fact and factnormalized tables
        model_fact_normalized = apps.get_model(app_label=self.app, model_name="factnormalized")

        model_group_name_ = dic["dimensions"]["person_group_dim"]["model"]
        model_person_group_dim = apps.get_model(app_label=app_, model_name=model_group_name_)

        model_name_ = dic["dimensions"]["person_dim"]["model"]
        model_person_dim = apps.get_model(app_label=app_, model_name=model_name_)

        model_name_ = dic["dimensions"]["gene_group_dim"]["model"]
        model_gene_group_dim = apps.get_model(app_label=app_, model_name=model_name_)

        model_name_ = dic["dimensions"]["gene_dim"]["model"]
        model_gene_dim = apps.get_model(app_label=app_, model_name=model_name_)

        gene_group_obj, is_created = model_gene_group_dim.objects.get_or_create(group_name="dep")
        gene_group_obj.save()
        #
        # qs = model_fact.objects.filter(gene_dim__gene_group_dim_group_name="dep",
        #                                person_dim__person_group_dim__group_name="Normaliaz")
        # df = pd.DataFrame(list(qs.values("gene_dim", "person_dim", "amount")))
        model_name_ = dic["fact"]["model"]
        model_fact = apps.get_model(app_label=app_, model_name=model_name_)

        wb = load_workbook(filename=file_path, read_only=False)
        sheet_names = wb.sheetnames
        for f in sheet_names:
            # print(f)
            pg_obj, is_created = model_person_group_dim.objects.get_or_create(group_name=f)
            pg_obj.save()

            ws = wb[f]
            data = ws.values
            columns = next(data)[0:]
            df = pd.DataFrame(data, columns=columns)
            # print(df)
            if f == "Normaliaz":
                self.log_debug("f=" + f)
                for index, row in df.iterrows():
                    # print(index, row["ID"])
                    try:
                        try:
                            if row["gender"] == "M":
                                row["gender"] = 0
                            else:
                                row["gender"] = 1
                            # print(row["gender"])
                        except Exception as ex:
                            pass
                        person_obj, is_create = model_person_dim.objects.get_or_create(person_code=row["ID"])
                        # print("'", "="*100,"\n",row["ID"])
                        person_obj.person_group_dim = pg_obj
                        try:
                            person_obj.gender = row["gender"]
                            person_obj.age_at_cdna = row["age_at_cDNA"]
                        except Exception as ex:
                            pass
                        person_obj.set_num = row["set_num"]
                        person_obj.save()
                        self.log_debug("SAVED " + str(row["ID"]))

                    except Exception as ex:
                        # pass
                        self.log_debug("Error 50-50-50 " + str(row["ID"]))
                        # print("Error: ", row["ID"], ex)

            if f == "Dependent":
                # print(columns[4:])
                for p  in columns[4:]:
                    gene_obj, is_created = model_gene_dim.objects.get_or_create(gene_group_dim=gene_group_obj,gene_code=p)
                    # print('\ngene_obj\n', gene_obj,'\n')
                    for index, row in df.iterrows():
                        person_obj = model_person_dim.objects.get(person_code=row["ID"])
                        fact_obj, is_created = model_fact.objects.get_or_create(gene_dim=gene_obj, person_dim=person_obj)
                        fact_obj_normalized, is_created = model_fact_normalized.objects.get_or_create(run_number = 100,
                                                                                                      gene_dim=gene_obj,
                                                                                                      person_dim=person_obj)
                        fact_obj.amount = row[p]
                        fact_obj.save()
                        fact_obj_normalized.amount = row[p]
                        fact_obj_normalized.run_number = 100
                        fact_obj_normalized.save()

            if f == "Model":
                for index, row in df.iterrows():
                    person_obj = model_person_dim.objects.get(person_code=row["ID"])
                    person_obj.person_group_dim = pg_obj
                    person_obj.save()

            self.log_debug("Done " + f)
        result = {"status": "ok"}
        return result

    def upload_vars_to_db(self, dic):
        # print("902555-1-3: \n", dic, "\n", "="*50)
        self.log_debug("=== upload_vars_to_db 100 ===")
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # file_path = "/home/amos/projects/development/academycity/data/ms/datasets/excel/raw_data/RAW DATA (607).xlsx"
        # print('file_path')
        # print(file_path)
        # print('file_path')
        # print('90121-  dic')
        dic = dic["cube_dic"]
        # print('90121-3 dic', dic)

        # dep variables uploaded to fact and factnormalized tables
        model_fact_normalized = apps.get_model(app_label=self.app, model_name="factnormalized")
        #
        model_group_name_ = dic["dimensions"]["person_group_dim"]["model"]
        model_person_group_dim = apps.get_model(app_label=app_, model_name=model_group_name_)
        #
        model_name_ = dic["dimensions"]["person_dim"]["model"]
        model_person_dim = apps.get_model(app_label=app_, model_name=model_name_)
        #
        model_name_ = dic["dimensions"]["gene_group_dim"]["model"]
        model_gene_group_dim = apps.get_model(app_label=app_, model_name=model_name_)
        #
        model_name_ = dic["dimensions"]["gene_dim"]["model"]
        model_gene_dim = apps.get_model(app_label=app_, model_name=model_name_)
        #
        gene_group_obj, is_created = model_gene_group_dim.objects.get_or_create(group_name="dep")
        gene_group_obj.save()
        #
        # qs = model_fact.objects.filter(gene_dim__gene_group_dim_group_name="dep",
        #                                person_dim__person_group_dim__group_name="Normaliaz")
        # df = pd.DataFrame(list(qs.values("gene_dim", "person_dim", "amount")))

        model_name_ = dic["fact"]["model"]
        model_fact = apps.get_model(app_label=app_, model_name=model_name_)

        # remove all dep variables
        qsg = model_gene_dim.objects.filter(gene_group_dim__group_name="dep")
        for q in qsg:
            q.delete()

        # convert all persons to normaliaz group
        pg_obj, is_created = model_person_group_dim.objects.get_or_create(group_name="Normaliaz")
        pg_obj.save()
        qs = model_person_dim.objects.all()
        for q in qs:
            q.person_group_dim = pg_obj
            q.save()

        # create group for every folder in excel
        # assign all the persons in that folder to that group
        wb = load_workbook(filename=file_path, read_only=False)
        sheet_names = wb.sheetnames
        # print("="*50)
        for f in sheet_names:
            self.log_debug("A0 f=" + f)
            pg_obj, is_created = model_person_group_dim.objects.get_or_create(group_name=f)
            pg_obj.save()

            ws = wb[f]
            data = ws.values
            columns = next(data)[0:]
            df = pd.DataFrame(data, columns=columns)
            # print(df)

        #
            if f == "Model":
                for p in columns[1:]:
                    gene_obj, is_created = model_gene_dim.objects.get_or_create(gene_group_dim=gene_group_obj,gene_code=p)
                    # print('\ngene_obj\n', gene_obj,'\n')

                    for index, row in df.iterrows():
                        try:
                            v = str(row[p])
                            if v == "" or v == "nan":
                                print("111 person=", row["ID"], "gene=", p, "v=", v)
                                print(float(v))
                                self.log_debug("A v=" + v + "=" + str(row["ID"]) + " gene=" + p + " ")
                                continue

                            v = float(row[p])

                            person_obj = model_person_dim.objects.get(person_code=row["ID"])
                            fact_obj, is_created = model_fact.objects.get_or_create(gene_dim=gene_obj,
                                                                                    person_dim=person_obj)
                            fact_obj_normalized, is_created = model_fact_normalized.objects.get_or_create(
                                run_number = 100,
                                gene_dim=gene_obj,
                                person_dim=person_obj)
                            fact_obj_normalized.run_number = 100
                            fact_obj.amount = row[p]
                            fact_obj.save()

                            fact_obj_normalized.amount = v
                            fact_obj_normalized.save()
                            self.log_debug("A1 saved v=" + str(v) + "=" + str(row["ID"]) + " gene=" + p + " ")
                        except Exception as ex:
                            self.log_debug("A10 ERROR v=" + str(v) + "=" + str(row["ID"]) + " gene=" + p + " ")
                            # print("", row[p], "error", ex)

                for index, row in df.iterrows():
                    try:
                        person_obj = model_person_dim.objects.get(person_code=row["ID"])
                        person_obj.person_group_dim = pg_obj
                        person_obj.save()
                        self.log_debug("A500 SAVED person_code=" + str(row["ID"]))
                    except Exception as ex:
                        self.log_debug("A11 ERROR person_code=" + str(row["ID"]) + " " + str(ex))

        self.log_debug("A200 Done")
        result = {"status": "ok"}
        return result

    def load_file_to_db(self, dic):
        print("90121-1: \n", dic, "="*50)
        self.clear_log_debug()
        self.log_debug("=== load_file_to_db 100 ===")
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # file_path = "/home/amos/projects/development/academycity/data/ms/datasets/excel/raw_data/RAW DATA (607).xlsx"
        # print('file_path')
        # print(file_path)
        # print('file_path')
        # print('90121-  dic')
        dic = dic["cube_dic"]
        # print('90121-3 dic', dic)
        model_name_ = dic["dimensions"]["person_dim"]["model"]
        model_person_dim = apps.get_model(app_label=app_, model_name=model_name_)

        model_name_ = dic["dimensions"]["person_group_dim"]["model"]
        model_person_group_dim = apps.get_model(app_label=app_, model_name=model_name_)

        model_name_ = dic["dimensions"]["gene_group_dim"]["model"]
        model_gene_group_dim = apps.get_model(app_label=app_, model_name=model_name_)

        model_name_ = dic["dimensions"]["gene_dim"]["model"]
        model_gene_dim = apps.get_model(app_label=app_, model_name=model_name_)

        model_name_ = dic["fact"]["model"]
        model_fact = apps.get_model(app_label=app_, model_name=model_name_)

        wb = load_workbook(filename=file_path, read_only=False)
        sheet_names = wb.sheetnames
        for f in sheet_names:
            self.log_debug("f=" + f)
            gene_group_dim_obj, is_created = model_gene_group_dim.objects.get_or_create(group_name=f)
            gene_group_dim_obj.save()
            ws = wb[f]
            data = ws.values
            # Get the first line in file as a header line
            columns = next(data)[0:]
            # print(columns)
            # Create a DataFrame based on the second and subsequent lines of data
            df = pd.DataFrame(data, columns=columns)
            df = df.reset_index()  # make sure indexes pair with number of rows
            # n__ = 0

            list_ = []
            for index, row in df.iterrows():
                # print("index", index)
                for j in range(1, len(columns)):
                    # print("j", j)
                    if row[1] is not None and str(row[1]) != "None" and str(row[1]) != "":
                        g_ = str(row[1])
                        # print("g_", g_)
                        gene_dim_obj, is_created = model_gene_dim.objects.get_or_create(
                            gene_group_dim=gene_group_dim_obj, gene_code=g_)
                        p_ = str(columns[j])
                        if p_ != "None" and p_ != "":
                            person_dim_obj, is_created = model_person_dim.objects.get_or_create(person_code=p_)
                        try:
                            # print(str(row[columns[j]]), "\n", "="*20)
                            v_ = float(str(row[columns[j]]))
                            if (v_ <= -0.000001) or (v_ > 0.000001):
                                fact_obj, is_created = model_fact.objects.get_or_create(gene_dim=gene_dim_obj,
                                                                                        person_dim=person_dim_obj)
                                fact_obj.amount = v_
                                fact_obj.save()
                        except Exception as ex:
                            # print("Error: ", ex, "\n")
                            if p_ not in list_:
                                list_.append(p_)
                                print(list_)
            # print("\n", "="*100, list_, "\n", "="*100)
            self.log_debug("Error List of patients " + str(list_))

            # print(f_, p_, v_)
            # print(n__, max_v, max_d)
        # print(max_v, max_d)
        # print('90121-6 fact')
        self.log_debug("=== Done load_file_to_db 101 ===")
        wb.close()
        result = {"status": "ok"}
        return result

    # =======
    # add the functions below to this function
    # def calculate_clusters(self, dic):
    #     # print("90921-0: \n", dic, "\n", "="*50)
    #     print(dic)
    #     # print('dic')
    #     self.clear_log_debug()
    #     self.log_debug("=== calculate_clusters 100 ===")
    #     app_ = dic["app"]
    #     self.log_debug("=== calculate_clusters 101 " + app_)
    #     model_name_ = dic["dimensions"]["person_dim"]["model"]
    #     # print(model_name_)
    #     model_person_dim = apps.get_model(app_label=app_, model_name=model_name_)
    #     model_name_ = dic["dimensions"]["gene_dim"]["model"]
    #     # print(model_name_)
    #     model_gene_dim = apps.get_model(app_label=app_, model_name=model_name_)
    #     model_name_ = dic["fact"]["model"]
    #     # print(model_name_)
    #     model_fact = apps.get_model(app_label=app_, model_name=model_name_)
    #     qs = model_fact.objects.filter(gene_dim__gene_group_dim__group_name="indep")
    #     df_s = pd.DataFrame(list(qs.values('gene_dim', 'person_dim', 'amount')))
    #     df_s.columns=['gene', 'person', 'amount']
    #     # print(df_s)
    #     df = df_s.pivot_table(values='amount', index='gene', columns=['person'], aggfunc='sum')
    #     # print(df)
    #     df = df.apply(lambda x: x.sort_values().values, axis=1, result_type='broadcast')
    #     # print(df)
    #     df["min"] = df.min(axis=1)
    #     df["max"] = df.max(axis=1)
    #     # print(df)
    #
    #     for index, row in df.iterrows():
    #         # print("="*100)
    #         # print(index)
    #         # print(row)
    #         clusters = self.get_gene_clusters(row)
    #         # print("number of clusters\n\n", len(clusters))
    #         obj = model_gene_dim.objects.get(id=index)
    #         obj.clusters = clusters
    #         # print("="*100)
    #         obj.save()
    #     result = {"status": "ok"}
    #     return result

    def fix_run_in_gene(self, dic):
        model_gene_dim = apps.get_model(app_label=app_, model_name="genedim")
        qs = model_gene_dim.objects.filter(gene_group_dim__group_name="indep")
        for q in qs:
            q.clusters = {"run_1": q.clusters}
            q.save()


    # Fix add run_number
    def calculate_clusters_a(self, dic):
        # print("90921-0: \n", dic, "\n", "="*50)
        app_ = dic["app"]
        df = dic["df"]
        nnn = str(dic["nnn"])
        run_number = str(dic["run_number"])
        # print(nnn, "\n", nnn, "\n", nnn, "\n", "="*10, "\n")
        self.log_debug("calculate_clusters_a A")

        model_gene_dim = apps.get_model(app_label=app_, model_name="genedim")
        # print(df)

        df["min"] = df.min(axis=1)
        df["max"] = df.max(axis=1)
        # print(nnn, "\n", df, "\n", df.shape)

        for index, row in df.iterrows():
            # print("="*100)
            # print(index)
            # print(row)
            self.log_debug("calculate_clusters_a A " + str(index))
            clusters = self.get_gene_clusters(row)
            # print("number of clusters\n\n", len(clusters))
            obj = model_gene_dim.objects.get(id=index)

            if run_number == "1" and nnn == "1":
                # print("both 1")
                dic_ = {"run_1": {1: clusters}}
            else:
                # print("one is not 1")
                dic_ = obj.clusters
                ss_ = "run_"+str(run_number)
                if not ss_ in dic_:
                    dic_[ss_] = {}
                dic_[ss_][nnn] = clusters

            # print(nnn, "\n\n", dic_)
            obj.clusters = dic_
            # print("="*100)
            obj.save()

        result = {"status": "ok"}
        return result

    # Not used  Not updated for run_number
    def calculate_clusters(self, dic):
        # print("90921-0: \n", dic, "\n", "="*50)
        print(dic)
        # print('dic')
        self.clear_log_debug()
        self.log_debug("=== calculate_clusters 100 ===")
        app_ = dic["app"]
        self.log_debug("=== calculate_clusters 101 " + app_)

        model_name_ = dic["fact"]["model"]
        # print(model_name_)
        model_fact = apps.get_model(app_label=app_, model_name=model_name_)

        nnn = 1
        l = [0, 131, 10000]
        while nnn <= 3:
            if nnn < 3:
                qs = model_fact.objects.filter(gene_dim__gene_group_dim__group_name="indep",
                                               person_dim__set_num__gt=l[nnn - 1],
                                               person_dim__set_num__lte=l[nnn])
            else:
                qs = model_fact.objects.filter(gene_dim__gene_group_dim__group_name="indep")

            df_s = pd.DataFrame(list(qs.values('gene_dim', 'person_dim', 'amount')))
            df_s.columns = ['gene', 'person', 'amount']
            # print(df_s)
            df = df_s.pivot_table(values='amount', index='gene', columns=['person'], aggfunc='sum')
            # print(df)
            df = df.apply(lambda x: x.sort_values().values, axis=1, result_type='broadcast')
            # print(df)

            # df["min"] = df.min(axis=1)
            # df["max"] = df.max(axis=1)
            #
            # # print(nnn, "\n", df, "\n", df.shape)
            # for index, row in df.iterrows():
            #     # print("="*100)
            #     # print(index)
            #     # print(row)
            #     clusters = self.get_gene_clusters(row)
            #     # print("number of clusters\n\n", len(clusters))
            #     obj = model_gene_dim.objects.get(id=index)
            #     if nnn == 1:
            #         dic_ = {1: clusters}
            #     else:
            #         dic_[nnn] = obj.clusters
            #     obj.clusters = dic_
            #     # print("="*100)
            #     obj.save()
            dic["nnn"] = nnn
            dic["df"] = df
            self.calculate_clusters_a(dic)
            nnn += 1

        result = {"status": "ok"}
        return result
    # --------

    def get_gene_clusters(self, row):
        # print('row')
        # print("="*50, "\n", row, row.index, "\n", "="*50)
        clusters = {1: {"entity": [], "entity_value": [], "centroid": -1.00}}
        # print(clusters)
        # print(range(len(row)-2))
        try:
            # print(row)
            d0_ = (row["max"] - row["min"])/20
            # print("1. d0_", d0_)
            cluster_n = 1
            # print("-"*30, "\n", d0_, "\n", "="*30)
            for j in row.index:
                if j in ["min", "max"]:
                    continue
                # print(j, row[j])
                if j == 1:
                    clusters[cluster_n]["entity"].append(j)
                    clusters[cluster_n]["entity_value"].append(float(row[j]))
                    clusters[cluster_n]["centroid"]=float(row[j])
                    cluster_n += 1
                    # print("-"*30, "\nj=", j, "\n", clusters, "\n", "="*30)
                else:
                    d_min = 1000000000000
                    c_min = -1
                    for c in clusters:
                        d_ = abs(float(row[j])-clusters[c]["centroid"])
                        if d_ < d_min:
                            d_min = d_
                            c_min = c
                    # print("j=", j, c_min)
                    if d_min < d0_:
                        clusters[c_min]["entity"].append(j)
                        clusters[c_min]["entity_value"].append(float(row[j]))
                    else:
                        clusters[cluster_n]={"entity": [], "entity_value": [], "centroid": -1.00}
                        clusters[cluster_n]["entity"].append(j)
                        clusters[cluster_n]["entity_value"].append(float(row[j]))
                        clusters[cluster_n]["centroid"]=float(row[j])
                        cluster_n += 1
        except Exception as ex:
            print("Error 90876-67: "+str(ex))

        # print('clusters 1', "\n", "="*100, "\n")
        # print(clusters)
        # print('clusters 1', "\n", "="*100, "\n")
        #
        n_ = 0
        l_sum = []
        while n_ < 100:
            # print("n_=", n_)
            sum_ = 0
            clusters_o = copy.deepcopy(clusters)
            # for c in clusters_o:
            #     print(clusters_o[c]["centroid"])
            clusters = self.converge_clusters_centroid(row, clusters_o, n_)
            for c in clusters:
                # print(c, clusters[c]["centroid"], clusters_o[c]["centroid"])
                sum_ += abs(clusters[c]["centroid"] - clusters_o[c]["centroid"])**2
            l_sum.append(sum_)
            if sum_ < 0.000001:
                break
            n_ += 1
        # print(n_, l_sum)
        for c in clusters:
            clusters[c]["centroid"] = round(clusters[c]["centroid"], 2)
        # print('clusters 2', "\n", "="*100, "\n")
        # print(clusters)
        # print('clusters 2', "\n", "="*100, "\n")
        return clusters

    def converge_clusters_centroid(self, row, clusters_o, n):
        clusters = copy.deepcopy(clusters_o)
        # print("="*10, "\n", n, "\n", "="*10)
        for c in clusters:
            # print("c=", c, "\n", clusters[c]["entity_value"])
            clusters[c]["centroid"] = mean(clusters[c]["entity_value"])
            clusters[c]["entity"]=[]
            clusters[c]["entity_value"]=[]
        for j in row.index:
            if j in ["min", "max"]:
                continue
            d_min = 1000000000000
            c_min = -1
            for c in clusters:
                d_ = abs(float(row[j]) - clusters[c]["centroid"])
                if d_ < d_min:
                    d_min = d_
                    c_min = c
            clusters[c_min]["entity"].append(j)
            clusters[c_min]["entity_value"].append(float(row[j]))
        clusters_o = copy.deepcopy(clusters)
        for c in clusters_o:
            if len(clusters_o[c]["entity_value"]) == 0:
                del clusters[c]
        return clusters
    #

    # NeedToDo to move to function calculate_clusters
    def create_homogeneous_genes_list(self, dic):
        self.clear_log_debug()
        self.log_debug("=== create_homogeneous ===")
        # print("90950-10: create_homogeneous_genes_list\n", dic, "\n", "=" * 50)
        app_ = dic["app"]
        data_name_ = dic["data_name"]
        group_ = dic["group"]
        number_of_patients_ = dic["number_of_patients"]
        # self.log_debug("number_of_patients_=="+str(number_of_patients_))
        model_name_ = dic["dimensions"]["gene_dim"]["model"]
        model_gene_dim = apps.get_model(app_label=app_, model_name=model_name_)
        qs_genes = model_gene_dim.objects.all()
        model_name_ = dic["dimensions"]["person_dim"]["model"]
        model_person_dim = apps.get_model(app_label=app_, model_name=model_name_)

        ll_g = []
        self.log_debug("= create_homogeneous 1 =")
        n_ = 0
        for q in qs_genes:
            n_ += 1
            # print("q: ", q, "\n", "=" * 50)
            clusters = q.clusters
            # print(clusters)
            ll_c = []
            for i in clusters:
                entity = clusters[i]["entity"]
                # print("i: ", i, "\nentity=", entity, "\n", "=" * 50)
                # print("number_of_patients_: ", str(number_of_patients_))
                # print("len(entity)=", str(len(entity)), "\n", "=" * 50)
                number_of_females = 0
                number_of_male = 0
                if len(entity) >= int(number_of_patients_):
                    for x in entity:
                        person_obj_ = model_person_dim.objects.get(id=x)
                        # self.log_debug("x="+str(x) +" gender="+str(person_obj_.gender))
                        if person_obj_.gender == 1:
                            number_of_females += 1
                        elif person_obj_.gender == 0:
                            number_of_male += 1
                        # print(number_of_females, number_of_male)
                    # self.log_debug("f=" + str(number_of_females) + " m " + str(number_of_male))
                    if number_of_females == 0 or number_of_male == 0:
                        ll_c.append(i)
                        # self.log_debug("cluster num:" + str(i))
                        # self.log_debug("cluster num:" + str(i)+" : ll_c: " +str(ll_c))
            # print("cluster number:", i, "\ll_c:", ll_c)
            if len(ll_c) > 0:
                ll_g.append(q.id)
                self.log_debug("gene w =" + str(q.id))
            if n_ % 100 == 0:
                self.log_debug("run gene =" + str(n_))

        # print("gene number:", q.id, "\ll_g:", ll_g)
        # print("done processing gene: " + str(ll_g))
        # print("ll_g : " + str(ll_g)[:25])
        self.log_debug("done processing gene: ")
        self.log_debug("ll_g : " + str(ll_g)[:25])

        try:
            model_general_data = apps.get_model(app_label="core", model_name="generaldata")
            obj, is_created = model_general_data.objects.get_or_create(app=app_, group=group_, data_name=data_name_)
            obj.data_json = {"data": ll_g}
            obj.save()
        except Exception as ex:
            print("Error 9026-67: "+str(ex))
            self.log_debug("900  Error saving genes: ")

        self.log_debug("saved gene data")

        result = {"status": "ok", "data": ll_g}
        return result

    # Create blocks
    # pop 7
    # count all people
    # (1) if 7/total people < 5% (make parameters T) then combine to the biggest (by nuber of peoples) block
    # (2) a get the lowest block it is 1  compare it to the pick of the other block if it is biger combine

    # ----- Batch Normalization -----
    def add_peaks_to_clusters(self, dic):
        print("90966-66: add_peaks_to_clusters\n", dic, "\n", "=" * 50)

        self.clear_log_debug()
        app_ = dic["app"]
        nnn = dic["nnn"]

        run_number = dic["run_number"]
        t_pop = dic["t_pop"]
        model_person_dim = apps.get_model(app_label=app_, model_name="persondim")
        model_gene_dim = apps.get_model(app_label=app_, model_name="genedim")

        # -----------------
        l = [0, 131, 10000]
        # -----------------
        if run_number == 1:
            model_fact = apps.get_model(app_label=app_, model_name='fact')
            qsf = model_fact.objects.filter(gene_dim__gene_group_dim__group_name="indep",
                                            person_dim__set_num__gt=l[nnn - 1],
                                            person_dim__set_num__lte=l[nnn])
        else:
            model_fact = apps.get_model(app_label=app_, model_name='factnormalized')
            run_number_ = run_number-1
            qsf = model_fact.objects.filter(run_number=run_number_,
                                            gene_dim__gene_group_dim__group_name="indep",
                                            person_dim__set_num__gt=l[nnn - 1],
                                            person_dim__set_num__lte=l[nnn])

        model_fact_normalized_temp = apps.get_model(app_label=app_, model_name='factnormalizedtemp')
        self.log_debug(str(nnn) + " A")
        qsp = model_person_dim.objects.filter(set_num__gt=l[nnn-1], set_num__lte=l[nnn]).all()
        # -------
        self.log_debug(str(nnn) + " B")

        df_f = pd.DataFrame(list(qsf.values('gene_dim', 'person_dim', 'amount')))
        df_f.columns = ['gene', 'person', 'amount']
        df = df_f.pivot_table(values='amount', index='gene', columns=['person'], aggfunc='sum')
        # print("AAAAA : ", nnn, " : l[nnn-1], l[nnn]\n", l[nnn-1], l[nnn], "\n\n", df, "\n\n", df.shape, "\n", "="*50, "\n\n")


        self.log_debug(str(nnn) + " C")
        dic["df"] = df.copy()
        self.log_debug(str(nnn) + " D")
        self.calculate_clusters_a(dic)
        # self.clear_log_debug()
        self.log_debug("calculate_clusters_a finished")
        self.log_debug(str(run_number) + " : " + str(nnn) + " E")
        # print("JJJ\n", nnn, "\n", df, "\n", df.shape)
        # --
        df_s = pd.DataFrame(list(qsp.values('id', 'person_code', 'set_num')))
        df_s.columns=['index', 'person_code', 'set_num']
        df_s = df_s.dropna(axis=0) # .reset_index()
        # print("df_s\n", df_s)

        self.log_debug(str(nnn) + " F")

        llb = df_s['set_num'].unique().tolist()

        dic_sets = {}

        for k in llb:
            # print("Start k=", k, "\n", "-"*20)
            df_sk = df_s[df_s['set_num']==k]
            dic_sets[k] = df_sk

        for k in dic_sets:
            # print(k)
            df_sk = dic_sets[k]
            llk = df_sk['index'].tolist()
            dfllk = df.loc[:, llk]  #
            dic_sets[k] = dfllk

            dfllk["min"] = dfllk.min(axis=1)
            dfllk["max"] = dfllk.max(axis=1)
            dic_sets[k] = dfllk

        #
        def take_key(elem):
            return elem[2]

        n__ = 0
        qsg = model_gene_dim.objects.filter(gene_group_dim__group_name="indep")
        nz = 0
        for o in qsg:
            nz+=1
            self.log_debug(str(nnn) + " : " + str(nz) + " : " + str(o.id))
            # print(str(nnn) + " : " + str(nz) + " : " + str(o.id))
            n__ += 1
            try:
                clusters_ = o.clusters["run_"+str(run_number)][str(nnn)]
                # print("AAA\nclusters_\n", clusters_)
            except Exception as ex:
                print(ex)

            ll = []
            clusters__ = {}
            for c in clusters_:
                ce = clusters_[c]["entity"]
                cd = clusters_[c]["centroid"]
                n = len(ce)
                ll.append([c, n, cd])
            ll.sort(key=take_key)
            ll_ = [j[1] for j in ll]
            ll__ = [[j[0], clusters_[j[0]]] for j in ll]
            n_clusters__ = 1
            for h in ll__:
                clusters__[n_clusters__] = h[1]
                n_clusters__ += 1

            dic_sets_o = {}

            for k in dic_sets:
                # print(k, "\n", o.id, "\n", dic_sets[k], "\n")
                dic_sets_o[k] = pd.DataFrame(dic_sets[k].loc[o.id])

            r_ = self.get_peaks({"cl_all":ll_, "t_pop": t_pop, "clusters": clusters__, "dic_sets_o": dic_sets_o,
                                     "model_fact_normalized": model_fact_normalized_temp,
                                     "model_person_dim": model_person_dim,
                                     "gene_obj":o, "run_number":run_number})
            # Should delete this ?
            if nnn == 2:
                r = r_["peak_array"]
                # print("\nSummary:\n", "="*50, "\n", r, "\n", "="*150)
                o.reduced_clusters = r
                o.save()

        self.log_debug("Done")
        result = {"status": "ok", "result": {"a": "a"}}
        # print(result)
        return result

    def add_peaks_to_clusters3(self, dic):
        print("90966-66-3: add_peaks_to_clusters3\n", dic, "\n", "=" * 50)
        self.clear_log_debug()
        self.log_debug("add_peaks_to_clusters3")
        app_ = dic["app"]
        nnn = dic["nnn"]
        run_number = dic["run_number"]
        t_pop = dic["t_pop"]
        model_person_dim = apps.get_model(app_label=app_, model_name="persondim")
        model_gene_dim = apps.get_model(app_label=app_, model_name="genedim")
        model_fact = apps.get_model(app_label=app_, model_name='fact')
        model_fact_normalized = apps.get_model(app_label=app_, model_name='factnormalized')
        model_fact_normalized_temp = apps.get_model(app_label=app_, model_name='factnormalizedtemp')
        # ------
        l = [0, 131, 10000]
        self.log_debug(str(nnn) + " A")
        qsp = model_person_dim.objects.all()
        # -------
        qsf = model_fact_normalized_temp.objects.filter(run_number=run_number,
                                                        gene_dim__gene_group_dim__group_name="indep")
        df_f = pd.DataFrame(list(qsf.values('gene_dim', 'person_dim', 'amount')))
        df_f.columns = ['gene', 'person', 'amount']
        self.log_debug(str(nnn) + " B")
        df = df_f.pivot_table(values='amount', index='gene', columns=['person'], aggfunc='sum')
        # print("ZZZZZ\n\n", df, "\n", df.shape)
        self.log_debug(str(nnn) + " C")

        dic["df"] = df.copy()
        self.log_debug(str(nnn) + " D")
        # self.calculate_clusters_a(dic)
        self.log_debug(str(nnn) + " E")
        # print("JJJ\n", nnn, "\n", df, "\n", df.shape)
        # --
        df_s = pd.DataFrame(list(qsp.values('id', 'person_code', 'set_num')))
        df_s.columns=['index', 'person_code', 'set_num']
        df_s = df_s.dropna(axis=0) # .reset_index()
        # print("df_s\n", df_s)
        self.log_debug(str(nnn) + " F")
        self.log_debug(str(df_s.shape))

        dic_sets = {}
        reference_set = 2
        # print("df_s\n", df_s)

        try:
            for k in [1, 2]:
                # print("k", k)
                df_sk = df_s[l[k-1] < df_s['set_num']]
                df_sk = df_sk[l[k] >= df_s['set_num']]
                llk = df_sk['index'].tolist()
                self.log_debug(str(k) + " \nlist\n" + str(llk)[:1000])
                # self.log_debug(str(k) + " \ndf.columns\n" + str(df.columns)[:1000])
                # print("llk\n", llk)

                dfllk = df.loc[:, llk]  #

                # print("dfllk\n", dfllk)
                dic_sets[k] = dfllk
                self.log_debug(str(k) + " \n" + str(dfllk.shape))
        except Exception as ex:
            # print(str(ex))
            self.log_debug("Error11-1 " + str(ex))
            self.log_debug(str(nnn) + " F1")
            # print(dic_sets)

        n__ = 0
        nz = 0
        try:
            qsg = model_gene_dim.objects.filter(gene_group_dim__group_name="indep")
        except Exception as ex:
            self.log_debug("Error11-  " + str(ex))

        self.log_debug(str(nnn) + " I")

        # print(qsg.count())

        for o in qsg:
            nz+=1
            self.log_debug("A1 " + str(run_number) + " : "  + str(nnn) + " : " + str(nz) + " : " + str(o.id))
            # print(str(nnn) + " : " + str(nz) + " : " + str(o.id))
            n__ += 1

            try:
                dic_sets_o = {}
                for k in dic_sets:
                    dic_sets_o[k] = pd.DataFrame(dic_sets[k].loc[o.id])
                # print("B")
                mr = float(dic_sets_o[reference_set].median())
            except Exception as ex:
                self.log_debug("Error11-5 " + str(run_number) + " : "  + str(nnn) + " : " + str(nz) + " : " + str(o.id) + " Error " + str(ex))

            # self.log_debug("A  " + str(nnn) + " : " + str(nz) + " : " + str(o.id))

            for k in dic_sets:
                if k != reference_set:
                    try:
                        # print("dic_sets_o[k] A\n", dic_sets_o[k])
                        mk = float(dic_sets_o[k].median())
                        f = mr/mk
                        dic_sets_o[k] = dic_sets_o[k].astype('float') * f
                        # print(mr, mk, f)
                        # print("dic_sets_o[k] B\n", dic_sets_o[k])
                    except Exception as ex:
                        self.log_debug("Error2  " + str(nnn) + " : " + str(nz) + " : " + str(o.id) + " Error " + str(ex))

                for index, row in dic_sets_o[k].iterrows():
                    try:
                        obj_p = model_person_dim.objects.get(id=index)
                        obj, is_created = model_fact_normalized.objects.get_or_create(run_number=run_number,
                                                                                      gene_dim=o, person_dim=obj_p)
                        obj.amount = float(row[dic_sets_o[k].columns[0]])
                        obj.run_number = run_number
                        obj.save()
                    except Exception as ex:
                        self.log_debug("Error33 " + str(nnn) + " : " + str(nz) + " : " + str(o.id) + " Error " + str(ex))
                        print("Error 1", ex)

        self.log_debug("Done add_peaks_to_clusters3")
        result = {"status": "ok", "result": {"a": "a"}}
        print(result)
        return result

    def get_peaks(self, dic):
        # print(" 90955-50-1: get_peaks\n", dic, "\n", "=" * 50)
        l = dic["cl_all"]
        clusters_ = dic["clusters"]
        dic_sets_o = dic["dic_sets_o"]
        gene_obj = dic["gene_obj"]
        model_fact_normalized = dic["model_fact_normalized"]
        model_person_dim = dic["model_person_dim"]
        run_number = dic["run_number"]
        t_pop = int(dic["t_pop"])/100
        t_pop = int(t_pop * sum(l))
        lb = 0
        ub = len(l)
        peak_array = {}
        def get_location_of_left_low(l_, lb_, ub_):
            i = ub_
            x = 0
            y = 0
            while i > lb_ and x < 2:
                if l_[i] < l_[i - 1]:
                    x += 1
                    y += 1
                elif l_[i] > l_[i - 1]:
                    x = 0
                    y = 0
                else:
                    y += 1
                i -= 1
            if x > 1:
                ret = i + y
            else:
                ret = i
            return ret

        def get_location_of_right_low(l_, lb_, ub_):
            i = lb_
            x = 0
            y = 0
            while i < (ub_-1) and x < 2:
                if l_[i] < l_[i + 1]:
                    x += 1
                    y += 1
                elif l_[i] > l_[i + 1]:
                    x = 0
                    y = 0
                else:
                    y += 1
                i += 1
            if x > 1:
                ret = i - y
            else:
                ret = i + 1
            return ret

        def get_global_high(l_, lb_, ub_):
            temp_ = 0
            ret = -1
            # print(range(lb_, ub_))
            for i in range(lb_, ub_):
                # print(i, ub_)
                if l_[i] > temp_:
                    temp_ = l_[i]
                    ret = i
            # print(ret, temp_)
            return ret

        def get_peaks_(l_, lb_, ub_, peak_array_):
            num_peaks_ = len(peak_array_)+1
            # print("G1", l_, "\nnum_peaks_=", num_peaks_, ", lb_=", lb_, ", ub_=", ub_)
            gh = get_global_high(l_, lb_, ub_)
            ll = get_location_of_left_low(l_, lb_, gh)
            rl = get_location_of_right_low(l_, gh, ub_)
            # print(" AAA Base 0: gh=", gh, ", Base 0: ll=", ll, ", Base 0: rl=", rl)
            peak_array[num_peaks_] = {}
            peak_array[num_peaks_]["peak"] = gh + 1
            peak_array[num_peaks_]["lb"] = ll + 1
            rl_ = rl + 1
            if rl_ > ub_:
                rl_ = ub_
            peak_array[num_peaks_]["ub"] = rl_
            peak_array[num_peaks_]["valid"] = True
            p = 0
            # print("AAA ", l_, "\n", "ll=", ll, "rl=", rl, "\n", "-"*100)
            z = rl + 1
            if z > ub_:
                z = ub_
            for k in range(ll, z):
                p += l_[k]
            peak_array[num_peaks_]["pop"] = p
            # print(" peak_array\n", peak_array)
            if (ll - lb_) > 0:
                # print("A ll - lb_", ll - lb_)
                get_peaks_(l_, lb_, ll+1, peak_array_)
            if (ub_ - rl - 1) > 0:
                # print("B ub_ - rl", ub_ - rl)
                get_peaks_(l_, rl, ub_, peak_array_)

        def two_rules_combination_2(ll, peaks):
            n_of_b = peaks["number_of_blocks"]
            # print("ll=", ll)
            # print("+"*20, "\n", peaks, "\n", "+"*20)
            z = 2
            while z <= n_of_b:
                # print("AAA\nA1=", peaks[z - 1]['ub'], "ub-3=: ", ll[peaks[z - 1]['ub'] - 3], "\npeaks[z]['peak']=",
                #       peaks[z]['peak'],
                #       "\nll[peaks[z]['peak']-1]=", ll[peaks[z]['peak'] - 1])
                # print("AAA\nA1 ll[peaks[z-1]['ub']-3]=", ll[peaks[z-1]['ub']-3], " > ", "\nA  ll[peaks[z]['peak']-1]=",
                #       ll[peaks[z]['peak']-1])
                # print("BBB\n B1 ", ll[peaks[z-1]['peak']-1],"\n < B  ",ll[peaks[z]['lb']+1])

                # if (ll[peaks[z-1]['ub']-3] > ll[peaks[z]['peak']-1]) or (ll[peaks[z-1]['peak']-1] < ll[peaks[z]['lb']+1]):
                if (ll[peaks[z-1]['ub']-2] > ll[peaks[z]['peak']-1]) or (ll[peaks[z-1]['peak']-1] < ll[peaks[z]['lb']]):
                    peaks[z - 1]["ub"] = peaks[z]["ub"]
                    peaks[z - 1]["pop"] += peaks[z]["pop"] - ll[peaks[z]["lb"]-1]
                    # del peaks[z]
                    peaks_ = {}
                    n_ = 1
                    for j in range(1, n_of_b+1):
                        if j != z:
                            peaks_[n_] = peaks[j].copy()
                            # print(n_, peaks_)
                            n_ += 1
                    peaks_["number_of_blocks"] = peaks["number_of_blocks"] - 1
                    peaks = peaks_.copy()
                    n_of_b = peaks["number_of_blocks"]
                else:
                    z += 1
                # print(peaks)
            return peaks
            # elif n_of_b == 3:
            #     z = 2
            #     if ll[peaks[z-1]['ub']-2] > ll[peaks[z]['peak']-1]:
            #         peaks[z - 1]["ub"] = peaks[z]["ub"]
            #         peaks[z - 1]["pop"] += peaks[z]["pop"] - ll[peaks[z]["lb"]-1]
            #         del peaks[z]
            #         n+=1
            #     elif ll[peaks[z-1]['peak']-1] < ll[peaks[z]['lb']]:
            #         # print("BBB")
            #         peaks[z]["lb"] = peaks[z-1]["lb"]
            #         peaks[z]["pop"] += peaks[z-1]["pop"] - ll[peaks[z-1]["ub"]-1]
            #         del peaks[z-1]
            # else:

        def two_rules_combination_1(ll, peaks, t):
            llb = []
            for b in peaks:
                if str(b).isnumeric():
                    if peaks[b]["pop"] < t:
                        llb.append(b)
            for h in llb:
                if ((h+1) in peaks) and ((h-1) in peaks):
                    if peaks[h+1]["pop"] > peaks[h-1]["pop"]:
                        peaks[h + 1]["lb"] = peaks[h]["lb"]
                        peaks[h + 1]["pop"] += peaks[h]["pop"] - ll[peaks[h]["ub"]-1]
                    else:
                        peaks[h - 1]["ub"] = peaks[h]["ub"]
                        peaks[h - 1]["pop"] += peaks[h]["pop"] - ll[peaks[h]["lb"]-1]
                elif (h+1) in peaks:
                    peaks[h + 1]["lb"] = peaks[h]["lb"]
                    peaks[h + 1]["pop"] += peaks[h]["pop"] - ll[peaks[h]["ub"]-1]
                elif (h-1) in peaks:
                    peaks[h - 1]["ub"] = peaks[h]["ub"]
                    peaks[h - 1]["pop"] += peaks[h]["pop"] - ll[peaks[h]["lb"]-1]
                del peaks[h]
                peaks["number_of_blocks"] -= 1
            # print("B Two rules: \nB1. Remove blocks lower then t_pop=", t, "\n", "-"*10, "\n", peaks, "\n", "-"*10)

            # n_of_b = peaks["number_of_blocks"]
            # if n_of_b > 1:
            #     print("+"*20, "\n", "+"*20, "\n", peaks, "\n", "+"*20, "\n", "+"*20)
            #     z=2
            #     # print(ll)
            #     n=0
            #     # print(n_of_b)
            #     peaks_ = {}
            #     while z <= n_of_b:
            #         if ll[peaks[z-1]['ub']-2] > ll[peaks[z]['peak']-1]:
            #             peaks[z - 1]["ub"] = peaks[z]["ub"]
            #             peaks[z - 1]["pop"] += peaks[z]["pop"] - ll[peaks[z]["lb"]-1]
            #             del peaks[z]
            #             n+=1
            #         elif ll[peaks[z-1]['peak']-1] < ll[peaks[z]['lb']]:
            #             # print("BBB")
            #             peaks[z]["lb"] = peaks[z-1]["lb"]
            #             peaks[z]["pop"] += peaks[z-1]["pop"] - ll[peaks[z-1]["ub"]-1]
            #             del peaks[z-1]
            #             n+=1
            #         n_ = 1
            #         for z in peaks:
            #             peaks_[n_] = peaks[z]
            #
            #         n_of_b -= 1
            #
            #     if n > 0:
            #         peaks["number_of_blocks"] -= n
            #         peaks_={}
            #         n_ = 0
            #         for z_ in peaks:
            #             if z_ != "number_of_blocks":
            #                 n_ += 1
            #                 peaks_[n] = peaks[z_].copy()
            #             else:
            #                 peaks_[z_] = peaks[z_]
            #         peaks = peaks_.copy()
            #         # print("aaa\n", peaks, "\n", "aaa")

        def get_block_median(total_sub_pop,  pop_05, hgr, hgl):
            # print("in:", total_sub_pop)
            if hgr + 1 <= ub_ and hgl - 1 >= lb_:
                if l[hgr + 1] > l[hgl - 1]:
                    total_sub_pop += l[hgr + 1]
                    hgr += 1
                else:
                    total_sub_pop += l[hgl - 1]
                    hgl -= 1
            elif hgr + 1 <= ub_:
                total_sub_pop += l[hgr + 1]
                hgr += 1
            elif hgl - 1 >= lb_:
                total_sub_pop += l[hgl - 1]
                hgl -= 1

            if total_sub_pop < pop_05:
                p, hgl, hgr = get_block_median(total_sub_pop, pop_05, hgr, hgl)
                return p, hgl, hgr

            # print("return: ", total_sub_pop, hgl, hgr)
            return total_sub_pop, hgl, hgr

        ##print("'", "="*100)
        ##print(" Find median of sets")
        ##print("'", "="*100)
        sets = {}
        for k in dic_sets_o:
            df = dic_sets_o[k]
            # if k == 22:

            # print("Set number= ", k)
            # print(df.columns[0])
            # print(df[int(df.columns[0])])

            # print(df[int(df.columns[0])].index)
            # print(" set=", k, "\n", "-"*10,"\n", df[int(df.columns[0])].astype(float).values.tolist())
            min_ = float(df.loc["min"])
            max_ = float(df.loc["max"])
            # print("\nset-min=", min_, "\set-max=", max_, "\n")
            # print("-"*100)

            for c_ in clusters_:
                c_min = min(clusters_[c_]['entity_value'])
                c_max = max(clusters_[c_]['entity_value'])

                # if k in [47, 75]:
                #     print(k, "\n min_=", min_, " max_=", max_)
                #     # print("\n\n", clusters_[c_]['entity_value'], "\n")
                #     print(k, "cluster number=", c_, "min=", c_min, "max=", c_max, "\n")

                if c_min <= min_ <= c_max:
                    min_c_ = c_
                    # if k in [47, 75]:
                    #     print("min cluster= ", c_)
                if c_min <= max_ <= c_max:
                    max_c_ = c_
                    # if k in [47, 75]:
                    #     print("max cluster= ", c_)

            # if k in [47, 75]:
            #     print(k, "min_=", min_, " max=", max_)
            #     print(k, " min and max clusters: min_c_= ", min_c_, " max_c_= ", max_c_, "\n", "-"*25)

            rll = []
            for c_ in clusters_:
                if int(max_c_) >= int(c_) >= int(min_c_) :
                    rll += clusters_[c_]['entity_value']

            # if k in [47, 75]:
            #     print(k, " values in the clusters between the min and the max=\n", rll, "\n number of values=", len(rll))
            #     break

            if len(rll) > 0:
                # print(" Median=", median(rll))
                sets[k] = (median(rll), df)
            else:
                print(k, " NONE")
            # print("-"*100)

        # print(sets)
        # --

        get_peaks_(l, lb, ub, peak_array)
        peak_ = sorted(peak_array.items(), key=lambda x: (x[1]["peak"], x[0]))
        n=0
        peak__ = {}
        for a in peak_:
            n+=1
            peak__[n] = a[1]
        peak_array = peak__.copy()
        peak_array["number_of_blocks"] = len(peak_array)
        ##print(" t_pop=", dic["t_pop"]+"%, t_pop=", t_pop)
        ##print(" A Create Blocks for l=", l, "sum=", sum(l), "\n","-"*100,"\n", peak_array, "\n", "-"*50, "\n Two_rules_combination")
        if peak_array["number_of_blocks"] > 1:
            two_rules_combination_1(l, peak_array, t_pop)
            # print(" After rule 1\n", "-"*10, "\n", peak_array, "\n", "-"*50)
        if peak_array["number_of_blocks"] > 1:
            # print(" Rule 2\n", "-"*10)
            peak_array = two_rules_combination_2(l, peak_array)
        ## print(" Final Blocks\n", "-"*10, "\n", peak_array, "\n")
        # print("AAAAAAAAAA")
        for b in peak_array:
            if b == "number_of_blocks":
                continue
            lb_ = peak_array[b]["lb"] - 1
            ub_ = peak_array[b]["ub"] - 1
            ##print("\nsupper cluster (block)=", b, "\n", peak_array[b], lb_, ub_, peak_array[b]["pop"], "\n", "'",  "="*50)

            hg = get_global_high(l, peak_array[b]["lb"]-1, peak_array[b]["ub"]-1)
            pop_05_ = 0.5*peak_array[b]["pop"]
            total_sub_pop_ = l[hg]
            # print("total_sub_pop_=", total_sub_pop_, "pop_05_=", pop_05_)

            hgl_ = hg
            hgr_ = hg
            p = ""
            if total_sub_pop_ < pop_05_:
                p, hgl_, hgr_ = get_block_median(total_sub_pop_, pop_05_, hgl_, hgr_)

            ##print("p=", p, "hgl_=", hgl_, "hgr_=", hgr_)
            all_entities_values = []

            hgl_ += 1
            hgr_ += 1
            for c_ in clusters_:
                if hgl_ <= int(c_) <= hgr_:
                    all_entities_values += clusters_[c_]["entity_value"]
                    # print(c_, clusters_[c_]["entity_value"],"\n")
            # print("genes values=", all_entities_values, "\nmedian= ", median(all_entities_values))
            ##print(" Compact Blocks:\n block:", b, " pop=", p, " left=", hgl_, " right=", hgr_, "Median=", median(all_entities_values))
            peak_array[b]["compact_block"] = [hgl_, hgr_, median(all_entities_values)]

        # print("peak_array\n", peak_array, "\nsets\n", sets)
        for s in sets:
            # if s == 174:
            #     print("set", s, sets[s])


            # ================================
            m = sets[s][0]
            # (median(rll), df)

            # A change we did on Sept 20 2024
            m = float(sets[s][1].median())
            # ================================

            # print(sets[s][1])
            # print("m=", m, "\n\nsm=", float(sets[s][1].median()))

            bs = 1000000000
            mcb_ = -1
            for b in peak_array:
                if b == "number_of_blocks":
                    continue
                mcb__ = abs(peak_array[b]["compact_block"][2]-m)
                if mcb__ <= bs:
                    mcb_ = peak_array[b]["compact_block"][2]
                    bs = mcb__
                # if s == 174:
                #     # print("set", s, sets[s])
                #     print("b", b, "m", m, "bv", peak_array[b]["compact_block"][2],
                #           "mcb__", mcb__, "bs", bs, "mcb_", mcb_)

            if mcb_ > -1:
                dfs = sets[s][1]
                # print("  peoples in this set:\n", dfs)
                for index, row in dfs.iterrows():
                    # print("ROW", index, row)
                    if index in ["min", "max"]:
                        continue
                    # print(int(dfs.columns[0]))

                    try:
                        obj_p = model_person_dim.objects.get(id=index)
                        # if s == 32:
                        # print("gene:", gene_obj.gene_code,
                        #       "set:", s,
                        #       "SetMedian=", round(100000*m)/100000,
                        #       "CompactBlockM=", round(100000*mcb_)/100000,
                        #       "Ratio=", round(100000*mcb_/m)/100000,
                        #       "person("+str(index)+")=", obj_p.person_code, "Amount=",
                        #       float(row[int(dfs.columns[0])]), ">> NAmount=",
                        #       round(100000*mcb_ * float(row[int(dfs.columns[0])])/m)/100000)
                        obj, is_created = model_fact_normalized.objects.get_or_create(run_number=run_number,
                                                                                      gene_dim=gene_obj,
                                                                                      person_dim=obj_p)
                    except Exception as ex:
                        print("Error 1", ex)

                    try:
                        obj.amount = mcb_ * float(row[int(dfs.columns[0])])/m
                    except Exception as ex:
                        print("Error 2", ex)
                    try:
                        obj.run_number = run_number
                        obj.save()
                    except Exception as ex:
                        print("Error 3", ex)

        result = {"status": "ok", "peak_array": peak_array, "sets": sets}
        return result

    # NeedToDo to move to function calculate_clusters
    def get_gene_structure(self, dic):
        # print("90950-10: \n", dic, "\n", "=" * 50)
        self.log_debug(str(dic))
        gene_id_ = int(dic["gene_id"])
        self.log_debug("gene_id=" + str(gene_id_))
        app_ = dic["app"]
        threshold = dic["threshold"]
        if threshold == "":
            threshold = "0.00"
        threshold = 0.7 + float(threshold)/100
        self.log_debug("threshold=" + str(threshold))
        model_name_ = dic["dimensions"]["gene_dim"]["model"]
        model_gene_dim = apps.get_model(app_label=app_, model_name=model_name_)
        model_name_ = dic["dimensions"]["person_dim"]["model"]
        model_person_dim = apps.get_model(app_label=app_, model_name=model_name_)
        gene_obj = model_gene_dim.objects.get(id=gene_id_)
        clusters = gene_obj.clusters
        # print(clusters)
        for c in clusters:
            # print(1, c)
            entity_list = clusters[c]['entity']
            qs = model_person_dim.objects.filter(id__in=entity_list).all()
            df = pd.DataFrame(list(qs.values()))
            dfm = df[df["gender"] == 0]
            dff = df[df["gender"] == 1]
            dfmb = dfm[dfm["age_at_cdna"] <= 30]["id"]
            dfma = dfm[dfm["age_at_cdna"] > 30]["id"]
            dffb = dff[dff["age_at_cdna"] <= 30]["id"]
            dffa = dff[dff["age_at_cdna"] > 30]["id"]
            # print("="*10)
            # print(c, dfmb.count(), dfma.count(), dffb.count(), dffa.count())
            clusters[c]["mb"] = int(dfmb.count())
            clusters[c]["ma"] = int(dfma.count())
            clusters[c]["fb"] = int(dffb.count())
            clusters[c]["fa"] = int(dffa.count())
            # print(100)

        reduced_clusters = gene_obj.reduced_clusters
        for b in reduced_clusters:
            if b == "number_of_blocks":
                continue
            reduced_clusters[b]["compact_block"][2] = round(100 * reduced_clusters[b]["compact_block"][2]) / 100
            reduced_clusters[b]["centroid"] = reduced_clusters[b]["compact_block"][2]

        self.log_debug("get_gene_structure 5")

        genes_temp = {}
        model_temp_var = apps.get_model(app_label=app_, model_name='tempvar')
        qs = model_temp_var.objects.filter(gene_dim__id = gene_id_, amount__gte=threshold).all().order_by("-amount")
        self.log_debug("get_gene_structure 7")
        n = 0
        for g in qs:
            genes_temp[n] = {"idx": g.temp.idx, "amount": float(g.amount), "sign":g.sign}
            n += 1

        # print(gene_id_,"\n", genes_temp)

        # print(reduced_clusters)
        self.log_debug("get_gene_structure 9")
        result = {"status": "ok", "clusters": clusters, "reduced_clusters": reduced_clusters, "genes_temp": genes_temp}
        # print(result)
        return result

    def calculate_pca(self, dic):
        print("90944-44: calculate_pca\n", dic, "\n", "=" * 50)
        self.log_debug(str(dic))
        app_ = dic["app"]
        run_number = dic["run_number"]
        model_person_dim = apps.get_model(app_label=app_, model_name="persondim")
        model_gene_dim = apps.get_model(app_label=app_, model_name="genedim")
        model_fact = apps.get_model(app_label=app_, model_name='fact')
        model_fact_normalized = apps.get_model(app_label=app_, model_name='factnormalized')
        model_fact_normalized_temp = apps.get_model(app_label=app_, model_name='factnormalizedtemp')
        # ------

        model_pca = apps.get_model(app_label=app_, model_name="pca")
        model_pca_data = apps.get_model(app_label=app_, model_name="pcadata")
        # ------
        nnn = 1
        l = [0, 131, 10000]
        re_dic = {}
        while nnn <= 3:
            self.log_debug("nnn="+str(nnn)+" A")
            re_dic[nnn] = {}
            if nnn == 1:
                qsf = model_fact.objects.filter(gene_dim__gene_group_dim__group_name="indep")
                df_f = pd.DataFrame(list(qsf.values('gene_dim', 'person_dim', 'amount')))
                df_f.columns = ['gene', 'person', 'amount']
                df = df_f.pivot_table(values='amount', index='person', columns=['gene'], aggfunc='sum')
                # print("AAAAA : ", nnn, "\n\n", df, "\n\n", df.shape, "\n", "="*50, "\n\n")
            elif nnn == 2:
                qsf = model_fact_normalized_temp.objects.filter(run_number=run_number,
                                                                gene_dim__gene_group_dim__group_name="indep")
                df_f = pd.DataFrame(list(qsf.values('gene_dim', 'person_dim', 'amount')))
                df_f.columns = ['gene', 'person', 'amount']
                df = df_f.pivot_table(values='amount', index='person', columns=['gene'], aggfunc='sum')
                # print("WWW\n\n", df, "\n", df.shape)
            else:
                # qsp = model_person_dim.objects.all()
                # -------
                qsf = model_fact_normalized.objects.filter(run_number=run_number,
                                                           gene_dim__gene_group_dim__group_name="indep")
                df_f = pd.DataFrame(list(qsf.values('gene_dim', 'person_dim', 'amount')))
                df_f.columns = ['gene', 'person', 'amount']
                df = df_f.pivot_table(values='amount', index='person', columns=['gene'], aggfunc='sum')
                # print("ZZZZZ\n\n", df, "\n", df.shape)

            self.log_debug("nnn="+str(nnn)+" B")
            pca = PCA(n_components=3)
            pca_data = pca.fit_transform(df)
            # print(nnn, "\n", pca_data)
            df_pca = pd.DataFrame(pca_data)
            cc = {0:"x", 1:"y", 2:"z"}
            df_pca.rename(columns=cc, inplace=True)
            # print("df_pca\n", df_pca)
            self.log_debug("nnn="+str(nnn)+" C")
            # --
            dfi = df.reset_index()
            dfi.index.name = "p"
            dfi = dfi.reset_index()
            dfi = dfi[["p", "person"]]
            dfi = dfi.set_index("person")
            # print(nnn, "\ndfi4\n", dfi)
            self.log_debug("nnn="+str(nnn)+" D")

            lll = [model_person_dim.objects.filter(set_num__gt=l[0], set_num__lte=l[1]).all(),
                   model_person_dim.objects.filter(set_num__gt=l[1], set_num__lte=l[2]).all()]

            n__ = 1
            self.log_debug("nnn="+str(nnn)+" G")
            for qsp in lll:
                # print(nnn, "\nLLLLL\n", n__)
                re_dic[nnn][n__] = {"x": [], "y": [], "z": []}
                df_p = pd.DataFrame(list(qsp.values('id')))
                llk = df_p['id'].tolist()
                # print("llk1\n", llk)
                dfi_ = dfi.loc[llk, :]
                # print("dfi_\n", dfi_)
                llk = dfi_['p'].tolist()
                # print("llk2\n", llk)
                df_pca_ = df_pca.loc[llk, :]
                # print(n__, "\ndf_pca_\n", df_pca_, "\n", df_pca_.shape)
                # -------
                self.log_debug("nnn="+str(nnn)+" H")
                for index, row in df_pca_.iterrows():
                    pca_obj, is_created = model_pca.objects.get_or_create(run_number=run_number,
                                                                          set=nnn, sub_set=n__, component=1)
                    pca_data_obj, is_created = model_pca_data.objects.get_or_create(pca=pca_obj, idx=index)
                    pca_data_obj.amount = round(1000*row["x"])/1000
                    pca_data_obj.save()
                    #
                    pca_obj, is_created = model_pca.objects.get_or_create(run_number=run_number,
                                                                          set=nnn, sub_set=n__, component=2)
                    pca_data_obj, is_created = model_pca_data.objects.get_or_create(pca=pca_obj, idx=index)
                    pca_data_obj.amount = round(1000*row["y"])/1000
                    pca_data_obj.save()
                    #
                    pca_obj, is_created = model_pca.objects.get_or_create(run_number=run_number,
                                                                          set=nnn, sub_set=n__, component=3)
                    pca_data_obj, is_created = model_pca_data.objects.get_or_create(pca=pca_obj, idx=index)
                    pca_data_obj.amount = round(1000*row["z"])/1000
                    pca_data_obj.save()
                    #
                    re_dic[nnn][n__]["x"].append(round(1000*row["x"])/1000)
                    re_dic[nnn][n__]["y"].append(round(1000*row["y"])/1000)
                    re_dic[nnn][n__]["z"].append(round(1000*row["z"])/1000)
                n__ += 1
            # print("PPPPPP", nnn)
            nnn += 1

        # print(re_dic)
        self.log_debug("End Process")
        result = {"status": "ok", "result": re_dic}
        # print(result)
        return result

    def get_batch_normalised_data(self, dic):

        def set_cols(df_):
            cc = {}
            for c in df_.columns:
                cc[c] = str(self.measures_name[self.measures_name['id'] == c].iloc[0][self.measure_name_]).strip()
            df_.rename(columns=cc, inplace=True)
            return df_

        def set_rows(df_):
            df_ = df_.T
            cc = {}
            for c in df_.columns:
                cc[c] = str(self.entities_name[self.entities_name['id'] == c].iloc[0][self.entity_name+'_'+self.entity_name_suffix]).strip()
            df_.rename(columns=cc, inplace=True)
            return df_.T

        print("90977-77: get_batch_normalised_data\n", dic, "\n", "=" * 50)
        self.log_debug(str(dic))
        app_ = dic["app"]
        run_number = dic["run_number"]
        model_person_dim = apps.get_model(app_label=app_, model_name="persondim")
        model_gene_dim = apps.get_model(app_label=app_, model_name="genedim")
        model_fact = apps.get_model(app_label=app_, model_name='fact')
        model_fact_normalized = apps.get_model(app_label=app_, model_name='factnormalized')
        model_fact_normalized_temp = apps.get_model(app_label=app_, model_name='factnormalizedtemp')
        # ------
        qsf = model_fact_normalized.objects.filter(run_number=run_number,
                                                   gene_dim__gene_group_dim__group_name="indep")
        df_f = pd.DataFrame(list(qsf.values('gene_dim', 'person_dim', 'amount')))
        df_f.columns = ['gene', 'person', 'amount']
        df = df_f.pivot_table(values='amount', index='person', columns=['gene'], aggfunc='sum')
        # print("AAAAA\n\n", df, "\n", df.shape)
        df = set_cols(df)
        df = set_rows(df)
        df = df.T
        # print("ZZZZZ\n\n", df, "\n", df.shape)

        save_to_file = os.path.join(self.PROJECT_MEDIA_DIR, "get_batch_normalised_data_"+str(run_number)+".xlsx")
        # print(save_to_file)

        self.log_debug(save_to_file)
        is_file = os.path.exists(save_to_file)
        if is_file:
            try:
                os.remove(save_to_file)
            except Exception as ex:
                print("90-90-90- 1 Error removing saved file " + save_to_file)
        is_file = os.path.exists(save_to_file)
        self.log_debug("Confirmed is_file = " + str(is_file))
        # print("  is_file = " + str(is_file))
        wb2  = Workbook()
        wb2.save(save_to_file)
        wb2.close()
        wb2  = None

        try:
            # print(save_to_file)
            with pd.ExcelWriter(save_to_file, engine='openpyxl', mode="a") as writer:
                df.to_excel(writer, sheet_name="data")
                self.log_debug("data saved to file:" + save_to_file)
        except Exception as ex:
            self.log_debug("Error save file 1 " + str(ex))
            # print(ex)

        try:
            writer.save()
        except Exception as ex:
            self.log_debug("Error save file 2 (ignore) " + str(ex))
            # print("Error Save", ex)

        try:
            wb = load_workbook(filename=save_to_file, read_only=False)
            del wb['Sheet']
            wb.save(save_to_file)
            wb.close()
            self.log_debug("After removal of sheet, data saved to file:" + save_to_file)
            # print("remove sheet")
        except Exception as ex:
            self.log_debug("Error save file 3 " + str(ex))
            # print(str(ex))

        self.log_debug("End Process")
        result = {"status": "ok"}
        # print(result)
        return result


    def get_pca(self, dic):
        print("90944-44: get_pca\n", dic, "\n", "=" * 50)
        run_number = dic["run_number"]
        self.log_debug(str(dic))
        app_ = dic["app"]
        # ------
        model_pca = apps.get_model(app_label=app_, model_name="pca")
        model_pca_data = apps.get_model(app_label=app_, model_name="pcadata")
        # ------
        pca_sq = model_pca.objects.filter(run_number=run_number)
        re_dic = {}
        for q in pca_sq:
            print(q.id, q.set, q.sub_set, q.component)
            if q.set not in re_dic:
                re_dic[q.set] = {}
            if q.sub_set not in re_dic[q.set]:
                re_dic[q.set][q.sub_set] = {}
            if q.component == 1:
                component = "x"
            elif q.component == 2:
                component = "y"
            elif q.component == 3:
                component = "z"
            if component not in re_dic[q.set][q.sub_set]:
                re_dic[q.set][q.sub_set][component] = []

            pca_data_sq = model_pca_data.objects.filter(pca__id=q.id).all().order_by('idx')
            for q_ in pca_data_sq:
                re_dic[q.set][q.sub_set][component].append(float(q_.amount))

        # print(re_dic)
        self.log_debug("End Process")
        result = {"status": "ok", "result": re_dic}
        print(result)
        return result
    # -----------------------------------------

