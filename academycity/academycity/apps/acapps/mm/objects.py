import warnings
import os
from django.conf import settings
from ..ml.basic_ml_objects import BaseDataProcessing, BasePotentialAlgo
import matplotlib as mpl
from django.apps import apps

import sys
import shutil
import pandas as pd
import numpy as np
import math
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
import pickle
from concurrent.futures import ThreadPoolExecutor as T
from statistics import mean, median
import copy

from openpyxl import Workbook, load_workbook
from ...core.utils import log_debug, clear_log_debug

from collections import OrderedDict
from operator import getitem

mpl.use('Agg')


class MMAlgo(object):
    def __init__(self, dic):  # to_data_path, target_field
        # print("90004-000-1 MMAlgo", dic, '\n', '-'*50)
        super(MMAlgo, self).__init__()
        # print("90004-000-2 MMAlgo", dic, '\n', '-'*50)
        # -----
        self.to_save_normalize = []
        self.to_save_similarity = []
        self.to_save_similarity_by_index = []

    def min_max_rule(self, row):
        row_ = row[0:2].copy()
        row_[:] = np.nan
        row_[0] = row.min()
        row_[1] = row.max()
        return row_


class MMDataProcessing(BaseDataProcessing, BasePotentialAlgo, MMAlgo):
    def __init__(self, dic):
        # print("90001-000-1 MMDataProcessing", dic, '\n', '-'*50)
        super().__init__(dic)
        # print("90001-000-2 MMDataProcessing", dic, '\n', '-'*50)

        self.file_similarity = os.path.join(self.PICKLE_PATH, "similarity"+".pkl")
        self.file_similarity_dic = os.path.join(self.PICKLE_PATH, "similarity_dic"+".pkl")

    def load_file_to_db(self, dic):
        # print("90121-1: \n", dic, "="*50)
        # print(dic)
        # print('dic')
        clear_log_debug()
        log_debug("=== load_file_to_db 100 ===")
        app_ = dic["app"]
        file_path = self.upload_file(dic)["file_path"]
        # file_path = "/home/amos/projects/development/academycity/data/ms/datasets/excel/raw_data/RAW DATA (607).xlsx"
        # print('file_path')
        print(file_path)
        # print('file_path')
        # print('90121-2 dic')

        dic = dic["cube_dic"]
        # print('90121-3 dic', dic)


        model_name_ = dic["dimensions"]["person_dim"]["model"]
        model_person_dim = apps.get_model(app_label=app_, model_name=model_name_)

        model_name_ = dic["dimensions"]["gene_group_dim"]["model"]
        model_gene_group_dim = apps.get_model(app_label=app_, model_name=model_name_)

        model_name_ = dic["dimensions"]["gene_dim"]["model"]
        model_gene_dim = apps.get_model(app_label=app_, model_name=model_name_)

        model_name_ = dic["fact"]["model"]
        model_fact = apps.get_model(app_label=app_, model_name=model_name_)

        # model_fact.truncate()
        # model_gene_dim.truncate()
        # model_person_dim.truncate()
        wb = load_workbook(filename=file_path, read_only=False)
        sheet_names = wb.sheetnames
        for f in sheet_names:
            gene_group_dim_obj, is_created = model_gene_group_dim.objects.get_or_create(group_name=f)
            gene_group_dim_obj.save()
            ws = wb[f]
            data = ws.values
            # Get the first line in file as a header line
            columns = next(data)[0:]
            # print(columns)
            # Create a DataFrame based on the second and subsequent lines of data
            df = pd.DataFrame(data, columns=columns)
            df = df.reset_index()  # make sure indexes pair with number of rows
            # n__ = 0

            list_ = []
            for index, row in df.iterrows():
                # print("index", index)
                for j in range(1, len(columns)):
                    # print("j", j)
                    if row[1] is not None and str(row[1]) != "None" and str(row[1]) != "":
                        g_ = str(row[1])
                        # print("g_", g_)
                        gene_dim_obj, is_created = model_gene_dim.objects.get_or_create(gene_group_dim=gene_group_dim_obj,
                                                                                        gene_code=g_)
                        p_ = str(columns[j])
                        if p_ != "None" and p_ != "":
                            person_dim_obj, is_created = model_person_dim.objects.get_or_create(person_code=p_)
                        try:
                            # print(str(row[columns[j]]), "\n", "="*20)
                            v_ = float(str(row[columns[j]]))
                            fact_obj, is_created = model_fact.objects.get_or_create(gene_dim=gene_dim_obj,
                                                                                    person_dim=person_dim_obj)
                            fact_obj.amount = v_
                            fact_obj.save()
                        except Exception as ex:
                            if p_ not in list_:
                                list_.append(p_)
                                print(list_)
            print("\n", "="*100, list_, "\n", "="*100)

            # print(f_, p_, v_)
            # print(n__, max_v, max_d)
        # print(max_v, max_d)
        # print('90121-6 fact')
        log_debug("=== load_file_to_db 101 ===")
        wb.close()
        result = {"status": "ok"}
        return result

