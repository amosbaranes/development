import warnings
import os
from django.conf import settings
import matplotlib as mpl
import numpy as np
import pandas as pd
from django.apps import apps
from openpyxl import Workbook, load_workbook
import math
import time
import shutil
from statistics import mean
import pickle
from django.db.models import Q
from ...core.utils import Debug, log_debug, clear_log_debug
#
import matplotlib.pyplot as plt
from sklearn import linear_model, neighbors
from sklearn import preprocessing
from sklearn import pipeline
import tarfile
import zipfile
from six.moves import urllib
import hashlib
from sklearn.model_selection import train_test_split, StratifiedShuffleSplit
import matplotlib.image as mpimg
from sklearn.metrics import mean_squared_error, mean_absolute_error
from sklearn.model_selection import cross_val_score
from scipy import stats
import joblib


class BaseDataProcessing(object):
    def __init__(self, dic):  # to_data_path, target_field
        try:
            # print("90002-000-1 BaseDataProcessing\n", dic, '\n', '-' * 50)
            super(BaseDataProcessing, self).__init__(dic)  # (dic)
            # print("90002-000-2 BaseDataProcessing\n", dic, '\n', '-' * 50)
        except Exception as ex:
            print("Error 90002-000-3 \n" + str(ex),"\n", "-"*50)

        # print("90002-001 BaseDataProcessing\n", dic, '\n', '-'*50)
        # print("\n90002 BaseDataProcessing", self.app)

        self.name = 'DataProcessing'
        self.uploaded_filename = None
        # print("90003-0 PBaseDataProcessing", dic, '-'*50)

        warnings.filterwarnings(action="ignore", message="^internal gelsd")
        # to make this notebook's output stable across runs

        self.RANDOM_STATE = 42
        np.random.seed(self.RANDOM_STATE)

        # To plot pretty figures
        mpl.rc('axes', labelsize=14)
        mpl.rc('xtick', labelsize=12)
        mpl.rc('ytick', labelsize=12)
        # Where to save the figures
        # --- Change to this when you start to use Django ---

        clear_log_debug()

        self.PROJECT_ROOT_DIR = os.path.join(settings.WEB_DIR, "data", dic["app"])

        # print(self.PROJECT_ROOT_DIR)
        # print('-'*50)
        os.makedirs(self.PROJECT_ROOT_DIR, exist_ok=True)
        self.TOPIC_ID = dic["topic_id"]  # "fundamentals"
        self.TO_DATA_PATH = os.path.join(self.PROJECT_ROOT_DIR, "datasets")
        os.makedirs(self.TO_DATA_PATH, exist_ok=True)
        self.TO_EXCEL = os.path.join(self.TO_DATA_PATH, "excel", self.TOPIC_ID)
        os.makedirs(self.TO_EXCEL, exist_ok=True)

        try:
            self.dependent_group = dic["dependent_group"]
            self.TO_EXCEL_OUTPUT = os.path.join(self.TO_EXCEL, "output", self.dependent_group)
        except Exception as ex:
            self.TO_EXCEL_OUTPUT = os.path.join(self.TO_EXCEL, "output")

        os.makedirs(self.TO_EXCEL_OUTPUT, exist_ok=True)
        self.IMAGES_PATH = os.path.join(self.PROJECT_ROOT_DIR, "images")
        os.makedirs(self.IMAGES_PATH, exist_ok=True)
        self.MODELS_PATH = os.path.join(self.PROJECT_ROOT_DIR, "models")
        os.makedirs(self.MODELS_PATH, exist_ok=True)
        self.PICKLE_PATH = os.path.join(self.PROJECT_ROOT_DIR, "pickle")
        os.makedirs(self.PICKLE_PATH, exist_ok=True)
        self.target_folder = None

        # self.TARGET_FIELD = target_field
        # self.DATA = None
        # self.TRAIN = None
        # self.TEST = None
        # self.TRAIN_TARGET = None
        # self.TRAIN_DATA = None
        # self.TEST_TARGET = None
        # self.TEST_DATA = None
        # self.train_data = None
        # self.test_data = None
        # self.num_attribs = None
        # self.extra_attribs = None
        # self.model = None
        # self.HASH = hashlib.md5
        # self.PIPELINE = None

        # print('-'*50)
        # print('9010 - End constructor parent')
        # print('-'*50)

    def upload_file(self, dic):
        # print("9002 BaseDataProcessing upload_file:")
        # print(dic)
        # print("upload_file:")

        upload_file_ = dic["request"].FILES['drive_file']
        result = {}

        log_debug("In upload_file.")

        # We can extend and add another property: data_folder
        # like topic_id. But, we need to add this property to: params in the core view
        # and use it here.
        # for example: if data_folder=excel we choose self.TO_EXCEL

        # print("target_folder = self.TO_"+dic["folder_type"].upper())
        self.target_folder = eval("self.TO_" + dic["folder_type"].upper())
        filename = dic["request"].POST['filename']
        # print(filename)
        self.uploaded_filename = filename
        file_path = os.path.join(self.target_folder, filename)
        with open(file_path, 'wb+') as destination:
            for c in upload_file_.chunks():
                destination.write(c)

        # print("9888-8 Uploaded\n", "-" * 30)
        result['file_path'] = file_path
        log_debug("End upload_file.")
        return result

    def get_general_data(self, dic):
        # print("9012-9012-2 BaseDataProcessing get_general_data:\n", dic, "\n", "="*50, "\n")
        result = {}
        n__=0
        for k in dic["dimensions"]:
            # print("\n", k, "\n", "="*50)
            dic_ = dic["dimensions"][k]
            s = k + ' = {}'
            try:
                exec(s)
                model_name_ = dic_["model"]
                model_ = apps.get_model(app_label=self.app, model_name=model_name_)
                p_key_field_name = model_._meta.pk.name
                # print("p_key_field_name", p_key_field_name)
                if p_key_field_name == "user":
                    p_key_field_name +="_id"
                # if p_key_field_name != "id":
                #     p_key_field_name +="_id"
                try:
                    s_ = ""
                    filters = dic_["filters"]
                    if filters:
                        for j in filters:
                            s_ += ".filter(" + j + "=" + str(filters[j]) + ")"
                            # print(s_)
                except Exception as ex:
                    pass
                    # print("Error 9012-9012-1: "+str(ex))

                sq = 'model_.objects'+ s_ +'.all()'
                # print("sq1:  ", sq)
                try:
                    qs = eval(sq)
                    # df_us = pd.DataFrame(list(qs.values()))
                    # print(df_us)
                    # print(qs)
                except Exception as ex:
                    print("er555-5 qs\n", str(ex))
                for r in qs:
                    try:
                        k_ = eval("r."+p_key_field_name)
                    except Exception as ex:
                        print(ex)

                    s = k + '["'+str(k_)+'"] = r.' + dic["dimensions"][k]["field_name"]
                    # print(s)
                    exec(s)
                # print(eval(k))
            except Exception as ex:
                print("err 1000: " + str(ex))

            # print('result[k] = ' + k)
            exec('result[k] = ' + k)
        # print(result)
        return result

    def clean_name(self, name):
        name_ = name.replace("-", "").replace(" ", "").replace("_", "")
        return name_

    def general_data(self, dic):
        action_ = dic['action']
        group_ = dic['group']
        data_name_ = dic['data_name']
        general_data_model = apps.get_model(app_label="core", model_name="generaldata")
        if action_ == "set":
            data_json_ = dic['data_json']
            obj, is_created = general_data_model.objects.get_or_create(app=self.app, group=group_, data_name=data_name_)
            obj.data_json = data_json_
            obj.save()
        else:
            obj = general_data_model.objects.get(app=self.app, group=group_, data_name=data_name_)
            return obj.data_json

    def get_next_number(self, dic):
        group_ = "general"
        data_name_ = "number"
        general_data_model = apps.get_model(app_label="core", model_name="generaldata")
        data_json_ = {"number": 1}
        return_number = 1
        obj, is_created = general_data_model.objects.get_or_create(app=self.app, group=group_, data_name=data_name_)
        if is_created:
            obj.data_json = data_json_
            obj.save()
        else:
            # obj = general_data_model.objects.get(app=self.app, group=group_, data_name=data_name_)
            data_json_ = obj.data_json
            return_number = data_json_["number"] + 1
            data_json_["number"] = return_number
            obj.data_json = data_json_
            obj.save()
        return return_number


class BasePotentialAlgo(object):
    def __init__(self, dic):  # to_data_path, target_field
        try:
            # print('-'*50, '\n', "90003-000-1 BasePotentialAlgo\n", dic, '\n', '-'*50)
            super(BasePotentialAlgo, self).__init__(dic)
            # print('-'*50, '\n', "90003-000-2 BasePotentialAlgo\n", dic, '\n', '-'*50)
        except Exception as ex:
            print("Error 90003-000-3 " + str(ex))
        # print('-'*50, '\n', "90003-000-3 BasePotentialAlgo\n", dic, '\n', '-'*50)
        # ---------------------
        self.app = dic["app"]
        # ---------------------
        self.rule_2 = 0.3
        self.rule_0 = 0.025
        self.missing_data_range = 0.3
        self.second_time_save = ''
        self.to_save = []
        self.to_save_all = []
        self.save_to_file = None
        self.df_index = None
        try:
            self.value_column = dic["value"]
        except Exception as ex:
            pass
        try:
            fact_model_name = dic["fact_model"]
            self.model_fact = apps.get_model(app_label=self.app, model_name=fact_model_name)
        except Exception as ex:
            pass

        measure_name = "measure_name"
        try:
            measure_name = dic["measure_name"]
        except Exception as ex:
            pass
        # print("measure_name=", measure_name)

        try:
            measure_model_name_ = dic["measure_model"]
            # print('-' * 50, '\n', "90003-000-41 BasePotentialAlgo\nmeasure_model_name_=", measure_model_name_, "\n", '-' * 50)
            model_measure_dim = apps.get_model(app_label=self.app, model_name=measure_model_name_)
            self.measures_name = pd.DataFrame(model_measure_dim.objects.all().values('id', measure_name))
        except Exception as ex:
            pass

        try:
            self.entity_name = dic["entity_name"]
            entity_model_name_ = dic["entity_model"]
            model_entity_dim = apps.get_model(app_label=self.app, model_name=entity_model_name_)
            entity_name_suffix = "name"
            try:
                entity_name_suffix = dic["entity_name_suffix"]
            except Exception as ex:
                pass
            try:
                self.entities_name = pd.DataFrame(model_entity_dim.objects.all().values('id', self.entity_name+'_'+entity_name_suffix))
            except Exception as ex:
                pass
            # print("self.entities_name\n",self.entities_name)
            # need to delete the next line
            # self.countries_name = pd.DataFrame(model_entity_dim.objects.all().values('id', self.entity_name+'_name'))
            try:
                # ---------------------
                year_ = str(dic["time_dim_value"])
                dic_ = {'filter_dim': 'time_dim', 'filter_value': year_,
                        'filter_amount': dic["population_filter_amount"],
                        'axes': [self.entity_name + '_dim', 'measure_dim'],
                        'measure_id': None, 'measure_name': dic["measure_name_convert"]}
                # print("90022-122-111 pre_process_data: \n", dic_, "\n", "="*50)
                self.df_entities_conversion_factor = self.get_dim_data_frame(dic_)["result"]
                # print(self.df_entities_conversion_factor)
                # ---------------------
            except Exception as ex:
                pass
        except Exception as ex:
            pass
            # print("Error 9001-222-22: "+str(ex))

        try:
            self.total_variables = dic["total_variables"]
            self.multiply_two_variables = dic["multiply_two_variables"]
            self.do_not_include_measures_objs = dic["do_not_include_measures_objs"]
            self.do_not_include_groups = dic["do_not_include_groups"]
        except Exception as ex:
            pass
        # print(self.measures_name)
        self.options = ["mm", "mx", "xm", "xx"]
        self.is_calculate_min_max = None
        # print('-'*50, '\n', "90003-000-10 BasePotentialAlgo\n", dic, '\n', '-'*50)

    def process_algo(self, dic):
        # ---- Assiting function -----------------------
        def get_culomns_names(self, l):
            cs = []
            for c in l:
                k_ = str(self.measures_name[self.measures_name["id"] == c]["measure_name"]).split("    ")[1].split("\n")[0]
                cs.append(k_)
            return cs
        # -------- Variables definition ----------------
        clear_log_debug()
        print("90011-111 process_algo: \n", dic, "\n", "="*50)
        log_debug("BasePotentialAlgo: process_algo 1: " + str(dic))

        # print("9015-1 BasePotentialAlgo process_algo\n", dic)
        # --- parames --------------------------
        output_fact_model_name_ = dic["output_fact_model"]
        relimp_fact_model_name_ = dic["relimp_fact_model"]
        range_model_model_name_ = dic["range_model"]
        year_ = dic["time_dim_value"]
        min_max_model_name_ = dic["min_max_model"]
        self.rule_2 = int(dic["rule_2"])/100
        self.rule_0 = float(dic["rule_0"])
        self.missing_data_range = int(dic["missing_data_range"])/100
        measure_model_name_ = dic["measure_model"]
        # ----------------------------------------
        model_range = apps.get_model(app_label=self.app, model_name=range_model_model_name_)
        model_output_fact = apps.get_model(app_label=self.app, model_name=output_fact_model_name_)
        model_relimp_fact = apps.get_model(app_label=self.app, model_name=relimp_fact_model_name_)

        # Add min_max algo
        # ----------------
        # try:
        #     self.is_calculate_min_max = eval(dic["is_calculate_min_max"])
        #     # print(self.is_calculate_min_max)
        # except Exception as ex:
        #     print("Error 9016: \n"+str(ex))
        # if not self.is_calculate_min_max:
        model_min_max = apps.get_model(app_label=self.app, model_name=min_max_model_name_)
        # -------------------------------------------------------------------------------
        # -- total population ---
        try:
            qs = model_min_max.objects.filter(measure_dim__measure_name="TotalPop",
                                                 time_dim_id=year_).all()[0]
            tot_pop_min = qs.min
            tot_pop_max = qs.max
            # print("population\nmin=", tot_pop_min, "max=", tot_pop_max)
        except Exception as ex:
            pass

        # else:
        #     self.calculate_min_max_cuts(dic)
        model_measure = apps.get_model(app_label=self.app, model_name=measure_model_name_)
        # print("90060-10 PotentialAlgo: \n", "="*50)
        wb2 = None
        groups = self.model_measure_group.objects.all()
        nn__ = 0
        sign_n1 = pd.DataFrame([[0, 0, 0, 0]])
        sign_n2 = pd.DataFrame([[0, 0, 0, 0]])
        sign_n1.columns = self.options
        sign_n2.columns = self.options
        similarity_n1 = pd.DataFrame([[0, 0, 0, 0]])
        similarity_n2 = pd.DataFrame([[0, 0, 0, 0]])
        similarity_n1.columns = self.options
        similarity_n2.columns = self.options
        contribution_n1 = pd.DataFrame([[0, 0, 0, 0]])
        contribution_n2 = pd.DataFrame([[0, 0, 0, 0]])
        adj_contribution_n1 = pd.DataFrame([[0, 0, 0, 0]])
        adj_contribution_n2 = pd.DataFrame([[0, 0, 0, 0]])
        # relimp_n1 = pd.DataFrame([[0, 0, 0, 0]])
        # relimp_n2 = pd.DataFrame([[0, 0, 0, 0]])
        group_d = ""

        # ---- PreProcessing - Normalization Stage ----
        ll_dfs = self.pre_process_data(dic)
        # print(ll_dfs)
        # ----------------------------------------------
        # ---------------------------------------------
        lll_groups = []
        for k in ll_dfs:
            group = k #.group_name
            print("="*50,"\n", group, "\n", "="*50)
            try:
                self.save_to_file = os.path.join(self.TO_EXCEL_OUTPUT, str(dic["time_dim_value"]) + "_" + group + "_o.xlsx")
                self.to_save = []
                print("file_path\n", self.save_to_file, "\n", "="*50)
                s = ""

                df = ll_dfs[group]
                df_columns = df.columns
                self.df_index = df.index
                # print(self.df_index)
                print("Before", df, "\n", "="*100)
                df_ = self.add_entity_to_df(df).sort_values(self.entity_name+'_name', ascending=True)

                print("After", df_, "\n", "="*100)
                self.to_save.append((df_.copy(), 'Data'))
                qs_mm = model_min_max.objects.filter(measure_dim__measure_group_dim__group_name=group,
                                                     time_dim_id=dic["time_dim_value"]).all()
                df_mm = pd.DataFrame(list(qs_mm.values('measure_dim', 'min', 'max')))

                print("="*200)
                print("="*200)
                print("df_mm\n", df_mm)
                print("df_mm.T\n", df_mm.T)
                print('df_mm.T.loc["min"]\n', pd.DataFrame(df_mm.T.loc["min"]))
                print('pd.DataFrame(df_mm.T.loc["min"]).T\n', pd.DataFrame(df_mm.T.loc["min"]).T)
                print('pd.DataFrame(df_mm.T.loc["min"]).T.reset_index()\n', pd.DataFrame(df_mm.T.loc["min"]).T.reset_index())
                print('pd.DataFrame(df_mm.T.loc["min"]).T.reset_index().drop(["index"]\n',
                      pd.DataFrame(df_mm.T.loc["min"]).T.reset_index().drop(['index'], axis=1))


                print("="*200)
                print("="*200)


                first_row = pd.DataFrame(df_mm.T.loc["min"]).T.reset_index().drop(['index'], axis=1)
                second_row = pd.DataFrame(df_mm.T.loc["max"]).T.reset_index().drop(['index'], axis=1)
                first_row.columns = df_mm.T.loc["measure_dim"]
                second_row.columns = df_mm.T.loc["measure_dim"]
                # print("1111 first_row\n", first_row, "\nsecond_row", "\n", second_row)
                for f in first_row:
                    if f in self.total_variables:
                        first_row[f] = first_row[f]/tot_pop_min
                        second_row[f] = second_row[f]/tot_pop_max
                diff_row = second_row.subtract(first_row, fill_value=None)
                diff_row.columns = first_row.columns
                # print("diff_row\n", diff_row)
                df = ll_dfs[group]
                df_columns = df.columns
                # print("df.columns\n", df.columns)
                # print("df\n", df)
                df_n1 = df.copy()
                try:
                    df_n1 = df_n1.astype(float)
                except Exception as ex:
                    print("1000: " + str(ex))
                for i, r in df_n1.iterrows():
                    for j in df_columns:
                        try:
                            if str(r[j]) != "nan":
                                z = (float(r[j]) - float(first_row[j].astype(float))) / float(diff_row[j].astype(float))
                                df_n1.loc[i][j] = z
                        except Exception as ex:
                            print("Error i " + str(i) + " " + str(ex))
                # print("11 df_n1\n", df_n1, "\n", "="*100)

                df_n1 = df_n1.apply(pd.to_numeric, errors='coerce').round(6)
                self.add_to_save(title='Normalized-1', a=df_n1, cols=None)
                #
                df_n2 = df_n1.copy()
                df_n2[df_n2 < 0] = 0
                df_n2[df_n2 > 1] = 1

                self.add_to_save(title='Normalized-2', a=df_n2, cols=None)
                # print("2 df_n2\n", df_n2, "\n", "="*100)
                #
                if len(df_n1.columns) < 2:
                    df_n1["max"] = df_n1[df_n1.columns[0]]  # df_n1["Birth Rate"]
                    df_n2["max"] = df_n2[df_n2.columns[0]]  # df_n2["Birth Rate"]
                    df_1_2 = pd.merge(left=df_n1, right=df_n2, left_index=True, right_index=True)
                    # print("2-1 df_1_2\n", df_1_2, "\n", "="*100)

                    # df_1_2.columns = ['min-n1', 'max-n1', 'min-n2', 'max-n2']
                    # cols = df_1_2.columns
                    # cols = cols.insert(0, self.entity_name+'_name')
                    # self.add_to_save(title='min-max', a=df_1_2, cols=cols)
                elif len(df_n1.columns) < 5:
                    # print("-1"*30)
                    # print(group)
                    # print("-000"*30)
                    df_n1 = df_n1.apply(lambda x: np.sort(x), axis=1, raw=True)
                    df_n2 = df_n2.apply(lambda x: np.sort(x), axis=1, raw=True)
                    df_n1["max"] = df_n1.max(axis=1)
                    df_n1["min"] = df_n1.min(axis=1)
                    df_n2["max"] = df_n2.max(axis=1)
                    df_n2["min"] = df_n2.min(axis=1)
                    # df_n1 = df_n1.drop(df_n1_columns, axis=1)
                    # df_n2 = df_n2.drop(df_n2_columns, axis=1)
                    df_n1 = df_n1[["min", "max"]]
                    df_n2 = df_n2[["min", "max"]]
                    df_1_2 = pd.merge(left=df_n1, right=df_n2, left_index=True, right_index=True)
                    # print("2-2 df_1_2\n", df_1_2, "\n", "="*100)
                    # df_1_2.columns = ['min-n1', 'max-n1', 'min-n2', 'max-n2']
                    # cols = df_1_2.columns
                    # cols = cols.insert(0, self.entity_name+'_name')
                    # df_ = self.add_entity_to_df(df_1_2, cols)
                    # print(df_)
                    # self.to_save.append((df_.copy(), 'min-max'))
                else:
                    zero_list = {}
                    one_list = {}
                    # print(df_n1)
                    for i in df_n1.index:
                        n0 = []
                        n1 = []
                        for num in df_n1.loc[i]:
                            if not pd.isna(num):
                                if num <= 0:
                                    n0.append(num)
                                elif num >= 1:
                                    n1.append(num)
                        if len(n0) > 0:
                            # print("n0\n", n0)
                            n0.sort(reverse=True)
                            # print("A n0\n", n0)
                            # print("="*10)
                            zero_list[i] = n0
                        if len(n1) > 0:
                            n1.sort()
                            one_list[i] = n1
                    # # #
                    a = df_n2.values
                    a_1 = a.copy()
                    a_1.sort(axis=1)
                    self.add_to_save(title='Sort L', a=a_1, cols=-1)

                    a_1 = self.clean_rows(a=a_1, j=1, side="L")
                    #
                    # print("1-1\n", a_1)
                    a_1m = pd.DataFrame(a_1)
                    a_1m = a_1m.apply(self.move_elements_to_right, axis=1)
                    a_1m = a_1m.apply(self.revers_elements_in_row, axis=1)
                    self.add_to_save(title='R arranged', a=a_1m, cols=-1)
                    #
                    a_1 = -1 * a_1.copy()
                    a_1.sort(axis=1)
                    # print("1-11\n", a_1)
                    self.add_to_save(title='Sort R', a=a_1, cols=-1)
                    a_1 = self.clean_rows(a=a_1, j=1, side="R")
                    a_1 = -1 * a_1.copy()
                    a_1.sort(axis=1)
                    #
                    # move the numbers to the right.
                    # a_1 = pd.DataFrame(a_1)
                    # a_1 = a_1.apply(self.revers_elements_in_row, axis=1)
                    # print(a_1)
                    #
                    self.add_to_save(title='Final R', a=a_1, cols=-1)
                    # print(a_1)
                    # a_1 = -1 * a_1
                    # a_1 = a_1.apply(self.revers_elements_in_row, axis=1)
                    # a_1.sort(axis=1)  #
                    # self.add_to_save(title='Final-1', a=a_1, cols=-1)
                    # # #
                    a_1 = pd.DataFrame(a_1)
                    a_1.dropna(how='all', axis=1, inplace=True)
                    # print("600000000-100-1")
                    #
                    # print("1-13\n", a_1)
                    a_1 = a_1.apply(self.twenty_rule, axis=1)
                    # print("600000000-100-12")
                    a_1 = a_1.apply(lambda x: np.sort(x), axis=1, raw=True)
                    # print("600000000-100-13")
                    self.add_to_save(title='Final-20-rule', a=a_1, cols=-1)

                    # print("600000000-100-2")

                    a_1 = a_1.apply(self.thirty_rule, axis=1)
                    self.add_to_save(title="Final-"+str(self.rule_2)+"-rule", a=a_1, cols=-1)
                    a_2 = a_1.copy()
                    #
                    # print("zero_list\n", zero_list, "\n", "="*100)
                    # print("one_list\n", one_list, "\n", "="*100)
                    # print("1-14\n", a_1)
                    ff = []
                    for j in a_1.index:
                        nn = list(a_1.loc[j])
                        # print("AAAA j=", j,"nn = ", nn)
                        if min(nn) == 0:
                            # print("BBB j=", j,"nn zero_list[j] =", zero_list[j], "\n", "-"*10)
                            nn = [zero_list[j].pop(0) if i == 0 else i for i in nn]
                            # print("CCC j=", j,"nn= ", nn, "-"*10)
                        if max(nn) == 1:
                            # print("DDD j=", j,"nn =", nn)
                            # print("DDD j=", j,"one_list[j]=", one_list[j], "\n", nn)
                            nn = [one_list[j].pop() if i == 1 else i for i in nn]
                            # print("DDDEE j=", j,"nn=", nn)
                        # nn.insert(0, j)
                        # print("EEEE j=", j,"nn=", nn)
                        nn.sort()
                        ff.append(nn)
                    self.add_to_save(title='Final-30-rule-n1', a=ff, cols=-1)
                    a_1 = pd.DataFrame(ff, index=list(self.df_index))
                    #
                    df_n1 = a_1.apply(self.min_max_rule, axis=1)
                    df_n1.columns = ['min-n1', 'max-n1']
                    #
                    df_n2 = a_2.apply(self.min_max_rule, axis=1)
                    df_n2.columns = ['min-n2', 'max-n2']
                    df_1_2 = pd.merge(left=df_n1, right=df_n2, left_index=True, right_index=True)
                    # print("2-3 df_1_2\n", df_1_2, "\n", "="*100)
                # print("3 df_n1\n", df_n1, "\n", "="*100)
                df_1_2.columns = ['min-n1', 'max-n1', 'min-n2', 'max-n2']
                cols = df_1_2.columns
                cols = cols.insert(0, self.entity_name+'_name')
                # print("4 df_n1\n", group, "\n", df_1_2, "\n", "="*100)
                self.add_to_save(title='min-max', a=df_1_2, cols=cols)
                # self.add_to_save(title='min-max', a=df_1_2, cols=-1)
                # print("50001-3-9")
                self.save_to_excel_()

                # print("50001-3-9-1")
            except Exception as ex:
                print("Error 50001-136-1: " + str(ex))
            df_n1_ = df_n1.copy()
            df_n1_.columns = ['m-' + group, 'x-' + group]
            df_n2_ = df_n2.copy()
            df_n2_.columns = ['m-' + group, 'x-' + group]

            if group == self.dependent_group:
                ss_n_mm = ""
                ss_n_xm = ""
                ss_n_mx = ""
                ss_n_xx = ""
                group_d = group
                df_n1_all = df_n1_
                df_n2_all = df_n2_
            else:
                # print(group_d, group)
                group_d, group, df_n1_all, df_n2_all, df_n1_, df_n2_, sign_n1, sign_n2, similarity_n1, similarity_n2 = \
                    self.create_similarity(group_d, group, df_n1_all, df_n2_all, df_n1_, df_n2_, sign_n1, sign_n2,
                                           similarity_n1, similarity_n2)
                ss_n_mm += '"' + group_d + '-' + group + '-mm",'
                ss_n_mx += '"' + group_d + '-' + group + '-mx",'
                ss_n_xm += '"' + group_d + '-' + group + '-xm",'
                ss_n_xx += '"' + group_d + '-' + group + '-xx",'
                # print(group, sign_n1)
                # print(group, sign_n2)

            lll_groups.append(group)

        # print("QQQQQQQQQQQQQQQQQ")
        # print("df_n1_all\n", df_n1_all)
        # print("RRRRRRRRRRRrRRR")

        ss_n_mm = ss_n_mm[:-1]
        ss_n_mx = ss_n_mx[:-1]
        ss_n_xm = ss_n_xm[:-1]
        ss_n_xx = ss_n_xx[:-1]

        df_n2_all_corr = df_n2_all.copy()
        # print("1 - df_n2_all_corr\n", df_n2_all_corr, "\n", ll_dfs)
        # print(df_n2_all_corr.columns)
        # Calculate corr
        for o in ["m", "x"]:
            s_corr = ""
            n__ = 0
            for g in ll_dfs:
                #print(g)
                if n__ >= 1:
                    s_corr += "'"+o+"-"+g+"',"
                    #print(0, n__, s_corr)
                n__ += 1
            # print(1, s_corr)
            s_corr = s_corr[:-1]
            # print(2, s_corr)

            df_corr = eval("df_n2_all_corr[["+s_corr+"]]")
            corr_ = df_corr.corr(method='pearson')
            # print("90050-30\n", corr_,"\n", type(corr_))
            self.add_to_save_all(title='corr-' + o, a=corr_, cols=-1)
        # print("2 - ")
        for n in ["1", "2"]:
            ll = []
            for k in self.options:
                s_ = "df_n" + n + "_all['d_" + k + "']=df_n" + n + "_all[[" + eval("ss_n_" + k) + "]].min(axis=1)"
                # print(s_)
                exec(s_)
                # print(eval("df_n" + n + "_all['d_" + k + "']"))
                s_ = "ll.append(1-df_n" + n + "_all[[" + eval("ss_n_" + k) + "]].min(axis=1).mean())"
                # print(s_)
                exec(s_)
            # print(ll)
            s_ = "similarity_n" + n + ".loc['SComb'] = ll"
            # print(s_)
            exec(s_)
            s_ = "sign_n" + n + ".drop([0], axis=0, inplace=True)"
            # print(s_)
            exec(s_)
            # print(eval("sign_n" + n))

            self.add_to_save_all(title='sign-n' + n, a=eval("sign_n" + n), cols=-1)
            s = "similarity_n" + n + ".drop([0], axis=0, inplace=True)"
            exec(s)
            self.add_to_save_all(title='similarity-n' + n, a=eval("similarity_n" + n), cols=-1)
        # print("90050-27\n")

        for n in ["1", "2"]:
            nn__ = 0
            llg = []
            llg_adj = []
            for k in lll_groups:
                try:
                    group = k #.group_name
                    # print("-"*10)
                    # print(group)
                    if nn__ > 0:
                        ll = []
                        for z in self.options:
                            s_ = "df_n" + n + "_all['dc_" + group + "_" + z + "'] = abs("
                            s_ += "df_n" + n + "_all['d_" + z + "'] - "
                            s_ += "df_n" + n + "_all['" + group_d + "-" + group + "-" + z + "'])"
                            exec(s_)
                            s__ = 'll.append(1-df_n' + n + '_all["dc_' + group + '_' + z + '"].mean())'
                            exec(s__)
                        s___ = "ll*similarity_n" + n + ".loc['SComb'].T"
                        ll_adj = eval(s___).tolist()
                        # if n == "1":
                        #     print(s___)
                        #     print("ll\n", ll, "\n", type(ll))
                        #     print("similarity_n1.loc['SComb'].T\n", similarity_n1.loc['SComb'].T)
                        # print("ll_adj\n", ll_adj, "\n", type(ll_adj))
                        #
                        llc = [(x - 0.7) if x > 0.7 else 0 for x in ll]
                        llg.append(llc)
                        # print("llg\n", llg)
                        llc_adj = [(x - 0.7) if x > 0.7 else 0 for x in ll_adj]
                        llg_adj.append(llc_adj)
                        # print("llg_adj\n", llg_adj)

                        exec("contribution_n" + n + ".loc[group] = ll")
                        exec("adj_contribution_n" + n + ".loc[group] = ll_adj")
                        # print("A   contribution_n" + n, eval("contribution_n"+n), "\n", ll)
                    else:
                        nn__ += 1
                except Exception as ex:
                    print("Error 50008-8: " + str(ex))

            # print(n, "\nCC\n", llg, "\n\n", llg_adj)
            df_temp = eval("df_n" + n + "_all")
            df_temp = df_temp.reset_index()
            df_temp = df_temp.rename(columns={'index': 'country_dim'})
            df_temp = self.add_entity_to_df(df_temp, cols=-1)
            df_temp = df_temp.sort_values(self.entity_name + '_name', ascending=True)
            self.add_to_save_all(title="all-n" + n, a=df_temp, cols=-1)

            exec("contribution_n" + n + ".columns = self.options")
            exec("adj_contribution_n" + n + ".columns = self.options")
            exec("contribution_n" + n + ".drop([0], axis=0, inplace=True)")
            exec("adj_contribution_n" + n + ".drop([0], axis=0, inplace=True)")
            # print("=1"*50)
            npg = np.array(llg)
            npg_adj = np.array(llg_adj)
            npgs = np.sum(llg, axis=0)
            npgs_adj = np.sum(llg_adj, axis=0)
            df_relimp = pd.DataFrame(npg/npgs, index=contribution_n1.index)

            # print("llg=\n", llg, "\nnpg=\n", npg, "\nllg_adj=", "\n", llg_adj, "npgs=\n", npgs, "\ndf_relimp=", df_relimp)

            df_relimp_adj = pd.DataFrame(npg_adj/npgs_adj, index=contribution_n1.index)

            df_relimp.columns = self.options
            df_relimp_adj.columns = self.options
            if n == "2":
                df_relimp_adj_2 = df_relimp_adj.copy(deep=True)
            df_relimp_adj.columns = self.options

            self.add_to_save_all(title='contribution-n' + n, a=eval("contribution_n" + n), cols=-1)
            self.add_to_save_all(title='adj_contribution-n' + n, a=eval("adj_contribution_n" + n), cols=-1)
            self.add_to_save_all(title='relimp-n' + n, a=df_relimp, cols=-1)
            self.add_to_save_all(title='adj_relimp-n' + n, a=df_relimp_adj, cols=-1)
            # print("=2"*50)

        np_sign_n2 = sign_n2.to_numpy()
        ll = [*range(np_sign_n2.shape[0])]

        mg_obj, is_created = self.model_measure_group.objects.get_or_create(group_name="Output")
        mm_obj, is_created = model_measure.objects.get_or_create(measure_name="Pot-"+self.dependent_group, measure_group_dim=mg_obj,
                                                                 measure_code=self.dependent_group, description="Potential for "+self.dependent_group)

        for i, r in df_relimp_adj_2.iterrows():
            for c in df_relimp_adj_2.columns:
                # print("i=", i, "c=", c, "float(r[c])", float(r[c]))
                if str(r[c]) != "nan":
                    range_obj = model_range.objects.get(range_name=str(c))
                    group_obj = self.model_measure_group.objects.get(group_name=i)
                    obj, is_created = model_relimp_fact.objects.get_or_create(range_dim=range_obj,
                                                                              time_dim_id=year_,
                                                                              measure_dim=mm_obj,
                                                                              measure_group_dim=group_obj)
                    obj.amount=round(100*float(r[c]))/100
                    obj.save()
        np_relimp_adj_2 = df_relimp_adj_2.to_numpy()
        np_signed_relimp_2 = np_relimp_adj_2*np_sign_n2
        np_signed_relimp = {"m": np_signed_relimp_2[np.ix_(ll, [0, 1, 2, 3])],
                            "x": np_signed_relimp_2[np.ix_(ll, [0, 1, 2, 3])]}

        df_n1_all_temp = df_n1_all.copy()
        df_n1_all_temp = df_n1_all_temp.reset_index()
        df_n1_all_temp = df_n1_all_temp.rename(columns={'index': 'country_dim'})
        # print("-"*10, "\n", "df_n1_all_temp", "\n", df_n1_all_temp, "-"*10, "\n")

        try:
            for k_ in ["m", "x"]:
                # print("-="*10, "\n" ,k_)
                df = pd.DataFrame()
                # df["country_dim"] = df_n1_all_temp["country_dim"]
                for group in ll_dfs:
                    if group != self.dependent_group:
                        # print("="*50)
                        # print(group)
                        df[k_ + "-" + group] = df_n1_all_temp[k_ + "-" + group]
                df.index=df_n1_all.index
                # print("-="*10, "\n" ,k_, "\n", df.head(100), "\n", df.tail(100))
                np_df = df.to_numpy()
                # print(np_df)
                if k_ == "m":
                    potential_m_z1 = None
                    potential_m_z2 = None
                    potential_m_z3 = None
                    potential_m_z4 = None
                    potential_m = None
                    np_df_clone = np_df.copy()
                    # print(np_df_clone)
                    # print("Before zzzzz\n")
                    # for row in np_df_clone:
                    #     print(row)

                    for z in range(np_signed_relimp[k_].shape[1]):
                        # for r in range(np_signed_relimp[k_].shape[0]):
                        #     if abs(np_signed_relimp[k_][..., z][r]) < 0.00001:
                        #         # print(r, np_signed_relimp[k_][..., z][r])
                        #         np_df_clone[:, r] = 0
                        # if z == 1:
                        #     print("zzzzzzzz1111\n")
                        #     for row in np_df_clone:
                        #         print(row)
                        res = np.argwhere(np.isnan(np_df_clone))
                        n_dic = {}
                        for row in res:
                            if row[0] not in n_dic:
                                n_dic[row[0]] = []
                            n_dic[row[0]].append(row[1])
                        # print(n_dic)
                        # print(np_signed_relimp[k_])
                        max_np = np.amax(np_signed_relimp[k_], axis=1)
                        # print(max_np)
                        for z_ in n_dic:
                            n_sum = 0
                            for h in n_dic[z_]:
                                n_sum += max_np[h] # np_signed_relimp[k_][..., z][h]
                            if n_sum < self.missing_data_range:
                                for h in n_dic[z_]:
                                    np_df_clone[z_, h] = 0
                                np_df_clone[z_] = np_df_clone[z_]*(1/(1-n_sum))

                        # if z == 1:
                        #     print("zzzzzzzz1111zzzzzzzz11111\n")
                        #     for row in np_df_clone:
                        #         print(row)

                        if z == 0:
                            potential_m_z1 = np.dot(np_df_clone, np.array([np_signed_relimp[k_][..., z]]).T)
                        elif z == 1:
                            potential_m_z2 = np.dot(np_df_clone, np.array([np_signed_relimp[k_][..., z]]).T)
                        elif z == 2:
                            potential_m_z3 = np.dot(np_df_clone, np.array([np_signed_relimp[k_][..., z]]).T)
                        elif z == 3:
                            # print("=411"*20)
                            # print(k_, "z=1", np_df_clone, "\n", np.array([np_signed_relimp[k_][..., z]]).T)
                            potential_m_z4 = np.dot(np_df_clone, np.array([np_signed_relimp[k_][..., z]]).T)
                            potential_m = np.column_stack((potential_m_z1, potential_m_z2, potential_m_z3, potential_m_z4))
                elif k_ == "x":
                    potential_x_z1 = None
                    potential_x_z2 = None
                    potential_x_z3 = None
                    potential_x_z4 = None
                    potential_x = None
                    np_df_clone = np_df.copy()
                    for z in range(np_signed_relimp[k_].shape[1]):
                        # for r in range(np_signed_relimp[k_].shape[0]):
                        #     if abs(np_signed_relimp[k_][..., z][r]) < 0.00001:
                        #         np_df_clone[:, r] = 0
                        res = np.argwhere(np.isnan(np_df_clone))
                        n_dic = {}
                        for row in res:
                            if row[0] not in n_dic:
                                n_dic[row[0]] = []
                            n_dic[row[0]].append(row[1])
                        max_np = np.amax(np_signed_relimp[k_], axis=1)
                        # print(max_np)
                        for z_ in n_dic:
                            n_sum = 0
                            for h in n_dic[z_]:
                                n_sum += max_np[h]  # np_signed_relimp[k_][..., z][h]
                            if n_sum < self.missing_data_range:
                                for h in n_dic[z_]:
                                    np_df_clone[z_, h] = 0
                                np_df_clone[z_] = np_df_clone[z_]*(1/(1-n_sum))
                        if z == 0:
                            potential_x_z1 = np.dot(np_df_clone, np.array([np_signed_relimp[k_][..., z]]).T)
                        elif z == 1:
                            potential_x_z2 = np.dot(np_df_clone, np.array([np_signed_relimp[k_][..., z]]).T)
                        elif z == 2:
                            potential_x_z3 = np.dot(np_df_clone, np.array([np_signed_relimp[k_][..., z]]).T)
                        elif z == 3:
                            # print(k_, "z=1", np_df_clone, "\n", np.array([np_signed_relimp[k_][..., z]]).T)
                            potential_x_z4 = np.dot(np_df_clone, np.array([np_signed_relimp[k_][..., z]]).T)
                            potential_x = np.column_stack((potential_x_z1, potential_x_z2, potential_x_z3, potential_x_z4))

            potential = np.concatenate((potential_m, potential_x), axis=1)
            df_potential = pd.DataFrame(potential, columns=["min_mm", "min_xm", "min_mx", "min_xx", "max_mm", "max_xm", "max_mx", "max_xx"])

            # print("df_potential100 df_potential100\n", df_potential, "\n", "="*100)

            df_potential.insert(0, 'country_dim', value=df_n1_all_temp['country_dim'])
            df_potential_cube = df_potential.copy()

            # print("df_potential200 df_potential200\n", df_potential, "\n", "="*100)

            df_potential = self.add_entity_to_df(df_potential, cols=-1)
            df_potential = df_potential.drop(['index'], axis=1)
            df_potential = df_potential.sort_values(self.entity_name+'_name', ascending=True)

            # print("df_potential df_potential\n", df_potential, "\n", "="*100)

            self.add_to_save_all(title='potential', a=df_potential, cols=-1)
        except Exception as ex:
            print(ex)

        # print("90050-28\n")
        self.save_to_excel_all_(dic["time_dim_value"])
        # print("90050-29\n")

        for i, r in df_potential_cube.iterrows():
            for c in df_potential_cube.columns[1:]:
                # print(c, float(r[c]), "="+str(r[c])+"=")
                if str(r[c]) != "nan":
                    l = c.split("_")
                    range_obj = model_range.objects.get(range_name=l[1])
                    obj, is_created = model_output_fact.objects.get_or_create(range_dim=range_obj,
                                                                              range_name=l[0],
                                                                              time_dim_id=year_,
                                                                              country_dim_id=int(r["country_dim"]),
                                                                              measure_dim=mm_obj)
                    obj.amount=round(100*float(r[c]))/100
                    obj.save()

        result = {"status": "ok"}
        return result

    def get_dim_data_frame(self, dic):
        # print("90066-106-11 PotentialAlgo get_dim_data_frame: \n", dic, "\n", "="*50)

        # dic =  {'app': 'avi', 'filter_dim': 'time_dim', 'filter_value': 2019,
        #         'axes': ['country_dim', 'measure_dim'],
        #         'model_name': 'worldbankfact', 'filter_amount': 1000000, 'measure_id': '19',
        #         'measure_name': 'TotalPop'}
        filter_dim = dic["filter_dim"]
        filter_value = dic["filter_value"]
        axes_ = dic["axes"]
        filter_amount = dic["filter_amount"]  # 1000000
        measure_id = dic["measure_id"]
        #
        return_dim = ""
        for k in axes_:
            if k not in ["measure_dim", filter_dim]:
                return_dim = k
        #
        # print(return_dim)
        if measure_id is None or measure_id == "":
            measure_name = dic["measure_name"]
            s = 'self.model_fact.objects.filter(measure_dim__measure_name="'+measure_name+'", '
        else:
            # print(measure_id)
            s = 'self.model_fact.objects.filter(measure_dim__id='+measure_id+', '
        s += filter_dim+'__id='+str(filter_value)+', '+self.value_column+'__gte='+str(filter_amount)+').all()'
        # print("1 \ns="+s)

        qs = eval(s)
        df = pd.DataFrame(list(qs.values(return_dim+"_id", self.value_column)))
        # print(df)
        df = df.rename(columns={return_dim+"_id": return_dim})
        # print("df df df\n", df)

        df = df.set_index(return_dim)
        # print(df)

        # print(df[return_dim+"_id"].tolist())
        # print(df.shape)
        if df.shape[0] > 0:
            result = {"status": "ok", "result": df}
        else:
            result = {"status": "ok", "result": -1}
        # print("result\n", result)
        return result

    def convert_total_to_per_capita(self, f, df_var):
        df_var = pd.DataFrame(df_var)
        # print("AAAAAAAAA\nf="+str(f)+"=", "\nself.df_entities_conversion_factor=\n", self.df_entities_conversion_factor, "\n" , df_var)
        try:
            df_ = df_var.merge(self.df_entities_conversion_factor, how='inner', left_index=True, right_index=True)
        except Exception as ex:
            print("Er 22-11", str(ex))
        df_["pc"] = df_[f]/df_["amount"].astype(float)
        # print(df_, "\n", "-"*10)
        df_ = df_.drop([f, 'amount'], axis=1)
        return df_

    def multiply_two_variables_f(self, f__, df_var1, f__2, df_var2):
        print("="*100)
        print("Need to check this fuction change using indexes.")
        df_var1 = df_var1.reset_index()
        df_var2 = df_var2.reset_index()
        # print("\n", df_var1, "\n", df_var2, "\n", "="*50)

        df_ = df_var1.merge(df_var2, how='inner', left_on=self.entity_name+'_dim', right_on=self.entity_name+'_dim')
        # print(df_)
        df_["m"] = df_[f__].astype(float)*df_[f__2].astype(float)
        df_ = df_.drop([f__, f__2], axis=1)
        # print(df_)
        df_ = df_.set_index(self.entity_name+'_dim')
        # print(df_)

        return df_

    def pre_process_data(self, dic):
        print("90033-133 pre_process_data: \n", dic, "\n", "="*50)
        year_ = str(dic["time_dim_value"])

        groups = self.model_measure_group.objects.filter(~Q(group_name__in=self.do_not_include_groups)).all()

        ll_groups = [self.dependent_group]
        for k in groups:
            group = k.group_name
            if group not in ll_groups and group not in self.do_not_include_groups:
                ll_groups.append(group)
        lll_groups = []  # this will have only the groups that do not have problems (have data)
        ll_dfs = {}  # includes all the df of all groups with data

        print("90-111-222-1 for k in groups\n", ll_groups)

        for k in ll_groups:
            # print("="*50, "\n", k, "\n", "="*50)
            try:
                # print("\n90-111-2-"+k, "\n", "-"*30)
                s = ""
                for v in dic["axes"]:
                    s += "'" + v + "',"
                s += "'" + self.value_column + "'"
                # print("s\n", s)
                # print(entity_list)
                # s_ = 'self.model_fact.objects.filter(measure_dim__measure_group_dim__group_name=k, time_dim_id=year_, ' + self.entity_name + '_dim_id__in=entity_list).filter(~Q(measure_dim__in=self.do_not_include_measures_objs)).all()'
                s_ = 'self.model_fact.objects.filter(measure_dim__measure_group_dim__group_name=k, time_dim_id=year_).filter(~Q(measure_dim__in=self.do_not_include_measures_objs)).all()'
                # print("s_= ", s_,"\ns= ",s)
                qs = eval(s_)
                # print(qs)

                s = "pd.DataFrame(list(qs.values(" + s + ")))"
                # print("eval(s)", s)
                df = eval(s)
                # print("1.  df =", k ,"\n", df)

                if df.shape[0] == 0:
                    continue
                lll_groups.append(k)
                try:
                    df = df.pivot(index=self.entity_name+"_dim", columns='measure_dim', values=self.value_column)
                except Exception as ex:
                    print(ex)
                # print("90-111-2-100\n", k ,"\n", df)

                ll_dfs[k] = df.apply(pd.to_numeric, errors='coerce').round(6)
                for f__ in ll_dfs[k]:
                    # print("90-111-3-"+k+" "+str(f__), self.total_variables, "\n", self.multiply_two_variables)
                    if f__ in self.total_variables:
                        ll_dfs[k][f__] = self.convert_total_to_per_capita(f__, ll_dfs[k][f__])
                    elif len(self.multiply_two_variables) > 0:
                        if f__ in self.multiply_two_variables["var1"]:
                            # print("two1 : ", f__, self.multiply_two_variables["var1"].index(f__))
                            idx__ = self.multiply_two_variables["var1"].index(f__)
                            f__var2 = self.multiply_two_variables["var2"][idx__]
                            f__var2_group = self.multiply_two_variables["var2_group"][idx__]
                            # print("two2: ", f__, f__var2, f__var2_group)
                            # print("two3: ", k, f__, f__var2, "\n")
                            # print("three3 : ")
                            ll_dfs[k][f__] = self.multiply_two_variables_f(f__, ll_dfs[k][f__], f__var2,
                                                                           ll_dfs[f__var2_group][f__var2])
                            # print("four4: ", k, f__, f__var2, "\n", ll_dfs[k][f__])

                # print("90-333-3-200\n", k ,"\n", ll_dfs[k])
            except Exception as ex:
                print("Error 50661-12 PotentialAlgo calculate_min_max_cuts: \n" + str(ex), "\n", "90-111-222-2-"+k+" "+f__)
        return ll_dfs

    # -------
    def normalize_similarity(self, dic):

        def normalize(n_dic):
            n_df = n_dic["df"]
            # print("A n_df\n", n_df)
            n_df = n_df.set_index('person_dim')
            # print("="*50, "\n", "="*50, "\n", "n_df\n", n_df)
            mm = n_dic["mm"]
            index = n_dic["index"]
            #
            df_mm_index = pd.DataFrame(data=[mm])
            # self.to_save_normalize.append((df_mm_index.copy(), "min_max_" + index_))
            #
            # print("\nindex", index)
            ii_ = index.split("_")
            oh = ii_[0]
            # oh_ = oh.split("-")
            ol = ii_[1]
            # ol_ = ol.split("-")
            ohi = ii_[2]
            ohi_ = ohi.split("-")
            hi__ = ohi_[1]
            oli = ii_[3]
            oli_ = oli.split("-")
            li__ = oli_[1]
            # print(oh_[0], oh_[1], ol_[0], ol_[1])
                # print( "="*20, "normalize:", ohi_[0], hi__, oli_[0], li__, "="*20, "\nmm", mm)
            # print("n_df", "\n", n_df, "\n")
            df_n1 = pd.DataFrame(index=n_df.index.copy())
            df_n2 = pd.DataFrame(index=n_df.index.copy())
            for xi in mm:
                if xi == "y":
                    mi = 1
                else:
                    mi = xi
                min_cut = mm[xi]["min_cut"]
                max_cut = mm[xi]["max_cut"]
                if min_cut == -1:
                    continue

                dff = pd.DataFrame(n_df.loc[:,mi].astype(float), index=n_df.index.copy())
                df_f = dff.copy()
                df_f = df_f.apply(lambda x: (x - min_cut) / (max_cut - min_cut))
                df_n1[mi] = df_f.copy()
                df_f[df_f < 0] = 0
                df_f[df_f > 1] = 1
                df_n2[mi] = df_f.copy()

            # self.to_save_normalize.append((df_n1.copy(), "n1_" + index))
            # self.to_save_normalize.append((df_n2.copy(), "n2_" + index))

            return df_n1.copy(), df_n2.copy()

        def similarity(index, n_df, dn):
            # print("="*50, "\nSIM_SIM for index = ", index, "\n", n_df)
            self.to_save_similarity.append((n_df.copy(), "n_" + index))
            df_d = pd.DataFrame()
            df_r = pd.DataFrame()
            df_ = pd.DataFrame()

            # print("n_df.columns\n", n_df.columns)

            for k in n_df.columns:
                if k == dn:
                    continue
                df_d[k] = abs(n_df[k] - n_df[dn])
                df_r[k] = abs(n_df[k] - (1 - n_df[dn]))
            # print("df_d\n", df_d, "\ndf_r\n", df_r)
            sd = df_d.sum()
            sr = df_r.sum()
            dfdm = df_d.mean()
            dfrm = df_r.mean()
            # print("df_d\n", df_d, "\n", "df_r\n", df_r)

            self.to_save_similarity.append((df_d.copy(), "df_d_" + index))
            self.to_save_similarity.append((df_r.copy(), "df_r_" + index))

            # print("sd=\n", sd, "\nsr=\n", sr, "\ndfdm=\n", dfdm, "\ndfrm=\n", dfrm)

            dfdm_ = []
            dfrm_ = []
            lls = []
            ll_dic = []
            # print(dfdm.index)

            for k in dfdm.index:
                # print("k\n", k)
                # print(dfdm[k])
                # print(dfrm[k])
                # print("---")

                dfdm_.append(dfdm[k])
                dfrm_.append(dfrm[k])

                if dfdm[k] < dfrm[k]:
                    # print("A")
                    df_[k] = df_d[k]
                    # print("AAA\n", 1 - dfdm[k])
                    lls.append(1 - dfdm[k])
                    ll_dic.append(1)
                else:
                    # print("B")
                    df_[k] = df_r[k]
                    # print("BBB\n", 1 - dfdm[k])
                    lls.append((1 - dfrm[k]))
                    ll_dic.append(-1)

            # print("df_\n", df_)

            self.to_save_similarity.append((df_.copy(), "df_" + index))
            # print("lls\n", lls, "\nll_dic\n", ll_dic)

            # print(dfdm_, "\n", dfrm_)
            # df_results = pd.DataFrame([dfdm_, dfrm_, lls, ll_dic], columns=df_.columns,
            #                           index=["d", "1-d", "similarity", "direction"])
            # print(index)
            is_=index.split("_")
            index__=""
            for k in is_:
                k_=str(int(float(k.split("-")[1])*100))
                if len(k_)==1:
                    k_="0"+k_
                index__ +=k_
            index = int(index__)
            df_results = pd.DataFrame([lls], columns=df_.columns, index=[index])

            # print("df_results\n", df_results)

            df_results_dic = pd.DataFrame([ll_dic], columns=df_.columns, index=[index])
            self.to_save_similarity.append((df_results.copy(), "df_results_" + str(index)))
            return df_results, df_results_dic, index

        # # #
        # print("90099-99-1000 BasePotentialAlgo normalize_similarity: \n", dic, "\n'", "="*100)
        log_debug("90099-99-1000 BasePotentialAlgo normalize_similarity:" + str(dic))

        method = dic["method"]
        dn_ = int(dic["dn"])
        cn = int(dic["cn"])  # 60406530
        log_debug("cn:" + str(cn))
        #
        first_high_group = 0.4
        first_low_group = 0.4
        step = 0.05
        n_step = int(first_high_group/step)
        #
        u_method = "median"
        l_method = "median"
        if method == "min":
            u_method = "min"
            l_method = "max"

        # ----
        model_temp = apps.get_model(app_label=self.app, model_name="temp")
        model_temp_var = apps.get_model(app_label=self.app, model_name="tempvar")
        # ----
        model_fact = apps.get_model(app_label=self.app, model_name="factnormalized")
        qs = model_fact.objects.filter(person_dim__person_group_dim__group_name="Model")
        df = pd.DataFrame(list(qs.values("gene_dim", "person_dim", "amount")))

        try:
            df = df.pivot(index="person_dim", columns='gene_dim', values='amount')
            df = df.sort_values(dn_, ascending=False)
            df = df.reset_index()
            # self.to_save_normalize.append((df.copy(), 'Data'))
        except Exception as ex:
            print(ex)

        # print(df)
        # print(df.columns)

        # print(df.head(56),"\n", df.tail(56))
        # print("'", "="*50)
        # print(df.columns[2:], len(df.columns[2:]))

        # print("'", "="*50)
        step_num = int(round(df.shape[0]*step))
        # print("step_num=", step_num)
        # print("'", "="*30)
        # print(range(int(first_high_group*100), 0, -int(step*100)))

        dic_hp = {}
        # int((first_low_group-step) * 100)

        c0_continue = 1
        c1_continue = 1
        c2_continue = 1
        c3_continue = 1
        c0=int(cn/1000000)
        c1=int(cn/10000)-c0*100
        c2=int(cn/100)-c0*10000-c1*100
        c3=cn-c0*1000000-c1*10000-c2*100
        # print(c0, c1, c2, c3)
        for l in range(int(first_low_group * 100), int(step*100), -int(step * 100)):
            if l != c1 and c1_continue == 1 and cn != 0:
                continue
            else:
                c1_continue = 0
            # print("l", l)
            # log_debug("l:" + str(l))
            l_ = l/100
            for h in range(int(first_high_group*100), int(step*100), -int(step*100)):
                if (100-h) != c0 and c0_continue == 1 and cn != 0:
                    continue
                else:
                    c0_continue = 0
                # print("h", 100-h)
                # log_debug("h:" + str(100-h))
                h_ = (100 - h)/100
                ## print("-"*30, "\n  h_=", h_, "l_=", l_,"\n","-"*30)

                ll = [h / 100, (100 - l) / 100]
                n = df.shape[0]
                h_cut = n*h/100
                l_cut = n*(100-l)/100

                # df_q = df.quantile(ll)
                # print("-" * 20)
                # print("df_q\n", df_q, "\n", df_q[["person_dim"]])
                # print("-" * 20)
                #
                # # print(df_q[["person_dim"]].iloc[0], "\n\n", df_q[["person_dim"]].iloc[1])
                # h_cut = (float(df_q[["person_dim"]].iloc[0]) - 1) * (df.shape[0] / (df.shape[0] - 1))
                # h_cut = int(round(h_cut))
                #
                # print(df.shape, "\nH cut index=", h_cut)
                # print("-" * 20)

                cond_h = df.index <= h_cut
                df_h_e = df[cond_h]
                    # print("Top records sorted by Y:\n", df_h_e.tail(56))
                    # # print(df_h_e.index)
                    # # print(len(df_h_e.index)-step_num-1)
                    #
                    #     # n_y = len(df_h_e.index) - step_num - 1
                    #     # print(n_y)
                    #     # y_max_cut = df_h_e.iloc[n_y][1]
                    #     # print("-" * 10, "     H", "     person_index=", df_h_e.iloc[n_y]["person_dim"], "     Y=",
                    #     #       y_max_cut)

                # l_cut = (float(df_q[["person_dim"]].iloc[1]) - 1) * (df.shape[0] / (df.shape[0] - 1))
                # l_cut = int(round(l_cut))

                cond_l = df.index > l_cut
                df_l_e = df[cond_l]
                    # print(df.shape, "\nL cut index=", l_cut)
                    # print("Low records sorted by Y:\n", df_l_e.head(56))
                    #     # print(df_l_e.index)
                    #
                    #     # y_min_cut = df_l_e.iloc[step_num][1]
                    #     # print("-" * 10, "     L", "     person_index=", df_l_e.iloc[step_num]["person_dim"], "     Y=",
                    #     #       y_min_cut)
                    #     # print(df_h_e.columns, len(df_h_e.columns))
                nn__ = 0
                nhi_ = 0
                for hi in range(h, int(step*100), -int(step*100)):
                    nhi_ += 1
                    hi_ = (100 - round(hi - step * 100)) / 100
                    # print(hi_, (100 - round(hi - step * 100)), c2, c2_continue, cn)
                    if (100 - round(hi - step * 100)) != c2 and c2_continue == 1 and cn != 0:
                        # print("QQQ")
                        continue
                    else:
                        c2_continue = 0
                        # print("PPPPP")

                    # print("hi", (100 - round(hi - step * 100)))
                    # print(len(df_h_e.index), step_num , (hi_ - h_)  ,step, step_num * ((hi_ - h_) / step))

                    # log_debug("hi:" + str(100 - round(hi - step * 100)))

                    n_y_h = int(len(df_h_e.index) - step_num*((hi_-h_)/step))

                    # print("h", h_, "hi", hi_, "n_y_h=", n_y_h)
                    # print(df_h_e)
                    # print(df_h_e.iloc[n_y_h])

                    y_max_cut = df_h_e.iloc[n_y_h][dn_]
                    # print("ZZZ\n", y_max_cut)

                    nli_ = 0
                    if cn == 0:
                        is_first_continue = 0
                    else:
                        is_first_continue = 1

                    # print("KKKKKK", range(l, int(step*100)), l)
                    for li in range(l, int(step*100), -int(step*100)):
                        if cn != 0:
                            if round(li - step*100) != c3 and c3_continue == 1:
                                # print("z")
                                continue
                            else:
                                c3_continue = 0
                                if is_first_continue == 1:
                                    is_first_continue = 0
                                    # print("W")
                                    continue

                            #  pull from database by cn
                            dic_hp = model_temp.objects.get(idx = cn).dic_hp
                            #

                        # print(2, hi, cn, li)
                        # print("li", round(li - step*100))

                        log_debug("In upload_file h=" + str(100 - h) + " l="+ str(l)
                                  +" hi="+str(100 - round(hi - step * 100))+" li="+str(round(li - step*100)))
                        nn__ += 1
                        nli_ += 1
                        li_ = round(li - step*100)/100
                        n_y_l = int(step_num*((l_-li_)/step))
                        # print("l", l_, "li",  li_, "n_y_l", n_y_l)
                        index_ = "h-"+str(h_) + "_" + "l-" + str(l_) + "_" + "hi-" + str(hi_) + "_" + "li-" + str(li_)
                        # print("-"*75, "\nindex", index_, "\n", "-"*40)
                        y_min_cut = df_l_e.iloc[n_y_l][dn_]
                            # print("  H", "  person_index=", df_h_e.iloc[n_y_h]["person_dim"], "  Y=",
                            #       y_max_cut, "           L", "  person_index=", df_l_e.iloc[n_y_l]["person_dim"], "  Y=",
                            #       y_min_cut)
                        dic_hp[index_] = {}
                        dic_hp[index_]["y"] = {"max_cut": float(y_max_cut), "min_cut": float(y_min_cut)}
                        # print(dic_hp)

                        # log_debug("In upload_file  1")

                        for gene_num in range(len(df_h_e.columns)-1, 1, -1):
                            # l
                            df_lx = df_l_e[[gene_num]].sort_values(gene_num, ascending=False)
                            df_lx = df_lx.reset_index()
                            # print(df_l_e[[gene_num]], "\n\n")
                            # print("df_lx\n", df_lx)
                            # print(df_lx)
                                # if nn__ < 10:
                                #     print("Internal Loop: Low range, variable(gene)=", gene_num, "\nsorted values:\n", df_lx)
                            # n_x = len(df_lx.index) - nli_ * step_num - 1
                            n_x = nli_ * step_num
                            # print("n_x=", n_x)
                            # print(df_lx.index)
                            lx = df_lx.iloc[n_x][gene_num]
                            li = pd.DataFrame(df_lx.iloc[n_x])
                            # print("Low person index=", li.columns[0], "value=", lx, "\n", "-"*30)
                            df_lx = df_lx.set_index('index')
                            # print(df_lx)
                            # h
                            df_hx = df_h_e[[gene_num]].sort_values(gene_num, ascending=False)
                            df_hx = df_hx.reset_index()
                            # print(df_h_e[[gene_num]], "\n\n")
                            # print(df_hx)
                                # if nn__ < 8:
                                #     print("Internal Loop: High range, variable(gene)=", gene_num, "\nsorted values:\n", df_hx)
                            n_x = len(df_hx.index) - nhi_ * step_num - 1
                            hx = df_hx.iloc[n_x][gene_num]
                            hi__ = pd.DataFrame(df_hx.iloc[n_x])
                            # print("High person index=", hi__.columns[0], "value=", hx, "\n", "-"*30)
                            df_hx = df_hx.set_index('index')

                            median_hx = float(df_hx.median())
                            median_lx = float(df_lx.median())
                            # print("Low median_lx", median_lx, " High median_hx", median_hx)

                            if median_lx > median_hx:
                                    # print("Convert Groups:\n AA max_cut", lx, "min_cut", hx)
                                n_x_l = len(df_lx.index) - nli_ * step_num - 1
                                lx_ = df_lx.iloc[n_x_l][gene_num]
                                n_x_h = nhi_ * step_num
                                hx_ = df_hx.iloc[n_x_h][gene_num]

                                lx = hx_
                                hx = lx_

                            if lx > hx:
                                    # print("BB ", "Max Cut=", hx, "Min Cut=", lx)
                                    # print("  >> MaxCut lower then MinCut. MinCut=MaxCut=-1 for this index:", index_)
                                dic_hp[index_][gene_num] = {"max_cut": -1, "min_cut": -1}
                                    # print(" CC max_cut", -1, "min_cut", -1)
                                # else:
                                #     dic_hp[index_][gene_num] = {"max_cut": float(lx), "min_cut": float(hx)}
                            else:
                                    # print("max_cut", hx, "min_cut", lx)
                                dic_hp[index_][gene_num] = {"max_cut": float(hx), "min_cut": float(lx)}

                        # log_debug("In upload_file  2")

                        ndic = {"df": df, "index": index_, "mm": dic_hp[index_]}
                        # print(ndic)
                        ndf_n1, ndf_n2 = normalize(ndic)
                        # print("ndf_n1\n", ndf_n1, "\nndf_n2\n", ndf_n2)
                        log_debug("In upload_file  3")
                        sim, ll_dic_, idx = similarity(index = index_, n_df = ndf_n2, dn=dn_)

                        dic_hp_obj, is_created = model_temp.objects.get_or_create(idx=idx)
                        dic_hp_obj.idx = idx
                        dic_hp_obj.dic_hp = dic_hp
                        dic_hp_obj.save()
                        log_debug("Saved = " + str(idx) )
                        for j in sim.columns:
                            temp_obj, is_created = model_temp_var.objects.get_or_create(temp=dic_hp_obj, var=j)
                            temp_obj.amount=float(sim.iloc[0][j])
                            temp_obj.sign=ll_dic_.iloc[0][j]
                            temp_obj.save()

                # print(dic_hp)
                # For normalization

                # self.save_to_excel_(save_to_file = os.path.join(self.TO_EXCEL_OUTPUT, str(h) + "_" + str(l) + "_normalization.xlsx"),
                #                     to_save = self.to_save_normalize)
                #
                # For similarity
                # self.save_to_excel_(save_to_file = os.path.join(self.TO_EXCEL_OUTPUT, str(h) + "_" + str(l) + "_similarity.xlsx"),
                #                     to_save = self.to_save_similarity)
                #
        print("A2000")
        result = {"status": "ok"}
        log_debug("Done normalize_similarity")
        return result

    def calculate_min_max_cuts(self, dic):
        # # #
        # print("90099-99-99 MMDataProcessing calculate_min_max_cuts: \n", dic, "\n'", "="*100)
        method = dic["method"]
        variables_model = dic["variables_model"]
        variables_model = apps.get_model(app_label=self.app, model_name=variables_model)
        model_temp_var = apps.get_model(app_label=self.app, model_name="tempvar")
        qs = model_temp_var.objects.all()
        #
        df = pd.DataFrame(list(qs.values("temp_id", "var", "amount")))
        df_similarity_ = df.pivot(index="temp_id", columns='var', values='amount').fillna(0)
        #
        df = pd.DataFrame(list(qs.values("temp_id", "var", "sign")))
        df_similarity_dic = df.pivot(index="temp_id", columns='var', values='sign').fillna(0)
        df_similarity_dic.columns = df_similarity_.columns
        #
        # print(df_similarity_, "\n\n", df_similarity_dic)
        schema = {'gene': 'int64', 'threshold': 'float64', 'score': 'float64'}
        df_scores_by_genes = pd.DataFrame(columns=schema).astype(schema)
        for n in range(0, len(df_similarity_.columns)):
            # print("gene=", df_similarity_.columns[n])
            df_similarity_[df_similarity_.columns[n]] = df_similarity_[df_similarity_.columns[n]].astype('float') * df_similarity_dic[df_similarity_.columns[n]].astype('float')
            df_similarity_[df_similarity_.columns[n]] -= 0.7
            df_ = df_similarity_.sort_values(df_similarity_.columns[n], ascending=False)
            df = df_.copy()
            df=df.reset_index(drop=True)
            df.index = df.index.set_names(['idx'])
            df=df.reset_index()
            # print("AAAA", df.columns[n+1], "\n", df)
            lg = [0, 0.05, 0.1, 0.15, 0.2]
            lp = [1, 2, 3, 4, 5]
            sum_t = 0
            for i in range(len(lg)):
                z = lg[i]
                try:
                    sum_ = round(100*lp[i]*df_similarity_[df_similarity_[df_similarity_.columns[n]] >= z][df_similarity_.columns[n]].sum())/100
                    # print("gene=", df_similarity_.columns[n], "threshold=", z, "       score=", sum_)

                    df_n = {'gene': int(df_similarity_.columns[n]), 'threshold': z, 'score': sum_}
                    df_scores_by_genes = df_scores_by_genes.append(df_n, ignore_index=True)

                    sum_t += sum_
                except Exception as ex:
                    print(ex)
        # print("\nGene Score by threshold:\n", df_scores_by_genes)
        df_total_scores = pd.DataFrame(df_scores_by_genes.groupby("gene")["score"].sum())
        for index, row in df_total_scores.iterrows():
            id_ = int(round(index))
            obj = variables_model.objects.get(id=id_)
            obj.score = row["score"]
            obj.save()
        log_debug("Done calculate_min_max_cuts")
        result = {"status": "ok"}
        return result
    # -------

    # def calculate_min_max_cuts(self, dic):
    #     # print("90066-100 PotentialAlgo calculate_min_max_cuts: \n", dic, "\n", "="*50)
    #     f = int(dic["dim_field"])
    #     method = dic["method"]
    #     u_method = "median"
    #     l_method = "median"
    #     if method == "min":
    #         u_method = "min"
    #         l_method = "max"
    #
    #     # print("90066-100-1 PotentialAlgo calculate_min_max_cuts: \n", "="*50)
    #     # print(self.dependent_group)
    #     year_ = str(dic["time_dim_value"])
    #     # print(year_)
    #     min_max_model_name_ = dic["min_max_model"]
    #     #
    #     # print("90066-100-2 PotentialAlgo calculate_min_max_cuts: \n", "="*50)
    #     ll_dfs = self.pre_process_data(dic)
    #     # print(ll_dfs)
    #
    #     # result = {"status": "ok"}
    #     # return result
    #     #
    #     # sign_ = pd.DataFrame([[0, 0, 0, 0]])
    #     # sign_.columns = self.options
    #     # similarity_ = pd.DataFrame([[0, 0, 0, 0]])
    #     # similarity_.columns = self.options
    #
    #     high_group_cut = [0.4, 0.35, 0.30, 0.25, 0.2, 0.15, 0.1]
    #     low_group_cut = [0.4, 0.35, 0.30, 0.25, 0.2, 0.15, 0.1]
    #     step_ = 10
    #     df_d = ll_dfs[self.dependent_group]
    #     # print(df_d)
    #     results = {}
    #     best_cut = {"bv": -1}
    #
    #     # print("90066-100-5 PotentialAlgo calculate_min_max_cuts: \n", "="*50)
    #
    #     for h in high_group_cut:
    #         hh = 1- h
    #         results[hh] = {}
    #         #
    #         # z = 5
    #         # steps_h = list(range(z, 100, z))
    #         # if h == high_group_cut[0]:
    #         #     steps_h.append(95)
    #         # print(h)
    #         steps_h = list(range(5, round(h * 100), 5))
    #         steps_h = [a/100 for a in steps_h]
    #         print("hhhhhhh", h, steps_h)
    #         #
    #         for ll in low_group_cut:
    #             results[hh][ll] = {}
    #             # z_ = 5
    #             # steps_l = list(range(z_, 100, z_))
    #             # if ll == low_group_cut[0]:
    #             #     steps_l.append(95)
    #             steps_l = list(range(5, round(ll * 100), 5))
    #             steps_l = [a/100 for a in steps_l]
    #             print("llllll", ll, steps_l)
    #             #
    #             # print(df_d)
    #             df_q = df_d.quantile([hh, ll])
    #             # print("df_d.shape", df_d.shape)
    #             # print("df_q", df_q)
    #
    #             results[hh][ll][f] = {}
    #             # results[hh][ll][f]["max_cut"] = df_q[f].iloc[0]
    #             # results[hh][ll][f]["min_cut"] = df_q[f].iloc[1]
    #             df_hi = df_d[df_d[f] >= df_q[f].iloc[0]].index
    #             df_li = df_d[df_d[f] <= df_q[f].iloc[1]].index
    #             # print("df", "\n", df_d[df_d[f] >= df_q[f].iloc[0]], "\n\n")
    #             # print("df", "\n", df_d[df_d[f] >= df_q[f].iloc[0]].index, "\n\n")
    #
    #             # print("df", "\n", float(df_d[df_d[f] >= df_q[f].iloc[0]][f].median()), "\n")
    #
    #             # results[hh][ll][f]["max_cut"] = float(df_d[df_d[f] >= df_q[f].iloc[0]][f].median())
    #             # results[hh][ll][f]["min_cut"] = float(df_d[df_d[f] <= df_q[f].iloc[1]][f].median())
    #
    #             for steph in steps_h:
    #                 for stepl in steps_l:
    #                     for g in ll_dfs:
    #                         # if g != self.dependent_group:
    #                         df = ll_dfs[g]
    #                         # print("\n",g, "\n",df.columns)
    #                         iih = []
    #                         iil = []
    #                         for j in df.index:
    #                             if j in df_hi:
    #                                 iih.append(j)
    #                             if j in df_li:
    #                                 iil.append(j)
    #                         dfh = df.loc[iih]
    #                         dfl = df.loc[iil]
    #
    #                         dfh_q = dfh.quantile(steph/h)
    #                         dfl_q = dfl.quantile((ll-stepl)/ll)
    #                         # print('dfh, dfh', 'steph ', steph, 'dfh_q ', dfh_q)
    #                         # , '\n', 'dfl', dfl, 'stepl', stepl, 'dfl_q', dfl_q)
    #
    #                         if steph not in results[hh][ll][f]:
    #                             results[hh][ll][f][steph] = {}
    #                         if stepl not in results[hh][ll][f][steph]:
    #                             results[hh][ll][f][steph][stepl] = {}
    #                             results[hh][ll][f][steph][stepl]["groups"] = {}
    #                         if g not in results[hh][ll][f][steph][stepl]["groups"]:
    #                             results[hh][ll][f][steph][stepl]["groups"][g] = {}
    #
    #                         for k in df.columns:
    #                             # print(steph, stepl, k, dfh_q[k], dfl_q[k])
    #                             dfh_ik = dfh[dfh[k] >= dfh_q[k]][k]
    #                             dfl_ik = dfl[dfl[k] <= dfl_q[k]][k]
    #                             # print('dfh[dfh[k] >= dfh_q[k]]', dfh[dfh[k] >= dfh_q[k]].index, 'dfh_ik.median()', dfh_ik.median(), 'dfl_ik.median()', dfl_ik.median())
    #                             # print('g, k, dfh_ik.shape', g, k, dfh_ik.shape)
    #                             if dfh_ik.shape == 0 or dfl_ik.shape == 0:
    #                                 continue
    #                             results[hh][ll][f][steph][stepl]["groups"][g][k] = eval('{"min_cut": dfl_ik.'+l_method+'(), "max_cut": dfh_ik.'+u_method+'()}')
    #                     # print(results)
    #                     dic = {"f":f, "steph":steph, "stepl":stepl,
    #                            "df_d": df_d, "ll_dfs": ll_dfs, "groups": results[hh][ll][f]}
    #                     bv = self.get_similarity(dic)
    #                     if bv > best_cut["bv"]:
    #                         best_cut = {"bv": bv, "hh": hh, "ll": ll, "steph":steph, "stepl":stepl, "f":f,
    #                                     "dependent_group": self.dependent_group, "year": year_,
    #                                     "f_groups":results[hh][ll][f][steph][stepl]["groups"]}
    #
    #                         # print('best_cut\n', 'hh', hh, 'll', ll, 'results 1 steph', steph, 'stepl', stepl, "dv", bv, "\n",
    #                         #       best_cut)
    #                     print('hh', hh, 'll', ll, 'steph', round(100*(steph+hh))/100, 'stepl', round(100*(ll-stepl))/100, 'dv', bv, 'best', best_cut['bv'])
    #
    #     # print("="*100)
    #     best_cut["df_d"] = df_d
    #     # print("final", best_cut, "="*100)
    #     #
    #     file_path = os.path.join(self.PICKLE_PATH, "result_"+str(f)+"_"+year_+".pkl")
    #     print(file_path)
    #     with open(file_path, 'wb') as handle:
    #         pickle.dump(results, handle, protocol=pickle.HIGHEST_PROTOCOL)
    #
    #     file_path = os.path.join(self.PICKLE_PATH, "best_cut_"+str(f)+"_"+year_+".pkl")
    #     print(file_path)
    #     with open(file_path, 'wb') as handle:
    #         pickle.dump(best_cut, handle, protocol=pickle.HIGHEST_PROTOCOL)
    #
    #     # # #
    #     print("=" * 100)
    #     #
    #
    #     result = {"status": "ok"}
    #     return result

    def create_total_pop(self, dic):
        print("900443-155 PotentialAlgo create_total_pop: \n", dic, "\n", "="*50)
        year= dic["year"]
        group = dic["group"]
        variable = dic["variable"]
        # print(self.entity_name)
        model_name_ = dic["dimensions"]["time_dim"]["model"]
        model_time_dim = apps.get_model(app_label=self.app, model_name=model_name_)

        model_name_ = dic["dimensions"][self.entity_name+"_dim"]["model"]
        model_entity_dim = apps.get_model(app_label=self.app, model_name=model_name_)

        model_name_ = "measuregroupdim"
        model_group_measure_dim = apps.get_model(app_label=self.self.app, model_name=model_name_)

        model_name_ = dic["dimensions"]["measure_dim"]["model"]
        model_measure_dim = apps.get_model(app_label=self.app, model_name=model_name_)

        model_name_ = dic["fact_model"]
        model_fact = apps.get_model(app_label=self.app, model_name=model_name_)

        t = model_time_dim.objects.get(id=year)
        g, is_created = model_group_measure_dim.objects.get_or_create(group_name=group)
        m, is_created = model_measure_dim.objects.get_or_create(measure_group_dim=g, measure_name=variable)

        qs = model_entity_dim.objects.all()
        for e in qs:
            # print(e)
            s = 'model_fact.objects.get_or_create(time_dim=t, ' + self.entity_name + '_dim=e, measure_dim=m)'
            # print(s)
            ef, is_created = eval(s)
            ef.amount=1
            ef.save()

        result = {"status": "ok"}
        return result

    def get_similarity(self, dic):
        # print("s="*50)
        f = dic["f"]
        df_d = dic["df_d"]
        steph = dic["steph"]
        stepl = dic["stepl"]
        ll_dfs = dic["ll_dfs"]
        f_groups = dic["groups"]
        groups = f_groups[steph][stepl]["groups"]
        dff = pd.DataFrame(df_d.loc[:, f].astype(float))
        df_f = dff.copy()

        max_cut = groups[self.dependent_group][f]["max_cut"]
        min_cut = groups[self.dependent_group][f]["min_cut"]

        # max_cut = f_groups["max_cut"]
        # min_cut = f_groups["min_cut"]

        df_f = df_f.apply(lambda x: (x - min_cut) / (max_cut - min_cut))
        df_f[df_f < 0] = 0
        df_f[df_f > 1] = 1
        dffn = df_f.copy()
        llg_ = []
        for g in groups:
            dfg = ll_dfs[g].copy()
            lls = []
            if len(groups[g]) == 0:
                return 0
            for f_g in groups[g]:
                max_cut = groups[g][f_g]["max_cut"]
                min_cut = groups[g][f_g]["min_cut"]
                dfg_f = dfg[f_g]
                dfg_f = dfg_f.apply(lambda x: (x - min_cut) / (max_cut - min_cut))
                dfg_f[dfg_f < 0] = 0
                dfg_f[dfg_f > 1] = 1
                dfm = pd.merge(left=dffn, how='outer', right=dfg_f, left_index=True, right_index=True)
                dfm['diffd'] = dfm.apply(lambda x: abs(x[dfm.columns[0]] - x[dfm.columns[1]]), axis=1)
                dfm['diffr'] = dfm.apply(lambda x: abs(x[dfm.columns[0]] - (1-x[dfm.columns[1]])), axis=1)
                s_d = dfm['diffd'].mean()
                s_r = dfm['diffr'].mean()
                if s_d < s_r:
                    s = (1 - s_d)-0.7
                else:
                    s = (1 - s_r)-0.7
                if s < 0:
                    s = 0
                lls.append(s)
            llg_.append(mean(lls))
        return mean(llg_)

    def update_min_max_cuts(self, dic):
        print("90055-156 PotentialAlgo update_min_max_cuts: \n", dic, "\n", "="*50)
        year_ = str(dic["time_dim_value"])
        f = str(dic["var_obj"])
        min_max_model_name_ = dic["min_max_model"]
        model_min_max = apps.get_model(app_label=self.app, model_name=min_max_model_name_)
        # model_measure_dim = apps.get_model(app_label=self.app, model_name="measuredim")
        file_path = os.path.join(self.PICKLE_PATH, "best_cut_"+f+"_"+year_+".pkl")
        print(file_path)
        with open(file_path, 'rb') as handle:
            best_cut = pickle.load(handle)
        # hh = best_cut["hh"]
        # ll = best_cut["ll"]
        # steph = best_cut["steph"]
        # stepl = best_cut["stepl"]
        file_path = os.path.join(self.PICKLE_PATH, "result_"+f+"_"+year_+".pkl")
        # print(file_path)
        with open(file_path, 'rb') as handle:
            results = pickle.load(handle)
        # print(results[hh][ll][int(f)])

        model_min_max.objects.filter(time_dim_id=year_).delete()

        # print("1111 best_cut\n", best_cut)
        f_groups = best_cut['f_groups']
        for g in f_groups:
            gfs = f_groups[g]
            # print("-"*60, "\n", g, "\n", gfs)
            for f in gfs:
                # print("-"*10, "\n", f, "\n", gfs[f])
                obj = model_min_max.objects.create(time_dim_id=year_, measure_dim_id=f,
                                                   min=round(100*gfs[f]["min_cut"])/100,
                                                   max=round(100*gfs[f]["max_cut"])/100)
        result = {"status": "ok"}
        return result

    def create_similarity(self, group_d, group, df_n1_all, df_n2_all, df_n1_, df_n2_, sign_n1, sign_n2,
                          similarity_n1, similarity_n2):

        # print("1 create_similarity df_n2_all\n", group, "\n", df_n2_all, df_n2_all.shape)
        # print("1 group", group, "\n", df_n1_)

        df_n1_all = pd.merge(left=df_n1_all, how='outer', right=df_n1_, left_index=True, right_index=True)
        df_n2_all = pd.merge(left=df_n2_all, how='outer', right=df_n2_, left_index=True, right_index=True)

        # print("2 group", group, "\n", df_n1_all)
        # print("2 create_similarity df_n2_all\n", group, "\n", df_n2_all, df_n2_all.shape)

        # print(df_n2_all.head(100))
        for n in ["1", "2"]:
            # print("normalization=", n)
            ll = []
            lls = []
            for k in self.options:
                # print(k[0], k[1])
                s_ = "abs(df_n" + n + "_all['" + k[0] + "-" + group_d + "'] - df_n" + n + "_all['" + k[1] + "-" + group + "'])"
                # print("direct:", s_)
                df_d = eval(s_)
                df_d_na=df_d.dropna()
                s_ = title='sim-n' + n + group + k
                df_d_na_ = self.add_entity_to_df1(df_d_na)
                self.add_to_save_all(title=s_, a=df_d_na_, cols=-1)
                s_d = df_d.sum()
                # print("s_d", s_d, df_d.mean())
                s_ = "abs(df_n" + n + "_all['" + k[0] + "-" + group_d + "'] - 1 + df_n" + n + "_all['" + k[1] + "-" + group + "'])"
                # print("revers:", s_)
                df_r = eval(s_)
                s_r = df_r.sum()
                # print("s_r", s_r, df_r.mean())
                dfdm = df_d.mean()
                dfrm = df_r.mean()
                # print("n=", n, "k=", k, "direct mean:", dfdm, "revers mean", dfrm)

                if dfdm < dfrm:
                    d_ = s_d
                    s_ = "df_n" + n + "_all['" + group_d + '-' + group + '-' + k + "'] = df_d"
                    exec(s_)
                    lls.append(1 - df_d.mean())
                    ll.append(1)
                else:
                    d_ = s_r
                    s_ = "df_n" + n + "_all['" + group_d + '-' + group + '-' + k + "'] = df_r"
                    exec(s_)
                    lls.append(1 - df_r.mean())
                    ll.append(-1)
                # print("="*50)
            # print("sign_n" + n + ".loc[group] = ll", ll)
            exec("sign_n" + n + ".loc[group] = ll")
            # print("similarity_n" + n + ".loc[group] = lls", lls)
            exec("similarity_n" + n + ".loc[group] = lls")
        return group_d, group, df_n1_all, df_n2_all, df_n1_, df_n2_, sign_n1, sign_n2, similarity_n1, similarity_n2

    def add_entity_to_df(self, df, cols=None):
        df_ = df.copy()
        # print("add_entity_to_df 1\n", df_)
        if cols is None:
            cols = [self.entity_name+'_name']
            df_c = df.columns
            for j in df_c:
                k = str(self.measures_name[self.measures_name["id"] == j]["measure_name"]).split("    ")[1].split("\n")[
                    0]
                cols.append(k)
        df_ = df_.reset_index()
        try:
            df_ = df_.merge(self.entities_name, how='inner', left_on=self.entity_name+'_dim', right_on='id')
            # print("add_entity_to_df 11\n", df_)

            df_ = df_.drop([self.entity_name+'_dim', 'id'], axis=1)
            # print("add_entity_to_df 12\n", df_)
        except Exception as ex:
            print("Error 90-90-88-1 add_entity_to_df: ", str(ex))
        try:
            c_ = df_.pop(self.entity_name+'_name')
            df_.insert(0, self.entity_name+'_name', c_)
            # print("CC\n", "\n", df_)
            # print("CC1\n", df_.columns, "\n", cols)
        except Exception as ex:
            print("Error 90-90-88-2 add_entity_to_df: ", str(ex))

        if isinstance(cols, pd.core.indexes.base.Index) or cols != -1:
            df_.columns = cols

        # print("add_entity_to_df 2\n", df_)
        return df_

    def add_entity_to_df1(self, df, cols=None):
        df_ = df.copy()
        df_ = df_.reset_index()
        df_.columns=["id", "amount"]
        # print(1000, "\n", "-"*10, "\nAA\n", "\n", df_, self.entities_name)
        df_ = self.entities_name.merge(df_, how='inner', left_on='id', right_on='id')
            # .drop(
            # [self.entity_name+'_dim', 'id'], axis=1)
        df_ = df_.sort_values(self.entity_name+'_name', ascending=True)
        # print("BB3\n", "\n", df_)
        return df_

    # It seems that I can delete this function
    def save_to_excel(self, df, folder):
        df2 = df.copy()
        # print(self.save_to_file + ' -Before sleep save_to_excel- ' + folder)
        total, used, free = shutil.disk_usage("/")
        # print(' total: ' + str(total) + ' used: ' + str(used) + ' free: ' + str(free))
        nnn = 0
        try:
            with pd.ExcelWriter(self.save_to_file, engine='openpyxl', mode='a') as writer_:
                df2.to_excel(writer_, sheet_name=folder)
                writer_.save()
                time.sleep(5)
            if self.second_time_save != '':
                print("save ok:", self.second_time_save)
            self.second_time_save = ''
            nnn = 1
        except Exception as ee:
            print(ee)
            time.sleep(5)
            self.save_to_excel(df2, folder)
            self.second_time_save = self.save_to_file
            nnn = 1
        finally:
            if nnn == 0:
                print(self.save_to_file + ' 55 finally -' + str(nnn) + ' - ' + folder)
                time.sleep(5)
                print(self.save_to_file + ' 551 finally -' + str(nnn) + ' - ' + folder)
                self.save_to_excel(df2, folder)
                self.second_time_save = self.save_to_file

    def save_to_excel_(self, save_to_file = None, to_save = None):
        wb2 = Workbook()
        if save_to_file is None:
            save_to_file = self.save_to_file
        if to_save is None:
            to_save = self.to_save

        wb2.save(save_to_file)
        wb2.close()
        wb2 = None
        log_debug("save_to_excel_ 11")
        # print("save_to_excel_\n", save_to_file)
        with pd.ExcelWriter(save_to_file, engine='openpyxl', mode="a") as writer_o:
            for d in to_save:
                try:
                    # print("d[0]\n", d[0])
                    # print("d[1]\n", d[1])
                    d[0].to_excel(writer_o, sheet_name=d[1])
                except Exception as ex:
                    print("9006-3 " + str(ex))
                    log_debug("9006-3 " + str(ex))
            writer_o.save()
        wb = load_workbook(filename=save_to_file, read_only=False)
        del wb['Sheet']
        log_debug("save_to_excel_ 16")
        try:
            wb.save(save_to_file)
        except Exception as ex:
            log_debug(str(ex))
        log_debug("save_to_excel_ 17")
        wb.close()
        log_debug("save_to_excel_ 18")

    def save_to_excel_all_(self, year):
        save_to_file_all = os.path.join(self.TO_EXCEL_OUTPUT, "all_" + str(year) + ".xlsx")
        # print("save_to_excel_all_", save_to_file_all)
        wb2 = Workbook()
        wb2.save(save_to_file_all)
        wb2.close()

        with pd.ExcelWriter(save_to_file_all, engine='openpyxl', mode="a") as writer:
            for d in self.to_save_all:
                try:
                    d[0].to_excel(writer, sheet_name=d[1])
                except Exception as ex:
                    print("9006-3 " + d[2] + str(ex))
            writer.save()

        # wb = load_workbook(filename=save_to_file_all, read_only=False)
        # del wb['Sheet']
        # wb.save(self.save_to_file)
        # wb.close()

    def add_to_save(self, title, a, cols):
        ai = pd.DataFrame(a)
        ai.index = self.df_index
        df_ = self.add_entity_to_df(ai, cols=cols)
        df_.dropna(how='all', axis=1, inplace=True)
        # print("A\n", df_)
        df_ = df_.sort_values(self.entity_name+'_name', ascending=True)
        # print("B\n", df_)
        self.to_save.append((df_.copy(), title))

    def add_to_save_all(self, title, a, cols):
        ai = pd.DataFrame(a)
        ai.dropna(how='all', axis=1, inplace=True)
        self.to_save_all.append((ai.copy(), title))

    def clean_rows(self, a, j, side="L"):
        # print("j", j)
        # a.sort(axis=1)
        # if j == 1:
        # self.add_to_save(title='Sort '+side+"-"+str(j), a=a, cols=-1)
        #
        a_ = pd.DataFrame(a.copy())
        len_b = a_.shape[0]
        a_ = a_.loc[a_.apply(lambda x: x.count(), axis=1) > 4]
        len_a = a_.shape[0]
        a_ = a_.to_numpy()
        # print(a.shape, "\n", a_.shape)
        if a_.shape[0] == 0:
            return a
        d = a[:, j:1 + j] - a[:, 0:1]
        d_ = a_[:, j:1 + j] - a_[:, 0:1]
        d_m = np.nanmean(d_, axis=0)
        # print("side=", side, " j=", j, " # of countries=", len_b, " # of countries in calculations=", len_a, d_m[0], " threshold=", self.rule_0)
        if d_m[0] < self.rule_0:
            #
            df_d = pd.DataFrame(d)
            df_d.index = self.df_index
            df_ = self.add_entity_to_df(df_d, cols=-1)
            self.to_save.append((df_.copy().sort_values(self.entity_name+'_name', ascending=True), 'D_' + side + ' - ' + str(j)))
            # print("d=\n", d)
            #
            b = pd.DataFrame(a.copy())
            b.loc[b.apply(lambda x: x.count(), axis=1) > 4, [j]] = np.nan
            #
            b.index = self.df_index
            df_ = self.add_entity_to_df(b, cols=-1)
            self.to_save.append((df_.copy().sort_values(self.entity_name+'_name', ascending=True), side + ' b ' + str(j)))
            #
            b = b.to_numpy()
            if j + 1 < b.shape[1]:
                b = self.clean_rows(b, j + 1, side)
        else:
            b = a.copy()
        return b

    def move_elements_to_right(self, row):
        n_nna = row.count()
        n_na = row.isna().sum().sum()
        n = n_na + n_nna
        # print("line 1 row=", n, n_na, n_nna)
        # print(row)
        row_c = row.copy()
        for j in range(n-1):
            try:
                if str(row_c[n-j-1:n-j].iloc[0]) == "nan":
                    for z in range(0, n-1-j):
                        if str(row_c[n-j-2-z:n-j-1-z].iloc[0]) != "nan":
                            row_c[n-j-1:n-j].iloc[0] = row_c[n-j-2-z:n-j-1-z].iloc[0]
                            row_c[n-j-2-z:n-j-1-z].iloc[0] = np.nan
                            break
            except Exception as ex:
                print("ex1: "+str(ex))
        # print("-"*100)
        # print(row)
        # print(row_c)
        # print("-"*100)
        return row_c

    def revers_elements_in_row(self, row):
        n_nna = row.count()
        n_na = row.isna().sum().sum()
        n = n_na + n_nna
        row_ = row.copy()
        for j in range(n):
            row_[j:j+1] = row[n-j-1:n-j]
        return row_

    def twenty_rule(self, row):
        # print("-1"*10, "\nrow:\n", row, "-2"*10, "\n")
        n_row = row.count()
        # print("n_row= ", n_row)
        if n_row < 5:
            n = 0
        elif n_row == 5:
            n = 1
        else:
            n = math.ceil(n_row * 0.2)
        min = n_row
        row_best = row
        if n == 0:
            return row_best

        # print("-4"*10)
        # print("n= ", n, "\n range(n+1)=", range(n+1))
        # print("-5"*10)
        for j in range(n + 1):
            row_c = row.copy()
            row_c[:j] = np.nan
            # print("-6"*10)
            # print(row_c)
            # print("-7"*10)
            row_c[n_row - (n - j):] = np.nan
            # print(row_c)
            # print("-8"*10)
            # print("max=", row_c.max(), "min=", row_c.min())
            # print("-9"*10)
            if row_c.max() - row_c.min() < min:
                min = row_c.max() - row_c.min()
                row_best = row_c.copy()
        return row_best

    def thirty_rule(self, row):
        n_row = row.count()
        row_best = row
        # print(abs(row.max() - row.min()), self.rule_2)
        if (n_row > 4) and abs(row.max() - row.min()) >= self.rule_2:
            n = 1
            min = n_row
            for j in range(n + 1):
                row_c = row.copy()
                row_c[:j] = np.nan
                row_c[n_row - (n - j):] = np.nan
                if (row_c.max() - row_c.min()) < min:
                    min = row_c.max() - row_c.min()
                    row_best = row_c.copy()
        if abs(row_best.max() - row_best.min()) >= self.rule_2:
            row_best[:] = np.nan
        return row_best

    def min_max_rule(self, row):
        row_ = row[0:2].copy()
        row_[:] = np.nan
        row_[0] = row.min()
        row_[1] = row.max()
        return row_

